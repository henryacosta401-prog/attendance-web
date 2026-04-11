import calendar
import os
import secrets
import shutil
import sqlite3
import textwrap
from io import BytesIO
from datetime import datetime, timedelta
from functools import wraps
from urllib.parse import quote

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, g, send_from_directory, jsonify, Response, abort,
    has_app_context
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from attendance_core.config import (
    ADMIN_ALERT_SCAN_TTL_SECONDS,
    ADMIN_PERMISSION_CODES,
    ADMIN_PERMISSION_LABELS,
    ADMIN_PERMISSION_OPTIONS,
    ADMIN_ROLE_PRESET_OPTIONS,
    ADMIN_ROLE_PRESETS,
    ADMIN_STATUS_CACHE_TTL_SECONDS,
    ALLOWED_EXTENSIONS,
    APP_TIMEZONE,
    ATTENDANCE_REQUEST_TYPES,
    BACKUP_FOLDER,
    BASE_DIR,
    BREAK_LIMIT_MINUTES,
    DATABASE_URL,
    DEFAULT_BACKUP_FOLDER,
    DEFAULT_BREAK_WINDOW_END,
    DEFAULT_BREAK_WINDOW_START,
    DEFAULT_INCIDENT_ERROR_TYPES,
    DEFAULT_PAID_LEAVE_DAYS,
    DEFAULT_SCHEDULE_DAYS,
    DEFAULT_SECRET_KEY,
    DEFAULT_SHIFT_END,
    DEFAULT_SHIFT_START,
    DEFAULT_SICK_LEAVE_DAYS,
    DEFAULT_SQLITE_DATABASE,
    DEFAULT_UPLOAD_FOLDER,
    DISCIPLINARY_ACTION_TYPES,
    DOCUMENT_EXTENSIONS,
    GOOGLE_CREDENTIALS_FILE,
    GOOGLE_SHEET_NAME,
    GOOGLE_SHEET_TAB,
    IMAGE_EXTENSIONS,
    INCIDENT_ACTION_STATUSES,
    INCIDENT_DISCIPLINARY_POLICY,
    LATE_GRACE_MINUTES,
    LEAVE_REQUEST_TYPES,
    LOGIN_MAX_ATTEMPTS,
    LOGIN_WINDOW_MINUTES,
    OPTION_CACHE_TTL_SECONDS,
    PERSISTENT_DISK_PATH,
    REPORT_CACHE_TTL_SECONDS,
    SCHEDULE_CHANGE_APPLY_TTL_SECONDS,
    SCHEDULE_SPECIAL_RULE_LABELS,
    SCHEDULE_SPECIAL_RULE_OPTIONS,
    SQLITE_DATABASE,
    UPLOAD_FOLDER,
    WEEKDAY_OPTIONS,
    get_configured_secret_key,
    is_production_environment,
)
from attendance_core.attendance import (
    combine_work_date_and_time,
    extract_clock_time,
    format_datetime_12h,
    format_time_12h,
    get_attendance_reference_datetime,
    get_overtime_reference_datetime,
    get_overbreak_minutes,
    get_schedule_code_for_date,
    get_schedule_day_codes,
    get_schedule_summary,
    get_shift_bounds_for_work_date,
    is_overbreak,
    normalize_history_reference,
    normalize_optional_clock_time,
    normalize_schedule_days,
    parse_break_limit_minutes,
    parse_datetime_local_input,
    parse_db_datetime,
    parse_optional_schedule_time,
    parse_shift_end,
    parse_shift_start,
    total_work_minutes,
)
from attendance_core.admin_access import (
    ADMIN_ENDPOINT_PERMISSIONS,
    admin_has_permission,
    describe_admin_permissions,
    get_admin_home_endpoint,
    get_admin_permission_codes,
    get_admin_role_preset_meta,
    get_home_endpoint_for_role,
    get_home_endpoint_for_user,
    infer_admin_role_preset,
    normalize_admin_permissions,
    normalize_admin_role_preset,
    row_get,
    sync_admin_role_preset,
)
from attendance_core.date_ranges import (
    get_admin_report_period_dates,
    get_payroll_period_dates,
    normalize_admin_report_filters,
    parse_iso_date,
    payroll_date_text,
)
from attendance_core.formatters import (
    format_currency,
    minutes_to_decimal_hours,
    minutes_to_hm,
)
from attendance_core.payroll import (
    build_employee_payslip_pdf_bytes,
    build_employee_payslip_pdf_filename,
    format_payroll_period_label,
    get_payroll_scope_label,
    recurring_rule_applies_to_period,
)
from attendance_core.reporting import (
    build_case_rows,
    build_report_highlights,
)
from attendance_core.scanner import resolve_client_ip
from attendance_core.workflows import (
    build_correction_change_summary,
    build_schedule_special_rule_label,
    calculate_suspension_end_date,
    describe_request_review_result,
    expand_request_dates,
    expand_suspension_dates,
    format_request_date_range,
    normalize_request_date_range,
    normalize_schedule_special_rule_type,
    resolve_correction_datetimes,
    schedule_preset_matches_department,
    split_datetime_to_time,
)

# Optional Postgres support
POSTGRES_ENABLED = False
try:
    import psycopg2
    import psycopg2.extras
    POSTGRES_ENABLED = True
except Exception:
    POSTGRES_ENABLED = False

# Optional Google Sheets sync
GOOGLE_SHEETS_ENABLED = False
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GOOGLE_SHEETS_ENABLED = True
except Exception:
    GOOGLE_SHEETS_ENABLED = False

app = Flask(__name__)
app.secret_key = get_configured_secret_key()
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10MB
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = is_production_environment()
_admin_employee_rows_cache = {"stamp": 0, "rows": []}
_options_cache = {
    "departments": {"stamp": 0, "rows": []},
    "employees": {"stamp": 0, "rows": []},
}
_reports_cache = {}
_schedule_change_apply_state = {"stamp": 0}
_admin_notification_scan_state = {}


# =========================
# TIME HELPERS
# =========================
def now_dt():
    return datetime.now(APP_TIMEZONE)


def now_str():
    return now_dt().strftime("%Y-%m-%d %H:%M:%S")


def today_str():
    return now_dt().strftime("%Y-%m-%d")


def now_timestamp():
    return int(now_dt().timestamp())


def ensure_csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_hex(32)
        session["_csrf_token"] = token
    return token


def is_scanner_api_request():
    return request.path in {"/scanner/scan", "/scanner/unlock"}


@app.context_processor
def inject_csrf_token():
    return {"csrf_token": ensure_csrf_token()}


@app.before_request
def verify_csrf_token():
    if request.method != "POST":
        return None

    session_token = session.get("_csrf_token")
    form_token = request.form.get("csrf_token", "")
    if not session_token or not form_token or not secrets.compare_digest(session_token, form_token):
        if is_scanner_api_request():
            return jsonify({"ok": False, "message": "Scanner session expired. Please log in again."}), 403
        flash("Your session expired or the form is no longer valid. Please try again.", "danger")
        return redirect(request.referrer or url_for("login"))

    return None


# =========================
# DATABASE HELPERS
# =========================
def using_postgres():
    return bool(DATABASE_URL) and POSTGRES_ENABLED


def get_db():
    if "db" not in g:
        if using_postgres():
            g.db = psycopg2.connect(DATABASE_URL)
        else:
            g.db = sqlite3.connect(SQLITE_DATABASE)
            g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def convert_query(query: str) -> str:
    return query.replace("?", "%s")


def normalize_db_row(row):
    if isinstance(row, sqlite3.Row):
        return dict(row)
    return row


def fetchone(query, params=()):
    if using_postgres():
        db = get_db() if has_app_context() else psycopg2.connect(DATABASE_URL)
        try:
            with db.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(convert_query(query), params)
                return cur.fetchone()
        finally:
            if not has_app_context():
                db.close()
    if has_app_context():
        cur = get_db().execute(query, params)
        return normalize_db_row(cur.fetchone())
    db = sqlite3.connect(SQLITE_DATABASE)
    db.row_factory = sqlite3.Row
    try:
        cur = db.execute(query, params)
        return normalize_db_row(cur.fetchone())
    finally:
        db.close()


def fetchall(query, params=()):
    if using_postgres():
        db = get_db() if has_app_context() else psycopg2.connect(DATABASE_URL)
        try:
            with db.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(convert_query(query), params)
                return cur.fetchall()
        finally:
            if not has_app_context():
                db.close()
    if has_app_context():
        cur = get_db().execute(query, params)
        return [normalize_db_row(row) for row in cur.fetchall()]
    db = sqlite3.connect(SQLITE_DATABASE)
    db.row_factory = sqlite3.Row
    try:
        cur = db.execute(query, params)
        return [normalize_db_row(row) for row in cur.fetchall()]
    finally:
        db.close()


def execute_db(query, params=(), commit=False):
    db = get_db()
    if using_postgres():
        with db.cursor() as cur:
            cur.execute(convert_query(query), params)
        if commit:
            db.commit()
    else:
        db.execute(query, params)
        if commit:
            db.commit()


def table_exists(table_name):
    db = get_db()
    if using_postgres():
        with db.cursor() as cur:
            cur.execute("""
                SELECT EXISTS (
                    SELECT 1
                    FROM information_schema.tables
                    WHERE table_schema = current_schema()
                      AND table_name = %s
                )
            """, (table_name,))
            row = cur.fetchone()
            return bool(row[0]) if row else False
    row = db.execute("""
        SELECT name
        FROM sqlite_master
        WHERE type = 'table' AND name = ?
    """, (table_name,)).fetchone()
    return bool(row)


def column_exists(table_name, column_name):
    if not table_exists(table_name):
        return False

    db = get_db()
    if using_postgres():
        with db.cursor() as cur:
            cur.execute("""
                SELECT EXISTS (
                    SELECT 1
                    FROM information_schema.columns
                    WHERE table_schema = current_schema()
                      AND table_name = %s
                      AND column_name = %s
                )
            """, (table_name, column_name))
            row = cur.fetchone()
            return bool(row[0]) if row else False

    rows = db.execute(f"PRAGMA table_info({table_name})").fetchall()
    return any(row[1] == column_name for row in rows)


def cursor_has_duplicate_employee_barcodes(cursor, postgres=False):
    trim_expr = "BTRIM(COALESCE(barcode_id, ''))" if postgres else "TRIM(COALESCE(barcode_id, ''))"
    cursor.execute(f"""
        SELECT 1
        FROM users
        WHERE role = 'employee'
          AND {trim_expr} <> ''
        GROUP BY {trim_expr}
        HAVING COUNT(*) > 1
        LIMIT 1
    """)
    return cursor.fetchone() is not None


def ensure_employee_barcode_unique_index_sqlite(cursor):
    if cursor_has_duplicate_employee_barcodes(cursor, postgres=False):
        return
    try:
        cursor.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_users_employee_barcode_unique
            ON users(TRIM(COALESCE(barcode_id, '')))
            WHERE role = 'employee' AND TRIM(COALESCE(barcode_id, '')) <> ''
        """)
    except (sqlite3.IntegrityError, sqlite3.OperationalError):
        pass


def ensure_employee_barcode_unique_index_postgres(cursor):
    if cursor_has_duplicate_employee_barcodes(cursor, postgres=True):
        return
    try:
        cursor.execute("SAVEPOINT employee_barcode_unique_index")
        cursor.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_users_employee_barcode_unique
            ON users ((BTRIM(COALESCE(barcode_id, ''))))
            WHERE role = 'employee' AND BTRIM(COALESCE(barcode_id, '')) <> ''
        """)
        cursor.execute("RELEASE SAVEPOINT employee_barcode_unique_index")
    except Exception:
        cursor.execute("ROLLBACK TO SAVEPOINT employee_barcode_unique_index")
        cursor.execute("RELEASE SAVEPOINT employee_barcode_unique_index")


def get_bootstrap_admin_password():
    return os.environ.get("BOOTSTRAP_ADMIN_PASSWORD", "").strip()


# =========================
# BASIC HELPERS
# =========================
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def is_image(filename):
    if not filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower() if "." in filename else ""
    return ext in IMAGE_EXTENSIONS


def get_client_ip():
    return resolve_client_ip(request.remote_addr)


def get_recent_login_attempts(ip_address):
    cutoff = (now_dt() - timedelta(minutes=LOGIN_WINDOW_MINUTES)).strftime("%Y-%m-%d %H:%M:%S")
    execute_db("DELETE FROM login_attempts WHERE attempted_at < ?", (cutoff,), commit=True)
    return fetchall("""
        SELECT attempted_at
        FROM login_attempts
        WHERE ip_address = ?
        ORDER BY attempted_at DESC
    """, (ip_address,))


def is_login_rate_limited(ip_address):
    return len(get_recent_login_attempts(ip_address)) >= LOGIN_MAX_ATTEMPTS


def register_login_failure(ip_address):
    execute_db("""
        INSERT INTO login_attempts (ip_address, attempted_at)
        VALUES (?, ?)
    """, (ip_address, now_str()), commit=True)


def clear_login_failures(ip_address):
    execute_db("DELETE FROM login_attempts WHERE ip_address = ?", (ip_address,), commit=True)


def login_required(role=None):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if "user_id" not in session:
                if is_scanner_api_request():
                    return jsonify({"ok": False, "message": "Scanner session expired. Please log in again."}), 401
                flash("Please log in first.", "warning")
                return redirect(url_for("login"))

            user = get_user_by_id(session["user_id"])
            if not user:
                session.clear()
                if is_scanner_api_request():
                    return jsonify({"ok": False, "message": "Scanner session expired. Please log in again."}), 401
                flash("Your session expired. Please log in again.", "warning")
                return redirect(url_for("login"))

            if int(user["is_active"] or 0) != 1:
                session.clear()
                if is_scanner_api_request():
                    return jsonify({"ok": False, "message": "This account is inactive. Please contact an administrator."}), 403
                flash("This account is inactive. Please contact an administrator.", "danger")
                return redirect(url_for("login"))

            session["role"] = user["role"]
            session["full_name"] = user["full_name"]

            if role and user["role"] != role:
                if is_scanner_api_request():
                    return jsonify({"ok": False, "message": "Scanner access denied."}), 403
                flash("Access denied.", "danger")
                return redirect(get_home_route_for_user(user))
            return f(*args, **kwargs)
        return wrapped
    return decorator


def get_home_route_for_role(role_name):
    return url_for(get_home_endpoint_for_role(role_name))


def get_home_route_for_user(user_row):
    return url_for(get_home_endpoint_for_user(user_row))


# =========================
# DATABASE INIT / MIGRATION
# =========================
def init_db():
    if using_postgres():
        init_postgres_db()
    else:
        init_sqlite_db()
    ensure_employee_schedule_history_seeded()
    apply_due_future_schedule_changes(force=True)


def init_sqlite_db():
    db = sqlite3.connect(SQLITE_DATABASE)
    cursor = db.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'employee',
            profile_image TEXT,
            department TEXT DEFAULT 'Stellar Seats',
            position TEXT DEFAULT 'Employee',
            emergency_contact_name TEXT,
            emergency_contact_phone TEXT,
            id_issue_date TEXT,
            id_expiration_date TEXT,
            barcode_id TEXT,
            hourly_rate REAL NOT NULL DEFAULT 0,
            sick_leave_days INTEGER NOT NULL DEFAULT 7,
            paid_leave_days INTEGER NOT NULL DEFAULT 7,
            sick_leave_used_manual INTEGER NOT NULL DEFAULT 0,
            paid_leave_used_manual INTEGER NOT NULL DEFAULT 0,
            whatsapp_number TEXT,
            schedule_days TEXT DEFAULT 'Mon,Tue,Wed,Thu,Fri',
            shift_start TEXT DEFAULT '09:00',
            shift_end TEXT DEFAULT '18:00',
            schedule_preset_id INTEGER,
            admin_permissions TEXT,
            admin_role_preset TEXT,
            break_window_start TEXT DEFAULT '12:00',
            break_window_end TEXT DEFAULT '12:15',
            break_limit_minutes INTEGER NOT NULL DEFAULT 15,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS incident_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            employee_name TEXT,
            report_department TEXT,
            error_type TEXT NOT NULL,
            incident_action TEXT DEFAULT 'Coaching',
            incident_date TEXT,
            report_date TEXT NOT NULL,
            message TEXT,
            status TEXT NOT NULL DEFAULT 'Open',
            admin_note TEXT,
            created_by INTEGER,
            reviewed_by INTEGER,
            reviewed_at TEXT,
            disciplinary_action_id INTEGER,
            policy_incident_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            work_date TEXT NOT NULL,
            time_in TEXT,
            time_out TEXT,
            status TEXT DEFAULT 'Offline',
            proof_file TEXT,
            notes TEXT,
            late_flag INTEGER NOT NULL DEFAULT 0,
            late_minutes INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS breaks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            attendance_id INTEGER,
            work_date TEXT NOT NULL,
            break_start TEXT,
            break_end TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (attendance_id) REFERENCES attendance (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            message TEXT NOT NULL,
            created_at TEXT NOT NULL,
            is_read INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS login_attempts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ip_address TEXT NOT NULL,
            attempted_at TEXT NOT NULL
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS disciplinary_actions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            action_type TEXT NOT NULL,
            action_date TEXT NOT NULL,
            duration_days INTEGER NOT NULL DEFAULT 1,
            end_date TEXT,
            details TEXT,
            incident_report_id INTEGER,
            error_type TEXT,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS schedule_presets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            department_scope TEXT,
            schedule_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri',
            shift_start TEXT NOT NULL DEFAULT '09:00',
            shift_end TEXT NOT NULL DEFAULT '18:00',
            break_limit_minutes INTEGER NOT NULL DEFAULT 15,
            notes TEXT,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS company_settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            id_signatory_name TEXT,
            id_signatory_title TEXT,
            id_signature_file TEXT,
            hr_signatory_name TEXT,
            hr_signatory_title TEXT,
            hr_signature_file TEXT,
            scanner_attendance_mode INTEGER NOT NULL DEFAULT 0,
            scanner_lock_timeout_seconds INTEGER NOT NULL DEFAULT 90,
            scanner_exit_pin_hash TEXT,
            overtime_multiplier REAL NOT NULL DEFAULT 1.25,
            last_external_backup_at TEXT,
            last_external_backup_by INTEGER,
            last_external_backup_note TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS correction_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            request_type TEXT NOT NULL,
            work_date TEXT NOT NULL,
            end_work_date TEXT,
            message TEXT,
            requested_time_in TEXT,
            requested_break_start TEXT,
            requested_break_end TEXT,
            requested_time_out TEXT,
            applied_changes TEXT,
            status TEXT NOT NULL DEFAULT 'Pending',
            admin_note TEXT,
            reviewed_by INTEGER,
            reviewed_at TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS activity_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            target_user_id INTEGER,
            action TEXT NOT NULL,
            details TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (target_user_id) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS scanner_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            scanner_user_id INTEGER,
            employee_user_id INTEGER,
            action_type TEXT,
            barcode_value TEXT,
            result_status TEXT NOT NULL,
            result_message TEXT,
            source_label TEXT,
            device_label TEXT,
            ip_address TEXT,
            user_agent TEXT,
            scanner_name_snapshot TEXT,
            scanner_username_snapshot TEXT,
            employee_name_snapshot TEXT,
            employee_department_snapshot TEXT,
            employee_position_snapshot TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (scanner_user_id) REFERENCES users (id),
            FOREIGN KEY (employee_user_id) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS overtime_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            attendance_id INTEGER,
            work_date TEXT NOT NULL,
            overtime_start TEXT NOT NULL,
            overtime_end TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (attendance_id) REFERENCES attendance (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll_adjustments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date_from TEXT NOT NULL,
            date_to TEXT NOT NULL,
            adjustment_type TEXT NOT NULL,
            label TEXT NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            notes TEXT,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll_recurring_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            adjustment_type TEXT NOT NULL,
            label TEXT NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            recurrence_type TEXT NOT NULL DEFAULT 'Every Payroll',
            start_date TEXT,
            end_date TEXT,
            notes TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date_from TEXT NOT NULL,
            date_to TEXT NOT NULL,
            department_filter TEXT,
            employee_filter TEXT,
            status TEXT NOT NULL DEFAULT 'Draft',
            notes TEXT,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            released_at TEXT,
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll_run_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            payroll_run_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            full_name TEXT,
            department TEXT,
            position TEXT,
            hourly_rate REAL NOT NULL DEFAULT 0,
            days_worked INTEGER NOT NULL DEFAULT 0,
            total_hours REAL NOT NULL DEFAULT 0,
            overtime_hours REAL NOT NULL DEFAULT 0,
            late_minutes INTEGER NOT NULL DEFAULT 0,
            break_minutes INTEGER NOT NULL DEFAULT 0,
            suspension_days INTEGER NOT NULL DEFAULT 0,
            suspension_pay REAL NOT NULL DEFAULT 0,
            gross_pay REAL NOT NULL DEFAULT 0,
            overtime_pay REAL NOT NULL DEFAULT 0,
            allowances REAL NOT NULL DEFAULT 0,
            deductions REAL NOT NULL DEFAULT 0,
            final_pay REAL NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            FOREIGN KEY (payroll_run_id) REFERENCES payroll_runs (id),
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll_run_item_adjustments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            payroll_run_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            source_kind TEXT,
            source_rule_id INTEGER,
            recurrence_type TEXT,
            adjustment_type TEXT NOT NULL,
            label TEXT NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            notes TEXT,
            created_by INTEGER,
            created_by_name TEXT,
            adjustment_created_at TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (payroll_run_id) REFERENCES payroll_runs (id),
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    """)

    # users migration
    existing_cols_users = [row[1] for row in cursor.execute("PRAGMA table_info(users)").fetchall()]
    if "department" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN department TEXT DEFAULT 'Stellar Seats'")
    if "position" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN position TEXT DEFAULT 'Employee'")
    if "emergency_contact_name" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN emergency_contact_name TEXT")
    if "emergency_contact_phone" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN emergency_contact_phone TEXT")
    if "id_issue_date" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN id_issue_date TEXT")
    if "id_expiration_date" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN id_expiration_date TEXT")
    if "barcode_id" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN barcode_id TEXT")
    if "hourly_rate" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN hourly_rate REAL NOT NULL DEFAULT 0")
    if "sick_leave_days" not in existing_cols_users:
        cursor.execute(f"ALTER TABLE users ADD COLUMN sick_leave_days INTEGER NOT NULL DEFAULT {DEFAULT_SICK_LEAVE_DAYS}")
    if "paid_leave_days" not in existing_cols_users:
        cursor.execute(f"ALTER TABLE users ADD COLUMN paid_leave_days INTEGER NOT NULL DEFAULT {DEFAULT_PAID_LEAVE_DAYS}")
    if "sick_leave_used_manual" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN sick_leave_used_manual INTEGER NOT NULL DEFAULT 0")
    if "paid_leave_used_manual" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN paid_leave_used_manual INTEGER NOT NULL DEFAULT 0")
    if "whatsapp_number" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN whatsapp_number TEXT")
    if "schedule_days" not in existing_cols_users:
        cursor.execute(f"ALTER TABLE users ADD COLUMN schedule_days TEXT DEFAULT '{DEFAULT_SCHEDULE_DAYS}'")
    if "shift_start" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN shift_start TEXT DEFAULT '09:00'")
    if "shift_end" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN shift_end TEXT DEFAULT '18:00'")
    if "schedule_preset_id" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN schedule_preset_id INTEGER")
    if "admin_permissions" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN admin_permissions TEXT")
    if "admin_role_preset" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN admin_role_preset TEXT")
    if "break_window_start" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN break_window_start TEXT DEFAULT '12:00'")
    if "break_window_end" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN break_window_end TEXT DEFAULT '12:15'")
    if "break_limit_minutes" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN break_limit_minutes INTEGER NOT NULL DEFAULT 15")
    if "is_active" not in existing_cols_users:
        cursor.execute("ALTER TABLE users ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
    ensure_employee_barcode_unique_index_sqlite(cursor)

    existing_cols_company_settings = [row[1] for row in cursor.execute("PRAGMA table_info(company_settings)").fetchall()]
    if not existing_cols_company_settings:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS company_settings (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                id_signatory_name TEXT,
                id_signatory_title TEXT,
                id_signature_file TEXT,
                hr_signatory_name TEXT,
                hr_signatory_title TEXT,
                hr_signature_file TEXT,
                scanner_attendance_mode INTEGER NOT NULL DEFAULT 0,
                scanner_lock_timeout_seconds INTEGER NOT NULL DEFAULT 90,
                scanner_exit_pin_hash TEXT,
                overtime_multiplier REAL NOT NULL DEFAULT 1.25,
                last_external_backup_at TEXT,
                last_external_backup_by INTEGER,
                last_external_backup_note TEXT
            )
        """)
        existing_cols_company_settings = [row[1] for row in cursor.execute("PRAGMA table_info(company_settings)").fetchall()]
    if "id_signatory_name" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN id_signatory_name TEXT")
    if "id_signatory_title" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN id_signatory_title TEXT")
    if "id_signature_file" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN id_signature_file TEXT")
    if "hr_signatory_name" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN hr_signatory_name TEXT")
    if "hr_signatory_title" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN hr_signatory_title TEXT")
    if "hr_signature_file" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN hr_signature_file TEXT")
    if "scanner_attendance_mode" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN scanner_attendance_mode INTEGER NOT NULL DEFAULT 0")
    if "scanner_lock_timeout_seconds" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN scanner_lock_timeout_seconds INTEGER NOT NULL DEFAULT 90")
    if "scanner_exit_pin_hash" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN scanner_exit_pin_hash TEXT")
    if "overtime_multiplier" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN overtime_multiplier REAL NOT NULL DEFAULT 1.25")
    if "last_external_backup_at" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN last_external_backup_at TEXT")
    if "last_external_backup_by" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN last_external_backup_by INTEGER")
    if "last_external_backup_note" not in existing_cols_company_settings:
        cursor.execute("ALTER TABLE company_settings ADD COLUMN last_external_backup_note TEXT")
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS schedule_presets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            department_scope TEXT,
            schedule_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri',
            shift_start TEXT NOT NULL DEFAULT '09:00',
            shift_end TEXT NOT NULL DEFAULT '18:00',
            break_limit_minutes INTEGER NOT NULL DEFAULT 15,
            notes TEXT,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employee_schedule_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            effective_at TEXT NOT NULL,
            department TEXT,
            position TEXT,
            schedule_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri',
            shift_start TEXT NOT NULL DEFAULT '09:00',
            shift_end TEXT NOT NULL DEFAULT '18:00',
            break_limit_minutes INTEGER NOT NULL DEFAULT 15,
            schedule_preset_id INTEGER,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employee_future_schedule_changes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            effective_date TEXT NOT NULL,
            department TEXT,
            position TEXT,
            schedule_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri',
            shift_start TEXT NOT NULL DEFAULT '09:00',
            shift_end TEXT NOT NULL DEFAULT '18:00',
            break_limit_minutes INTEGER NOT NULL DEFAULT 15,
            schedule_preset_id INTEGER,
            notes TEXT,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            applied_at TEXT,
            applied_by INTEGER,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (created_by) REFERENCES users (id),
            FOREIGN KEY (applied_by) REFERENCES users (id)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS schedule_special_dates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            special_date TEXT NOT NULL UNIQUE,
            rule_type TEXT NOT NULL DEFAULT 'holiday',
            label TEXT NOT NULL,
            notes TEXT,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    """)
    existing_cols_schedule_presets = [row[1] for row in cursor.execute("PRAGMA table_info(schedule_presets)").fetchall()]
    if "department_scope" not in existing_cols_schedule_presets:
        cursor.execute("ALTER TABLE schedule_presets ADD COLUMN department_scope TEXT")
    if "schedule_days" not in existing_cols_schedule_presets:
        cursor.execute(f"ALTER TABLE schedule_presets ADD COLUMN schedule_days TEXT NOT NULL DEFAULT '{DEFAULT_SCHEDULE_DAYS}'")
    if "shift_start" not in existing_cols_schedule_presets:
        cursor.execute(f"ALTER TABLE schedule_presets ADD COLUMN shift_start TEXT NOT NULL DEFAULT '{DEFAULT_SHIFT_START}'")
    if "shift_end" not in existing_cols_schedule_presets:
        cursor.execute(f"ALTER TABLE schedule_presets ADD COLUMN shift_end TEXT NOT NULL DEFAULT '{DEFAULT_SHIFT_END}'")
    if "break_limit_minutes" not in existing_cols_schedule_presets:
        cursor.execute("ALTER TABLE schedule_presets ADD COLUMN break_limit_minutes INTEGER NOT NULL DEFAULT 15")
    if "notes" not in existing_cols_schedule_presets:
        cursor.execute("ALTER TABLE schedule_presets ADD COLUMN notes TEXT")
    if "created_by" not in existing_cols_schedule_presets:
        cursor.execute("ALTER TABLE schedule_presets ADD COLUMN created_by INTEGER")
    if "created_at" not in existing_cols_schedule_presets:
        cursor.execute("ALTER TABLE schedule_presets ADD COLUMN created_at TEXT")
    if "updated_at" not in existing_cols_schedule_presets:
        cursor.execute("ALTER TABLE schedule_presets ADD COLUMN updated_at TEXT")
    existing_cols_schedule_history = [row[1] for row in cursor.execute("PRAGMA table_info(employee_schedule_history)").fetchall()]
    if "department" not in existing_cols_schedule_history:
        cursor.execute("ALTER TABLE employee_schedule_history ADD COLUMN department TEXT")
    if "position" not in existing_cols_schedule_history:
        cursor.execute("ALTER TABLE employee_schedule_history ADD COLUMN position TEXT")
    if "schedule_days" not in existing_cols_schedule_history:
        cursor.execute(f"ALTER TABLE employee_schedule_history ADD COLUMN schedule_days TEXT NOT NULL DEFAULT '{DEFAULT_SCHEDULE_DAYS}'")
    if "shift_start" not in existing_cols_schedule_history:
        cursor.execute(f"ALTER TABLE employee_schedule_history ADD COLUMN shift_start TEXT NOT NULL DEFAULT '{DEFAULT_SHIFT_START}'")
    if "shift_end" not in existing_cols_schedule_history:
        cursor.execute(f"ALTER TABLE employee_schedule_history ADD COLUMN shift_end TEXT NOT NULL DEFAULT '{DEFAULT_SHIFT_END}'")
    if "break_limit_minutes" not in existing_cols_schedule_history:
        cursor.execute("ALTER TABLE employee_schedule_history ADD COLUMN break_limit_minutes INTEGER NOT NULL DEFAULT 15")
    if "schedule_preset_id" not in existing_cols_schedule_history:
        cursor.execute("ALTER TABLE employee_schedule_history ADD COLUMN schedule_preset_id INTEGER")
    if "created_by" not in existing_cols_schedule_history:
        cursor.execute("ALTER TABLE employee_schedule_history ADD COLUMN created_by INTEGER")
    if "created_at" not in existing_cols_schedule_history:
        cursor.execute("ALTER TABLE employee_schedule_history ADD COLUMN created_at TEXT")
    existing_cols_future_schedule_changes = [row[1] for row in cursor.execute("PRAGMA table_info(employee_future_schedule_changes)").fetchall()]
    if "department" not in existing_cols_future_schedule_changes:
        cursor.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN department TEXT")
    if "position" not in existing_cols_future_schedule_changes:
        cursor.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN position TEXT")
    if "schedule_days" not in existing_cols_future_schedule_changes:
        cursor.execute(f"ALTER TABLE employee_future_schedule_changes ADD COLUMN schedule_days TEXT NOT NULL DEFAULT '{DEFAULT_SCHEDULE_DAYS}'")
    if "shift_start" not in existing_cols_future_schedule_changes:
        cursor.execute(f"ALTER TABLE employee_future_schedule_changes ADD COLUMN shift_start TEXT NOT NULL DEFAULT '{DEFAULT_SHIFT_START}'")
    if "shift_end" not in existing_cols_future_schedule_changes:
        cursor.execute(f"ALTER TABLE employee_future_schedule_changes ADD COLUMN shift_end TEXT NOT NULL DEFAULT '{DEFAULT_SHIFT_END}'")
    if "break_limit_minutes" not in existing_cols_future_schedule_changes:
        cursor.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN break_limit_minutes INTEGER NOT NULL DEFAULT 15")
    if "schedule_preset_id" not in existing_cols_future_schedule_changes:
        cursor.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN schedule_preset_id INTEGER")
    if "notes" not in existing_cols_future_schedule_changes:
        cursor.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN notes TEXT")
    if "created_by" not in existing_cols_future_schedule_changes:
        cursor.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN created_by INTEGER")
    if "created_at" not in existing_cols_future_schedule_changes:
        cursor.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN created_at TEXT")
    if "applied_at" not in existing_cols_future_schedule_changes:
        cursor.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN applied_at TEXT")
    if "applied_by" not in existing_cols_future_schedule_changes:
        cursor.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN applied_by INTEGER")
    existing_cols_schedule_special_dates = [row[1] for row in cursor.execute("PRAGMA table_info(schedule_special_dates)").fetchall()]
    if "special_date" not in existing_cols_schedule_special_dates:
        cursor.execute("ALTER TABLE schedule_special_dates ADD COLUMN special_date TEXT")
    if "rule_type" not in existing_cols_schedule_special_dates:
        cursor.execute("ALTER TABLE schedule_special_dates ADD COLUMN rule_type TEXT NOT NULL DEFAULT 'holiday'")
    if "label" not in existing_cols_schedule_special_dates:
        cursor.execute("ALTER TABLE schedule_special_dates ADD COLUMN label TEXT")
    if "notes" not in existing_cols_schedule_special_dates:
        cursor.execute("ALTER TABLE schedule_special_dates ADD COLUMN notes TEXT")
    if "created_by" not in existing_cols_schedule_special_dates:
        cursor.execute("ALTER TABLE schedule_special_dates ADD COLUMN created_by INTEGER")
    if "created_at" not in existing_cols_schedule_special_dates:
        cursor.execute("ALTER TABLE schedule_special_dates ADD COLUMN created_at TEXT")
    if "updated_at" not in existing_cols_schedule_special_dates:
        cursor.execute("ALTER TABLE schedule_special_dates ADD COLUMN updated_at TEXT")

    cursor.execute("""
        INSERT OR IGNORE INTO company_settings (
            id, id_signatory_name, id_signatory_title, id_signature_file,
            scanner_attendance_mode, scanner_lock_timeout_seconds, scanner_exit_pin_hash, overtime_multiplier
        )
        VALUES (1, 'Kirk Danny Fernandez', 'Head Of Operations', NULL, 0, 90, NULL, 1.25)
    """)

    # attendance migration
    existing_cols_att = [row[1] for row in cursor.execute("PRAGMA table_info(attendance)").fetchall()]
    if "late_flag" not in existing_cols_att:
        cursor.execute("ALTER TABLE attendance ADD COLUMN late_flag INTEGER NOT NULL DEFAULT 0")
    if "late_minutes" not in existing_cols_att:
        cursor.execute("ALTER TABLE attendance ADD COLUMN late_minutes INTEGER NOT NULL DEFAULT 0")

    cursor.execute("CREATE INDEX IF NOT EXISTS idx_login_attempts_ip_attempted_at ON login_attempts(ip_address, attempted_at)")
    try:
        cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_attendance_one_open_per_user ON attendance(user_id) WHERE time_in IS NOT NULL AND time_out IS NULL")
    except sqlite3.IntegrityError:
        pass
    try:
        cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_breaks_one_open_per_attendance ON breaks(attendance_id) WHERE attendance_id IS NOT NULL AND break_end IS NULL")
    except sqlite3.IntegrityError:
        pass
    try:
        cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_overtime_one_open_per_user ON overtime_sessions(user_id) WHERE overtime_end IS NULL")
    except sqlite3.IntegrityError:
        pass

    # incident reports migration (this fixes your current error)
    existing_cols_incident = [row[1] for row in cursor.execute("PRAGMA table_info(incident_reports)").fetchall()]
    if existing_cols_incident:
        if "error_type" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN error_type TEXT")
        if "report_date" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN report_date TEXT")
        if "message" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN message TEXT")
        if "employee_name" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN employee_name TEXT")
        if "report_department" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN report_department TEXT")
        if "incident_action" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN incident_action TEXT DEFAULT 'Coaching'")
        if "incident_date" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN incident_date TEXT")
        if "status" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN status TEXT NOT NULL DEFAULT 'Open'")
        if "admin_note" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN admin_note TEXT")
        if "created_by" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN created_by INTEGER")
        if "reviewed_by" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN reviewed_by INTEGER")
        if "reviewed_at" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN reviewed_at TEXT")
        if "created_at" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN created_at TEXT")
        if "disciplinary_action_id" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN disciplinary_action_id INTEGER")
        if "policy_incident_count" not in existing_cols_incident:
            cursor.execute("ALTER TABLE incident_reports ADD COLUMN policy_incident_count INTEGER NOT NULL DEFAULT 0")

    existing_cols_disciplinary = [row[1] for row in cursor.execute("PRAGMA table_info(disciplinary_actions)").fetchall()]
    if existing_cols_disciplinary:
        if "duration_days" not in existing_cols_disciplinary:
            cursor.execute("ALTER TABLE disciplinary_actions ADD COLUMN duration_days INTEGER NOT NULL DEFAULT 1")
        if "end_date" not in existing_cols_disciplinary:
            cursor.execute("ALTER TABLE disciplinary_actions ADD COLUMN end_date TEXT")
        if "details" not in existing_cols_disciplinary:
            cursor.execute("ALTER TABLE disciplinary_actions ADD COLUMN details TEXT")
        if "incident_report_id" not in existing_cols_disciplinary:
            cursor.execute("ALTER TABLE disciplinary_actions ADD COLUMN incident_report_id INTEGER")
        if "error_type" not in existing_cols_disciplinary:
            cursor.execute("ALTER TABLE disciplinary_actions ADD COLUMN error_type TEXT")
        if "created_by" not in existing_cols_disciplinary:
            cursor.execute("ALTER TABLE disciplinary_actions ADD COLUMN created_by INTEGER")
        if "created_at" not in existing_cols_disciplinary:
            cursor.execute("ALTER TABLE disciplinary_actions ADD COLUMN created_at TEXT")

    existing_cols_corrections = [row[1] for row in cursor.execute("PRAGMA table_info(correction_requests)").fetchall()]
    if existing_cols_corrections:
        if "end_work_date" not in existing_cols_corrections:
            cursor.execute("ALTER TABLE correction_requests ADD COLUMN end_work_date TEXT")
        if "requested_time_in" not in existing_cols_corrections:
            cursor.execute("ALTER TABLE correction_requests ADD COLUMN requested_time_in TEXT")
        if "requested_break_start" not in existing_cols_corrections:
            cursor.execute("ALTER TABLE correction_requests ADD COLUMN requested_break_start TEXT")
        if "requested_break_end" not in existing_cols_corrections:
            cursor.execute("ALTER TABLE correction_requests ADD COLUMN requested_break_end TEXT")
        if "requested_time_out" not in existing_cols_corrections:
            cursor.execute("ALTER TABLE correction_requests ADD COLUMN requested_time_out TEXT")
        if "applied_changes" not in existing_cols_corrections:
            cursor.execute("ALTER TABLE correction_requests ADD COLUMN applied_changes TEXT")
        if "status" not in existing_cols_corrections:
            cursor.execute("ALTER TABLE correction_requests ADD COLUMN status TEXT NOT NULL DEFAULT 'Pending'")
        if "admin_note" not in existing_cols_corrections:
            cursor.execute("ALTER TABLE correction_requests ADD COLUMN admin_note TEXT")
        if "reviewed_by" not in existing_cols_corrections:
            cursor.execute("ALTER TABLE correction_requests ADD COLUMN reviewed_by INTEGER")
        if "reviewed_at" not in existing_cols_corrections:
            cursor.execute("ALTER TABLE correction_requests ADD COLUMN reviewed_at TEXT")

    existing_cols_scanner_logs = [row[1] for row in cursor.execute("PRAGMA table_info(scanner_logs)").fetchall()]
    if not existing_cols_scanner_logs:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS scanner_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                scanner_user_id INTEGER,
                employee_user_id INTEGER,
                action_type TEXT,
                barcode_value TEXT,
                result_status TEXT NOT NULL,
                result_message TEXT,
                source_label TEXT,
                device_label TEXT,
                ip_address TEXT,
                user_agent TEXT,
                scanner_name_snapshot TEXT,
                scanner_username_snapshot TEXT,
                employee_name_snapshot TEXT,
                employee_department_snapshot TEXT,
                employee_position_snapshot TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (scanner_user_id) REFERENCES users (id),
                FOREIGN KEY (employee_user_id) REFERENCES users (id)
            )
        """)
        existing_cols_scanner_logs = [row[1] for row in cursor.execute("PRAGMA table_info(scanner_logs)").fetchall()]
    if "scanner_user_id" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN scanner_user_id INTEGER")
    if "employee_user_id" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN employee_user_id INTEGER")
    if "action_type" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN action_type TEXT")
    if "barcode_value" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN barcode_value TEXT")
    if "result_status" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN result_status TEXT NOT NULL DEFAULT 'error'")
    if "result_message" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN result_message TEXT")
    if "source_label" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN source_label TEXT")
    if "device_label" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN device_label TEXT")
    if "ip_address" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN ip_address TEXT")
    if "user_agent" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN user_agent TEXT")
    if "scanner_name_snapshot" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN scanner_name_snapshot TEXT")
    if "scanner_username_snapshot" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN scanner_username_snapshot TEXT")
    if "employee_name_snapshot" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN employee_name_snapshot TEXT")
    if "employee_department_snapshot" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN employee_department_snapshot TEXT")
    if "employee_position_snapshot" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN employee_position_snapshot TEXT")
    if "created_at" not in existing_cols_scanner_logs:
        cursor.execute("ALTER TABLE scanner_logs ADD COLUMN created_at TEXT")

    existing_cols_activity_logs = [row[1] for row in cursor.execute("PRAGMA table_info(activity_logs)").fetchall()]
    if "target_user_id" not in existing_cols_activity_logs:
        cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_user_id INTEGER")
    cursor.execute("UPDATE activity_logs SET target_user_id = user_id WHERE target_user_id IS NULL")

    existing_cols_overtime = [row[1] for row in cursor.execute("PRAGMA table_info(overtime_sessions)").fetchall()]
    if not existing_cols_overtime:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS overtime_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                attendance_id INTEGER,
                work_date TEXT NOT NULL,
                overtime_start TEXT NOT NULL,
                overtime_end TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users (id),
                FOREIGN KEY (attendance_id) REFERENCES attendance (id)
            )
        """)
        existing_cols_overtime = [row[1] for row in cursor.execute("PRAGMA table_info(overtime_sessions)").fetchall()]
    if "user_id" not in existing_cols_overtime:
        cursor.execute("ALTER TABLE overtime_sessions ADD COLUMN user_id INTEGER")
    if "attendance_id" not in existing_cols_overtime:
        cursor.execute("ALTER TABLE overtime_sessions ADD COLUMN attendance_id INTEGER")
    if "work_date" not in existing_cols_overtime:
        cursor.execute("ALTER TABLE overtime_sessions ADD COLUMN work_date TEXT")
    if "overtime_start" not in existing_cols_overtime:
        cursor.execute("ALTER TABLE overtime_sessions ADD COLUMN overtime_start TEXT")
    if "overtime_end" not in existing_cols_overtime:
        cursor.execute("ALTER TABLE overtime_sessions ADD COLUMN overtime_end TEXT")
    if "created_at" not in existing_cols_overtime:
        cursor.execute("ALTER TABLE overtime_sessions ADD COLUMN created_at TEXT")

    existing_cols_payroll_adjustments = [row[1] for row in cursor.execute("PRAGMA table_info(payroll_adjustments)").fetchall()]
    if not existing_cols_payroll_adjustments:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS payroll_adjustments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                date_from TEXT NOT NULL,
                date_to TEXT NOT NULL,
                adjustment_type TEXT NOT NULL,
                label TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                notes TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users (id),
                FOREIGN KEY (created_by) REFERENCES users (id)
            )
        """)

    existing_cols_payroll_runs = [row[1] for row in cursor.execute("PRAGMA table_info(payroll_runs)").fetchall()]
    if not existing_cols_payroll_runs:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS payroll_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date_from TEXT NOT NULL,
                date_to TEXT NOT NULL,
                department_filter TEXT,
                employee_filter TEXT,
                status TEXT NOT NULL DEFAULT 'Draft',
                notes TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                released_at TEXT,
                FOREIGN KEY (created_by) REFERENCES users (id)
            )
        """)

    existing_cols_payroll_run_items = [row[1] for row in cursor.execute("PRAGMA table_info(payroll_run_items)").fetchall()]
    if not existing_cols_payroll_run_items:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS payroll_run_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                payroll_run_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                full_name TEXT,
                department TEXT,
                position TEXT,
                hourly_rate REAL NOT NULL DEFAULT 0,
                days_worked INTEGER NOT NULL DEFAULT 0,
                total_hours REAL NOT NULL DEFAULT 0,
                overtime_hours REAL NOT NULL DEFAULT 0,
                late_minutes INTEGER NOT NULL DEFAULT 0,
                break_minutes INTEGER NOT NULL DEFAULT 0,
                suspension_days INTEGER NOT NULL DEFAULT 0,
                suspension_pay REAL NOT NULL DEFAULT 0,
                gross_pay REAL NOT NULL DEFAULT 0,
                overtime_pay REAL NOT NULL DEFAULT 0,
                allowances REAL NOT NULL DEFAULT 0,
                deductions REAL NOT NULL DEFAULT 0,
                final_pay REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                FOREIGN KEY (payroll_run_id) REFERENCES payroll_runs (id),
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        """)

    existing_cols_payroll_run_item_adjustments = [row[1] for row in cursor.execute("PRAGMA table_info(payroll_run_item_adjustments)").fetchall()]
    if not existing_cols_payroll_run_item_adjustments:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS payroll_run_item_adjustments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                payroll_run_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                source_kind TEXT,
                source_rule_id INTEGER,
                recurrence_type TEXT,
                adjustment_type TEXT NOT NULL,
                label TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                notes TEXT,
                created_by INTEGER,
                created_by_name TEXT,
                adjustment_created_at TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (payroll_run_id) REFERENCES payroll_runs (id),
                FOREIGN KEY (user_id) REFERENCES users (id),
                FOREIGN KEY (created_by) REFERENCES users (id)
            )
        """)
    else:
        if "source_kind" not in existing_cols_payroll_run_item_adjustments:
            cursor.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN source_kind TEXT")
        if "source_rule_id" not in existing_cols_payroll_run_item_adjustments:
            cursor.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN source_rule_id INTEGER")
        if "recurrence_type" not in existing_cols_payroll_run_item_adjustments:
            cursor.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN recurrence_type TEXT")

    existing_cols_payroll_recurring_rules = [row[1] for row in cursor.execute("PRAGMA table_info(payroll_recurring_rules)").fetchall()]
    if not existing_cols_payroll_recurring_rules:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS payroll_recurring_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                adjustment_type TEXT NOT NULL,
                label TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                recurrence_type TEXT NOT NULL DEFAULT 'Every Payroll',
                start_date TEXT,
                end_date TEXT,
                notes TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_by INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users (id),
                FOREIGN KEY (created_by) REFERENCES users (id)
            )
        """)
    else:
        if "recurrence_type" not in existing_cols_payroll_recurring_rules:
            cursor.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN recurrence_type TEXT NOT NULL DEFAULT 'Every Payroll'")
        if "start_date" not in existing_cols_payroll_recurring_rules:
            cursor.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN start_date TEXT")
        if "end_date" not in existing_cols_payroll_recurring_rules:
            cursor.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN end_date TEXT")
        if "notes" not in existing_cols_payroll_recurring_rules:
            cursor.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN notes TEXT")
        if "is_active" not in existing_cols_payroll_recurring_rules:
            cursor.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
        if "created_by" not in existing_cols_payroll_recurring_rules:
            cursor.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN created_by INTEGER")
        if "created_at" not in existing_cols_payroll_recurring_rules:
            cursor.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN created_at TEXT")
        if "updated_at" not in existing_cols_payroll_recurring_rules:
            cursor.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN updated_at TEXT")

    db.commit()

    admin = cursor.execute("SELECT * FROM users WHERE username = ?", ("admin",)).fetchone()
    if not admin:
        bootstrap_password = get_bootstrap_admin_password()
        if is_production_environment() and not bootstrap_password:
            db.close()
            raise RuntimeError(
                "BOOTSTRAP_ADMIN_PASSWORD must be set before first production startup when the admin account does not exist."
            )
        cursor.execute("""
            INSERT INTO users (
                full_name, username, password_hash, role,
                profile_image, department, position, shift_start, is_active, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "Administrator",
            "admin",
            generate_password_hash(bootstrap_password or "admin123"),
            "admin",
            None,
            "Stellar Seats",
            "Administrator",
            DEFAULT_SHIFT_START,
            1,
            now_str()
        ))
        db.commit()

    db.close()


def init_postgres_db():
    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS incident_reports (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                employee_name TEXT,
                report_department TEXT,
                error_type TEXT NOT NULL,
                incident_action TEXT DEFAULT 'Coaching',
                incident_date TEXT,
                report_date TEXT NOT NULL,
                message TEXT,
                status TEXT NOT NULL DEFAULT 'Open',
                admin_note TEXT,
                created_by INTEGER,
                reviewed_by INTEGER,
                reviewed_at TEXT,
                disciplinary_action_id INTEGER,
                policy_incident_count INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS disciplinary_actions (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                action_type TEXT NOT NULL,
                action_date TEXT NOT NULL,
                duration_days INTEGER NOT NULL DEFAULT 1,
                end_date TEXT,
                details TEXT,
                incident_report_id INTEGER,
                error_type TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL
            )
        """)

        cur.execute("ALTER TABLE incident_reports ADD COLUMN IF NOT EXISTS employee_name TEXT")
        cur.execute("ALTER TABLE incident_reports ADD COLUMN IF NOT EXISTS report_department TEXT")
        cur.execute("ALTER TABLE incident_reports ADD COLUMN IF NOT EXISTS incident_action TEXT DEFAULT 'Coaching'")
        cur.execute("ALTER TABLE incident_reports ADD COLUMN IF NOT EXISTS incident_date TEXT")
        cur.execute("ALTER TABLE incident_reports ADD COLUMN IF NOT EXISTS status TEXT NOT NULL DEFAULT 'Open'")
        cur.execute("ALTER TABLE incident_reports ADD COLUMN IF NOT EXISTS admin_note TEXT")
        cur.execute("ALTER TABLE incident_reports ADD COLUMN IF NOT EXISTS reviewed_by INTEGER")
        cur.execute("ALTER TABLE incident_reports ADD COLUMN IF NOT EXISTS reviewed_at TEXT")
        cur.execute("ALTER TABLE incident_reports ADD COLUMN IF NOT EXISTS disciplinary_action_id INTEGER")
        cur.execute("ALTER TABLE incident_reports ADD COLUMN IF NOT EXISTS policy_incident_count INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE disciplinary_actions ADD COLUMN IF NOT EXISTS duration_days INTEGER NOT NULL DEFAULT 1")
        cur.execute("ALTER TABLE disciplinary_actions ADD COLUMN IF NOT EXISTS end_date TEXT")
        cur.execute("ALTER TABLE disciplinary_actions ADD COLUMN IF NOT EXISTS details TEXT")
        cur.execute("ALTER TABLE disciplinary_actions ADD COLUMN IF NOT EXISTS incident_report_id INTEGER")
        cur.execute("ALTER TABLE disciplinary_actions ADD COLUMN IF NOT EXISTS error_type TEXT")
        cur.execute("ALTER TABLE disciplinary_actions ADD COLUMN IF NOT EXISTS created_by INTEGER")
        cur.execute("ALTER TABLE disciplinary_actions ADD COLUMN IF NOT EXISTS created_at TEXT")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                full_name TEXT NOT NULL,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'employee',
                profile_image TEXT,
                department TEXT DEFAULT 'Stellar Seats',
                position TEXT DEFAULT 'Employee',
                emergency_contact_name TEXT,
                emergency_contact_phone TEXT,
                id_issue_date TEXT,
                id_expiration_date TEXT,
                barcode_id TEXT,
                hourly_rate NUMERIC(12, 2) NOT NULL DEFAULT 0,
                sick_leave_days INTEGER NOT NULL DEFAULT 7,
                paid_leave_days INTEGER NOT NULL DEFAULT 7,
                sick_leave_used_manual INTEGER NOT NULL DEFAULT 0,
                paid_leave_used_manual INTEGER NOT NULL DEFAULT 0,
                whatsapp_number TEXT,
                schedule_days TEXT DEFAULT 'Mon,Tue,Wed,Thu,Fri',
                shift_start TEXT DEFAULT '09:00',
                shift_end TEXT DEFAULT '18:00',
                schedule_preset_id INTEGER,
                admin_permissions TEXT,
                admin_role_preset TEXT,
                break_window_start TEXT DEFAULT '12:00',
                break_window_end TEXT DEFAULT '12:15',
                break_limit_minutes INTEGER NOT NULL DEFAULT 15,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            )
        """)

        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS whatsapp_number TEXT")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS emergency_contact_name TEXT")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS emergency_contact_phone TEXT")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS id_issue_date TEXT")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS id_expiration_date TEXT")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS barcode_id TEXT")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS hourly_rate NUMERIC(12, 2) NOT NULL DEFAULT 0")
        cur.execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS sick_leave_days INTEGER NOT NULL DEFAULT {DEFAULT_SICK_LEAVE_DAYS}")
        cur.execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS paid_leave_days INTEGER NOT NULL DEFAULT {DEFAULT_PAID_LEAVE_DAYS}")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS sick_leave_used_manual INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS paid_leave_used_manual INTEGER NOT NULL DEFAULT 0")
        cur.execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS schedule_days TEXT DEFAULT '{DEFAULT_SCHEDULE_DAYS}'")
        cur.execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS shift_end TEXT DEFAULT '{DEFAULT_SHIFT_END}'")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS schedule_preset_id INTEGER")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS admin_permissions TEXT")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS admin_role_preset TEXT")
        cur.execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS break_window_start TEXT DEFAULT '{DEFAULT_BREAK_WINDOW_START}'")
        cur.execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS break_window_end TEXT DEFAULT '{DEFAULT_BREAK_WINDOW_END}'")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS break_limit_minutes INTEGER NOT NULL DEFAULT 15")
        ensure_employee_barcode_unique_index_postgres(cur)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS schedule_presets (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                department_scope TEXT,
                schedule_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri',
                shift_start TEXT NOT NULL DEFAULT '09:00',
                shift_end TEXT NOT NULL DEFAULT '18:00',
                break_limit_minutes INTEGER NOT NULL DEFAULT 15,
                notes TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE schedule_presets ADD COLUMN IF NOT EXISTS department_scope TEXT")
        cur.execute("ALTER TABLE schedule_presets ADD COLUMN IF NOT EXISTS schedule_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri'")
        cur.execute("ALTER TABLE schedule_presets ADD COLUMN IF NOT EXISTS shift_start TEXT NOT NULL DEFAULT '09:00'")
        cur.execute("ALTER TABLE schedule_presets ADD COLUMN IF NOT EXISTS shift_end TEXT NOT NULL DEFAULT '18:00'")
        cur.execute("ALTER TABLE schedule_presets ADD COLUMN IF NOT EXISTS break_limit_minutes INTEGER NOT NULL DEFAULT 15")
        cur.execute("ALTER TABLE schedule_presets ADD COLUMN IF NOT EXISTS notes TEXT")
        cur.execute("ALTER TABLE schedule_presets ADD COLUMN IF NOT EXISTS created_by INTEGER")
        cur.execute("ALTER TABLE schedule_presets ADD COLUMN IF NOT EXISTS created_at TEXT")
        cur.execute("ALTER TABLE schedule_presets ADD COLUMN IF NOT EXISTS updated_at TEXT")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS employee_schedule_history (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                effective_at TEXT NOT NULL,
                department TEXT,
                position TEXT,
                schedule_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri',
                shift_start TEXT NOT NULL DEFAULT '09:00',
                shift_end TEXT NOT NULL DEFAULT '18:00',
                break_limit_minutes INTEGER NOT NULL DEFAULT 15,
                schedule_preset_id INTEGER,
                created_by INTEGER,
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE employee_schedule_history ADD COLUMN IF NOT EXISTS department TEXT")
        cur.execute("ALTER TABLE employee_schedule_history ADD COLUMN IF NOT EXISTS position TEXT")
        cur.execute("ALTER TABLE employee_schedule_history ADD COLUMN IF NOT EXISTS schedule_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri'")
        cur.execute("ALTER TABLE employee_schedule_history ADD COLUMN IF NOT EXISTS shift_start TEXT NOT NULL DEFAULT '09:00'")
        cur.execute("ALTER TABLE employee_schedule_history ADD COLUMN IF NOT EXISTS shift_end TEXT NOT NULL DEFAULT '18:00'")
        cur.execute("ALTER TABLE employee_schedule_history ADD COLUMN IF NOT EXISTS break_limit_minutes INTEGER NOT NULL DEFAULT 15")
        cur.execute("ALTER TABLE employee_schedule_history ADD COLUMN IF NOT EXISTS schedule_preset_id INTEGER")
        cur.execute("ALTER TABLE employee_schedule_history ADD COLUMN IF NOT EXISTS created_by INTEGER")
        cur.execute("ALTER TABLE employee_schedule_history ADD COLUMN IF NOT EXISTS created_at TEXT")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS employee_future_schedule_changes (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                effective_date TEXT NOT NULL,
                department TEXT,
                position TEXT,
                schedule_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri',
                shift_start TEXT NOT NULL DEFAULT '09:00',
                shift_end TEXT NOT NULL DEFAULT '18:00',
                break_limit_minutes INTEGER NOT NULL DEFAULT 15,
                schedule_preset_id INTEGER,
                notes TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL,
                applied_at TEXT,
                applied_by INTEGER
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS schedule_special_dates (
                id SERIAL PRIMARY KEY,
                special_date TEXT NOT NULL UNIQUE,
                rule_type TEXT NOT NULL DEFAULT 'holiday',
                label TEXT NOT NULL,
                notes TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS department TEXT")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS position TEXT")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS schedule_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri'")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS shift_start TEXT NOT NULL DEFAULT '09:00'")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS shift_end TEXT NOT NULL DEFAULT '18:00'")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS break_limit_minutes INTEGER NOT NULL DEFAULT 15")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS schedule_preset_id INTEGER")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS notes TEXT")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS created_by INTEGER")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS created_at TEXT")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS applied_at TEXT")
        cur.execute("ALTER TABLE employee_future_schedule_changes ADD COLUMN IF NOT EXISTS applied_by INTEGER")
        cur.execute("ALTER TABLE schedule_special_dates ADD COLUMN IF NOT EXISTS special_date TEXT")
        cur.execute("ALTER TABLE schedule_special_dates ADD COLUMN IF NOT EXISTS rule_type TEXT NOT NULL DEFAULT 'holiday'")
        cur.execute("ALTER TABLE schedule_special_dates ADD COLUMN IF NOT EXISTS label TEXT")
        cur.execute("ALTER TABLE schedule_special_dates ADD COLUMN IF NOT EXISTS notes TEXT")
        cur.execute("ALTER TABLE schedule_special_dates ADD COLUMN IF NOT EXISTS created_by INTEGER")
        cur.execute("ALTER TABLE schedule_special_dates ADD COLUMN IF NOT EXISTS created_at TEXT")
        cur.execute("ALTER TABLE schedule_special_dates ADD COLUMN IF NOT EXISTS updated_at TEXT")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS attendance (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                work_date TEXT NOT NULL,
                time_in TEXT,
                time_out TEXT,
                status TEXT DEFAULT 'Offline',
                proof_file TEXT,
                notes TEXT,
                late_flag INTEGER NOT NULL DEFAULT 0,
                late_minutes INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS breaks (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                attendance_id INTEGER,
                work_date TEXT NOT NULL,
                break_start TEXT,
                break_end TEXT,
                created_at TEXT NOT NULL
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS notifications (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                message TEXT NOT NULL,
                created_at TEXT NOT NULL,
                is_read INTEGER NOT NULL DEFAULT 0
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS login_attempts (
                id SERIAL PRIMARY KEY,
                ip_address TEXT NOT NULL,
                attempted_at TEXT NOT NULL
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS activity_logs (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                target_user_id INTEGER,
                action TEXT NOT NULL,
                details TEXT,
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE activity_logs ADD COLUMN IF NOT EXISTS target_user_id INTEGER")
        cur.execute("UPDATE activity_logs SET target_user_id = user_id WHERE target_user_id IS NULL")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS scanner_logs (
                id SERIAL PRIMARY KEY,
                scanner_user_id INTEGER,
                employee_user_id INTEGER,
                action_type TEXT,
                barcode_value TEXT,
                result_status TEXT NOT NULL,
                result_message TEXT,
                source_label TEXT,
                device_label TEXT,
                ip_address TEXT,
                user_agent TEXT,
                scanner_name_snapshot TEXT,
                scanner_username_snapshot TEXT,
                employee_name_snapshot TEXT,
                employee_department_snapshot TEXT,
                employee_position_snapshot TEXT,
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS scanner_user_id INTEGER")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS employee_user_id INTEGER")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS action_type TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS barcode_value TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS result_status TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS result_message TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS source_label TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS device_label TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS ip_address TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS user_agent TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS scanner_name_snapshot TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS scanner_username_snapshot TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS employee_name_snapshot TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS employee_department_snapshot TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS employee_position_snapshot TEXT")
        cur.execute("ALTER TABLE scanner_logs ADD COLUMN IF NOT EXISTS created_at TEXT")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS correction_requests (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                request_type TEXT NOT NULL,
                work_date TEXT NOT NULL,
                end_work_date TEXT,
                message TEXT,
                requested_time_in TEXT,
                requested_break_start TEXT,
                requested_break_end TEXT,
                requested_time_out TEXT,
                applied_changes TEXT,
                status TEXT NOT NULL DEFAULT 'Pending',
                admin_note TEXT,
                reviewed_by INTEGER,
                reviewed_at TEXT,
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE correction_requests ADD COLUMN IF NOT EXISTS end_work_date TEXT")
        cur.execute("ALTER TABLE correction_requests ADD COLUMN IF NOT EXISTS requested_time_in TEXT")
        cur.execute("ALTER TABLE correction_requests ADD COLUMN IF NOT EXISTS requested_break_start TEXT")
        cur.execute("ALTER TABLE correction_requests ADD COLUMN IF NOT EXISTS requested_break_end TEXT")
        cur.execute("ALTER TABLE correction_requests ADD COLUMN IF NOT EXISTS requested_time_out TEXT")
        cur.execute("ALTER TABLE correction_requests ADD COLUMN IF NOT EXISTS applied_changes TEXT")
        cur.execute("ALTER TABLE correction_requests ADD COLUMN IF NOT EXISTS status TEXT NOT NULL DEFAULT 'Pending'")
        cur.execute("ALTER TABLE correction_requests ADD COLUMN IF NOT EXISTS admin_note TEXT")
        cur.execute("ALTER TABLE correction_requests ADD COLUMN IF NOT EXISTS reviewed_by INTEGER")
        cur.execute("ALTER TABLE correction_requests ADD COLUMN IF NOT EXISTS reviewed_at TEXT")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS company_settings (
                id INTEGER PRIMARY KEY,
                id_signatory_name TEXT,
                id_signatory_title TEXT,
                id_signature_file TEXT,
                hr_signatory_name TEXT,
                hr_signatory_title TEXT,
                hr_signature_file TEXT,
                scanner_attendance_mode INTEGER NOT NULL DEFAULT 0,
                scanner_lock_timeout_seconds INTEGER NOT NULL DEFAULT 90,
                scanner_exit_pin_hash TEXT,
                overtime_multiplier REAL NOT NULL DEFAULT 1.25,
                last_external_backup_at TEXT,
                last_external_backup_by INTEGER,
                last_external_backup_note TEXT
            )
        """)
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS id_signatory_name TEXT")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS id_signatory_title TEXT")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS id_signature_file TEXT")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS hr_signatory_name TEXT")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS hr_signatory_title TEXT")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS hr_signature_file TEXT")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS scanner_attendance_mode INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS scanner_lock_timeout_seconds INTEGER NOT NULL DEFAULT 90")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS scanner_exit_pin_hash TEXT")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS overtime_multiplier REAL NOT NULL DEFAULT 1.25")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS last_external_backup_at TEXT")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS last_external_backup_by INTEGER")
        cur.execute("ALTER TABLE company_settings ADD COLUMN IF NOT EXISTS last_external_backup_note TEXT")
        cur.execute("""
            INSERT INTO company_settings (
                id, id_signatory_name, id_signatory_title, id_signature_file,
                scanner_attendance_mode, scanner_lock_timeout_seconds, scanner_exit_pin_hash, overtime_multiplier
            )
            VALUES (1, 'Kirk Danny Fernandez', 'Head Of Operations', NULL, 0, 90, NULL, 1.25)
            ON CONFLICT (id) DO NOTHING
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS overtime_sessions (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                attendance_id INTEGER,
                work_date TEXT NOT NULL,
                overtime_start TEXT NOT NULL,
                overtime_end TEXT,
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE overtime_sessions ADD COLUMN IF NOT EXISTS user_id INTEGER")
        cur.execute("ALTER TABLE overtime_sessions ADD COLUMN IF NOT EXISTS attendance_id INTEGER")
        cur.execute("ALTER TABLE overtime_sessions ADD COLUMN IF NOT EXISTS work_date TEXT")
        cur.execute("ALTER TABLE overtime_sessions ADD COLUMN IF NOT EXISTS overtime_start TEXT")
        cur.execute("ALTER TABLE overtime_sessions ADD COLUMN IF NOT EXISTS overtime_end TEXT")
        cur.execute("ALTER TABLE overtime_sessions ADD COLUMN IF NOT EXISTS created_at TEXT")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS payroll_adjustments (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                date_from TEXT NOT NULL,
                date_to TEXT NOT NULL,
                adjustment_type TEXT NOT NULL,
                label TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                notes TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE payroll_adjustments ADD COLUMN IF NOT EXISTS user_id INTEGER")
        cur.execute("ALTER TABLE payroll_adjustments ADD COLUMN IF NOT EXISTS date_from TEXT")
        cur.execute("ALTER TABLE payroll_adjustments ADD COLUMN IF NOT EXISTS date_to TEXT")
        cur.execute("ALTER TABLE payroll_adjustments ADD COLUMN IF NOT EXISTS adjustment_type TEXT")
        cur.execute("ALTER TABLE payroll_adjustments ADD COLUMN IF NOT EXISTS label TEXT")
        cur.execute("ALTER TABLE payroll_adjustments ADD COLUMN IF NOT EXISTS amount REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_adjustments ADD COLUMN IF NOT EXISTS notes TEXT")
        cur.execute("ALTER TABLE payroll_adjustments ADD COLUMN IF NOT EXISTS created_by INTEGER")
        cur.execute("ALTER TABLE payroll_adjustments ADD COLUMN IF NOT EXISTS created_at TEXT")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS payroll_recurring_rules (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                adjustment_type TEXT NOT NULL,
                label TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                recurrence_type TEXT NOT NULL DEFAULT 'Every Payroll',
                start_date TEXT,
                end_date TEXT,
                notes TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_by INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS user_id INTEGER")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS adjustment_type TEXT")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS label TEXT")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS amount REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS recurrence_type TEXT NOT NULL DEFAULT 'Every Payroll'")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS start_date TEXT")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS end_date TEXT")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS notes TEXT")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS is_active INTEGER NOT NULL DEFAULT 1")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS created_by INTEGER")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS created_at TEXT")
        cur.execute("ALTER TABLE payroll_recurring_rules ADD COLUMN IF NOT EXISTS updated_at TEXT")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS payroll_runs (
                id SERIAL PRIMARY KEY,
                date_from TEXT NOT NULL,
                date_to TEXT NOT NULL,
                department_filter TEXT,
                employee_filter TEXT,
                status TEXT NOT NULL DEFAULT 'Draft',
                notes TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                released_at TEXT
            )
        """)
        cur.execute("ALTER TABLE payroll_runs ADD COLUMN IF NOT EXISTS date_from TEXT")
        cur.execute("ALTER TABLE payroll_runs ADD COLUMN IF NOT EXISTS date_to TEXT")
        cur.execute("ALTER TABLE payroll_runs ADD COLUMN IF NOT EXISTS department_filter TEXT")
        cur.execute("ALTER TABLE payroll_runs ADD COLUMN IF NOT EXISTS employee_filter TEXT")
        cur.execute("ALTER TABLE payroll_runs ADD COLUMN IF NOT EXISTS status TEXT NOT NULL DEFAULT 'Draft'")
        cur.execute("ALTER TABLE payroll_runs ADD COLUMN IF NOT EXISTS notes TEXT")
        cur.execute("ALTER TABLE payroll_runs ADD COLUMN IF NOT EXISTS created_by INTEGER")
        cur.execute("ALTER TABLE payroll_runs ADD COLUMN IF NOT EXISTS created_at TEXT")
        cur.execute("ALTER TABLE payroll_runs ADD COLUMN IF NOT EXISTS updated_at TEXT")
        cur.execute("ALTER TABLE payroll_runs ADD COLUMN IF NOT EXISTS released_at TEXT")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS payroll_run_items (
                id SERIAL PRIMARY KEY,
                payroll_run_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                full_name TEXT,
                department TEXT,
                position TEXT,
                hourly_rate REAL NOT NULL DEFAULT 0,
                days_worked INTEGER NOT NULL DEFAULT 0,
                total_hours REAL NOT NULL DEFAULT 0,
                overtime_hours REAL NOT NULL DEFAULT 0,
                late_minutes INTEGER NOT NULL DEFAULT 0,
                break_minutes INTEGER NOT NULL DEFAULT 0,
                suspension_days INTEGER NOT NULL DEFAULT 0,
                suspension_pay REAL NOT NULL DEFAULT 0,
                gross_pay REAL NOT NULL DEFAULT 0,
                overtime_pay REAL NOT NULL DEFAULT 0,
                allowances REAL NOT NULL DEFAULT 0,
                deductions REAL NOT NULL DEFAULT 0,
                final_pay REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS payroll_run_id INTEGER")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS user_id INTEGER")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS full_name TEXT")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS department TEXT")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS position TEXT")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS hourly_rate REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS days_worked INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS total_hours REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS overtime_hours REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS late_minutes INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS break_minutes INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS suspension_days INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS suspension_pay REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS gross_pay REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS overtime_pay REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS allowances REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS deductions REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS final_pay REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_items ADD COLUMN IF NOT EXISTS created_at TEXT")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS payroll_run_item_adjustments (
                id SERIAL PRIMARY KEY,
                payroll_run_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                source_kind TEXT,
                source_rule_id INTEGER,
                recurrence_type TEXT,
                adjustment_type TEXT NOT NULL,
                label TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                notes TEXT,
                created_by INTEGER,
                created_by_name TEXT,
                adjustment_created_at TEXT,
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS payroll_run_id INTEGER")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS user_id INTEGER")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS source_kind TEXT")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS source_rule_id INTEGER")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS recurrence_type TEXT")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS adjustment_type TEXT")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS label TEXT")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS amount REAL NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS notes TEXT")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS created_by INTEGER")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS created_by_name TEXT")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS adjustment_created_at TEXT")
        cur.execute("ALTER TABLE payroll_run_item_adjustments ADD COLUMN IF NOT EXISTS created_at TEXT")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_login_attempts_ip_attempted_at ON login_attempts(ip_address, attempted_at)")
        try:
            cur.execute("SAVEPOINT attendance_open_index")
            cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_attendance_one_open_per_user ON attendance(user_id) WHERE time_in IS NOT NULL AND time_out IS NULL")
            cur.execute("RELEASE SAVEPOINT attendance_open_index")
        except Exception:
            cur.execute("ROLLBACK TO SAVEPOINT attendance_open_index")
            cur.execute("RELEASE SAVEPOINT attendance_open_index")
        try:
            cur.execute("SAVEPOINT break_open_index")
            cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_breaks_one_open_per_attendance ON breaks(attendance_id) WHERE attendance_id IS NOT NULL AND break_end IS NULL")
            cur.execute("RELEASE SAVEPOINT break_open_index")
        except Exception:
            cur.execute("ROLLBACK TO SAVEPOINT break_open_index")
            cur.execute("RELEASE SAVEPOINT break_open_index")
        try:
            cur.execute("SAVEPOINT overtime_open_index")
            cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_overtime_one_open_per_user ON overtime_sessions(user_id) WHERE overtime_end IS NULL")
            cur.execute("RELEASE SAVEPOINT overtime_open_index")
        except Exception:
            cur.execute("ROLLBACK TO SAVEPOINT overtime_open_index")
            cur.execute("RELEASE SAVEPOINT overtime_open_index")

    db.commit()

    admin = fetchone("SELECT * FROM users WHERE username = ?", ("admin",))
    if not admin:
        bootstrap_password = get_bootstrap_admin_password()
        if is_production_environment() and not bootstrap_password:
            raise RuntimeError(
                "BOOTSTRAP_ADMIN_PASSWORD must be set before first production startup when the admin account does not exist."
            )
        execute_db("""
            INSERT INTO users (
                full_name, username, password_hash, role,
                profile_image, department, position, shift_start, break_limit_minutes, is_active, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "Administrator",
            "admin",
            generate_password_hash(bootstrap_password or "admin123"),
            "admin",
            None,
            "Stellar Seats",
            "Administrator",
            DEFAULT_SHIFT_START,
            BREAK_LIMIT_MINUTES,
            1,
            now_str()
        ), commit=True)


os.makedirs(os.path.dirname(SQLITE_DATABASE), exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# =========================
# APP HELPERS
# =========================
def create_notification(user_id, title, message):
    execute_db("""
        INSERT INTO notifications (user_id, title, message, created_at, is_read)
        VALUES (?, ?, ?, ?, 0)
    """, (user_id, title, message, now_str()), commit=True)


def normalize_incident_error_type(error_type, new_error_type=""):
    raw_value = (error_type or "").strip()
    custom_value = " ".join((new_error_type or "").strip().split())
    if raw_value == "__new__":
        return custom_value
    return " ".join(raw_value.split())


def get_incident_error_type_options():
    saved_types = [
        row["error_type"]
        for row in fetchall("""
            SELECT DISTINCT TRIM(error_type) AS error_type
            FROM incident_reports
            WHERE error_type IS NOT NULL AND TRIM(error_type) != ''
            ORDER BY TRIM(error_type)
        """)
    ]
    options = []
    seen = set()
    for error_type in [*DEFAULT_INCIDENT_ERROR_TYPES, *saved_types]:
        normalized = " ".join((error_type or "").strip().split())
        key = normalized.lower()
        if not normalized or key in seen:
            continue
        seen.add(key)
        options.append(normalized)
    return options


def count_repeated_incidents(user_id, error_type):
    normalized_error_type = (error_type or "").strip().lower()
    if not user_id or not normalized_error_type:
        return 0
    row = fetchone("""
        SELECT COUNT(*) AS cnt
        FROM incident_reports
        WHERE user_id = ?
          AND LOWER(TRIM(COALESCE(error_type, ''))) = ?
    """, (user_id, normalized_error_type))
    return int(row["cnt"] or 0) if row else 0


def get_incident_policy_action(incident_count, *, exact_threshold=True):
    try:
        count = int(incident_count or 0)
    except (TypeError, ValueError):
        count = 0
    if exact_threshold:
        return INCIDENT_DISCIPLINARY_POLICY.get(count, "")
    if count >= 5:
        return INCIDENT_DISCIPLINARY_POLICY[5]
    return INCIDENT_DISCIPLINARY_POLICY.get(count, "")


def build_incident_policy_details(report, incident_count, policy_action):
    error_type = report["error_type"] if report else ""
    report_date = report["report_date"] or report["incident_date"] if report else ""
    message = (report["message"] or "").strip() if report else ""
    parts = [
        f"Policy trigger: {incident_count} repeated {error_type} incident(s).",
        f"Source incident #{report['id']} dated {report_date}.",
    ]
    if policy_action == "Termination":
        parts.append("This logs the termination policy step only; process final HR/account action separately.")
    if message:
        parts.append(f"Incident details: {message}")
    return " ".join(parts)


def create_incident(user_id, error_type, report_date, message, admin_id, incident_action, report_department):
    user = get_user_by_id(user_id)
    created_at = now_str()

    execute_db("""
        INSERT INTO incident_reports (
            user_id,
            employee_name,
            report_department,
            error_type,
            incident_action,
            incident_date,
            report_date,
            message,
            status,
            admin_note,
            created_by,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Open', NULL, ?, ?)
    """, (
        user_id,
        user["full_name"] if user else "",
        report_department,
        error_type,
        incident_action,
        report_date,
        report_date,
        message,
        admin_id,
        created_at
    ), commit=True)
    return fetchone("""
        SELECT *
        FROM incident_reports
        WHERE user_id = ? AND created_by = ? AND created_at = ?
        ORDER BY id DESC
        LIMIT 1
    """, (user_id, admin_id, created_at))


def create_disciplinary_action(user_id, action_type, action_date, details, created_by, duration_days=1, incident_report_id=None, error_type=""):
    duration_days = max(int(duration_days or 1), 1)
    end_date = calculate_suspension_end_date(action_date, duration_days) if action_type == "Suspension" else action_date
    created_at = now_str()
    execute_db("""
        INSERT INTO disciplinary_actions (
            user_id, action_type, action_date, duration_days, end_date, details,
            incident_report_id, error_type, created_by, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        user_id,
        action_type,
        action_date,
        duration_days,
        end_date,
        details,
        incident_report_id,
        error_type,
        created_by,
        created_at
    ), commit=True)
    return fetchone("""
        SELECT *
        FROM disciplinary_actions
        WHERE user_id = ? AND created_by = ? AND created_at = ?
        ORDER BY id DESC
        LIMIT 1
    """, (user_id, created_by, created_at))


def sync_incident_policy(report_id, actor_id, allow_create=True):
    report = fetchone("""
        SELECT r.*, u.full_name, u.department
        FROM incident_reports r
        LEFT JOIN users u ON u.id = r.user_id
        WHERE r.id = ?
    """, (report_id,))
    if not report:
        return {"created_action": None, "policy_action": "", "incident_count": 0, "message": "Incident not found."}

    incident_count = count_repeated_incidents(report["user_id"], report["error_type"])
    policy_action = get_incident_policy_action(incident_count, exact_threshold=True)
    display_policy_action = policy_action or get_incident_policy_action(incident_count, exact_threshold=False)
    execute_db("""
        UPDATE incident_reports
        SET policy_incident_count = ?, incident_action = ?
        WHERE id = ?
    """, (incident_count, display_policy_action, report_id), commit=True)

    if not allow_create or not policy_action:
        return {
            "created_action": None,
            "policy_action": display_policy_action,
            "incident_count": incident_count,
            "message": "",
        }

    if report["disciplinary_action_id"]:
        return {
            "created_action": get_disciplinary_action_by_id(report["disciplinary_action_id"]),
            "policy_action": policy_action,
            "incident_count": incident_count,
            "message": "Incident already has a linked disciplinary action.",
        }

    action_date = report["report_date"] or report["incident_date"] or today_str()
    duration_days = 1
    conflict = find_conflicting_disciplinary_action(report["user_id"], policy_action, action_date, duration_days)
    if conflict:
        return {
            "created_action": None,
            "policy_action": policy_action,
            "incident_count": incident_count,
            "message": conflict["conflict_reason"],
        }

    created_action = create_disciplinary_action(
        user_id=report["user_id"],
        action_type=policy_action,
        action_date=action_date,
        details=build_incident_policy_details(report, incident_count, policy_action),
        created_by=actor_id,
        duration_days=duration_days,
        incident_report_id=report_id,
        error_type=report["error_type"],
    )
    if created_action:
        execute_db("""
            UPDATE incident_reports
            SET disciplinary_action_id = ?
            WHERE id = ?
        """, (created_action["id"], report_id), commit=True)
        create_notification(
            report["user_id"],
            "Incident Policy Step Logged",
            f"{policy_action} was logged after {incident_count} repeated {report['error_type']} incident(s)."
        )
    return {
        "created_action": created_action,
        "policy_action": policy_action,
        "incident_count": incident_count,
        "message": "",
    }


def get_disciplinary_action_by_id(action_id):
    return fetchone("""
        SELECT *
        FROM disciplinary_actions
        WHERE id = ?
    """, (action_id,))


def log_activity(user_id, action, details="", target_user_id=None):
    target_user_id = target_user_id or user_id
    execute_db("""
        INSERT INTO activity_logs (user_id, target_user_id, action, details, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (user_id, target_user_id, action, details, now_str()), commit=True)


def log_scanner_activity(scanner_user_id, action_type, barcode_value, result_status, result_message,
                         employee_user_id=None, source_label="Scanner kiosk", device_label="Tablet camera kiosk",
                         ip_address=None, user_agent=None, scanner_name_snapshot=None,
                         scanner_username_snapshot=None, employee_name_snapshot=None,
                         employee_department_snapshot=None, employee_position_snapshot=None):
    execute_db("""
        INSERT INTO scanner_logs (
            scanner_user_id, employee_user_id, action_type, barcode_value, result_status,
            result_message, source_label, device_label, ip_address, user_agent,
            scanner_name_snapshot, scanner_username_snapshot, employee_name_snapshot,
            employee_department_snapshot, employee_position_snapshot, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        scanner_user_id,
        employee_user_id,
        action_type,
        barcode_value,
        result_status,
        result_message,
        source_label,
        device_label,
        ip_address,
        user_agent,
        scanner_name_snapshot,
        scanner_username_snapshot,
        employee_name_snapshot,
        employee_department_snapshot,
        employee_position_snapshot,
        now_str()
    ), commit=True)


def find_employee_barcode_matches(barcode_id):
    cleaned = (barcode_id or "").strip()
    result = {
        "cleaned": cleaned,
        "matches": [],
        "is_duplicate": False,
        "match_type": "none",
    }
    if not cleaned:
        return result

    direct_matches = fetchall("""
        SELECT *
        FROM users
        WHERE role = 'employee' AND TRIM(COALESCE(barcode_id, '')) = ?
        ORDER BY id ASC
        LIMIT 2
    """, (cleaned,))
    if direct_matches:
        result["matches"] = direct_matches
        result["is_duplicate"] = len(direct_matches) > 1
        result["match_type"] = "barcode"
        return result

    if cleaned.upper().startswith("EMP-"):
        suffix = cleaned[4:].strip()
        if suffix.isdigit():
            employee = fetchone("""
                SELECT *
                FROM users
                WHERE role = 'employee' AND id = ?
                ORDER BY id DESC LIMIT 1
            """, (int(suffix),))
            if employee:
                result["matches"] = [employee]
                result["match_type"] = "employee_id"
    return result


def get_user_by_id(user_id):
    return fetchone("SELECT * FROM users WHERE id = ?", (user_id,))


def get_user_by_barcode(barcode_id):
    lookup = find_employee_barcode_matches(barcode_id)
    return lookup["matches"][0] if len(lookup["matches"]) == 1 else None


def get_scanner_log_select_expressions(log_alias="sl", scanner_alias="scanner", employee_alias="employee"):
    scanner_name_expr = f"{scanner_alias}.full_name"
    scanner_username_expr = f"{scanner_alias}.username"
    employee_name_expr = f"{employee_alias}.full_name"
    employee_department_expr = f"{employee_alias}.department"
    employee_position_expr = f"{employee_alias}.position"

    if column_exists("scanner_logs", "scanner_name_snapshot"):
        scanner_name_expr = f"COALESCE({log_alias}.scanner_name_snapshot, {scanner_name_expr})"
    if column_exists("scanner_logs", "scanner_username_snapshot"):
        scanner_username_expr = f"COALESCE({log_alias}.scanner_username_snapshot, {scanner_username_expr})"
    if column_exists("scanner_logs", "employee_name_snapshot"):
        employee_name_expr = f"COALESCE({log_alias}.employee_name_snapshot, {employee_name_expr})"
    if column_exists("scanner_logs", "employee_department_snapshot"):
        employee_department_expr = f"COALESCE({log_alias}.employee_department_snapshot, {employee_department_expr})"
    if column_exists("scanner_logs", "employee_position_snapshot"):
        employee_position_expr = f"COALESCE({log_alias}.employee_position_snapshot, {employee_position_expr})"

    return {
        "scanner_name": scanner_name_expr,
        "scanner_username": scanner_username_expr,
        "employee_name": employee_name_expr,
        "employee_department": employee_department_expr,
        "employee_position": employee_position_expr,
    }


def get_company_settings():
    settings = fetchone("SELECT * FROM company_settings WHERE id = 1")
    if settings:
        return dict(settings)
    return {
        "id": 1,
        "id_signatory_name": "Kirk Danny Fernandez",
        "id_signatory_title": "Head Of Operations",
        "id_signature_file": None,
        "hr_signatory_name": "",
        "hr_signatory_title": "Human Resources Manager",
        "hr_signature_file": None,
        "scanner_attendance_mode": 0,
        "scanner_lock_timeout_seconds": 90,
        "scanner_exit_pin_hash": None,
        "overtime_multiplier": 1.25,
        "last_external_backup_at": None,
        "last_external_backup_by": None,
        "last_external_backup_note": None,
    }


def admin_can_permission(permission_code):
    if session.get("role") != "admin" or not session.get("user_id"):
        return False
    return admin_has_permission(get_user_by_id(session["user_id"]), permission_code)


@app.before_request
def enforce_admin_endpoint_permissions():
    if session.get("role") != "admin" or not session.get("user_id"):
        return None
    permission_code = ADMIN_ENDPOINT_PERMISSIONS.get(request.endpoint or "")
    if not permission_code:
        return None
    admin_user = get_user_by_id(session["user_id"])
    if admin_has_permission(admin_user, permission_code):
        return None
    flash(f"Your admin account does not have access to {ADMIN_PERMISSION_LABELS.get(permission_code, 'this section')}.", "danger")
    return redirect(get_home_route_for_user(admin_user))


@app.before_request
def apply_pending_schedule_changes_before_request():
    try:
        apply_due_future_schedule_changes()
    except Exception:
        app.logger.exception("Failed to apply pending employee schedule changes.")
    return None


def get_admin_accounts():
    return fetchall("""
        SELECT *
        FROM users
        WHERE role = 'admin'
        ORDER BY id ASC
    """)


def invalidate_reports_cache():
    _reports_cache.clear()


def invalidate_option_caches():
    for key in _options_cache.keys():
        _options_cache[key]["stamp"] = 0
        _options_cache[key]["rows"] = []


def invalidate_schedule_change_apply_state():
    _schedule_change_apply_state["stamp"] = 0
    _schedule_change_apply_state["date"] = ""


def should_run_admin_notification_scan(cache_key, ttl_seconds=ADMIN_ALERT_SCAN_TTL_SECONDS):
    cache_entry = _admin_notification_scan_state.setdefault(cache_key, {"stamp": 0})
    cache_age = now_timestamp() - int(cache_entry.get("stamp") or 0)
    if cache_age <= ttl_seconds:
        return False
    cache_entry["stamp"] = now_timestamp()
    return True


def get_cached_rows(cache_key, ttl_seconds, builder):
    cache_entry = _options_cache.setdefault(cache_key, {"stamp": 0, "rows": []})
    cache_age = now_timestamp() - int(cache_entry.get("stamp") or 0)
    if cache_age > ttl_seconds or not cache_entry.get("rows"):
        cache_entry["rows"] = list(builder())
        cache_entry["stamp"] = now_timestamp()
    return cache_entry["rows"]


def get_schedule_presets(department_filter=""):
    sql = """
        SELECT sp.*, creator.full_name AS created_by_name
        FROM schedule_presets sp
        LEFT JOIN users creator ON creator.id = sp.created_by
        WHERE 1 = 1
    """
    params = []
    if department_filter:
        sql += " AND (COALESCE(sp.department_scope, '') = '' OR sp.department_scope = ?)"
        params.append(department_filter)
    sql += " ORDER BY sp.name ASC"
    return fetchall(sql, params)


def invalidate_schedule_special_rule_cache():
    _options_cache["schedule_special_dates"] = {"stamp": 0, "rows": []}


def enrich_schedule_special_rule(row):
    item = dict(row)
    item["rule_type"] = normalize_schedule_special_rule_type(item.get("rule_type"))
    item["rule_type_label"] = SCHEDULE_SPECIAL_RULE_LABELS[item["rule_type"]]
    item["label"] = build_schedule_special_rule_label(item["rule_type"], item.get("label"))
    item["notes"] = (item.get("notes") or "").strip()
    item["display_label"] = item["label"]
    effective_date = parse_iso_date(item.get("special_date"))
    if effective_date:
        item["date_label"] = effective_date.strftime("%b %d, %Y")
        day_delta = (effective_date - now_dt().date()).days
        item["days_until"] = day_delta
        if day_delta == 0:
            item["countdown_label"] = "Today"
        elif day_delta == 1:
            item["countdown_label"] = "Tomorrow"
        elif day_delta > 1:
            item["countdown_label"] = f"In {day_delta} days"
        else:
            item["countdown_label"] = f"{abs(day_delta)} day(s) ago"
    else:
        item["date_label"] = item.get("special_date") or ""
        item["days_until"] = None
        item["countdown_label"] = ""
    return item


def get_schedule_special_dates(include_past=False, limit=60):
    rows = get_cached_rows(
        "schedule_special_dates",
        OPTION_CACHE_TTL_SECONDS,
        lambda: [dict(row) for row in fetchall("""
            SELECT ssd.*, creator.full_name AS created_by_name
            FROM schedule_special_dates ssd
            LEFT JOIN users creator ON creator.id = ssd.created_by
            ORDER BY ssd.special_date ASC, ssd.id ASC
        """)]
    )
    today_value = today_str()
    filtered = []
    for row in rows:
        if not include_past and row.get("special_date") and row["special_date"] < today_value:
            continue
        filtered.append(enrich_schedule_special_rule(row))
    if limit:
        return filtered[:max(int(limit or 1), 1)]
    return filtered


def get_schedule_special_rule_for_date(date_str):
    lookup = None
    if has_app_context():
        lookup = getattr(g, "_schedule_special_rule_lookup", None)
        if lookup is None:
            lookup = {row["special_date"]: row for row in get_schedule_special_dates(include_past=True, limit=0)}
            g._schedule_special_rule_lookup = lookup
    else:
        lookup = {row["special_date"]: row for row in get_schedule_special_dates(include_past=True, limit=0)}
    return lookup.get(date_str)


def get_schedule_special_rule_map(date_from="", date_to=""):
    rows = get_schedule_special_dates(include_past=True, limit=0)
    result = {}
    for row in rows:
        special_date = row.get("special_date") or ""
        if date_from and special_date < date_from:
            continue
        if date_to and special_date > date_to:
            continue
        result[special_date] = row
    return result


def get_schedule_preset(preset_id):
    raw_value = str(preset_id or "").strip()
    if not raw_value.isdigit():
        return None
    return fetchone("""
        SELECT *
        FROM schedule_presets
        WHERE id = ?
    """, (int(raw_value),))


def resolve_schedule_assignment(form, fallback_user=None):
    preset = get_schedule_preset(form.get("schedule_preset_id", ""))
    if preset:
        return {
            "schedule_preset_id": preset["id"],
            "schedule_days": normalize_schedule_days(get_schedule_day_codes(preset["schedule_days"] or DEFAULT_SCHEDULE_DAYS)),
            "shift_start": parse_shift_start(preset["shift_start"] or DEFAULT_SHIFT_START),
            "shift_end": parse_shift_end(preset["shift_end"] or DEFAULT_SHIFT_END),
            "break_limit_minutes": parse_break_limit_minutes(preset["break_limit_minutes"] if preset["break_limit_minutes"] is not None else BREAK_LIMIT_MINUTES),
            "schedule_source_label": preset["name"],
        }

    default_shift_start = fallback_user["shift_start"] if fallback_user else DEFAULT_SHIFT_START
    default_shift_end = fallback_user["shift_end"] if fallback_user else DEFAULT_SHIFT_END
    default_break_limit = fallback_user["break_limit_minutes"] if fallback_user and fallback_user["break_limit_minutes"] is not None else BREAK_LIMIT_MINUTES
    return {
        "schedule_preset_id": None,
        "schedule_days": normalize_schedule_days(form.getlist("schedule_days")),
        "shift_start": parse_shift_start(form.get("shift_start", default_shift_start)),
        "shift_end": parse_shift_end(form.get("shift_end", default_shift_end)),
        "break_limit_minutes": parse_break_limit_minutes(form.get("break_limit_minutes", default_break_limit)),
        "schedule_source_label": "Custom Manual Schedule",
    }


def apply_schedule_history_snapshot(base_user, history_row):
    if not base_user:
        return None
    context_user = dict(base_user)
    if not history_row:
        return context_user
    history_row = dict(history_row)
    context_user["department"] = history_row.get("department") if history_row.get("department") is not None else context_user.get("department")
    context_user["position"] = history_row.get("position") if history_row.get("position") is not None else context_user.get("position")
    context_user["schedule_days"] = history_row.get("schedule_days") or context_user.get("schedule_days") or DEFAULT_SCHEDULE_DAYS
    context_user["shift_start"] = history_row.get("shift_start") or context_user.get("shift_start") or DEFAULT_SHIFT_START
    context_user["shift_end"] = history_row.get("shift_end") or context_user.get("shift_end") or DEFAULT_SHIFT_END
    if history_row.get("break_limit_minutes") is not None:
        context_user["break_limit_minutes"] = history_row.get("break_limit_minutes")
    if "schedule_preset_id" in history_row.keys():
        context_user["schedule_preset_id"] = history_row.get("schedule_preset_id")
    return context_user


def get_employee_schedule_history_row(user_id, reference_datetime=None, reference_date=""):
    ref_text = normalize_history_reference(reference_datetime=reference_datetime, reference_date=reference_date)
    return fetchone("""
        SELECT *
        FROM employee_schedule_history
        WHERE user_id = ?
          AND effective_at <= ?
        ORDER BY effective_at DESC, id DESC
        LIMIT 1
    """, (user_id, ref_text))


def get_effective_employee_context(user_row=None, user_id=None, reference_datetime=None, reference_date=""):
    base_user = None
    if user_row:
        base_user = dict(user_row)
    elif user_id:
        fetched = get_user_by_id(user_id)
        base_user = dict(fetched) if fetched else None

    if not base_user:
        return None

    history_row = get_employee_schedule_history_row(
        base_user["id"],
        reference_datetime=reference_datetime,
        reference_date=reference_date
    )
    return apply_schedule_history_snapshot(base_user, history_row)


def get_effective_employee_context_map(user_rows, reference_datetime=None, reference_date=""):
    rows = [dict(row) for row in (user_rows or [])]
    if not rows or not table_exists("employee_schedule_history"):
        return {int(row["id"]): dict(row) for row in rows}

    user_ids = []
    seen_ids = set()
    for row in rows:
        row_id = int(row["id"])
        if row_id in seen_ids:
            continue
        seen_ids.add(row_id)
        user_ids.append(row_id)

    placeholders = ", ".join(["?"] * len(user_ids))
    history_rows = fetchall(f"""
        SELECT *
        FROM employee_schedule_history
        WHERE user_id IN ({placeholders})
          AND effective_at <= ?
        ORDER BY user_id ASC, effective_at DESC, id DESC
    """, tuple(user_ids) + (normalize_history_reference(reference_datetime=reference_datetime, reference_date=reference_date),))

    history_lookup = {}
    for row in history_rows:
        row_dict = dict(row)
        history_lookup.setdefault(int(row_dict["user_id"]), row_dict)

    return {
        int(row["id"]): apply_schedule_history_snapshot(row, history_lookup.get(int(row["id"])))
        for row in rows
    }


def record_employee_schedule_history(user_row, actor_id=None, effective_at=None, commit=False):
    if not user_row:
        return
    user_row = dict(user_row)
    if user_row.get("role") != "employee":
        return

    snapshot = {
        "department": (user_row.get("department") or "").strip() or None,
        "position": (user_row.get("position") or "").strip() or None,
        "schedule_days": normalize_schedule_days(user_row.get("schedule_days") or DEFAULT_SCHEDULE_DAYS),
        "shift_start": parse_shift_start(user_row.get("shift_start") or DEFAULT_SHIFT_START),
        "shift_end": parse_shift_end(user_row.get("shift_end") or DEFAULT_SHIFT_END),
        "break_limit_minutes": parse_break_limit_minutes(
            user_row.get("break_limit_minutes") if user_row.get("break_limit_minutes") is not None else BREAK_LIMIT_MINUTES
        ),
        "schedule_preset_id": user_row.get("schedule_preset_id"),
    }
    history_effective_at = normalize_history_reference(reference_datetime=effective_at)
    latest_row = fetchone("""
        SELECT *
        FROM employee_schedule_history
        WHERE user_id = ?
        ORDER BY effective_at DESC, id DESC
        LIMIT 1
    """, (user_row["id"],))
    if latest_row:
        latest_row = dict(latest_row)
        latest_snapshot = {
            "department": (latest_row.get("department") or "").strip() or None,
            "position": (latest_row.get("position") or "").strip() or None,
            "schedule_days": normalize_schedule_days(latest_row.get("schedule_days") or DEFAULT_SCHEDULE_DAYS),
            "shift_start": parse_shift_start(latest_row.get("shift_start") or DEFAULT_SHIFT_START),
            "shift_end": parse_shift_end(latest_row.get("shift_end") or DEFAULT_SHIFT_END),
            "break_limit_minutes": parse_break_limit_minutes(
                latest_row.get("break_limit_minutes") if latest_row.get("break_limit_minutes") is not None else BREAK_LIMIT_MINUTES
            ),
            "schedule_preset_id": latest_row.get("schedule_preset_id"),
        }
        if latest_snapshot == snapshot:
            return

    execute_db("""
        INSERT INTO employee_schedule_history (
            user_id, effective_at, department, position, schedule_days, shift_start,
            shift_end, break_limit_minutes, schedule_preset_id, created_by, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        user_row["id"],
        history_effective_at,
        snapshot["department"],
        snapshot["position"],
        snapshot["schedule_days"],
        snapshot["shift_start"],
        snapshot["shift_end"],
        snapshot["break_limit_minutes"],
        snapshot["schedule_preset_id"],
        actor_id,
        now_str()
    ), commit=commit)


def ensure_employee_schedule_history_seeded():
    if not table_exists("employee_schedule_history"):
        return
    rows = fetchall("""
        SELECT *
        FROM users
        WHERE role = 'employee'
        ORDER BY id ASC
    """)
    for row in rows:
        row = dict(row)
        existing_history = fetchone("""
            SELECT id
            FROM employee_schedule_history
            WHERE user_id = ?
            LIMIT 1
        """, (row["id"],))
        if existing_history:
            continue
        record_employee_schedule_history(
            row,
            actor_id=None,
            effective_at=row.get("created_at") or now_str()
        )
    db = g.get("db")
    if db is not None:
        db.commit()


def get_recent_employee_schedule_history(user_id, limit=6):
    raw_limit = max(int(limit or 1), 1)
    rows = fetchall("""
        SELECT h.*, preset.name AS preset_name, creator.full_name AS created_by_name
        FROM employee_schedule_history h
        LEFT JOIN schedule_presets preset ON preset.id = h.schedule_preset_id
        LEFT JOIN users creator ON creator.id = h.created_by
        WHERE h.user_id = ?
        ORDER BY h.effective_at DESC, h.id DESC
        LIMIT ?
    """, (user_id, raw_limit))
    items = []
    for row in rows:
        item = dict(row)
        item["schedule_summary"] = get_schedule_summary(item.get("schedule_days") or DEFAULT_SCHEDULE_DAYS)
        item["window_summary"] = f"{item.get('shift_start') or DEFAULT_SHIFT_START} - {item.get('shift_end') or DEFAULT_SHIFT_END}"
        items.append(item)
    return items


def get_future_schedule_changes(user_id=None, include_applied=False, limit=50):
    sql = """
        SELECT fsc.*, u.full_name, u.department AS current_department, u.position AS current_position,
               preset.name AS preset_name, creator.full_name AS created_by_name, applier.full_name AS applied_by_name
        FROM employee_future_schedule_changes fsc
        JOIN users u ON u.id = fsc.user_id
        LEFT JOIN schedule_presets preset ON preset.id = fsc.schedule_preset_id
        LEFT JOIN users creator ON creator.id = fsc.created_by
        LEFT JOIN users applier ON applier.id = fsc.applied_by
        WHERE u.role = 'employee'
    """
    params = []
    if user_id:
        sql += " AND fsc.user_id = ?"
        params.append(int(user_id))
    if not include_applied:
        sql += " AND fsc.applied_at IS NULL"
    sql += " ORDER BY fsc.effective_date ASC, fsc.id ASC"
    if limit:
        sql += " LIMIT ?"
        params.append(max(int(limit or 1), 1))

    items = []
    for row in fetchall(sql, tuple(params)):
        item = dict(row)
        item["schedule_summary"] = get_schedule_summary(item.get("schedule_days") or DEFAULT_SCHEDULE_DAYS)
        item["window_summary"] = f"{item.get('shift_start') or DEFAULT_SHIFT_START} - {item.get('shift_end') or DEFAULT_SHIFT_END}"
        item["effective_label"] = item.get("effective_date") or ""
        effective_date_value = parse_iso_date(item.get("effective_date"))
        if effective_date_value:
            day_delta = (effective_date_value - now_dt().date()).days
            item["days_until_effective"] = day_delta
            if day_delta == 0:
                item["countdown_label"] = "Applies today"
            elif day_delta == 1:
                item["countdown_label"] = "Applies tomorrow"
            elif day_delta > 1:
                item["countdown_label"] = f"Applies in {day_delta} days"
            else:
                item["countdown_label"] = f"Effective {abs(day_delta)} day(s) ago"
        else:
            item["days_until_effective"] = None
            item["countdown_label"] = ""
        item["department_label"] = item.get("department") or item.get("current_department") or "Unassigned"
        item["position_label"] = item.get("position") or item.get("current_position") or "Employee"
        items.append(item)
    return items


def build_future_schedule_change_map(user_ids):
    cleaned_ids = []
    seen_ids = set()
    for raw_id in user_ids or []:
        try:
            parsed = int(raw_id)
        except Exception:
            continue
        if parsed in seen_ids:
            continue
        seen_ids.add(parsed)
        cleaned_ids.append(parsed)
    if not cleaned_ids:
        return {}

    placeholders = ", ".join(["?"] * len(cleaned_ids))
    rows = fetchall(f"""
        SELECT fsc.*, preset.name AS preset_name
        FROM employee_future_schedule_changes fsc
        LEFT JOIN schedule_presets preset ON preset.id = fsc.schedule_preset_id
        WHERE fsc.applied_at IS NULL
          AND fsc.user_id IN ({placeholders})
        ORDER BY fsc.user_id ASC, fsc.effective_date ASC, fsc.id ASC
    """, tuple(cleaned_ids))

    change_map = {}
    for row in rows:
        item = dict(row)
        item["schedule_summary"] = get_schedule_summary(item.get("schedule_days") or DEFAULT_SCHEDULE_DAYS)
        item["window_summary"] = f"{item.get('shift_start') or DEFAULT_SHIFT_START} - {item.get('shift_end') or DEFAULT_SHIFT_END}"
        effective_date_value = parse_iso_date(item.get("effective_date"))
        if effective_date_value:
            day_delta = (effective_date_value - now_dt().date()).days
            if day_delta == 0:
                item["countdown_label"] = "Today"
            elif day_delta == 1:
                item["countdown_label"] = "Tomorrow"
            elif day_delta > 1:
                item["countdown_label"] = f"In {day_delta} days"
            else:
                item["countdown_label"] = f"{abs(day_delta)} day(s) ago"
        else:
            item["countdown_label"] = ""
        change_map.setdefault(int(item["user_id"]), item)
    return change_map


def queue_future_schedule_change(user_row, schedule_assignment, effective_date, actor_id=None, notes="", department=None, position=None, commit=False):
    if not user_row:
        raise ValueError("Employee not found.")
    user_row = dict(user_row)
    if user_row.get("role") != "employee":
        raise ValueError("Only employee schedules can be queued.")

    effective_date_value = parse_iso_date(effective_date)
    if not effective_date_value:
        raise ValueError("Choose a valid effective date for the future schedule change.")
    if effective_date_value <= now_dt().date():
        raise ValueError("Future schedule changes must use a date after today.")

    normalized_assignment = {
        "schedule_preset_id": schedule_assignment.get("schedule_preset_id"),
        "schedule_days": normalize_schedule_days(schedule_assignment.get("schedule_days") or DEFAULT_SCHEDULE_DAYS),
        "shift_start": parse_shift_start(schedule_assignment.get("shift_start") or DEFAULT_SHIFT_START),
        "shift_end": parse_shift_end(schedule_assignment.get("shift_end") or DEFAULT_SHIFT_END),
        "break_limit_minutes": parse_break_limit_minutes(schedule_assignment.get("break_limit_minutes") if schedule_assignment.get("break_limit_minutes") is not None else BREAK_LIMIT_MINUTES),
    }
    notes_text = (notes or "").strip() or None
    department_value = (department or "").strip() or None
    position_value = (position or "").strip() or None

    existing = fetchone("""
        SELECT id
        FROM employee_future_schedule_changes
        WHERE user_id = ?
          AND effective_date = ?
          AND applied_at IS NULL
        ORDER BY id DESC
        LIMIT 1
    """, (user_row["id"], effective_date_value.strftime("%Y-%m-%d")))
    if existing:
        execute_db("""
            UPDATE employee_future_schedule_changes
            SET department = ?, position = ?, schedule_days = ?, shift_start = ?, shift_end = ?,
                break_limit_minutes = ?, schedule_preset_id = ?, notes = ?, created_by = ?, created_at = ?
            WHERE id = ?
        """, (
            department_value,
            position_value,
            normalized_assignment["schedule_days"],
            normalized_assignment["shift_start"],
            normalized_assignment["shift_end"],
            normalized_assignment["break_limit_minutes"],
            normalized_assignment["schedule_preset_id"],
            notes_text,
            actor_id,
            now_str(),
            existing["id"]
        ), commit=commit)
        invalidate_schedule_change_apply_state()
        return int(existing["id"])

    execute_db("""
        INSERT INTO employee_future_schedule_changes (
            user_id, effective_date, department, position, schedule_days, shift_start,
            shift_end, break_limit_minutes, schedule_preset_id, notes, created_by, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        user_row["id"],
        effective_date_value.strftime("%Y-%m-%d"),
        department_value,
        position_value,
        normalized_assignment["schedule_days"],
        normalized_assignment["shift_start"],
        normalized_assignment["shift_end"],
        normalized_assignment["break_limit_minutes"],
        normalized_assignment["schedule_preset_id"],
        notes_text,
        actor_id,
        now_str()
    ), commit=commit)
    invalidate_schedule_change_apply_state()
    created_row = fetchone("""
        SELECT id
        FROM employee_future_schedule_changes
        WHERE user_id = ?
          AND effective_date = ?
          AND applied_at IS NULL
        ORDER BY id DESC
        LIMIT 1
    """, (user_row["id"], effective_date_value.strftime("%Y-%m-%d")))
    return int(created_row["id"]) if created_row else 0


def apply_due_future_schedule_changes(force=False):
    if not table_exists("employee_future_schedule_changes"):
        return 0
    today_value = today_str()
    if _schedule_change_apply_state.get("date") != today_value:
        invalidate_schedule_change_apply_state()
        _schedule_change_apply_state["date"] = today_value
    if not force:
        cache_age = now_timestamp() - int(_schedule_change_apply_state.get("stamp") or 0)
        if cache_age < SCHEDULE_CHANGE_APPLY_TTL_SECONDS:
            return 0
    _schedule_change_apply_state["stamp"] = now_timestamp()
    _schedule_change_apply_state["date"] = today_value

    due_rows = [
        dict(row)
        for row in fetchall("""
            SELECT *
            FROM employee_future_schedule_changes
            WHERE applied_at IS NULL
              AND effective_date <= ?
            ORDER BY effective_date ASC, id ASC
        """, (today_value,))
    ]
    if not due_rows:
        return 0

    db = get_db()
    applied_count = 0
    try:
        for change in due_rows:
            employee = get_user_by_id(change["user_id"])
            employee = dict(employee) if employee else None
            if not employee or employee.get("role") != "employee":
                execute_db("""
                    UPDATE employee_future_schedule_changes
                    SET applied_at = ?, applied_by = ?
                    WHERE id = ?
                """, (now_str(), change.get("created_by"), change["id"]))
                continue

            updated_values = {
                "department": change.get("department") if change.get("department") is not None else employee.get("department"),
                "position": change.get("position") if change.get("position") is not None else employee.get("position"),
                "schedule_days": normalize_schedule_days(change.get("schedule_days") or employee.get("schedule_days") or DEFAULT_SCHEDULE_DAYS),
                "shift_start": parse_shift_start(change.get("shift_start") or employee.get("shift_start") or DEFAULT_SHIFT_START),
                "shift_end": parse_shift_end(change.get("shift_end") or employee.get("shift_end") or DEFAULT_SHIFT_END),
                "break_limit_minutes": parse_break_limit_minutes(change.get("break_limit_minutes") if change.get("break_limit_minutes") is not None else employee.get("break_limit_minutes")),
                "schedule_preset_id": change.get("schedule_preset_id"),
            }

            execute_db("""
                UPDATE users
                SET department = ?, position = ?, schedule_days = ?, shift_start = ?, shift_end = ?,
                    break_limit_minutes = ?, schedule_preset_id = ?
                WHERE id = ?
            """, (
                updated_values["department"],
                updated_values["position"],
                updated_values["schedule_days"],
                updated_values["shift_start"],
                updated_values["shift_end"],
                updated_values["break_limit_minutes"],
                updated_values["schedule_preset_id"],
                employee["id"]
            ))

            updated_employee = get_user_by_id(employee["id"])
            if updated_employee:
                record_employee_schedule_history(
                    updated_employee,
                    actor_id=change.get("created_by"),
                    effective_at=normalize_history_reference(reference_date=change["effective_date"])
                )
            execute_db("""
                UPDATE employee_future_schedule_changes
                SET applied_at = ?, applied_by = ?
                WHERE id = ?
            """, (now_str(), change.get("created_by"), change["id"]))
            log_activity(
                change.get("created_by") or employee["id"],
                "APPLY FUTURE SCHEDULE CHANGE",
                f"Applied scheduled shift update for {employee['full_name']} effective {change['effective_date']}.",
                target_user_id=employee["id"]
            )
            applied_count += 1
        db.commit()
    except Exception:
        db.rollback()
        raise

    if applied_count:
        invalidate_admin_employee_rows_cache()
    return applied_count


def is_scanner_attendance_mode_enabled():
    settings = get_company_settings()
    return int(settings.get("scanner_attendance_mode") or 0) == 1


def get_scanner_lock_timeout_seconds():
    settings = get_company_settings()
    raw_value = settings.get("scanner_lock_timeout_seconds", 90)
    try:
        timeout = int(raw_value or 90)
    except (TypeError, ValueError):
        timeout = 90
    return max(min(timeout, 900), 15)


def get_overtime_multiplier():
    settings = get_company_settings()
    raw_value = settings.get("overtime_multiplier", 1.25)
    try:
        multiplier = float(raw_value or 1.25)
    except (TypeError, ValueError):
        multiplier = 1.25
    return max(min(multiplier, 5.0), 1.0)


def has_scanner_exit_pin():
    settings = get_company_settings()
    return bool((settings.get("scanner_exit_pin_hash") or "").strip())


def verify_scanner_exit_pin(pin_value):
    settings = get_company_settings()
    stored_hash = (settings.get("scanner_exit_pin_hash") or "").strip()
    if not stored_hash:
        return True
    if not pin_value:
        return False
    return check_password_hash(stored_hash, pin_value)


def get_manual_attendance_block_message():
    if not is_scanner_attendance_mode_enabled():
        return ""
    return "Attendance is recorded through the company scanner kiosk. Please use the scanner station for time in, breaks, time out, and overtime."


def get_scanner_account():
    return fetchone("""
        SELECT *
        FROM users
        WHERE role = 'scanner'
        ORDER BY id ASC
        LIMIT 1
    """)


def get_attendance_by_id(attendance_id):
    return fetchone("SELECT * FROM attendance WHERE id = ?", (attendance_id,))


def get_today_attendance(user_id):
    return fetchone("""
        SELECT * FROM attendance
        WHERE user_id = ? AND work_date = ?
        ORDER BY id DESC LIMIT 1
    """, (user_id, today_str()))


def get_active_attendance(user_id):
    return fetchone("""
        SELECT * FROM attendance
        WHERE user_id = ? AND time_in IS NOT NULL AND time_out IS NULL
        ORDER BY work_date DESC, id DESC LIMIT 1
    """, (user_id,))


def get_current_attendance(user_id):
    active_attendance = get_active_attendance(user_id)
    if active_attendance:
        return active_attendance
    return get_today_attendance(user_id)


def get_latest_attendance_with_time_in(user_id):
    return fetchone("""
        SELECT *
        FROM attendance
        WHERE user_id = ?
          AND time_in IS NOT NULL
        ORDER BY work_date DESC, id DESC
        LIMIT 1
    """, (user_id,))


def get_open_break_for_attendance(attendance_id):
    return fetchone("""
        SELECT * FROM breaks
        WHERE attendance_id = ? AND break_end IS NULL
        ORDER BY id DESC LIMIT 1
    """, (attendance_id,))


def get_open_break(user_id, attendance_row=None):
    attendance = attendance_row or get_current_attendance(user_id)
    if not attendance:
        return None
    return get_open_break_for_attendance(attendance["id"])


def auto_close_stale_overtime_session(user_id, overtime_row, actor_id=None, source_label="System"):
    if not overtime_row:
        return None
    overtime_row = dict(overtime_row)
    if overtime_row.get("overtime_end"):
        return overtime_row

    work_date = ((overtime_row.get("work_date") or "") or (overtime_row.get("overtime_start") or "")[:10]).strip()
    if not work_date or work_date >= today_str():
        return overtime_row

    overtime_start_dt = parse_db_datetime(overtime_row.get("overtime_start"))
    if overtime_start_dt:
        overtime_age = now_dt().replace(tzinfo=None) - overtime_start_dt
        # Allow valid overtime to continue after midnight and only force-close sessions
        # that have clearly been abandoned well into the next day.
        if overtime_age < timedelta(hours=18):
            return overtime_row

    forced_end_dt = datetime.strptime(f"{work_date} 23:59:59", "%Y-%m-%d %H:%M:%S")
    if overtime_start_dt and forced_end_dt < overtime_start_dt:
        forced_end_dt = overtime_start_dt

    execute_db("""
        UPDATE overtime_sessions
        SET overtime_end = ?
        WHERE id = ? AND overtime_end IS NULL
    """, (forced_end_dt.strftime("%Y-%m-%d %H:%M:%S"), overtime_row["id"]), commit=True)
    invalidate_admin_employee_rows_cache()

    user = get_user_by_id(user_id)
    employee_name = user["full_name"] if user else f"User {user_id}"
    create_notification(
        user_id,
        "Overtime Auto-Closed",
        f"Your previous overtime session from {work_date} was closed automatically."
    )
    log_activity(
        actor_id or user_id,
        "AUTO CLOSE OVERTIME",
        f"{source_label} automatically closed stale overtime for {employee_name} from {work_date}.",
        target_user_id=user_id
    )
    return None


def get_open_overtime_session(user_id, auto_close_stale=True, actor_id=None, source_label="System"):
    row = fetchone("""
        SELECT *
        FROM overtime_sessions
        WHERE user_id = ? AND overtime_end IS NULL
        ORDER BY id DESC
        LIMIT 1
    """, (user_id,))
    if row and auto_close_stale:
        return auto_close_stale_overtime_session(user_id, row, actor_id=actor_id, source_label=source_label)
    return row


def get_overtime_session_by_id(session_id):
    return fetchone("SELECT * FROM overtime_sessions WHERE id = ?", (session_id,))


def get_overtime_sessions_in_range(user_id, date_from, date_to):
    return fetchall("""
        SELECT *
        FROM overtime_sessions
        WHERE user_id = ?
          AND work_date BETWEEN ? AND ?
        ORDER BY work_date ASC, id ASC
    """, (user_id, date_from, date_to))


def overtime_minutes_for_session(row):
    if not row:
        return 0
    row = dict(row)
    if not row.get("overtime_start") or not row.get("overtime_end"):
        return 0
    try:
        start = datetime.strptime(row["overtime_start"], "%Y-%m-%d %H:%M:%S")
        end = datetime.strptime(row["overtime_end"], "%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        return 0
    return max(int((end - start).total_seconds() // 60), 0)


def auto_close_stale_attendance(user_row, attendance_row, actor_id=None, source_label="System"):
    if not user_row or not attendance_row:
        return attendance_row
    if not attendance_row["time_in"] or attendance_row["time_out"]:
        return attendance_row

    effective_user = get_effective_employee_context(
        user_row=user_row,
        reference_datetime=get_attendance_reference_datetime(attendance_row),
        reference_date=attendance_row["work_date"]
    )
    _, shift_end_dt = get_shift_bounds_for_work_date(effective_user, attendance_row["work_date"])
    stale_cutoff_dt = shift_end_dt + timedelta(minutes=LATE_GRACE_MINUTES)
    if now_dt() < stale_cutoff_dt:
        return attendance_row

    time_in_dt = parse_db_datetime(attendance_row["time_in"])
    forced_time_out_dt = shift_end_dt.replace(tzinfo=None)
    if time_in_dt and forced_time_out_dt < time_in_dt:
        forced_time_out_dt = time_in_dt
    forced_time_out = forced_time_out_dt.strftime("%Y-%m-%d %H:%M:%S")

    open_break = get_open_break_for_attendance(attendance_row["id"])
    if open_break:
        forced_break_end_dt = forced_time_out_dt
        break_start_dt = parse_db_datetime(open_break["break_start"])
        if break_start_dt and forced_break_end_dt < break_start_dt:
            forced_break_end_dt = break_start_dt
        execute_db("""
            UPDATE breaks
            SET break_end = ?
            WHERE id = ?
        """, (forced_break_end_dt.strftime("%Y-%m-%d %H:%M:%S"), open_break["id"]), commit=True)

    execute_db("""
        UPDATE attendance
        SET time_out = ?, status = ?, updated_at = ?
        WHERE id = ?
    """, (forced_time_out, "Timed Out", now_str(), attendance_row["id"]), commit=True)
    invalidate_admin_employee_rows_cache()

    log_activity(
        actor_id or user_row["id"],
        "AUTO CLOSE STALE ATTENDANCE",
        f"{source_label} auto-closed stale attendance for {user_row['full_name']} dated {attendance_row['work_date']}",
        target_user_id=user_row["id"]
    )
    return get_attendance_by_id(attendance_row["id"])


def get_attendance_override_block_message(override_status, action_key, attendance_row=None):
    if not override_status:
        return None

    attendance_time_in = None
    attendance_time_out = None
    if attendance_row:
        if hasattr(attendance_row, "keys"):
            keys = attendance_row.keys()
            attendance_time_in = attendance_row["time_in"] if "time_in" in keys else None
            attendance_time_out = attendance_row["time_out"] if "time_out" in keys else None
        else:
            attendance_time_in = attendance_row.get("time_in")
            attendance_time_out = attendance_row.get("time_out")
    has_open_attendance = bool(attendance_time_in and not attendance_time_out)
    if has_open_attendance and action_key in {"end_break", "time_out"}:
        return None

    if hasattr(override_status, "keys"):
        override_end_date = override_status["end_date"] if "end_date" in override_status.keys() else None
        override_type = override_status["type"]
    else:
        override_end_date = override_status.get("end_date")
        override_type = override_status["type"]
    end_date = override_end_date or today_str()
    if override_type == "Suspension":
        return f"Employee is suspended until {end_date}."
    return f"Employee is on {override_type} until {end_date}."


def perform_attendance_action(user_id, action_type, actor_id=None, source_label="System"):
    user = get_user_by_id(user_id)
    if not user or user["role"] != "employee":
        return False, "Employee not found.", None

    if user["is_active"] != 1:
        return False, "Employee account is inactive.", user

    attendance = get_current_attendance(user_id)
    attendance = auto_close_stale_attendance(user, attendance, actor_id=actor_id, source_label=source_label)
    open_break = get_open_break(user_id, attendance)
    action_key = (action_type or "").strip().lower()
    override_status = get_employee_override_status_for_date(user_id, today_str())
    override_block_message = get_attendance_override_block_message(override_status, action_key, attendance)
    if override_block_message:
        return False, override_block_message, user

    if action_key == "time_in":
        if attendance and attendance["time_in"] and not attendance["time_out"]:
            return False, "Employee is already timed in.", user

        execute_db("""
            INSERT INTO attendance (
                user_id, work_date, time_in, time_out, status, proof_file, notes,
                late_flag, late_minutes, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user_id,
            today_str(),
            now_str(),
            None,
            "Timed In",
            None,
            f"{source_label} action",
            *calculate_late_info(now_str(), parse_shift_start(user["shift_start"] if user else DEFAULT_SHIFT_START)),
            now_str(),
            now_str()
        ), commit=True)

        latest_attendance = get_current_attendance(user_id)
        shift_start = parse_shift_start(user["shift_start"] if user else DEFAULT_SHIFT_START)
        if latest_attendance and latest_attendance["late_flag"]:
            create_notification(
                user_id,
                "Late Time-In",
                f"You timed in late by {latest_attendance['late_minutes']} minute(s). Shift start: {shift_start} ET."
            )
        else:
            create_notification(user_id, "Timed In", f"You timed in at {now_str()} ET.")
        log_activity(actor_id or user_id, "KIOSK TIME IN", f"{source_label} time in for {user['full_name']}")
        invalidate_admin_employee_rows_cache()
        return True, "Time in successful.", user

    if action_key == "start_break":
        if not attendance or not attendance["time_in"] or attendance["time_out"]:
            return False, "Employee must be timed in first.", user
        if open_break:
            return False, "Employee is already on break.", user

        break_limit_minutes = get_employee_break_limit(user)
        used_break_minutes = total_break_minutes(attendance["id"])
        if used_break_minutes >= break_limit_minutes:
            return False, f"Daily break limit already used ({break_limit_minutes} minutes).", user

        execute_db("""
            INSERT INTO breaks (user_id, attendance_id, work_date, break_start, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (user_id, attendance["id"], attendance["work_date"], now_str(), now_str()), commit=True)
        execute_db("""
            UPDATE attendance
            SET status = ?, updated_at = ?
            WHERE id = ?
        """, ("On Break", now_str(), attendance["id"]), commit=True)
        remaining_break = max(break_limit_minutes - used_break_minutes, 0)
        create_notification(user_id, "Break Started", f"You started break at {now_str()} ET. Remaining break allowance: {remaining_break} minute(s).")
        log_activity(actor_id or user_id, "KIOSK BREAK START", f"{source_label} break start for {user['full_name']}")
        invalidate_admin_employee_rows_cache()
        return True, "Break started.", user

    if action_key == "end_break":
        if not open_break:
            return False, "No active break found.", user

        execute_db("""
            UPDATE breaks
            SET break_end = ?
            WHERE id = ?
        """, (now_str(), open_break["id"]), commit=True)

        if attendance:
            execute_db("""
                UPDATE attendance
                SET status = ?, updated_at = ?
                WHERE id = ?
            """, ("Timed In", now_str(), attendance["id"]), commit=True)

        total_break = total_break_minutes(attendance["id"]) if attendance else 0
        break_limit_minutes = get_employee_break_limit(user)
        create_notification(user_id, "Break Ended", f"You ended break at {now_str()} ET.")
        if attendance and is_overbreak(total_break, break_limit_minutes):
            create_notification(
                user_id,
                "Break Limit Exceeded",
                f"Your total break time for today is {minutes_to_hm(total_break)}, which is over your {break_limit_minutes} minute limit."
            )
        log_activity(actor_id or user_id, "KIOSK BREAK END", f"{source_label} break end for {user['full_name']}")
        invalidate_admin_employee_rows_cache()
        return True, "Break ended.", user

    if action_key == "time_out":
        if not attendance or not attendance["time_in"]:
            return False, "Employee is not timed in.", user
        if attendance["time_out"]:
            recorded_time_out_dt = parse_db_datetime(attendance["time_out"])
            _, shift_end_dt = get_shift_bounds_for_work_date(user, attendance["work_date"])
            shift_end_naive = shift_end_dt.replace(tzinfo=None)
            if recorded_time_out_dt and recorded_time_out_dt == shift_end_naive and now_dt().replace(tzinfo=None) > shift_end_naive:
                return False, "Regular shift already closed at the scheduled end. Use Overtime Start for work after midnight.", user
            return False, "Employee is already timed out.", user

        if open_break:
            execute_db("""
                UPDATE breaks
                SET break_end = ?
                WHERE id = ?
            """, (now_str(), open_break["id"]), commit=True)

        execute_db("""
            UPDATE attendance
            SET time_out = ?, status = ?, updated_at = ?
            WHERE id = ?
        """, (now_str(), "Timed Out", now_str(), attendance["id"]), commit=True)

        updated_attendance = get_attendance_by_id(attendance["id"])
        total_break = total_break_minutes(attendance["id"]) if attendance else 0
        break_limit_minutes = get_employee_break_limit(user)
        ok, msg = append_attendance_to_google_sheet(user, updated_attendance)
        create_notification(user_id, "Timed Out", f"You timed out at {now_str()} ET.")
        if attendance and is_overbreak(total_break, break_limit_minutes):
            create_notification(
                user_id,
                "Break Limit Exceeded",
                f"Your total break time for today is {minutes_to_hm(total_break)}, which is over your {break_limit_minutes} minute limit."
            )
        log_activity(actor_id or user_id, "KIOSK TIME OUT", f"{source_label} time out for {user['full_name']}. Sheets sync: {msg if ok else 'Skipped/Failed'}")
        invalidate_admin_employee_rows_cache()
        return True, "Time out successful.", user

    if action_key == "overtime_start":
        open_overtime = get_open_overtime_session(user_id, actor_id=actor_id, source_label=source_label)
        if open_overtime:
            return False, "Employee already has an active overtime session.", user

        latest_reference = get_active_attendance(user_id) or get_latest_attendance_with_time_in(user_id)
        if not latest_reference or not latest_reference["time_in"]:
            return False, "Employee must complete the regular shift before starting overtime.", user

        latest_reference_was_active = not latest_reference.get("time_out")
        latest_reference = auto_close_stale_attendance(user, latest_reference, actor_id=actor_id, source_label=source_label)
        if not latest_reference or not latest_reference["time_out"]:
            return False, "Employee must complete the regular shift before starting overtime.", user

        effective_user = get_effective_employee_context(
            user_row=user,
            reference_datetime=get_attendance_reference_datetime(latest_reference),
            reference_date=latest_reference["work_date"]
        )
        _, shift_end_dt = get_shift_bounds_for_work_date(effective_user, latest_reference["work_date"])
        shift_end_naive = shift_end_dt.replace(tzinfo=None)
        regular_time_out_dt = parse_db_datetime(latest_reference["time_out"])

        if not latest_reference_was_active and regular_time_out_dt and regular_time_out_dt < shift_end_naive:
            return False, "Regular shift must reach the scheduled end before overtime can start.", user

        overtime_start_dt = regular_time_out_dt or shift_end_naive

        if overtime_start_dt < shift_end_naive:
            overtime_start_dt = shift_end_naive

        if now_dt().replace(tzinfo=None) < shift_end_naive:
            return False, "Overtime can only start after the regular shift ends.", user

        try:
            execute_db("""
                INSERT INTO overtime_sessions (user_id, attendance_id, work_date, overtime_start, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, (
                user_id,
                latest_reference["id"],
                overtime_start_dt.strftime("%Y-%m-%d"),
                overtime_start_dt.strftime("%Y-%m-%d %H:%M:%S"),
                now_str()
            ), commit=True)
        except sqlite3.IntegrityError:
            return False, "Employee already has an active overtime session.", user
        except Exception as exc:
            if POSTGRES_ENABLED and exc.__class__.__name__ == "IntegrityError":
                return False, "Employee already has an active overtime session.", user
            raise
        create_notification(
            user_id,
            "Overtime Started",
            f"Overtime started at {overtime_start_dt.strftime('%Y-%m-%d %H:%M:%S')} ET."
        )
        log_activity(
            actor_id or user_id,
            "KIOSK OVERTIME START",
            f"{source_label} overtime start for {user['full_name']} after regular shift ending {shift_end_naive.strftime('%Y-%m-%d %H:%M:%S')}",
            target_user_id=user_id
        )
        invalidate_admin_employee_rows_cache()
        return True, "Overtime started.", user

    if action_key == "overtime_end":
        open_overtime = get_open_overtime_session(user_id, actor_id=actor_id, source_label=source_label)
        if not open_overtime:
            return False, "No active overtime session found.", user

        execute_db("""
            UPDATE overtime_sessions
            SET overtime_end = ?
            WHERE id = ?
        """, (now_str(), open_overtime["id"]), commit=True)
        closed_session = get_overtime_session_by_id(open_overtime["id"])
        overtime_minutes = overtime_minutes_for_session(closed_session)
        create_notification(user_id, "Overtime Ended", f"Overtime ended at {now_str()} ET.")
        log_activity(actor_id or user_id, "KIOSK OVERTIME END", f"{source_label} overtime end for {user['full_name']} ({minutes_to_hm(overtime_minutes)})", target_user_id=user_id)
        invalidate_admin_employee_rows_cache()
        return True, "Overtime ended.", user

    return False, "Invalid attendance action.", user


def get_user_live_status(user_id):
    if get_open_overtime_session(user_id):
        return "On Overtime"

    attendance = get_current_attendance(user_id)
    open_break = get_open_break(user_id, attendance)

    if not attendance:
        return "Offline"

    if attendance["time_in"] and not attendance["time_out"]:
        if open_break:
            return "On Break"
        return "Timed In"

    if attendance["time_out"]:
        return "Timed Out"

    return "Offline"


def get_schedule_window_summary(user_row, reference_datetime=None, reference_date=""):
    if not user_row:
        return f"{DEFAULT_SHIFT_START} - {DEFAULT_SHIFT_END}"

    effective_user = get_effective_employee_context(
        user_row=user_row,
        reference_datetime=reference_datetime,
        reference_date=reference_date
    )
    shift_start = parse_shift_start(effective_user["shift_start"] if effective_user["shift_start"] else DEFAULT_SHIFT_START)
    shift_end = parse_shift_end(effective_user["shift_end"] if effective_user["shift_end"] else DEFAULT_SHIFT_END)
    return f"{shift_start} - {shift_end}"


def get_today_schedule_code():
    return WEEKDAY_OPTIONS[now_dt().weekday()][0]


def is_scheduled_on_date(user_row, date_str):
    special_rule = get_schedule_special_rule_for_date(date_str)
    if special_rule and special_rule.get("rule_type") in {"holiday", "rest_day"}:
        return False
    return get_schedule_code_for_date(date_str) in get_schedule_day_codes(
        user_row["schedule_days"] if user_row else DEFAULT_SCHEDULE_DAYS
    )


def is_scheduled_today(user_row):
    return is_scheduled_on_date(user_row, today_str())


def get_approved_leave_for_date(user_id, work_date):
    if not user_id or not work_date:
        return None
    return fetchone("""
        SELECT *
        FROM correction_requests
        WHERE user_id = ?
          AND work_date <= ?
          AND COALESCE(end_work_date, work_date) >= ?
          AND status = 'Approved'
          AND request_type IN ('Sick Leave', 'Paid Leave')
        ORDER BY id DESC LIMIT 1
    """, (user_id, work_date, work_date))


def get_employee_override_status_for_date(user_id, work_date):
    suspension = get_suspension_for_date(user_id, work_date)
    if suspension:
        suspension = dict(suspension)
        return {
            "type": "Suspension",
            "label": "Suspended",
            "details": suspension.get("details") or "",
            "end_date": suspension.get("end_date") or suspension.get("action_date") or work_date,
        }
    leave = get_approved_leave_for_date(user_id, work_date)
    if leave:
        leave = dict(leave)
        return {
            "type": leave["request_type"],
            "label": leave["request_type"],
            "details": leave.get("message") or leave.get("admin_note") or "",
            "end_date": leave.get("end_work_date") or leave["work_date"],
        }
    return None


def is_absent_today(user_row, attendance_row):
    if not user_row or user_row["is_active"] != 1 or attendance_row:
        return False
    if not is_scheduled_today(user_row):
        return False
    if has_app_context():
        if get_approved_leave_for_date(user_row["id"], today_str()):
            return False
        if get_suspension_for_date(user_row["id"], today_str()):
            return False

    shift_dt, _ = get_shift_bounds_for_work_date(user_row, today_str())
    return now_dt() >= (shift_dt + timedelta(minutes=LATE_GRACE_MINUTES))


def is_missing_timeout_today(user_row, attendance_row):
    if not user_row or user_row["is_active"] != 1 or not attendance_row:
        return False
    if not attendance_row["time_in"] or attendance_row["time_out"]:
        return False

    effective_user = get_effective_employee_context(
        user_row=user_row,
        reference_datetime=get_attendance_reference_datetime(attendance_row),
        reference_date=attendance_row["work_date"]
    )
    _, shift_end_dt = get_shift_bounds_for_work_date(effective_user, attendance_row["work_date"])
    return now_dt() >= (shift_end_dt + timedelta(minutes=LATE_GRACE_MINUTES))


def is_undertime_record(user_row, attendance_row):
    if not user_row or not attendance_row:
        return False
    if not attendance_row.get("time_in") or not attendance_row.get("time_out"):
        return False
    if attendance_row.get("source_type") != "attendance":
        return False

    actual_time_in = parse_db_datetime(attendance_row.get("time_in"))
    actual_time_out = parse_db_datetime(attendance_row.get("time_out"))
    if not actual_time_in or not actual_time_out:
        return False

    effective_user = get_effective_employee_context(
        user_row=user_row,
        reference_datetime=get_attendance_reference_datetime(attendance_row),
        reference_date=attendance_row["work_date"]
    )
    shift_start_dt, shift_end_dt = get_shift_bounds_for_work_date(effective_user, attendance_row["work_date"])
    shift_start_naive = shift_start_dt.replace(tzinfo=None)
    shift_end_naive = shift_end_dt.replace(tzinfo=None)
    worked_minutes = max(int((actual_time_out - actual_time_in).total_seconds() // 60), 0)
    scheduled_minutes = max(int((shift_end_naive - shift_start_naive).total_seconds() // 60), 0)
    if worked_minutes <= 0 or scheduled_minutes <= 0:
        return False
    if worked_minutes > scheduled_minutes:
        return False
    return shift_start_naive <= actual_time_out < shift_end_naive


def get_scheduled_shift_minutes(user_row, work_date):
    if user_row and not hasattr(user_row, "get") and hasattr(user_row, "keys"):
        user_row = dict(user_row)
    effective_user = user_row
    lookup_user_id = None
    if user_row:
        if hasattr(user_row, "get"):
            if user_row.get("user_id"):
                lookup_user_id = user_row.get("user_id")
            elif user_row.get("role") == "employee" and user_row.get("id"):
                lookup_user_id = user_row.get("id")
            if lookup_user_id:
                effective_user = get_effective_employee_context(
                    user_row=user_row if user_row.get("role") == "employee" else None,
                    user_id=lookup_user_id,
                    reference_datetime=user_row.get("time_in") or user_row.get("created_at"),
                    reference_date=work_date
                ) or user_row
    shift_start_dt, shift_end_dt = get_shift_bounds_for_work_date(effective_user, work_date)
    return max(int((shift_end_dt - shift_start_dt).total_seconds() // 60), 0)


def is_suspicious_work_duration(user_row, attendance_row):
    if not user_row or not attendance_row:
        return False
    if not attendance_row.get("time_in") or not attendance_row.get("time_out"):
        return False

    time_in_dt = parse_db_datetime(attendance_row.get("time_in"))
    time_out_dt = parse_db_datetime(attendance_row.get("time_out"))
    if not time_in_dt or not time_out_dt or time_out_dt < time_in_dt:
        return True

    worked_minutes = max(int((time_out_dt - time_in_dt).total_seconds() // 60), 0)
    scheduled_minutes = get_scheduled_shift_minutes(user_row, attendance_row["work_date"])

    if worked_minutes > (18 * 60):
        return True
    if scheduled_minutes > 0 and worked_minutes > (scheduled_minutes + 240):
        return True
    return False


def collect_attendance_diagnostics(user_row, attendance_row):
    issues = []
    warnings = []
    if not user_row or not attendance_row:
        return issues, warnings

    time_in_dt = parse_db_datetime(attendance_row.get("time_in"))
    time_out_dt = parse_db_datetime(attendance_row.get("time_out"))
    break_rows = get_break_rows(attendance_row.get("id"))

    if attendance_row.get("time_in") and not time_in_dt:
        issues.append("Invalid time in format")
    if attendance_row.get("time_out") and not time_out_dt:
        issues.append("Invalid time out format")
    if time_in_dt and time_out_dt and time_out_dt < time_in_dt:
        issues.append("Time out earlier than time in")
    if is_suspicious_work_duration(user_row, attendance_row):
        issues.append("Suspicious work duration")

    for index, break_row in enumerate(break_rows, start=1):
        break_start_dt = parse_db_datetime(break_row.get("break_start"))
        break_end_dt = parse_db_datetime(break_row.get("break_end"))
        label = f"Break {index}"
        if break_row.get("break_start") and not break_start_dt:
            issues.append(f"{label} has invalid start")
        if break_row.get("break_end") and not break_end_dt:
            issues.append(f"{label} has invalid end")
        if break_start_dt and break_end_dt and break_end_dt < break_start_dt:
            issues.append(f"{label} ends before it starts")
        if time_in_dt and break_start_dt and break_start_dt < time_in_dt:
            issues.append(f"{label} starts before time in")
        if time_out_dt and break_end_dt and break_end_dt > time_out_dt:
            issues.append(f"{label} ends after time out")

    if len(break_rows) > 1:
        warnings.append(f"{len(break_rows)} break rows attached")

    return issues, warnings


def get_attendance_context(user_id, work_date):
    attendance = fetchone("""
        SELECT *
        FROM attendance
        WHERE user_id = ? AND work_date = ?
        ORDER BY id DESC LIMIT 1
    """, (user_id, work_date))

    break_row = None
    if attendance:
        break_row = fetchone("""
            SELECT *
            FROM breaks
            WHERE attendance_id = ?
            ORDER BY id ASC LIMIT 1
        """, (attendance["id"],))

    return attendance, break_row


def get_attendance_context_by_row(attendance):
    if not attendance:
        return None, None
    break_row = fetchone("""
        SELECT *
        FROM breaks
        WHERE attendance_id = ?
        ORDER BY id ASC LIMIT 1
    """, (attendance["id"],))
    return attendance, break_row


def get_matching_attendance_context_for_request(user_id, work_date, request_type="", requested_time_out=None):
    attendance, break_row = get_attendance_context(user_id, work_date)
    if request_type != "Undertime":
        return attendance, break_row
    requested_clock_time = extract_clock_time(requested_time_out)
    user_row = get_user_by_id(user_id)
    candidate_rows = fetchall("""
        SELECT *
        FROM attendance
        WHERE user_id = ?
        ORDER BY work_date DESC, id DESC
        LIMIT 10
    """, (user_id,))
    if attendance and not any(candidate["id"] == attendance["id"] for candidate in candidate_rows):
        candidate_rows.insert(0, attendance)

    best_candidate = None
    best_score = None
    for candidate in candidate_rows:
        start_dt = parse_db_datetime(candidate["time_in"])
        if not start_dt:
            continue
        candidate_time_out = combine_work_date_and_time(candidate["work_date"], requested_clock_time, not_before=candidate["time_in"]) if requested_clock_time else None
        candidate_time_out_dt = parse_db_datetime(candidate_time_out)
        if not candidate_time_out_dt:
            continue
        effective_user = get_effective_employee_context(
            user_row=user_row,
            reference_datetime=get_attendance_reference_datetime(candidate),
            reference_date=candidate["work_date"]
        )
        _, shift_end_dt = get_shift_bounds_for_work_date(effective_user, candidate["work_date"])
        shift_end_naive = shift_end_dt.replace(tzinfo=None)
        existing_time_out_dt = parse_db_datetime(candidate["time_out"]) or shift_end_naive
        if candidate_time_out_dt < start_dt or candidate_time_out_dt > shift_end_naive:
            continue
        if existing_time_out_dt and candidate_time_out_dt > existing_time_out_dt:
            continue

        score = abs(int((existing_time_out_dt - candidate_time_out_dt).total_seconds())) if existing_time_out_dt else 0
        if best_score is None or score < best_score:
            best_candidate = candidate
            best_score = score

    if best_candidate:
        return get_attendance_context_by_row(best_candidate)
    if attendance:
        return attendance, break_row
    return None, None


def calculate_late_info(time_in_str, shift_start):
    if not time_in_str:
        return 0, 0

    shift_start = parse_shift_start(shift_start)
    time_in_dt = datetime.strptime(time_in_str, "%Y-%m-%d %H:%M:%S")
    shift_dt = datetime.strptime(
        f"{time_in_dt.strftime('%Y-%m-%d')} {shift_start}:00",
        "%Y-%m-%d %H:%M:%S"
    )

    late_threshold = shift_dt + timedelta(minutes=LATE_GRACE_MINUTES)

    if time_in_dt >= late_threshold:
        late_minutes = int((time_in_dt - shift_dt).total_seconds() // 60)
        return 1, late_minutes

    return 0, 0


def total_break_minutes(attendance_id, include_open=False):
    breaks_rows = get_break_rows(attendance_id)

    total_minutes = 0
    for br in breaks_rows:
        if br["break_start"] and (br["break_end"] or include_open):
            start = datetime.strptime(br["break_start"], "%Y-%m-%d %H:%M:%S")
            end = datetime.strptime(br["break_end"] or now_str(), "%Y-%m-%d %H:%M:%S")
            total_minutes += int((end - start).total_seconds() // 60)
    return total_minutes


def get_break_rows(attendance_id):
    if not attendance_id:
        return []
    return [
        dict(row) for row in fetchall("""
            SELECT * FROM breaks
            WHERE attendance_id = ?
            ORDER BY id ASC
        """, (attendance_id,))
    ]


def build_break_sessions(attendance_id):
    sessions = []
    for br in get_break_rows(attendance_id):
        sessions.append({
            "start": br.get("break_start"),
            "end": br.get("break_end"),
            "start_display": format_datetime_12h(br.get("break_start")) if br.get("break_start") else "-",
            "end_display": format_datetime_12h(br.get("break_end")) if br.get("break_end") else "Open",
        })
    return sessions


def summarize_break_sessions(break_sessions):
    if not break_sessions:
        return "No break sessions recorded."
    parts = []
    for index, session in enumerate(break_sessions, start=1):
        parts.append(f"Break {index}: {session['start_display']} -> {session['end_display']}")
    return " | ".join(parts)


def get_backup_files(limit=15):
    if not os.path.isdir(BACKUP_FOLDER):
        return []
    entries = []
    for name in os.listdir(BACKUP_FOLDER):
        full_path = os.path.join(BACKUP_FOLDER, name)
        if not os.path.isfile(full_path):
            continue
        stat = os.stat(full_path)
        entries.append({
            "name": name,
            "path": full_path,
            "size_bytes": stat.st_size,
            "size_kb": round(stat.st_size / 1024, 1),
            "modified_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        })
    entries.sort(key=lambda item: item["modified_at"], reverse=True)
    return entries[:limit]


def path_is_inside(base_path, candidate_path):
    if not base_path or not candidate_path:
        return False
    try:
        base_real = os.path.realpath(base_path)
        candidate_real = os.path.realpath(candidate_path)
        return os.path.commonpath([base_real, candidate_real]) == base_real
    except ValueError:
        return False


def record_external_backup_marker(note="", actor_id=None):
    settings = get_company_settings()
    backup_note = (note or "").strip()
    if len(backup_note) > 240:
        backup_note = backup_note[:237] + "..."
    backup_at = now_str()
    execute_db("""
        INSERT INTO company_settings (
            id, id_signatory_name, id_signatory_title, id_signature_file,
            scanner_attendance_mode, scanner_lock_timeout_seconds, scanner_exit_pin_hash,
            overtime_multiplier, last_external_backup_at, last_external_backup_by,
            last_external_backup_note
        )
        VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            last_external_backup_at = excluded.last_external_backup_at,
            last_external_backup_by = excluded.last_external_backup_by,
            last_external_backup_note = excluded.last_external_backup_note
    """, (
        settings.get("id_signatory_name") or "Kirk Danny Fernandez",
        settings.get("id_signatory_title") or "Head Of Operations",
        settings.get("id_signature_file"),
        int(settings.get("scanner_attendance_mode") or 0),
        int(settings.get("scanner_lock_timeout_seconds") or 90),
        settings.get("scanner_exit_pin_hash"),
        float(settings.get("overtime_multiplier") or 1.25),
        backup_at,
        actor_id,
        backup_note,
    ), commit=True)
    return {
        "backup_at": backup_at,
        "note": backup_note,
    }


def get_backup_recovery_snapshot():
    backups = get_backup_files(limit=25)
    latest_backup = backups[0] if backups else None
    upload_count = 0
    if os.path.isdir(app.config["UPLOAD_FOLDER"]):
        upload_count = len([
            name for name in os.listdir(app.config["UPLOAD_FOLDER"])
            if os.path.isfile(os.path.join(app.config["UPLOAD_FOLDER"], name))
        ])

    count_specs = {
        "employee_accounts": "SELECT COUNT(*) AS cnt FROM users WHERE role = 'employee'",
        "admin_accounts": "SELECT COUNT(*) AS cnt FROM users WHERE role = 'admin'",
        "attendance_rows": "SELECT COUNT(*) AS cnt FROM attendance",
        "scanner_logs": "SELECT COUNT(*) AS cnt FROM scanner_logs",
        "payroll_runs": "SELECT COUNT(*) AS cnt FROM payroll_runs",
        "future_schedule_changes": "SELECT COUNT(*) AS cnt FROM employee_future_schedule_changes WHERE applied_at IS NULL",
        "schedule_special_rules": "SELECT COUNT(*) AS cnt FROM schedule_special_dates",
    }
    counts = {}
    for key, query in count_specs.items():
        row = fetchone(query)
        counts[key] = int((dict(row)["cnt"] if row else 0) or 0)

    settings = get_company_settings()
    backup_by_name = ""
    backup_by_id = settings.get("last_external_backup_by")
    if backup_by_id:
        try:
            backup_user = get_user_by_id(int(backup_by_id))
        except (TypeError, ValueError):
            backup_user = None
        if backup_user:
            backup_by_name = backup_user.get("full_name") or backup_user.get("username") or ""
    upload_folder = app.config["UPLOAD_FOLDER"]

    return {
        "generated_at": now_str(),
        "environment": "Postgres" if using_postgres() else "SQLite",
        "backup_supported": not using_postgres(),
        "backup_count": len(backups),
        "latest_backup": latest_backup,
        "upload_count": upload_count,
        "storage": {
            "database_label": "Render Postgres" if using_postgres() else "SQLite file",
            "upload_folder": upload_folder,
            "backup_folder": BACKUP_FOLDER,
            "persistent_disk_path": PERSISTENT_DISK_PATH or "",
            "uploads_on_persistent_disk": path_is_inside(PERSISTENT_DISK_PATH, upload_folder) if PERSISTENT_DISK_PATH else False,
        },
        "external_backup": {
            "last_at": settings.get("last_external_backup_at"),
            "last_by": backup_by_name,
            "note": settings.get("last_external_backup_note"),
            "recorded": bool(settings.get("last_external_backup_at")),
        },
        "counts": counts,
    }


def create_sqlite_backup():
    if using_postgres():
        raise ValueError("Automatic backup copy is only available for SQLite right now.")
    if not os.path.exists(SQLITE_DATABASE):
        raise ValueError("SQLite database file was not found.")
    backup_name = f"attendance-backup-{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.db"
    backup_path = os.path.join(BACKUP_FOLDER, backup_name)
    shutil.copy2(SQLITE_DATABASE, backup_path)
    return backup_path


def remove_orphaned_proof_uploads():
    if not os.path.isdir(app.config["UPLOAD_FOLDER"]):
        return 0

    protected_files = {
        row["profile_image"]
        for row in fetchall("""
            SELECT profile_image
            FROM users
            WHERE profile_image IS NOT NULL AND TRIM(profile_image) != ''
        """)
    }
    company_settings = get_company_settings()
    signatory_file = (company_settings.get("id_signature_file") or "").strip() if company_settings else ""
    if signatory_file:
        protected_files.add(signatory_file)
    removed = 0
    for name in os.listdir(app.config["UPLOAD_FOLDER"]):
        full_path = os.path.join(app.config["UPLOAD_FOLDER"], name)
        if not os.path.isfile(full_path):
            continue
        if name in protected_files:
            continue
        try:
            os.remove(full_path)
            removed += 1
        except OSError:
            continue
    return removed


def perform_go_live_reset():
    backup_path = None
    backup_supported = not using_postgres()
    if not using_postgres():
        backup_path = create_sqlite_backup()

    db = get_db()
    if using_postgres():
        with db.cursor() as cur:
            cur.execute("""
                TRUNCATE TABLE
                    breaks,
                    attendance,
                    correction_requests,
                    notifications,
                    activity_logs,
                    scanner_logs,
                    overtime_sessions,
                    payroll_adjustments,
                    payroll_run_item_adjustments,
                    payroll_run_items,
                    payroll_runs,
                    employee_future_schedule_changes,
                    login_attempts,
                    incident_reports,
                    disciplinary_actions
                RESTART IDENTITY
            """)
        db.commit()
    else:
        cur = db.cursor()
        for table_name in [
            "breaks",
            "attendance",
            "correction_requests",
            "notifications",
            "activity_logs",
            "scanner_logs",
            "overtime_sessions",
            "payroll_adjustments",
            "payroll_run_item_adjustments",
            "payroll_run_items",
            "payroll_runs",
            "employee_future_schedule_changes",
            "login_attempts",
            "incident_reports",
            "disciplinary_actions",
        ]:
            cur.execute(f"DELETE FROM {table_name}")
        try:
            cur.execute("""
                DELETE FROM sqlite_sequence
                WHERE name IN ('breaks', 'attendance', 'correction_requests', 'notifications', 'activity_logs', 'scanner_logs', 'overtime_sessions', 'payroll_adjustments', 'payroll_run_item_adjustments', 'payroll_run_items', 'payroll_runs', 'employee_future_schedule_changes', 'login_attempts', 'incident_reports', 'disciplinary_actions')
            """)
        except sqlite3.OperationalError:
            pass
        db.commit()

    removed_uploads = remove_orphaned_proof_uploads()
    invalidate_schedule_change_apply_state()
    return {
        "backup_path": backup_path,
        "backup_supported": backup_supported,
        "removed_uploads": removed_uploads,
    }


def get_log_cleanup_summary():
    summary = {
        "activity_logs": 0,
        "scanner_logs": 0,
        "login_attempts": 0,
        "read_notifications": 0,
    }
    count_queries = {
        "activity_logs": "SELECT COUNT(*) AS cnt FROM activity_logs",
        "scanner_logs": "SELECT COUNT(*) AS cnt FROM scanner_logs",
        "login_attempts": "SELECT COUNT(*) AS cnt FROM login_attempts",
        "read_notifications": "SELECT COUNT(*) AS cnt FROM notifications WHERE COALESCE(is_read, 0) = 1",
    }
    for key, query in count_queries.items():
        row = fetchone(query)
        summary[key] = int((dict(row)["cnt"] if row else 0) or 0)
    summary["total_cleanup_targets"] = (
        summary["activity_logs"]
        + summary["scanner_logs"]
        + summary["login_attempts"]
        + summary["read_notifications"]
    )
    return summary


def perform_log_retention_cleanup(retention_days):
    retention_days = int(retention_days or 0)
    if retention_days <= 0:
        raise ValueError("Retention days must be greater than zero.")

    cutoff_dt = now_dt() - timedelta(days=retention_days)
    cutoff_text = cutoff_dt.strftime("%Y-%m-%d %H:%M:%S")
    db = get_db()
    removed_counts = {
        "activity_logs": 0,
        "scanner_logs": 0,
        "login_attempts": 0,
        "read_notifications": 0,
    }

    delete_specs = [
        ("activity_logs", "DELETE FROM activity_logs WHERE created_at < ?"),
        ("scanner_logs", "DELETE FROM scanner_logs WHERE created_at < ?"),
        ("login_attempts", "DELETE FROM login_attempts WHERE attempted_at < ?"),
        ("read_notifications", "DELETE FROM notifications WHERE COALESCE(is_read, 0) = 1 AND created_at < ?"),
    ]

    try:
        if using_postgres():
            with db.cursor() as cur:
                for key, query in delete_specs:
                    cur.execute(convert_query(query), (cutoff_text,))
                    removed_counts[key] = max(cur.rowcount or 0, 0)
            db.commit()
        else:
            for key, query in delete_specs:
                cur = db.execute(query, (cutoff_text,))
                removed_counts[key] = max(cur.rowcount or 0, 0)
            db.commit()
    except Exception:
        db.rollback()
        raise

    removed_counts["cutoff_text"] = cutoff_text
    removed_counts["retention_days"] = retention_days
    removed_counts["total_removed"] = (
        removed_counts["activity_logs"]
        + removed_counts["scanner_logs"]
        + removed_counts["login_attempts"]
        + removed_counts["read_notifications"]
    )
    return removed_counts


def perform_log_cleanup_for_date_range(date_from_value, date_to_value):
    start_date = parse_iso_date(date_from_value)
    end_date = parse_iso_date(date_to_value)

    if not start_date or not end_date:
        raise ValueError("Choose both cleanup start and cleanup end dates.")

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    range_start_text = datetime.combine(start_date, datetime.min.time()).strftime("%Y-%m-%d %H:%M:%S")
    range_end_exclusive_text = datetime.combine(end_date + timedelta(days=1), datetime.min.time()).strftime("%Y-%m-%d %H:%M:%S")

    db = get_db()
    removed_counts = {
        "activity_logs": 0,
        "scanner_logs": 0,
        "login_attempts": 0,
        "read_notifications": 0,
    }

    delete_specs = [
        ("activity_logs", "DELETE FROM activity_logs WHERE created_at >= ? AND created_at < ?"),
        ("scanner_logs", "DELETE FROM scanner_logs WHERE created_at >= ? AND created_at < ?"),
        ("login_attempts", "DELETE FROM login_attempts WHERE attempted_at >= ? AND attempted_at < ?"),
        ("read_notifications", "DELETE FROM notifications WHERE COALESCE(is_read, 0) = 1 AND created_at >= ? AND created_at < ?"),
    ]

    try:
        if using_postgres():
            with db.cursor() as cur:
                for key, query in delete_specs:
                    cur.execute(convert_query(query), (range_start_text, range_end_exclusive_text))
                    removed_counts[key] = max(cur.rowcount or 0, 0)
            db.commit()
        else:
            for key, query in delete_specs:
                cur = db.execute(query, (range_start_text, range_end_exclusive_text))
                removed_counts[key] = max(cur.rowcount or 0, 0)
            db.commit()
    except Exception:
        db.rollback()
        raise

    removed_counts["date_from"] = start_date.strftime("%Y-%m-%d")
    removed_counts["date_to"] = end_date.strftime("%Y-%m-%d")
    removed_counts["total_removed"] = (
        removed_counts["activity_logs"]
        + removed_counts["scanner_logs"]
        + removed_counts["login_attempts"]
        + removed_counts["read_notifications"]
    )
    return removed_counts


def get_employee_break_limit(user_row, reference_datetime=None, reference_date=""):
    if not user_row:
        return BREAK_LIMIT_MINUTES
    if not hasattr(user_row, "get") and hasattr(user_row, "keys"):
        user_row = dict(user_row)

    effective_user = user_row
    if hasattr(user_row, "get"):
        lookup_user_id = user_row.get("user_id") or (user_row.get("id") if user_row.get("role") == "employee" else None)
        if lookup_user_id:
            effective_user = get_effective_employee_context(
                user_row=user_row if user_row.get("role") == "employee" else None,
                user_id=lookup_user_id,
                reference_datetime=reference_datetime or user_row.get("time_in") or user_row.get("created_at"),
                reference_date=reference_date or user_row.get("work_date") or today_str()
            ) or user_row
    return parse_break_limit_minutes(effective_user["break_limit_minutes"])


def save_uploaded_file(file_obj, prefix="file", allowed_exts=None):
    if not file_obj or not file_obj.filename:
        return None
    allowed_exts = set(allowed_exts or ALLOWED_EXTENSIONS)
    ext = file_obj.filename.rsplit(".", 1)[1].lower() if "." in file_obj.filename else ""
    if ext not in allowed_exts:
        return None

    safe_name = secure_filename(file_obj.filename)
    filename = f"{prefix}_{now_timestamp()}_{secrets.token_hex(6)}_{safe_name}"
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file_obj.save(filepath)
    return filename


def uploaded_file_exists(filename):
    if not filename:
        return False
    return os.path.exists(os.path.join(app.config["UPLOAD_FOLDER"], filename))


def delete_uploaded_file_if_unused(filename):
    cleaned = (filename or "").strip()
    if not cleaned:
        return
    if fetchone("""
        SELECT id
        FROM users
        WHERE profile_image = ?
        ORDER BY id DESC LIMIT 1
    """, (cleaned,)):
        return
    if fetchone("""
        SELECT id
        FROM attendance
        WHERE proof_file = ?
        ORDER BY id DESC LIMIT 1
    """, (cleaned,)):
        return

    file_path = os.path.join(app.config["UPLOAD_FOLDER"], cleaned)
    if os.path.isfile(file_path):
        try:
            os.remove(file_path)
        except OSError:
            pass


def static_file_exists(filename):
    if not filename:
        return False
    return os.path.exists(os.path.join(app.static_folder, filename))


def can_access_uploaded_file(user_row, filename):
    if not user_row or not filename:
        return False
    if user_row["role"] == "admin":
        return True
    if user_row["role"] == "scanner":
        return filename.startswith("profile_") and is_image(filename)
    if user_row["profile_image"] == filename:
        return True

    proof_row = fetchone("""
        SELECT id
        FROM attendance
        WHERE user_id = ? AND proof_file = ?
        ORDER BY id DESC LIMIT 1
    """, (user_row["id"], filename))
    return bool(proof_row)


def get_avatar_initials(name):
    parts = [part.strip() for part in (name or "").split() if part.strip()]
    if not parts:
        return "U"
    initials = "".join(part[0] for part in parts[:2]).upper()
    return initials or "U"


def get_employee_card_number(user_row):
    if not user_row:
        return ""
    barcode_value = user_row["barcode_id"] if "barcode_id" in user_row.keys() else ""
    if barcode_value and str(barcode_value).strip():
        return str(barcode_value).strip()
    user_id = user_row["id"] if "id" in user_row.keys() else 0
    return f"EMP-{int(user_id or 0):04d}"


CODE128_PATTERNS = [
    "212222", "222122", "222221", "121223", "121322", "131222", "122213", "122312", "132212", "221213",
    "221312", "231212", "112232", "122132", "122231", "113222", "123122", "123221", "223211", "221132",
    "221231", "213212", "223112", "312131", "311222", "321122", "321221", "312212", "322112", "322211",
    "212123", "212321", "232121", "111323", "131123", "131321", "112313", "132113", "132311", "211313",
    "231113", "231311", "112133", "112331", "132131", "113123", "113321", "133121", "313121", "211331",
    "231131", "213113", "213311", "213131", "311123", "311321", "331121", "312113", "312311", "332111",
    "314111", "221411", "431111", "111224", "111422", "121124", "121421", "141122", "141221", "112214",
    "112412", "122114", "122411", "142112", "142211", "241211", "221114", "413111", "241112", "134111",
    "111242", "121142", "121241", "114212", "124112", "124211", "411212", "421112", "421211", "212141",
    "214121", "412121", "111143", "111341", "131141", "114113", "114311", "411113", "411311", "113141",
    "114131", "311141", "411131", "211412", "211214", "211232", "2331112"
]


def generate_code128_svg_markup(value, module_width=2, height=88):
    raw_value = str(value or "").strip()
    if not raw_value:
        return ""
    if any(ord(ch) < 32 or ord(ch) > 126 for ch in raw_value):
        return ""

    code_values = [104] + [ord(ch) - 32 for ch in raw_value]
    checksum_total = 104
    for index, code in enumerate(code_values[1:], start=1):
        checksum_total += code * index
    code_values.append(checksum_total % 103)
    code_values.append(106)

    quiet_zone = 10 * module_width
    x = quiet_zone
    rects = []
    for code in code_values:
        pattern = CODE128_PATTERNS[code]
        for idx, width_char in enumerate(pattern):
            segment_width = int(width_char) * module_width
            if idx % 2 == 0:
                rects.append(f'<rect x="{x}" y="0" width="{segment_width}" height="{height}" fill="#0f172a" />')
            x += segment_width
    total_width = x + quiet_zone
    text_y = height + 18
    safe_label = (
        raw_value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )
    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{total_width}" height="{height + 26}" '
        f'viewBox="0 0 {total_width} {height + 26}" preserveAspectRatio="xMidYMin meet" '
        f'role="img" aria-label="Barcode {safe_label}" style="display:block;margin:24px auto;background:#ffffff;">'
        f'<rect width="{total_width}" height="{height + 26}" fill="#ffffff" rx="8" ry="8" />'
        + "".join(rects) +
        f'<text x="{total_width / 2}" y="{text_y}" text-anchor="middle" font-family="Inter, Arial, sans-serif" '
        f'font-size="14" font-weight="700" fill="#0f172a">{safe_label}</text>'
        '</svg>'
    )


def generate_code128_svg_data_uri(value, module_width=2, height=88):
    svg = generate_code128_svg_markup(value, module_width=module_width, height=height)
    if not svg:
        return ""
    return f"data:image/svg+xml;charset=utf-8,{quote(svg)}"


def get_department_options():
    return get_cached_rows(
        "departments",
        OPTION_CACHE_TTL_SECONDS,
        lambda: [
            row["department"]
            for row in fetchall("""
                SELECT DISTINCT department
                FROM users
                WHERE role = 'employee' AND department IS NOT NULL AND TRIM(department) != ''
                ORDER BY department ASC
            """)
        ]
    )


def get_employee_options():
    return get_cached_rows(
        "employees",
        OPTION_CACHE_TTL_SECONDS,
        lambda: fetchall("""
            SELECT id, full_name, department, position
            FROM users
            WHERE role = 'employee'
            ORDER BY full_name ASC
        """)
    )


def get_leave_usage_rows(user_id=None, year=None, department=""):
    target_year = int(year or now_dt().year)
    sql = """
        SELECT c.user_id, c.request_type, c.work_date, c.end_work_date, u.full_name, u.username, u.department,
               u.sick_leave_days, u.paid_leave_days
        FROM correction_requests c
        JOIN users u ON u.id = c.user_id
        WHERE c.status = 'Approved'
          AND c.request_type IN ('Sick Leave', 'Paid Leave')
    """
    params = []
    if user_id:
        sql += " AND c.user_id = ?"
        params.append(user_id)
    if department:
        sql += " AND COALESCE(u.department, '') = ?"
        params.append(department)
    sql += " ORDER BY c.work_date DESC, c.id DESC"
    rows = []
    for row in fetchall(sql, params):
        item = dict(row)
        for leave_date in expand_request_dates(item["work_date"], item.get("end_work_date")):
            parsed_date = parse_iso_date(leave_date)
            if parsed_date and parsed_date.year == target_year:
                expanded = dict(item)
                expanded["leave_date"] = leave_date
                expanded["work_date"] = leave_date
                rows.append(expanded)
    return rows


def get_request_day_count(row_or_work_date, end_work_date=""):
    if isinstance(row_or_work_date, dict):
        return len(expand_request_dates(row_or_work_date.get("work_date"), row_or_work_date.get("end_work_date")))
    return len(expand_request_dates(row_or_work_date, end_work_date))

def has_overlapping_leave_request(user_id, request_type, work_date, end_work_date, exclude_id=None):
    start_str, end_str = normalize_request_date_range(work_date, end_work_date)
    sql = """
        SELECT id
        FROM correction_requests
        WHERE user_id = ?
          AND request_type = ?
          AND status IN ('Pending', 'Approved')
          AND work_date <= ?
          AND COALESCE(end_work_date, work_date) >= ?
    """
    params = [user_id, request_type, end_str, start_str]
    if exclude_id:
        sql += " AND id != ?"
        params.append(exclude_id)
    sql += " ORDER BY id DESC LIMIT 1"
    return fetchone(sql, params)


def get_overlap_days(start_a, end_a, start_b, end_b):
    left = max(start_a, start_b)
    right = min(end_a, end_b)
    if right < left:
        return []
    total_days = (right - left).days
    return [(left + timedelta(days=offset)).strftime("%Y-%m-%d") for offset in range(total_days + 1)]


def find_conflicting_disciplinary_action(user_id, action_type, action_date, duration_days=1, exclude_id=None):
    target_start = parse_iso_date(action_date)
    target_end = parse_iso_date(
        calculate_suspension_end_date(action_date, duration_days) if action_type == "Suspension" else action_date,
        target_start
    )
    if not target_start or not target_end:
        return None

    sql = """
        SELECT *
        FROM disciplinary_actions
        WHERE user_id = ?
    """
    params = [user_id]
    if exclude_id:
        sql += " AND id != ?"
        params.append(exclude_id)
    sql += " ORDER BY action_date DESC, id DESC"

    for row in fetchall(sql, params):
        item = dict(row)
        row_start = parse_iso_date(item["action_date"])
        row_end = parse_iso_date(item.get("end_date") or item["action_date"], row_start)
        if not row_start or not row_end:
            continue
        overlapping_days = get_overlap_days(target_start, target_end, row_start, row_end)
        if item["action_type"] == "Suspension" and action_type == "Suspension" and overlapping_days:
            item["conflict_reason"] = f"Overlaps existing suspension on {format_request_date_range(overlapping_days[0], overlapping_days[-1])}."
            return item
        if item["action_date"] == action_date and item["action_type"] == action_type:
            item["conflict_reason"] = f"{item['action_type']} already exists on {action_date}."
            return item
    return None


def get_disciplinary_actions(action_type="", user_id="", department="", date_from="", date_to=""):
    sql = """
        SELECT d.*, u.full_name, u.username, u.department, u.break_limit_minutes,
               creator.full_name AS created_by_name,
               incident.error_type AS incident_error_type,
               incident.policy_incident_count AS linked_incident_count
        FROM disciplinary_actions d
        JOIN users u ON u.id = d.user_id
        LEFT JOIN users creator ON creator.id = d.created_by
        LEFT JOIN incident_reports incident ON incident.id = d.incident_report_id
        WHERE 1=1
    """
    params = []
    if action_type:
        sql += " AND d.action_type = ?"
        params.append(action_type)
    if user_id:
        sql += " AND d.user_id = ?"
        params.append(user_id)
    if department:
        sql += " AND COALESCE(u.department, '') = ?"
        params.append(department)
    if date_from:
        sql += " AND COALESCE(d.end_date, d.action_date) >= ?"
        params.append(date_from)
    if date_to:
        sql += " AND d.action_date <= ?"
        params.append(date_to)
    sql += " ORDER BY d.action_date DESC, d.id DESC"

    rows = [dict(row) for row in fetchall(sql, params)]
    today = today_str()
    for row in rows:
        if row["action_type"] == "Suspension":
            action_end = row["end_date"] or row["action_date"]
            if today < row["action_date"]:
                row["status_label"] = "Upcoming"
            elif row["action_date"] <= today <= action_end:
                row["status_label"] = "Active"
            else:
                row["status_label"] = "Completed"
        else:
            row["status_label"] = "Logged"
    return rows


def get_suspension_for_date(user_id, work_date):
    if not user_id or not work_date:
        return None
    return fetchone("""
        SELECT *
        FROM disciplinary_actions
        WHERE user_id = ?
          AND action_type = 'Suspension'
          AND action_date <= ?
          AND COALESCE(end_date, action_date) >= ?
        ORDER BY id DESC LIMIT 1
    """, (user_id, work_date, work_date))


def get_pending_leave_requests(user_id=None, department="", year=None):
    sql = """
        SELECT c.*, u.full_name, u.username, u.department
        FROM correction_requests c
        JOIN users u ON u.id = c.user_id
        WHERE c.status = 'Pending'
          AND c.request_type IN ('Sick Leave', 'Paid Leave')
    """
    params = []
    if user_id:
        sql += " AND c.user_id = ?"
        params.append(user_id)
    if department:
        sql += " AND COALESCE(u.department, '') = ?"
        params.append(department)
    if year:
        sql += " AND c.work_date <= ? AND COALESCE(c.end_work_date, c.work_date) >= ?"
        params.extend([f"{int(year)}-12-31", f"{int(year)}-01-01"])
    sql += " ORDER BY c.work_date ASC, c.id ASC"
    rows = []
    for row in fetchall(sql, params):
        item = dict(row)
        item["requested_days"] = get_request_day_count(item)
        item["display_date_range"] = format_request_date_range(item["work_date"], item.get("end_work_date"))
        try:
            created_date = datetime.strptime((item.get("created_at") or now_str())[:19], "%Y-%m-%d %H:%M:%S").date()
            item["age_days"] = max((now_dt().date() - created_date).days, 0)
        except Exception:
            item["age_days"] = 0
        rows.append(item)
    return rows


def get_leave_balance_summary(user_row, year=None):
    if not user_row:
        return {
            "year": now_dt().year,
            "sick_total": DEFAULT_SICK_LEAVE_DAYS,
            "sick_used": 0,
            "sick_remaining": DEFAULT_SICK_LEAVE_DAYS,
            "paid_total": DEFAULT_PAID_LEAVE_DAYS,
            "paid_used": 0,
            "paid_remaining": DEFAULT_PAID_LEAVE_DAYS,
        }

    target_year = int(year or now_dt().year)
    approved_rows = get_leave_usage_rows(user_id=user_row["id"], year=target_year)
    sick_used_in_app = len([row for row in approved_rows if row["request_type"] == "Sick Leave"])
    paid_used_in_app = len([row for row in approved_rows if row["request_type"] == "Paid Leave"])
    sick_used_manual = int(user_row["sick_leave_used_manual"] if user_row["sick_leave_used_manual"] is not None else 0)
    paid_used_manual = int(user_row["paid_leave_used_manual"] if user_row["paid_leave_used_manual"] is not None else 0)
    sick_total = int(user_row["sick_leave_days"] if user_row["sick_leave_days"] is not None else DEFAULT_SICK_LEAVE_DAYS)
    paid_total = int(user_row["paid_leave_days"] if user_row["paid_leave_days"] is not None else DEFAULT_PAID_LEAVE_DAYS)
    sick_used = sick_used_manual + sick_used_in_app
    paid_used = paid_used_manual + paid_used_in_app

    return {
        "year": target_year,
        "sick_total": sick_total,
        "sick_used": sick_used,
        "sick_remaining": max(sick_total - sick_used, 0),
        "paid_total": paid_total,
        "paid_used": paid_used,
        "paid_remaining": max(paid_total - paid_used, 0),
        "sick_used_manual": sick_used_manual,
        "paid_used_manual": paid_used_manual,
        "sick_used_in_app": sick_used_in_app,
        "paid_used_in_app": paid_used_in_app,
    }


def build_leave_dashboard_rows(year=None, department="", user_id=None):
    target_year = int(year or now_dt().year)
    employees_sql = """
        SELECT *
        FROM users
        WHERE role = 'employee'
    """
    params = []
    if department:
        employees_sql += " AND COALESCE(department, '') = ?"
        params.append(department)
    if user_id:
        employees_sql += " AND id = ?"
        params.append(user_id)
    employees_sql += " ORDER BY full_name ASC"
    employees = fetchall(employees_sql, params)

    approved_rows = get_leave_usage_rows(user_id=user_id, year=target_year, department=department)
    pending_rows = get_pending_leave_requests(user_id=user_id, department=department, year=target_year)
    approved_map = {}
    pending_map = {}

    for row in approved_rows:
        stats = approved_map.setdefault(row["user_id"], {"Sick Leave": 0, "Paid Leave": 0})
        stats[row["request_type"]] = stats.get(row["request_type"], 0) + 1

    for row in pending_rows:
        stats = pending_map.setdefault(row["user_id"], {"Sick Leave": 0, "Paid Leave": 0})
        stats[row["request_type"]] = stats.get(row["request_type"], 0) + int(row.get("requested_days") or 0)

    rows = []
    for employee in employees:
        approved = approved_map.get(employee["id"], {})
        pending = pending_map.get(employee["id"], {})
        sick_total = int(employee["sick_leave_days"] if employee["sick_leave_days"] is not None else DEFAULT_SICK_LEAVE_DAYS)
        paid_total = int(employee["paid_leave_days"] if employee["paid_leave_days"] is not None else DEFAULT_PAID_LEAVE_DAYS)
        sick_used_manual = int(employee["sick_leave_used_manual"] if employee["sick_leave_used_manual"] is not None else 0)
        paid_used_manual = int(employee["paid_leave_used_manual"] if employee["paid_leave_used_manual"] is not None else 0)
        sick_used_in_app = approved.get("Sick Leave", 0)
        paid_used_in_app = approved.get("Paid Leave", 0)
        sick_used = sick_used_manual + sick_used_in_app
        paid_used = paid_used_manual + paid_used_in_app
        rows.append({
            "user_id": employee["id"],
            "full_name": employee["full_name"],
            "username": employee["username"],
            "department": employee["department"] or "",
            "position": employee["position"] or "",
            "sick_total": sick_total,
            "sick_used": sick_used,
            "sick_remaining": max(sick_total - sick_used, 0),
            "paid_total": paid_total,
            "paid_used": paid_used,
            "paid_remaining": max(paid_total - paid_used, 0),
            "sick_used_manual": sick_used_manual,
            "paid_used_manual": paid_used_manual,
            "sick_used_in_app": sick_used_in_app,
            "paid_used_in_app": paid_used_in_app,
            "pending_sick": pending.get("Sick Leave", 0),
            "pending_paid": pending.get("Paid Leave", 0),
            "pending_total": pending.get("Sick Leave", 0) + pending.get("Paid Leave", 0),
        })

    return rows


def parse_positive_int(value, default):
    try:
        parsed = int(str(value).strip())
        return parsed if parsed > 0 else default
    except Exception:
        return default


def parse_non_negative_int(value, default):
    try:
        parsed = int(str(value).strip())
        return parsed if parsed >= 0 else default
    except Exception:
        return default


def parse_money_value(value, default=0.0):
    try:
        cleaned = str(value).strip().replace(",", "")
        if cleaned == "":
            return default
        parsed = round(float(cleaned), 2)
        return parsed if parsed >= 0 else default
    except Exception:
        return default


def notify_admins_for_leave_and_disciplinary_events():
    if not should_run_admin_notification_scan("leave_and_disciplinary"):
        return

    leave_rows = build_leave_dashboard_rows(year=now_dt().year)
    for row in leave_rows:
        if row["sick_remaining"] <= 0:
            create_admin_alert_once(
                "Sick Leave Exhausted",
                f"{row['full_name']} has no sick leave remaining for {now_dt().year}."
            )
        if row["paid_remaining"] <= 0:
            create_admin_alert_once(
                "Paid Leave Exhausted",
                f"{row['full_name']} has no paid leave remaining for {now_dt().year}."
            )

    for row in get_pending_leave_requests():
        if int(row.get("age_days") or 0) >= 3:
            create_admin_alert_once(
                "Pending Leave Aging",
                f"{row['full_name']}'s {row['request_type']} request ({row['display_date_range']}) has been pending for {row['age_days']} day(s)."
            )

    for row in get_disciplinary_actions(action_type="Suspension", date_from=today_str(), date_to=today_str()):
        create_admin_alert_once(
            "Suspension Starts Today",
            f"{row['full_name']}'s suspension starts today and runs through {row.get('end_date') or row['action_date']}."
        )


def summarize_employee_admin_changes(before_row, after_values):
    if not before_row:
        return "Created employee record."
    labels = {
        "department": "Department",
        "position": "Position",
        "emergency_contact_name": "Emergency Contact Name",
        "emergency_contact_phone": "Emergency Contact Number",
        "id_issue_date": "ID Issue Date",
        "id_expiration_date": "ID Expiration Date",
        "barcode_id": "Barcode ID",
        "hourly_rate": "Hourly Rate",
        "sick_leave_days": "Sick Leave Allotment",
        "paid_leave_days": "Paid Leave Allotment",
        "sick_leave_used_manual": "Sick Leave Already Used",
        "paid_leave_used_manual": "Paid Leave Already Used",
        "schedule_preset_id": "Schedule Preset",
        "shift_start": "Shift Start",
        "shift_end": "Shift End",
        "break_limit_minutes": "Break Limit",
        "is_active": "Account Status",
    }
    changes = []
    for key, label in labels.items():
        before = before_row[key] if key in before_row.keys() else None
        after = after_values.get(key)
        if str(before) != str(after):
            changes.append(f"{label}: {before} -> {after}")
    return "; ".join(changes) if changes else "No tracked employee settings changed."


def workbook_to_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def autosize_workbook_sheet(sheet):
    width_map = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width_map[cell.column_letter] = max(width_map.get(cell.column_letter, 0), len(str(cell.value)))
    for column_letter, width in width_map.items():
        sheet.column_dimensions[column_letter].width = min(max(width + 2, 12), 40)


def append_workbook_rows(workbook, title, rows):
    sheet = workbook.create_sheet(title=title[:31])
    normalized_rows = [dict(row) for row in (rows or [])]
    if not normalized_rows:
        sheet.append(["Message"])
        sheet.append(["No rows exported for this sheet."])
        autosize_workbook_sheet(sheet)
        return sheet

    headers = list(normalized_rows[0].keys())
    sheet.append(headers)
    for row in normalized_rows:
        sheet.append([row.get(header) for header in headers])
    autosize_workbook_sheet(sheet)
    return sheet


def get_recovery_pack_company_settings():
    settings = dict(get_company_settings() or {})
    scanner_pin_hash = settings.pop("scanner_exit_pin_hash", None)
    settings["scanner_exit_pin_configured"] = 1 if scanner_pin_hash else 0
    return settings


def build_recovery_pack_workbook():
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise ValueError("Recovery pack export requires openpyxl.") from exc

    recovery_snapshot = get_backup_recovery_snapshot()
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Recovery Pack", "Stellar Seats Attendance"])
    overview.append(["Generated At", now_str()])
    overview.append(["Environment", "Postgres" if using_postgres() else "SQLite"])
    overview.append(["Purpose", "Operational recovery reference and export workbook"])
    overview.append(["Note", "This workbook is not a one-click restore. Keep it with upload/file backups."])
    overview.append(["Last External Backup Noted", recovery_snapshot["external_backup"]["last_at"] or "Not noted in app"])
    overview.append(["External Backup Note", recovery_snapshot["external_backup"]["note"] or ""])
    overview.append(["Employee Accounts", recovery_snapshot["counts"]["employee_accounts"]])
    overview.append(["Admin Accounts", recovery_snapshot["counts"]["admin_accounts"]])
    overview.append(["Attendance Rows", recovery_snapshot["counts"]["attendance_rows"]])
    overview.append(["Scanner Logs", recovery_snapshot["counts"]["scanner_logs"]])
    overview.append(["Payroll Runs", recovery_snapshot["counts"]["payroll_runs"]])
    overview.append(["Pending Schedule Changes", recovery_snapshot["counts"]["future_schedule_changes"]])
    overview.append(["Holiday / Rest-Day Rules", recovery_snapshot["counts"]["schedule_special_rules"]])
    overview.append(["Uploaded Files", recovery_snapshot["upload_count"]])
    autosize_workbook_sheet(overview)

    guide = workbook.create_sheet(title="Recovery Guide")
    guide.append(["Step", "Action"])
    guide.append(["1", "Download the recovery pack before any major cleanup, reset, or policy change."])
    guide.append(["2", "Create an external Postgres backup or provider snapshot if production is using Render/Postgres."])
    guide.append(["3", "Keep uploaded files, barcode assets, and signature images together with this workbook."])
    guide.append(["4", "If rebuilding production, restore users and settings first, then attendance/payroll/log data."])
    guide.append(["5", "Use the workbook sheets as the source of truth for schedule presets, payroll rules, and historical workflows."])
    autosize_workbook_sheet(guide)

    append_workbook_rows(workbook, "Users", fetchall("""
        SELECT id, full_name, username, role, department, position, barcode_id,
               hourly_rate, schedule_days, shift_start, shift_end, schedule_preset_id,
               admin_permissions, admin_role_preset, is_active, created_at
        FROM users
        ORDER BY role ASC, full_name ASC
    """))
    append_workbook_rows(workbook, "Company Settings", [get_recovery_pack_company_settings()])
    append_workbook_rows(workbook, "Schedule Presets", fetchall("""
        SELECT sp.*, creator.full_name AS created_by_name
        FROM schedule_presets sp
        LEFT JOIN users creator ON creator.id = sp.created_by
        ORDER BY sp.name ASC
    """))
    append_workbook_rows(workbook, "Future Schedule Changes", fetchall("""
        SELECT fsc.*, u.full_name, preset.name AS preset_name, creator.full_name AS created_by_name
        FROM employee_future_schedule_changes fsc
        LEFT JOIN users u ON u.id = fsc.user_id
        LEFT JOIN schedule_presets preset ON preset.id = fsc.schedule_preset_id
        LEFT JOIN users creator ON creator.id = fsc.created_by
        ORDER BY fsc.effective_date ASC, fsc.id ASC
    """))
    append_workbook_rows(workbook, "Schedule Special Dates", fetchall("""
        SELECT ssd.*, creator.full_name AS created_by_name
        FROM schedule_special_dates ssd
        LEFT JOIN users creator ON creator.id = ssd.created_by
        ORDER BY ssd.special_date ASC, ssd.id ASC
    """))
    append_workbook_rows(workbook, "Attendance", fetchall("SELECT * FROM attendance ORDER BY work_date DESC, id DESC"))
    append_workbook_rows(workbook, "Breaks", fetchall("SELECT * FROM breaks ORDER BY work_date DESC, id DESC"))
    append_workbook_rows(workbook, "Corrections", fetchall("SELECT * FROM correction_requests ORDER BY created_at DESC, id DESC"))
    append_workbook_rows(workbook, "Scanner Logs", fetchall("SELECT * FROM scanner_logs ORDER BY created_at DESC, id DESC"))
    append_workbook_rows(workbook, "Overtime", fetchall("SELECT * FROM overtime_sessions ORDER BY created_at DESC, id DESC"))
    append_workbook_rows(workbook, "Payroll Runs", fetchall("SELECT * FROM payroll_runs ORDER BY updated_at DESC, id DESC"))
    append_workbook_rows(workbook, "Payroll Items", fetchall("SELECT * FROM payroll_run_items ORDER BY payroll_run_id DESC, id DESC"))
    append_workbook_rows(workbook, "Payroll Item Adjustments", fetchall("SELECT * FROM payroll_run_item_adjustments ORDER BY payroll_run_id DESC, id DESC"))
    append_workbook_rows(workbook, "Payroll Adjustments", fetchall("SELECT * FROM payroll_adjustments ORDER BY created_at DESC, id DESC"))
    append_workbook_rows(workbook, "Recurring Rules", fetchall("SELECT * FROM payroll_recurring_rules ORDER BY updated_at DESC, id DESC"))
    append_workbook_rows(workbook, "Incident Reports", fetchall("SELECT * FROM incident_reports ORDER BY created_at DESC, id DESC"))
    append_workbook_rows(workbook, "Disciplinary", fetchall("SELECT * FROM disciplinary_actions ORDER BY created_at DESC, id DESC"))
    return workbook




def get_cached_admin_reports_data(date_from, date_to, department_filter=""):
    cache_key = (date_from.strftime("%Y-%m-%d"), date_to.strftime("%Y-%m-%d"), (department_filter or "").strip())
    cache_entry = _reports_cache.get(cache_key)
    cache_age = now_timestamp() - int(cache_entry["stamp"]) if cache_entry else REPORT_CACHE_TTL_SECONDS + 1
    if cache_entry and cache_age <= REPORT_CACHE_TTL_SECONDS:
        return cache_entry["data"]
    report_data = build_admin_reports_data(date_from, date_to, department_filter=department_filter)
    _reports_cache[cache_key] = {"stamp": now_timestamp(), "data": report_data}
    return report_data


def build_admin_reports_data(date_from, date_to, department_filter=""):
    date_from_text = date_from.strftime("%Y-%m-%d")
    date_to_text = date_to.strftime("%Y-%m-%d")
    employees_sql = """
        SELECT id, full_name, department, position, is_active
        FROM users
        WHERE role = 'employee'
    """
    employees_sql += " ORDER BY full_name ASC"
    employees = [dict(row) for row in fetchall(employees_sql)]
    employee_lookup = {int(employee["id"]): employee for employee in employees}
    employee_context_cache = {}
    employee_context_map_for_end_date = get_effective_employee_context_map(employees, reference_date=date_to_text)

    def get_report_employee_context(user_row=None, user_id=None, reference_datetime=None, reference_date="", employee=None):
        employee_row = user_row or employee
        resolved_user_id = int(user_id or (employee_row["id"] if employee_row else 0) or 0)
        if not resolved_user_id:
            return None
        reference_key = normalize_history_reference(reference_datetime=reference_datetime, reference_date=reference_date)
        cache_key = (resolved_user_id, reference_key)
        if cache_key not in employee_context_cache:
            employee_context_cache[cache_key] = get_effective_employee_context(
                user_row=employee_row,
                user_id=resolved_user_id,
                reference_datetime=reference_datetime,
                reference_date=reference_date
            )
        return employee_context_cache[cache_key]

    attendance_sql = """
        SELECT a.*, u.full_name
        FROM attendance a
        JOIN users u ON u.id = a.user_id
        WHERE u.role = 'employee'
          AND a.work_date BETWEEN ? AND ?
    """
    attendance_params = [date_from_text, date_to_text]
    attendance_sql += " ORDER BY a.work_date ASC, a.id ASC"
    attendance_rows = [dict(row) for row in fetchall(attendance_sql, attendance_params)]
    attendance_break_minutes_map = get_break_minutes_map(
        [int(row["id"]) for row in attendance_rows if row.get("id")],
        include_open=False
    )

    overtime_sql = """
        SELECT o.*, u.full_name
        FROM overtime_sessions o
        JOIN users u ON u.id = o.user_id
        WHERE o.work_date BETWEEN ? AND ?
    """
    overtime_params = [date_from_text, date_to_text]
    overtime_sql += " ORDER BY o.work_date ASC, o.id ASC"
    overtime_rows = [dict(row) for row in fetchall(overtime_sql, overtime_params)]

    department_map = {}
    employee_counts = {}
    department_employee_ids = {}
    report_employee_ids = set()
    for employee in employees:
        effective_employee = employee_context_map_for_end_date.get(int(employee["id"]), employee)
        department_name = (effective_employee.get("department") if effective_employee else employee.get("department")) or "Unassigned"
        if department_filter and department_name != department_filter:
            continue
        employee_counts[department_name] = employee_counts.get(department_name, 0) + 1
        department_map.setdefault(department_name, {
            "department": department_name,
            "employee_count": 0,
            "attendance_days": 0,
            "attendance_hours": 0.0,
            "late_punches": 0,
            "overtime_hours": 0.0,
        })

    daily_map = {}
    employee_metric_map = {}
    total_hours = 0.0
    late_punches = 0
    total_break_minutes_value = 0
    attendance_days_count = 0
    for attendance in attendance_rows:
        employee = employee_lookup.get(int(attendance["user_id"]))
        effective_employee = get_report_employee_context(
            user_row=employee,
            user_id=attendance["user_id"],
            reference_datetime=get_attendance_reference_datetime(attendance),
            reference_date=attendance["work_date"]
        )
        department_name = (effective_employee.get("department") if effective_employee else "") or "Unassigned"
        if department_filter and department_name != department_filter:
            continue
        attendance_days_count += 1
        minutes_worked = max(total_work_minutes(attendance), 0)
        break_minutes = attendance_break_minutes_map.get(int(attendance["id"]), 0)
        hours_worked = round(minutes_worked / 60, 2)
        total_hours += hours_worked
        total_break_minutes_value += break_minutes
        late_punches += int(attendance.get("late_flag") or 0)
        report_employee_ids.add(int(attendance["user_id"]))
        department_employee_ids.setdefault(department_name, set()).add(int(attendance["user_id"]))
        employee_metrics = employee_metric_map.setdefault(int(attendance["user_id"]), {
            "user_id": int(attendance["user_id"]),
            "full_name": attendance.get("full_name") or (effective_employee.get("full_name") if effective_employee else ""),
            "department": department_name,
            "position": (effective_employee.get("position") if effective_employee else "") or "",
            "attendance_days": 0,
            "attendance_hours": 0.0,
            "late_punches": 0,
            "overtime_hours": 0.0,
        })
        employee_metrics["attendance_days"] += 1
        employee_metrics["attendance_hours"] = round(employee_metrics["attendance_hours"] + hours_worked, 2)
        employee_metrics["late_punches"] += int(attendance.get("late_flag") or 0)
        dept_row = department_map.setdefault(department_name, {
            "department": department_name,
            "employee_count": 0,
            "attendance_days": 0,
            "attendance_hours": 0.0,
            "late_punches": 0,
            "overtime_hours": 0.0,
        })
        dept_row["attendance_days"] += 1
        dept_row["attendance_hours"] = round(dept_row["attendance_hours"] + hours_worked, 2)
        dept_row["late_punches"] += int(attendance.get("late_flag") or 0)

        day_row = daily_map.setdefault(attendance["work_date"], {
            "work_date": attendance["work_date"],
            "attendance_days": 0,
            "attendance_hours": 0.0,
            "late_punches": 0,
            "overtime_hours": 0.0,
            "break_minutes": 0,
        })
        day_row["attendance_days"] += 1
        day_row["attendance_hours"] = round(day_row["attendance_hours"] + hours_worked, 2)
        day_row["late_punches"] += int(attendance.get("late_flag") or 0)
        day_row["break_minutes"] += break_minutes

    total_overtime_hours = 0.0
    for overtime in overtime_rows:
        overtime_hours = round(overtime_minutes_for_session(overtime) / 60, 2)
        employee = employee_lookup.get(int(overtime["user_id"]))
        effective_employee = get_report_employee_context(
            user_row=employee,
            user_id=overtime["user_id"],
            reference_datetime=get_overtime_reference_datetime(overtime),
            reference_date=overtime["work_date"]
        )
        department_name = (effective_employee.get("department") if effective_employee else "") or "Unassigned"
        if department_filter and department_name != department_filter:
            continue
        total_overtime_hours += overtime_hours
        report_employee_ids.add(int(overtime["user_id"]))
        department_employee_ids.setdefault(department_name, set()).add(int(overtime["user_id"]))
        employee_metrics = employee_metric_map.setdefault(int(overtime["user_id"]), {
            "user_id": int(overtime["user_id"]),
            "full_name": overtime.get("full_name") or (effective_employee.get("full_name") if effective_employee else ""),
            "department": department_name,
            "position": (effective_employee.get("position") if effective_employee else "") or "",
            "attendance_days": 0,
            "attendance_hours": 0.0,
            "late_punches": 0,
            "overtime_hours": 0.0,
        })
        employee_metrics["overtime_hours"] = round(employee_metrics["overtime_hours"] + overtime_hours, 2)
        dept_row = department_map.setdefault(department_name, {
            "department": department_name,
            "employee_count": 0,
            "attendance_days": 0,
            "attendance_hours": 0.0,
            "late_punches": 0,
            "overtime_hours": 0.0,
        })
        dept_row["overtime_hours"] = round(dept_row["overtime_hours"] + overtime_hours, 2)
        day_row = daily_map.setdefault(overtime["work_date"], {
            "work_date": overtime["work_date"],
            "attendance_days": 0,
            "attendance_hours": 0.0,
            "late_punches": 0,
            "overtime_hours": 0.0,
            "break_minutes": 0,
        })
        day_row["overtime_hours"] = round(day_row["overtime_hours"] + overtime_hours, 2)

    for department_name, count in employee_counts.items():
        department_map.setdefault(department_name, {
            "department": department_name,
            "employee_count": 0,
            "attendance_days": 0,
            "attendance_hours": 0.0,
            "late_punches": 0,
            "overtime_hours": 0.0,
        })
        department_map[department_name]["employee_count"] = count
    for department_name, user_ids in department_employee_ids.items():
        department_map.setdefault(department_name, {
            "department": department_name,
            "employee_count": 0,
            "attendance_days": 0,
            "attendance_hours": 0.0,
            "late_punches": 0,
            "overtime_hours": 0.0,
        })
        department_map[department_name]["employee_count"] = max(
            int(department_map[department_name].get("employee_count") or 0),
            len(user_ids)
        )

    correction_rows = [
        dict(row)
        for row in fetchall("""
            SELECT c.*, u.full_name, u.department, u.position
            FROM correction_requests c
            JOIN users u ON u.id = c.user_id
            WHERE c.work_date <= ?
              AND COALESCE(c.end_work_date, c.work_date) >= ?
              AND u.role = 'employee'
            ORDER BY c.work_date ASC, c.id ASC
        """, (date_to_text, date_from_text))
    ]
    pending_corrections_count = 0
    pending_leave_requests_count = 0
    leave_summary_map = {}
    correction_summary_map = {}
    for request_row in correction_rows:
        effective_employee = get_report_employee_context(
            user_row=employee_lookup.get(int(request_row["user_id"])),
            user_id=request_row["user_id"],
            reference_date=request_row["work_date"]
        )
        department_name = (effective_employee.get("department") if effective_employee else request_row.get("department")) or "Unassigned"
        if department_filter and department_name != department_filter:
            continue
        report_employee_ids.add(int(request_row["user_id"]))
        department_employee_ids.setdefault(department_name, set()).add(int(request_row["user_id"]))
        if request_row["request_type"] in {"Sick Leave", "Paid Leave"}:
            summary_key = (request_row["request_type"], request_row["status"])
            leave_summary_map[summary_key] = leave_summary_map.get(summary_key, 0) + 1
            if request_row["status"] == "Pending":
                pending_leave_requests_count += 1
        else:
            correction_key = (request_row["request_type"], request_row["status"])
            correction_summary_map[correction_key] = correction_summary_map.get(correction_key, 0) + 1
            if request_row["status"] == "Pending":
                pending_corrections_count += 1
    leave_summary_rows = [
        {
            "request_type": request_type,
            "status": status,
            "request_count": request_count,
        }
        for (request_type, status), request_count in sorted(
            leave_summary_map.items(),
            key=lambda item: (item[0][0], item[0][1])
        )
    ]
    incident_row = fetchone("""
        SELECT COUNT(*) AS cnt
        FROM incident_reports r
        LEFT JOIN users u ON u.id = r.user_id
        WHERE COALESCE(r.report_date, r.incident_date) BETWEEN ? AND ?
          {department_clause}
    """.replace("{department_clause}", "AND COALESCE(r.report_department, u.department, '') = ?" if department_filter else ""), (
        date_from_text,
        date_to_text,
            *((department_filter,) if department_filter else ())
    ))
    incident_rows = [
        dict(row)
        for row in fetchall("""
            SELECT r.error_type, r.status, COUNT(*) AS total_count
            FROM incident_reports r
            LEFT JOIN users u ON u.id = r.user_id
            WHERE COALESCE(r.report_date, r.incident_date) BETWEEN ? AND ?
              {department_clause}
            GROUP BY r.error_type, r.status
            ORDER BY r.error_type ASC, r.status ASC
        """.replace("{department_clause}", "AND COALESCE(r.report_department, u.department, '') = ?" if department_filter else ""), (
            date_from_text,
            date_to_text,
            *((department_filter,) if department_filter else ())
        ))
    ]
    incident_summary_map = {}
    for row in incident_rows:
        summary_row = incident_summary_map.setdefault(row["error_type"] or "Unspecified", {
            "error_type": row["error_type"] or "Unspecified",
            "open_count": 0,
            "reviewed_count": 0,
            "resolved_count": 0,
            "total_count": 0,
        })
        status_text = (row["status"] or "Open").strip()
        if status_text == "Resolved":
            summary_row["resolved_count"] += int(row["total_count"] or 0)
        elif status_text == "Reviewed":
            summary_row["reviewed_count"] += int(row["total_count"] or 0)
        else:
            summary_row["open_count"] += int(row["total_count"] or 0)
        summary_row["total_count"] += int(row["total_count"] or 0)
    released_runs = [
        dict(row)
        for row in fetchall("""
            SELECT
                pr.*,
                creator.full_name AS created_by_name,
                COUNT(pri.id) AS payroll_rows,
                COUNT(DISTINCT pri.user_id) AS employee_count,
                COALESCE(SUM(pri.final_pay), 0) AS released_total,
                COALESCE(SUM(pri.overtime_pay), 0) AS overtime_pay_total,
                COALESCE(SUM(pri.allowances), 0) AS allowances_total,
                COALESCE(SUM(pri.deductions), 0) AS deductions_total
            FROM payroll_runs pr
            LEFT JOIN users creator ON creator.id = pr.created_by
            JOIN payroll_run_items pri ON pri.payroll_run_id = pr.id
            WHERE pr.status = 'Released'
              AND pr.date_from <= ?
              AND pr.date_to >= ?
              {department_clause}
            GROUP BY pr.id, creator.full_name
            ORDER BY COALESCE(pr.released_at, pr.updated_at) DESC, pr.id DESC
            LIMIT 12
        """.replace(
            "{department_clause}",
            "AND COALESCE(pri.department, '') = ?" if department_filter else ""
        ), (
            date_to_text,
            date_from_text,
            *((department_filter,) if department_filter else ())
        ))
    ]

    for run in released_runs:
        run["period_label"] = f"{run['date_from']} to {run['date_to']}"
        run["scope_label"] = get_payroll_scope_label(
            employee_filter=run.get("employee_filter"),
            department_filter=run.get("department_filter"),
        )
        run["payroll_rows"] = int(run.get("payroll_rows") or 0)
        run["employee_count"] = int(run.get("employee_count") or 0)
        run["released_total"] = round(float(run.get("released_total") or 0), 2)
        run["overtime_pay_total"] = round(float(run.get("overtime_pay_total") or 0), 2)
        run["allowances_total"] = round(float(run.get("allowances_total") or 0), 2)
        run["deductions_total"] = round(float(run.get("deductions_total") or 0), 2)
        run["average_final_pay"] = round(
            run["released_total"] / run["payroll_rows"],
            2
        ) if run["payroll_rows"] else 0.0

    released_payroll_row = fetchone("""
        SELECT
            COUNT(DISTINCT pr.id) AS run_count,
            COUNT(*) AS payroll_rows,
            COUNT(DISTINCT pri.user_id) AS employee_count,
            COALESCE(SUM(pri.final_pay), 0) AS released_total,
            COALESCE(SUM(pri.overtime_pay), 0) AS overtime_pay_total,
            COALESCE(SUM(pri.allowances), 0) AS allowances_total,
            COALESCE(SUM(pri.deductions), 0) AS deductions_total
        FROM payroll_runs pr
        JOIN payroll_run_items pri ON pri.payroll_run_id = pr.id
        WHERE pr.status = 'Released'
          AND pr.date_from <= ?
          AND pr.date_to >= ?
          {department_clause}
    """.replace("{department_clause}", "AND COALESCE(pri.department, '') = ?" if department_filter else ""), (
        date_to_text,
        date_from_text,
        *((department_filter,) if department_filter else ())
    ))

    payroll_department_rows = [
        dict(row)
        for row in fetchall("""
            SELECT
                COALESCE(pri.department, 'Unassigned') AS department,
                COUNT(*) AS payroll_rows,
                COUNT(DISTINCT pri.user_id) AS employee_count,
                COALESCE(SUM(pri.final_pay), 0) AS final_pay_total,
                COALESCE(SUM(pri.overtime_pay), 0) AS overtime_pay_total,
                COALESCE(SUM(pri.allowances), 0) AS allowances_total,
                COALESCE(SUM(pri.deductions), 0) AS deductions_total
            FROM payroll_runs pr
            JOIN payroll_run_items pri ON pri.payroll_run_id = pr.id
            WHERE pr.status = 'Released'
              AND pr.date_from <= ?
              AND pr.date_to >= ?
              {department_clause}
            GROUP BY COALESCE(pri.department, 'Unassigned')
            ORDER BY final_pay_total DESC, department ASC
        """.replace("{department_clause}", "AND COALESCE(pri.department, '') = ?" if department_filter else ""), (
            date_to_text,
            date_from_text,
            *((department_filter,) if department_filter else ())
        ))
    ]

    leave_rows = [
        {
            "request_type": row["request_type"],
            "status": row["status"],
            "request_count": int(row["request_count"] or 0),
        }
        for row in leave_summary_rows
    ]
    correction_rows_summary = [
        {
            "request_type": request_type,
            "status": status,
            "request_count": int(request_count or 0),
        }
        for (request_type, status), request_count in sorted(
            correction_summary_map.items(),
            key=lambda item: (item[0][0], item[0][1])
        )
    ]
    department_rows = sorted(department_map.values(), key=lambda item: (item["department"] or "").lower())
    for row in department_rows:
        row["avg_hours_per_day"] = round((row["attendance_hours"] / row["attendance_days"]), 2) if row["attendance_days"] else 0.0
    daily_rows = [daily_map[key] for key in sorted(daily_map.keys())]
    top_employee_rows = sorted(
        employee_metric_map.values(),
        key=lambda item: (-item["late_punches"], -item["overtime_hours"], -item["attendance_hours"], item["full_name"] or "")
    )[:10]
    incident_summary_rows = sorted(incident_summary_map.values(), key=lambda item: (-item["total_count"], item["error_type"]))

    report_employee_count = len(report_employee_ids) if report_employee_ids else sum(employee_counts.values())
    active_report_days = len(daily_rows)
    leave_requests_total = sum(row["request_count"] for row in leave_rows)
    pending_leave_requests = pending_leave_requests_count
    correction_requests_total = sum(correction_summary_map.values())
    incident_total_count = sum(int(row.get("total_count") or 0) for row in incident_summary_rows)
    incident_open_count = sum(int(row.get("open_count") or 0) for row in incident_summary_rows)
    incident_reviewed_count = sum(int(row.get("reviewed_count") or 0) for row in incident_summary_rows)
    incident_resolved_count = sum(int(row.get("resolved_count") or 0) for row in incident_summary_rows)
    incident_follow_up_count = incident_open_count + incident_reviewed_count
    released_payroll_runs_count = int(released_payroll_row["run_count"] or 0) if released_payroll_row else 0
    released_payroll_rows_count = int(released_payroll_row["payroll_rows"] or 0) if released_payroll_row else 0
    released_payroll_employee_count = int(released_payroll_row["employee_count"] or 0) if released_payroll_row else 0
    released_payroll_total = round(float(released_payroll_row["released_total"] or 0), 2) if released_payroll_row else 0.0
    released_payroll_overtime_total = round(float(released_payroll_row["overtime_pay_total"] or 0), 2) if released_payroll_row else 0.0
    released_payroll_allowances_total = round(float(released_payroll_row["allowances_total"] or 0), 2) if released_payroll_row else 0.0
    released_payroll_deductions_total = round(float(released_payroll_row["deductions_total"] or 0), 2) if released_payroll_row else 0.0

    for row in department_rows:
        row["attendance_share_percent"] = round((row["attendance_hours"] / total_hours) * 100, 1) if total_hours else 0.0
        row["overtime_share_percent"] = round((row["overtime_hours"] / total_overtime_hours) * 100, 1) if total_overtime_hours else 0.0

    for row in payroll_department_rows:
        row["payroll_rows"] = int(row.get("payroll_rows") or 0)
        row["employee_count"] = int(row.get("employee_count") or 0)
        row["final_pay_total"] = round(float(row.get("final_pay_total") or 0), 2)
        row["overtime_pay_total"] = round(float(row.get("overtime_pay_total") or 0), 2)
        row["allowances_total"] = round(float(row.get("allowances_total") or 0), 2)
        row["deductions_total"] = round(float(row.get("deductions_total") or 0), 2)
        row["average_final_pay"] = round(
            row["final_pay_total"] / row["payroll_rows"],
            2
        ) if row["payroll_rows"] else 0.0

    trend_highlights, department_highlights = build_report_highlights(daily_rows, department_rows)
    case_rows = build_case_rows(
        pending_leave_requests,
        leave_requests_total,
        pending_corrections_count,
        correction_requests_total,
        incident_follow_up_count,
        incident_resolved_count,
        incident_total_count,
    )

    summary = {
        "employee_count": report_employee_count,
        "attendance_days": attendance_days_count,
        "attendance_hours": round(total_hours, 2),
        "avg_hours_per_day": round(total_hours / active_report_days, 2) if active_report_days else 0.0,
        "avg_hours_per_employee": round(total_hours / report_employee_count, 2) if report_employee_count else 0.0,
        "late_punches": late_punches,
        "late_rate_percent": round((late_punches / attendance_days_count) * 100, 1) if attendance_days_count else 0.0,
        "overtime_hours": round(total_overtime_hours, 2),
        "overtime_share_percent": round((total_overtime_hours / total_hours) * 100, 1) if total_hours else 0.0,
        "break_minutes": total_break_minutes_value,
        "break_hours": round(total_break_minutes_value / 60, 2),
        "correction_requests": correction_requests_total,
        "pending_corrections": pending_corrections_count,
        "leave_requests": leave_requests_total,
        "pending_leave_requests": pending_leave_requests,
        "incident_reports": int(incident_row["cnt"] or 0) if incident_row else 0,
        "incident_follow_ups": incident_follow_up_count,
        "follow_up_count": pending_leave_requests + pending_corrections_count + incident_follow_up_count,
        "released_payroll_runs": released_payroll_runs_count,
        "released_payroll_rows": released_payroll_rows_count,
        "released_payroll_employee_count": released_payroll_employee_count,
        "released_payroll_total": released_payroll_total,
        "released_payroll_average": round(
            released_payroll_total / released_payroll_rows_count,
            2
        ) if released_payroll_rows_count else 0.0,
        "released_payroll_overtime_total": released_payroll_overtime_total,
        "released_payroll_allowances_total": released_payroll_allowances_total,
        "released_payroll_deductions_total": released_payroll_deductions_total,
        "active_report_days": active_report_days,
    }

    return {
        "summary": summary,
        "trend_highlights": trend_highlights,
        "department_highlights": department_highlights,
        "case_rows": case_rows,
        "department_rows": department_rows,
        "daily_rows": daily_rows,
        "leave_rows": leave_rows,
        "correction_rows": correction_rows_summary,
        "incident_rows": incident_summary_rows,
        "top_employee_rows": top_employee_rows,
        "payroll_department_rows": payroll_department_rows,
        "released_runs": released_runs,
    }


def get_payroll_adjustments(date_from, date_to, department_filter="", employee_filter=""):
    date_from_text = payroll_date_text(date_from)
    date_to_text = payroll_date_text(date_to)
    sql = """
        SELECT
            pa.*,
            u.full_name AS employee_name,
            u.department AS employee_department,
            u.position AS employee_position,
            creator.full_name AS created_by_name
        FROM payroll_adjustments pa
        JOIN users u ON u.id = pa.user_id
        LEFT JOIN users creator ON creator.id = pa.created_by
        WHERE pa.date_from = ?
          AND pa.date_to = ?
    """
    params = [date_from_text, date_to_text]
    if department_filter:
        sql += " AND COALESCE(u.department, '') = ?"
        params.append(department_filter)
    if employee_filter:
        sql += " AND pa.user_id = ?"
        params.append(int(employee_filter))
    sql += " ORDER BY pa.created_at DESC, pa.id DESC"
    return fetchall(sql, tuple(params))


def get_payroll_recurring_rules(department_filter="", employee_filter="", include_inactive=True):
    sql = """
        SELECT
            prr.*,
            u.full_name AS employee_name,
            u.department AS employee_department,
            u.position AS employee_position,
            creator.full_name AS created_by_name
        FROM payroll_recurring_rules prr
        JOIN users u ON u.id = prr.user_id
        LEFT JOIN users creator ON creator.id = prr.created_by
        WHERE 1 = 1
    """
    params = []
    if not include_inactive:
        sql += " AND COALESCE(prr.is_active, 1) = 1"
    if department_filter:
        sql += " AND COALESCE(u.department, '') = ?"
        params.append(department_filter)
    if employee_filter:
        sql += " AND prr.user_id = ?"
        params.append(int(employee_filter))
    sql += " ORDER BY COALESCE(prr.is_active, 1) DESC, u.full_name ASC, prr.label ASC, prr.id DESC"
    rows = [dict(row) for row in fetchall(sql, tuple(params))]
    for row in rows:
        monthly_anchor = parse_iso_date(row.get("start_date"))
        row["amount"] = round(float(row.get("amount") or 0), 2)
        row["display_amount"] = format_currency(row["amount"])
        row["status_label"] = "Active" if int(row.get("is_active") or 0) == 1 else "Inactive"
        row["active_range_label"] = "Open-ended"
        if row.get("start_date") and row.get("end_date"):
            row["active_range_label"] = f"{row['start_date']} to {row['end_date']}"
        elif row.get("start_date"):
            row["active_range_label"] = f"Starting {row['start_date']}"
        elif row.get("end_date"):
            row["active_range_label"] = f"Until {row['end_date']}"
        row["recurrence_description"] = (
            "Applies to every payroll period in range."
            if row.get("recurrence_type") != "Monthly"
            else f"Applies once per month when the payroll period includes day {(monthly_anchor.day if monthly_anchor else 1)}."
        )
    return rows


def build_effective_payroll_adjustments(date_from, date_to, department_filter="", employee_filter=""):
    date_from_text = payroll_date_text(date_from)
    date_to_text = payroll_date_text(date_to)
    period_from = parse_iso_date(date_from_text)
    period_to = parse_iso_date(date_to_text, period_from)

    effective_rows = []
    for row in [dict(item) for item in get_payroll_adjustments(date_from_text, date_to_text, department_filter=department_filter, employee_filter=employee_filter)]:
        row["source_kind"] = "Manual"
        row["source_rule_id"] = None
        row["recurrence_type"] = ""
        effective_rows.append(row)

    if not period_from or not period_to:
        return effective_rows

    recurring_rules = get_payroll_recurring_rules(
        department_filter=department_filter,
        employee_filter=employee_filter,
        include_inactive=False
    )
    for rule in recurring_rules:
        if not recurring_rule_applies_to_period(rule, period_from, period_to):
            continue
        effective_rows.append({
            "id": None,
            "user_id": rule["user_id"],
            "date_from": date_from_text,
            "date_to": date_to_text,
            "adjustment_type": rule["adjustment_type"],
            "label": rule["label"],
            "amount": rule["amount"],
            "notes": rule.get("notes") or "",
            "created_by": rule.get("created_by"),
            "created_at": rule.get("updated_at") or rule.get("created_at") or now_str(),
            "created_by_name": rule.get("created_by_name") or "Administrator",
            "employee_name": rule.get("employee_name"),
            "employee_department": rule.get("employee_department"),
            "employee_position": rule.get("employee_position"),
            "source_kind": "Recurring Rule",
            "source_rule_id": rule["id"],
            "recurrence_type": rule.get("recurrence_type") or "Every Payroll",
        })

    effective_rows.sort(key=lambda item: (item.get("employee_name") or "", item.get("adjustment_type") or "", item.get("label") or "", item.get("created_at") or ""))
    return effective_rows


def get_payroll_run_item_adjustments(payroll_run_id, user_id=None):
    sql = """
        SELECT *
        FROM payroll_run_item_adjustments
        WHERE payroll_run_id = ?
    """
    params = [payroll_run_id]
    if user_id is not None:
        sql += " AND user_id = ?"
        params.append(int(user_id))
    sql += " ORDER BY adjustment_created_at ASC, created_at ASC, id ASC"
    rows = [dict(row) for row in fetchall(sql, tuple(params))]
    for row in rows:
        row["amount"] = round(float(row.get("amount") or 0), 2)
        row["is_deduction"] = row.get("adjustment_type") == "Deduction"
        row["display_amount"] = format_currency(row["amount"])
        row["source_kind"] = row.get("source_kind") or "Manual"
        row["recurrence_type"] = row.get("recurrence_type") or ""
    return rows


def get_payroll_run(date_from, date_to, department_filter="", employee_filter=""):
    return fetchone("""
        SELECT pr.*, creator.full_name AS created_by_name
        FROM payroll_runs pr
        LEFT JOIN users creator ON creator.id = pr.created_by
        WHERE pr.date_from = ?
          AND pr.date_to = ?
          AND COALESCE(pr.department_filter, '') = ?
          AND COALESCE(pr.employee_filter, '') = ?
        ORDER BY pr.id DESC
        LIMIT 1
    """, (
        payroll_date_text(date_from),
        payroll_date_text(date_to),
        department_filter or "",
        str(employee_filter or "")
    ))


def get_payroll_recurring_rule(rule_id):
    if not str(rule_id or "").strip().isdigit():
        return None
    rows = get_payroll_recurring_rules(include_inactive=True)
    target_id = int(rule_id)
    for row in rows:
        if int(row["id"]) == target_id:
            return row
    return None


def get_payroll_run_item_count(payroll_run_id):
    row = fetchone("SELECT COUNT(*) AS cnt FROM payroll_run_items WHERE payroll_run_id = ?", (payroll_run_id,))
    return int(row["cnt"] or 0) if row else 0


def get_payroll_run_release_check(payroll_run_id):
    row = fetchone("""
        SELECT
            COUNT(*) AS item_count,
            SUM(CASE WHEN COALESCE(hourly_rate, 0) <= 0 THEN 1 ELSE 0 END) AS missing_rates
        FROM payroll_run_items
        WHERE payroll_run_id = ?
    """, (payroll_run_id,))
    row_data = dict(row) if row else {}
    item_count = int(row_data.get("item_count") or 0)
    missing_rates = int(row_data.get("missing_rates") or 0)
    can_release = item_count > 0 and missing_rates == 0
    blocked_reason = ""
    if item_count <= 0:
        blocked_reason = "No employee rows were saved in this snapshot."
    elif missing_rates > 0:
        blocked_reason = f"{missing_rates} employee row(s) are still missing hourly rates."
    return {
        "item_count": item_count,
        "missing_rates": missing_rates,
        "can_release": can_release,
        "release_block_reason": blocked_reason,
    }


def enrich_admin_payroll_run(row):
    if not row:
        return None

    item = dict(row)
    item.update(get_payroll_run_release_check(item["id"]))
    item["period_label"] = format_payroll_period_label(item.get("date_from"), item.get("date_to"))
    item["released_display"] = format_datetime_12h(item.get("released_at")) if item.get("released_at") else ""
    item["updated_display"] = format_datetime_12h(item.get("updated_at")) if item.get("updated_at") else ""

    employee_name = ""
    if (item.get("employee_filter") or "").isdigit():
        employee = get_user_by_id(int(item["employee_filter"]))
        employee_name = employee["full_name"] if employee else ""
    item["employee_name"] = employee_name
    return item


def get_recent_payroll_runs(limit=8):
    rows = fetchall("""
        SELECT pr.*, creator.full_name AS created_by_name
        FROM payroll_runs pr
        LEFT JOIN users creator ON creator.id = pr.created_by
        ORDER BY pr.id DESC
        LIMIT ?
    """, (limit,))
    enriched = []
    for row in rows:
        enriched.append(enrich_admin_payroll_run(row))
    return enriched


def enrich_employee_payroll_item(row, current_user_id=None):
    if not row:
        return None

    item = dict(row)
    item["period_label"] = format_payroll_period_label(item.get("date_from"), item.get("date_to"))
    released_source = item.get("released_at") or item.get("updated_at") or item.get("created_at")
    item["released_display"] = format_datetime_12h(released_source) if released_source else ""
    item["adjustment_balance"] = round(float(item.get("allowances") or 0) - float(item.get("deductions") or 0), 2)
    item["total_compensation"] = round(float(item.get("gross_pay") or 0) + float(item.get("overtime_pay") or 0), 2)

    item["scope_label"] = get_payroll_scope_label(
        employee_filter=item.get("employee_filter"),
        department_filter=item.get("department_filter"),
        current_user_id=current_user_id,
    )

    item["created_by_name"] = item.get("created_by_name") or "Administrator"
    return item


def get_employee_released_payroll_runs(user_id, limit=24):
    rows = [
        enrich_employee_payroll_item(row, current_user_id=user_id)
        for row in fetchall("""
            SELECT
                pr.id AS payroll_run_id,
                pr.date_from,
                pr.date_to,
                pr.department_filter,
                pr.employee_filter,
                pr.status,
                pr.notes,
                pr.created_by,
                pr.created_at,
                pr.updated_at,
                pr.released_at,
                creator.full_name AS created_by_name,
                pri.user_id,
                pri.full_name,
                pri.department,
                pri.position,
                pri.hourly_rate,
                pri.days_worked,
                pri.total_hours,
                pri.overtime_hours,
                pri.late_minutes,
                pri.break_minutes,
                pri.suspension_days,
                pri.suspension_pay,
                pri.gross_pay,
                pri.overtime_pay,
                pri.allowances,
                pri.deductions,
                pri.final_pay,
                pri.created_at AS item_created_at
            FROM payroll_run_items pri
            JOIN payroll_runs pr ON pr.id = pri.payroll_run_id
            LEFT JOIN users creator ON creator.id = pr.created_by
            WHERE pri.user_id = ?
              AND pr.status = 'Released'
            ORDER BY pr.date_to DESC, COALESCE(pr.released_at, pr.updated_at, pr.created_at) DESC, pr.id DESC
        """, (user_id,))
    ]

    deduped = []
    seen_periods = set()
    for item in rows:
        period_key = (item.get("date_from"), item.get("date_to"))
        if period_key in seen_periods:
            continue
        seen_periods.add(period_key)
        deduped.append(item)
        if len(deduped) >= limit:
            break
    return deduped


def get_employee_released_payroll_item(user_id, payroll_run_id):
    row = fetchone("""
        SELECT
            pr.id AS payroll_run_id,
            pr.date_from,
            pr.date_to,
            pr.department_filter,
            pr.employee_filter,
            pr.status,
            pr.notes,
            pr.created_by,
            pr.created_at,
            pr.updated_at,
            pr.released_at,
            creator.full_name AS created_by_name,
            pri.user_id,
            pri.full_name,
            pri.department,
            pri.position,
            pri.hourly_rate,
            pri.days_worked,
            pri.total_hours,
            pri.overtime_hours,
            pri.late_minutes,
            pri.break_minutes,
            pri.suspension_days,
            pri.suspension_pay,
            pri.gross_pay,
            pri.overtime_pay,
            pri.allowances,
            pri.deductions,
            pri.final_pay,
            pri.created_at AS item_created_at
        FROM payroll_run_items pri
        JOIN payroll_runs pr ON pr.id = pri.payroll_run_id
        LEFT JOIN users creator ON creator.id = pr.created_by
        WHERE pri.user_id = ?
          AND pr.id = ?
          AND pr.status = 'Released'
        LIMIT 1
    """, (user_id, payroll_run_id))
    if not row:
        return None
    item = enrich_employee_payroll_item(row, current_user_id=user_id)
    item["adjustment_entries"] = get_payroll_run_item_adjustments(item["payroll_run_id"], user_id=user_id)
    item["allowance_entries"] = [entry for entry in item["adjustment_entries"] if entry["adjustment_type"] == "Allowance"]
    item["deduction_entries"] = [entry for entry in item["adjustment_entries"] if entry["adjustment_type"] == "Deduction"]
    item["has_adjustment_entries"] = bool(item["adjustment_entries"])
    item["missing_adjustment_detail"] = (
        not item["adjustment_entries"]
        and (float(item.get("allowances") or 0) > 0 or float(item.get("deductions") or 0) > 0)
    )
    return item


def shift_month(year, month, delta):
    month_index = (year * 12 + (month - 1)) + delta
    return month_index // 12, (month_index % 12) + 1


def parse_calendar_month(year_value="", month_value=""):
    today = now_dt().date()
    try:
        year = int(year_value)
    except Exception:
        year = today.year
    try:
        month = int(month_value)
    except Exception:
        month = today.month

    if year < 2000 or year > 2100:
        year = today.year
    if month < 1 or month > 12:
        month = today.month
    return year, month


def build_employee_attendance_calendar(user_row, year, month):
    month_start = datetime(year, month, 1).date()
    _, days_in_month = calendar.monthrange(year, month)
    month_end = datetime(year, month, days_in_month).date()
    date_from_text = month_start.strftime("%Y-%m-%d")
    date_to_text = month_end.strftime("%Y-%m-%d")
    today_value = now_dt().date()
    today_text = today_value.strftime("%Y-%m-%d")
    special_rule_map = get_schedule_special_rule_map(date_from_text, date_to_text)
    attendance_map = {}
    for row in fetchall("""
        SELECT *
        FROM attendance
        WHERE user_id = ?
          AND work_date BETWEEN ? AND ?
        ORDER BY work_date DESC, id DESC
    """, (user_row["id"], date_from_text, date_to_text)):
        row = dict(row)
        attendance_map.setdefault(row["work_date"], row)

    request_map = {}
    for row in [dict(item) for item in get_approved_special_requests(user_id=user_row["id"], date_from=date_from_text, date_to=date_to_text)]:
        request_dates = expand_request_dates(row["work_date"], row.get("end_work_date")) if row["request_type"] in LEAVE_REQUEST_TYPES else [row["work_date"]]
        for request_work_date in request_dates:
            matched_attendance, _ = get_matching_attendance_context_for_request(
                row["user_id"],
                request_work_date,
                request_type=row["request_type"],
                requested_time_out=row.get("requested_time_out")
            )
            target_work_date = matched_attendance["work_date"] if matched_attendance else request_work_date
            if target_work_date < date_from_text or target_work_date > date_to_text:
                continue
            existing = request_map.get(target_work_date)
            if existing and existing.get("request_type") in LEAVE_REQUEST_TYPES:
                continue
            if row["request_type"] in LEAVE_REQUEST_TYPES or not existing:
                request_map[target_work_date] = row

    suspension_map = {}
    for row in [dict(item) for item in get_disciplinary_actions(action_type="Suspension", user_id=user_row["id"])]:
        for suspension_date in expand_suspension_dates(row):
            if suspension_date < date_from_text or suspension_date > date_to_text:
                continue
            suspension_map[suspension_date] = row

    counts = {
        "worked": 0,
        "late": 0,
        "leave": 0,
        "suspension": 0,
        "undertime": 0,
        "absent": 0,
        "off_day": 0,
    }
    highlights = []

    def push_highlight(entry):
        if not entry.get("highlight"):
            return
        highlights.append({
            "date": entry["date"],
            "label": entry["label"],
            "details": entry["details"],
            "tone": entry["tone"],
        })

    def build_entry(work_date):
        attendance_row = attendance_map.get(work_date)
        special_request = request_map.get(work_date)
        effective_user = get_effective_employee_context(user_row=user_row, reference_date=work_date) or dict(user_row)
        break_limit_minutes = parse_break_limit_minutes(
            effective_user.get("break_limit_minutes")
            if effective_user.get("break_limit_minutes") is not None
            else BREAK_LIMIT_MINUTES
        )

        if attendance_row:
            row = dict(attendance_row)
            row["source_type"] = "attendance"
            row["request_type"] = special_request["request_type"] if special_request else ""
            row["admin_note"] = special_request["admin_note"] if special_request else ""
            row["message"] = special_request["message"] if special_request else ""
            row["requested_time_out"] = special_request["requested_time_out"] if special_request else ""
            item = enrich_history_record(row, break_limit_minutes, employee_row=effective_user)
            state_key = "worked"
            label = "Present"
            tone = "blue"

            if item["record_type"] == "Suspension":
                state_key, label, tone = "suspension", "Suspended", "red"
            elif item["record_type"] in LEAVE_REQUEST_TYPES:
                state_key = "leave"
                label = item["record_type"]
                tone = "yellow" if item["record_type"] == "Sick Leave" else "green"
            elif item["record_type"] == "Undertime":
                state_key, label, tone = "undertime", "Undertime", "yellow"
            elif row.get("late_flag") == 1:
                state_key, label, tone = "late", "Late", "yellow"
            elif row.get("time_in") and not row.get("time_out") and work_date < today_text:
                state_key, label, tone = "absent", "Missing Time Out", "red"
            elif item["display_status"] == "On Break":
                state_key, label, tone = "worked", "On Break", "blue"
            elif item["display_status"] == "Timed In":
                state_key, label, tone = "worked", "Timed In", "blue"

            details = []
            if row.get("time_in"):
                details.append(f"In {format_time_12h(row['time_in'])}")
            if row.get("time_out"):
                details.append(f"Out {format_time_12h(row['time_out'])}")
            if item["work_minutes"]:
                details.append(minutes_to_hm(item["work_minutes"]))
            if row.get("late_flag") == 1:
                details.append(f"{int(row.get('late_minutes') or 0)} min late")
            if item["over_break_minutes"] > 0:
                details.append(f"Over break {item['over_break_minutes']}m")
            if not details and item["request_note"]:
                details.append(item["request_note"])

            return {
                "date": work_date,
                "label": label,
                "tone": tone,
                "state_key": state_key,
                "details": " | ".join(details) if details else (item["display_status"] or item["record_type"]),
                "highlight": state_key in {"late", "leave", "suspension", "undertime", "absent"} or item["over_break_minutes"] > 0,
            }

        if work_date in suspension_map:
            suspension = suspension_map[work_date]
            return {
                "date": work_date,
                "label": "Suspended",
                "tone": "red",
                "state_key": "suspension",
                "details": suspension.get("details") or "Suspension in effect for this date.",
                "highlight": True,
            }

        if special_request and special_request["request_type"] in LEAVE_REQUEST_TYPES:
            note = special_request.get("admin_note") or special_request.get("message") or ""
            return {
                "date": work_date,
                "label": special_request["request_type"],
                "tone": "yellow" if special_request["request_type"] == "Sick Leave" else "green",
                "state_key": "leave",
                "details": note or f"Approved {special_request['request_type'].lower()} request.",
                "highlight": True,
            }

        if special_request and special_request["request_type"] == "Undertime":
            synthetic_row = {
                "id": None,
                "user_id": special_request["user_id"],
                "work_date": work_date,
                "time_in": None,
                "time_out": special_request.get("requested_time_out"),
                "status": "Approved Undertime",
                "proof_file": None,
                "late_flag": 0,
                "late_minutes": 0,
                "request_type": "Undertime",
                "admin_note": special_request.get("admin_note") or "",
                "message": special_request.get("message") or "",
                "requested_time_out": special_request.get("requested_time_out") or "",
                "source_type": "request",
            }
            item = enrich_history_record(synthetic_row, break_limit_minutes, employee_row=effective_user)
            return {
                "date": work_date,
                "label": "Undertime",
                "tone": "yellow",
                "state_key": "undertime",
                "details": item["request_note"] or "Approved undertime request.",
                "highlight": True,
            }

        special_rule = special_rule_map.get(work_date)
        if special_rule:
            return {
                "date": work_date,
                "label": special_rule["rule_type_label"],
                "tone": "gray",
                "state_key": special_rule["rule_type"],
                "details": special_rule["notes"] or special_rule["display_label"],
                "highlight": False,
            }

        scheduled = is_scheduled_on_date(effective_user, work_date)
        if scheduled:
            if work_date < today_text or (work_date == today_text and is_absent_today(effective_user, None)):
                return {
                    "date": work_date,
                    "label": "Absent",
                    "tone": "red",
                    "state_key": "absent",
                    "details": "Scheduled day with no attendance record.",
                    "highlight": True,
                }
            if work_date == today_text:
                return {
                    "date": work_date,
                    "label": "Awaiting Scan",
                    "tone": "gray",
                    "state_key": "scheduled",
                    "details": "Scheduled for today. Attendance has not been recorded yet.",
                    "highlight": False,
                }
            return {
                "date": work_date,
                "label": "Scheduled",
                "tone": "gray",
                "state_key": "scheduled",
                "details": f"Shift {effective_user['shift_start'] or DEFAULT_SHIFT_START} - {effective_user['shift_end'] or DEFAULT_SHIFT_END}",
                "highlight": False,
            }

        return {
            "date": work_date,
            "label": "Off Day",
            "tone": "gray",
            "state_key": "off_day",
            "details": "Not part of your scheduled workdays.",
            "highlight": False,
        }

    weeks = []
    calendar_weeks = calendar.Calendar(firstweekday=0).monthdatescalendar(year, month)
    for week in calendar_weeks:
        week_entries = []
        for day in week:
            work_date = day.strftime("%Y-%m-%d")
            in_month = day.month == month
            entry = {
                "date": work_date,
                "day_number": day.day,
                "in_month": in_month,
                "is_today": work_date == today_text,
                "label": "",
                "tone": "gray",
                "state_key": "outside",
                "details": "",
                "highlight": False,
            }
            if in_month:
                entry.update(build_entry(work_date))
                if entry["state_key"] in {"worked", "late", "undertime"}:
                    counts["worked"] += 1
                elif entry["state_key"] == "leave":
                    counts["leave"] += 1
                elif entry["state_key"] == "suspension":
                    counts["suspension"] += 1
                elif entry["state_key"] == "absent":
                    counts["absent"] += 1
                elif entry["state_key"] in {"off_day", "holiday", "rest_day"}:
                    counts["off_day"] += 1

                if entry["state_key"] == "late":
                    counts["late"] += 1
                if entry["state_key"] == "undertime":
                    counts["undertime"] += 1
                push_highlight(entry)
            week_entries.append(entry)
        weeks.append(week_entries)

    prev_year, prev_month = shift_month(year, month, -1)
    next_year, next_month = shift_month(year, month, 1)
    highlights.sort(key=lambda item: item["date"], reverse=True)

    return {
        "month_label": month_start.strftime("%B %Y"),
        "weeks": weeks,
        "counts": counts,
        "highlights": highlights[:10],
        "prev_year": prev_year,
        "prev_month": prev_month,
        "next_year": next_year,
        "next_month": next_month,
        "year": year,
        "month": month,
        "weekday_labels": [label for _, label in WEEKDAY_OPTIONS],
    }


def save_payroll_run_snapshot(date_from, date_to, department_filter="", employee_filter="", status="Draft", notes="", actor_id=None):
    status = status if status in {"Draft", "Released"} else "Draft"
    date_from_text = payroll_date_text(date_from)
    date_to_text = payroll_date_text(date_to)
    rows = build_payroll_rows(date_from, date_to, department_filter=department_filter, employee_filter=employee_filter)
    snapshot_adjustments = build_effective_payroll_adjustments(
        date_from_text,
        date_to_text,
        department_filter=department_filter,
        employee_filter=employee_filter
    )
    timestamp = now_str()
    existing = get_payroll_run(date_from_text, date_to_text, department_filter, employee_filter)
    if existing and existing["status"] == "Released" and status != "Released":
        raise ValueError("This payroll period is already released. Use Release Payroll to refresh it instead of saving it back to draft.")
    create_new_run = bool(existing and existing["status"] == "Released" and status == "Released")

    db = get_db()

    def run(query, params=()):
        if using_postgres():
            with db.cursor() as cur:
                cur.execute(convert_query(query), params)
        else:
            db.execute(query, params)

    try:
        released_at = timestamp if status == "Released" else None
        if existing and not create_new_run:
            run("""
                UPDATE payroll_runs
                SET status = ?, notes = ?, updated_at = ?, released_at = ?
                WHERE id = ?
            """, (
                status,
                notes or None,
                timestamp,
                released_at,
                existing["id"]
            ))
            payroll_run_id = existing["id"]
            run("DELETE FROM payroll_run_item_adjustments WHERE payroll_run_id = ?", (payroll_run_id,))
            run("DELETE FROM payroll_run_items WHERE payroll_run_id = ?", (payroll_run_id,))
        else:
            run("""
                INSERT INTO payroll_runs (
                    date_from, date_to, department_filter, employee_filter,
                    status, notes, created_by, created_at, updated_at, released_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                date_from_text,
                date_to_text,
                department_filter or "",
                str(employee_filter or ""),
                status,
                notes or None,
                actor_id,
                timestamp,
                timestamp,
                released_at
            ))
            created = get_payroll_run(date_from_text, date_to_text, department_filter, employee_filter)
            payroll_run_id = created["id"]

        for row in rows:
            run("""
                INSERT INTO payroll_run_items (
                    payroll_run_id, user_id, full_name, department, position,
                    hourly_rate, days_worked, total_hours, overtime_hours,
                    late_minutes, break_minutes, suspension_days, suspension_pay,
                    gross_pay, overtime_pay, allowances, deductions, final_pay, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                payroll_run_id,
                row["user_id"],
                row["full_name"],
                row["department"],
                row["position"],
                row["hourly_rate"],
                row["days_worked"],
                row["total_hours"],
                row["overtime_hours"],
                row["late_minutes"],
                row["break_minutes"],
                row["suspension_days"],
                row["suspension_pay"],
                row["gross_pay"],
                row["overtime_pay"],
                row["allowances"],
                row["deductions"],
                row["final_pay"],
                timestamp
            ))

        for adjustment in snapshot_adjustments:
            run("""
                INSERT INTO payroll_run_item_adjustments (
                    payroll_run_id, user_id, source_kind, source_rule_id, recurrence_type,
                    adjustment_type, label, amount, notes, created_by,
                    created_by_name, adjustment_created_at, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                payroll_run_id,
                adjustment["user_id"],
                adjustment.get("source_kind") or "Manual",
                adjustment.get("source_rule_id"),
                adjustment.get("recurrence_type") or None,
                adjustment["adjustment_type"],
                adjustment["label"],
                round(float(adjustment["amount"] or 0), 2),
                adjustment.get("notes") or None,
                adjustment.get("created_by"),
                adjustment.get("created_by_name") or "Administrator",
                adjustment.get("created_at") or timestamp,
                timestamp
            ))
        db.commit()
    except Exception:
        db.rollback()
        raise

    return get_payroll_run(date_from_text, date_to_text, department_filter, employee_filter), rows


def build_payroll_stats(payroll_rows):
    return {
        "employees": len(payroll_rows),
        "paid_employees": len([row for row in payroll_rows if row["gross_pay"] > 0]),
        "missing_rates": len([row for row in payroll_rows if row["has_rate"] == 0]),
        "total_hours": round(sum(row["total_hours"] for row in payroll_rows), 2),
        "total_overtime_hours": round(sum(row["overtime_hours"] for row in payroll_rows), 2),
        "total_overtime_pay": round(sum(row["overtime_pay"] for row in payroll_rows), 2),
        "total_gross": round(sum(row["gross_pay"] for row in payroll_rows), 2),
        "total_allowances": round(sum(row["allowances"] for row in payroll_rows), 2),
        "total_deductions": round(sum(row["deductions"] for row in payroll_rows), 2),
        "total_final_pay": round(sum(row["final_pay"] for row in payroll_rows), 2),
        "total_net_estimate": round(sum(row["net_pay_estimate"] for row in payroll_rows), 2),
        "suspension_days": sum(row["suspension_days"] for row in payroll_rows),
        "suspension_pay": round(sum(row["suspension_pay"] for row in payroll_rows), 2),
        "overtime_multiplier": get_overtime_multiplier(),
    }


def normalize_payroll_filters(period_value, date_from_value="", date_to_value="", department_filter="", employee_filter=""):
    period = (period_value or "this_month").strip() or "this_month"
    department_filter = (department_filter or "").strip()
    employee_filter = (employee_filter or "").strip()
    if employee_filter and not employee_filter.isdigit():
        employee_filter = ""
    date_from, date_to = get_payroll_period_dates(period, date_from_value, date_to_value)
    return {
        "period": period,
        "department_filter": department_filter,
        "employee_filter": employee_filter,
        "date_from": date_from,
        "date_to": date_to,
        "date_from_text": payroll_date_text(date_from),
        "date_to_text": payroll_date_text(date_to),
    }


def payroll_filter_redirect_args(source):
    normalized = normalize_payroll_filters(
        source.get("period", "this_month"),
        source.get("date_from", ""),
        source.get("date_to", ""),
        source.get("department", ""),
        source.get("employee_id", "")
    )
    return {
        "period": normalized["period"],
        "date_from": normalized["date_from_text"],
        "date_to": normalized["date_to_text"],
        "department": normalized["department_filter"],
        "employee_id": normalized["employee_filter"],
    }


def get_payroll_employee_filter_label(employee_filter):
    if str(employee_filter or "").isdigit():
        employee = get_user_by_id(int(employee_filter))
        return employee["full_name"] if employee else str(employee_filter)
    return "All Employees"


def paginate_items(items, page, page_size):
    total = len(items)
    page_size = page_size if page_size in {10, 25, 50, 100} else 25
    total_pages = max((total + page_size - 1) // page_size, 1)
    page = max(min(page, total_pages), 1)
    start = (page - 1) * page_size
    end = start + page_size
    return {
        "items": items[start:end],
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": total_pages,
        "start_index": start + 1 if total else 0,
        "end_index": min(end, total),
        "has_prev": page > 1,
        "has_next": page < total_pages,
    }


def get_admin_users():
    return fetchall("""
        SELECT id, full_name
        FROM users
        WHERE role = 'admin' AND is_active = 1
        ORDER BY full_name ASC
    """)


def create_admin_alert_once(title, message):
    today_start = f"{today_str()} 00:00:00"
    for admin in get_admin_users():
        existing = fetchone("""
            SELECT id
            FROM notifications
            WHERE user_id = ? AND title = ? AND message = ? AND created_at >= ?
            ORDER BY id DESC LIMIT 1
        """, (admin["id"], title, message, today_start))
        if not existing:
            create_notification(admin["id"], title, message)


def build_payroll_rows(date_from, date_to, department_filter="", employee_filter=""):
    date_from_text = payroll_date_text(date_from)
    date_to_text = payroll_date_text(date_to)
    employees_sql = """
        SELECT *
        FROM users
        WHERE role = 'employee'
    """
    params = []
    if department_filter:
        employees_sql += " AND department = ?"
        params.append(department_filter)
    if employee_filter:
        employees_sql += " AND id = ?"
        params.append(employee_filter)
    employees_sql += " ORDER BY full_name ASC"

    employees = fetchall(employees_sql, params)
    employee_map = {}
    for employee in employees:
        employee_map[employee["id"]] = {
            "user_id": employee["id"],
            "full_name": employee["full_name"],
            "username": employee["username"],
            "department": employee["department"] or "",
            "position": employee["position"] or "",
            "hourly_rate": float(employee["hourly_rate"] or 0),
            "days_worked": 0,
            "total_minutes": 0,
            "late_minutes": 0,
            "break_minutes": 0,
            "overtime_minutes": 0,
            "overtime_hours": 0,
            "overtime_pay": 0,
            "gross_pay": 0,
            "allowances": 0,
            "deductions": 0,
            "net_pay_estimate": 0,
            "final_pay": 0,
            "suspension_days": 0,
            "suspension_hours": 0,
            "suspension_pay": 0,
            "has_rate": 1 if float(employee["hourly_rate"] or 0) > 0 else 0,
            "is_active": employee["is_active"],
            "schedule_days": employee["schedule_days"] or DEFAULT_SCHEDULE_DAYS,
            "shift_start": employee["shift_start"] or DEFAULT_SHIFT_START,
            "shift_end": employee["shift_end"] or DEFAULT_SHIFT_END,
        }

    attendance_rows = fetchall("""
        SELECT a.*, u.department
        FROM attendance a
        JOIN users u ON u.id = a.user_id
        WHERE u.role = 'employee'
          AND a.work_date BETWEEN ? AND ?
          AND a.time_in IS NOT NULL
          AND a.time_out IS NOT NULL
        ORDER BY a.work_date ASC, a.id ASC
    """, (date_from_text, date_to_text))

    for attendance in attendance_rows:
        summary = employee_map.get(attendance["user_id"])
        if not summary:
            continue
        minutes_worked = max(total_work_minutes(attendance), 0)
        summary["days_worked"] += 1
        summary["total_minutes"] += minutes_worked
        summary["late_minutes"] += int(attendance["late_minutes"] or 0)
        summary["break_minutes"] += total_break_minutes(attendance["id"])

    overtime_rows = fetchall("""
        SELECT *
        FROM overtime_sessions
        WHERE work_date BETWEEN ? AND ?
    """, (date_from_text, date_to_text))
    overtime_multiplier = get_overtime_multiplier()
    for overtime in overtime_rows:
        summary = employee_map.get(overtime["user_id"])
        if not summary:
            continue
        overtime_minutes = overtime_minutes_for_session(overtime)
        summary["overtime_minutes"] += overtime_minutes

    suspension_rows = fetchall("""
        SELECT *
        FROM disciplinary_actions
        WHERE action_type = 'Suspension'
          AND action_date <= ?
          AND COALESCE(end_date, action_date) >= ?
    """, (date_to_text, date_from_text))
    for suspension in suspension_rows:
        summary = employee_map.get(suspension["user_id"])
        if not summary:
            continue
        employee_record = get_user_by_id(suspension["user_id"])
        for suspension_date in expand_suspension_dates(suspension):
            if suspension_date < date_from_text or suspension_date > date_to_text:
                continue
            employee_stub = get_effective_employee_context(
                user_row=employee_record,
                user_id=suspension["user_id"],
                reference_date=suspension_date
            ) or {
                "schedule_days": summary["schedule_days"],
                "shift_start": summary["shift_start"],
                "shift_end": summary["shift_end"],
            }
            if not is_scheduled_on_date(employee_stub, suspension_date):
                continue
            shift_minutes = get_scheduled_shift_minutes(employee_stub, suspension_date)
            summary["suspension_days"] += 1
            summary["suspension_hours"] += round(shift_minutes / 60, 2)

    adjustment_totals = {}
    for adjustment in build_effective_payroll_adjustments(date_from_text, date_to_text, department_filter=department_filter, employee_filter=employee_filter):
        totals = adjustment_totals.setdefault(adjustment["user_id"], {"allowances": 0.0, "deductions": 0.0})
        amount = round(float(adjustment["amount"] or 0), 2)
        if adjustment["adjustment_type"] == "Allowance":
            totals["allowances"] += amount
        else:
            totals["deductions"] += amount

    for summary in employee_map.values():
        summary["total_hours"] = round(summary["total_minutes"] / 60, 2)
        summary["overtime_hours"] = round(summary["overtime_minutes"] / 60, 2)
        summary["gross_pay"] = round(summary["total_hours"] * summary["hourly_rate"], 2)
        summary["overtime_pay"] = round(summary["overtime_hours"] * summary["hourly_rate"] * overtime_multiplier, 2)
        adjustment_summary = adjustment_totals.get(summary["user_id"], {"allowances": 0.0, "deductions": 0.0})
        summary["allowances"] = round(adjustment_summary["allowances"], 2)
        summary["deductions"] = round(adjustment_summary["deductions"], 2)
        summary["final_pay"] = round(summary["gross_pay"] + summary["overtime_pay"] + summary["allowances"] - summary["deductions"], 2)
        summary["net_pay_estimate"] = summary["final_pay"]
        summary["suspension_hours"] = round(summary["suspension_hours"], 2)
        summary["suspension_pay"] = round(summary["suspension_hours"] * summary["hourly_rate"], 2)
        summary["status_label"] = "Ready" if summary["has_rate"] else "Missing Rate"

    return sorted(employee_map.values(), key=lambda item: (-item["final_pay"], item["full_name"].lower()))


# =========================
# GOOGLE SHEETS SYNC (OPTIONAL)
# =========================
def append_attendance_to_google_sheet(user_row, attendance_row):
    if not GOOGLE_SHEETS_ENABLED:
        return False, "Google Sheets libraries not installed."

    if not os.path.exists(GOOGLE_CREDENTIALS_FILE):
        return False, "attendance-credentials.json not found."

    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_file(GOOGLE_CREDENTIALS_FILE, scopes=scopes)
        client = gspread.authorize(creds)
        sheet = client.open(GOOGLE_SHEET_NAME)

        try:
            ws = sheet.worksheet(GOOGLE_SHEET_TAB)
        except Exception:
            ws = sheet.add_worksheet(title=GOOGLE_SHEET_TAB, rows=1000, cols=20)

        values = ws.get_all_values()
        if not values:
            ws.append_row([
                "Employee ID", "Full Name", "Username", "Department", "Position", "Shift Start",
                "Work Date", "Time In", "Time Out", "Status",
                "Late", "Late Minutes", "Break Minutes", "Work Minutes",
                "Notes", "Proof File"
            ])

        ws.append_row([
            user_row["id"],
            user_row["full_name"],
            user_row["username"],
            user_row["department"] or "",
            user_row["position"] or "",
            user_row["shift_start"] or DEFAULT_SHIFT_START,
            attendance_row["work_date"] or "",
            attendance_row["time_in"] or "",
            attendance_row["time_out"] or "",
            attendance_row["status"] or "",
            "YES" if attendance_row["late_flag"] else "NO",
            attendance_row["late_minutes"] or 0,
            total_break_minutes(attendance_row["id"]),
            total_work_minutes(attendance_row),
            attendance_row["notes"] or "",
            attendance_row["proof_file"] or ""
        ])
        return True, "Synced to Google Sheets."
    except Exception as e:
        return False, str(e)


# =========================
# TEMPLATE GLOBALS
# =========================
@app.context_processor
def inject_globals():
    user = None
    unread_count = 0
    latest_notifications = []

    if session.get("user_id"):
        user = get_user_by_id(session["user_id"])
        if user:
            unread = fetchone("""
                SELECT COUNT(*) AS cnt FROM notifications
                WHERE user_id = ? AND is_read = 0
            """, (session["user_id"],))
            unread_count = unread["cnt"] if unread else 0
            latest_notifications = fetchall("""
                SELECT *
                FROM notifications
                WHERE user_id = ?
                ORDER BY id DESC
                LIMIT 6
            """, (session["user_id"],))

    return dict(
        current_user=user,
        company_settings=get_company_settings(),
        unread_count=unread_count,
        latest_notifications=latest_notifications,
        admin_can=admin_can_permission,
        admin_permission_options=ADMIN_PERMISSION_OPTIONS,
        admin_role_preset_options=ADMIN_ROLE_PRESET_OPTIONS,
        describe_admin_permissions=describe_admin_permissions,
        get_admin_role_preset_meta=get_admin_role_preset_meta,
        is_image=is_image,
        static_file_exists=static_file_exists,
        uploaded_file_exists=uploaded_file_exists,
        get_avatar_initials=get_avatar_initials,
        get_employee_card_number=get_employee_card_number,
        generate_code128_svg_data_uri=generate_code128_svg_data_uri,
        format_datetime_12h=format_datetime_12h,
        format_time_12h=format_time_12h,
        format_currency=format_currency
    )


# =========================
# AUTH / HOME
# =========================
@app.route("/health")
def health():
    return jsonify({"status": "ok", "time": now_str()}), 200


@app.route("/")
def home():
    if "user_id" in session:
        user = get_user_by_id(session["user_id"])
        if not user:
            session.clear()
            return redirect(url_for("login"))
        return redirect(get_home_route_for_user(user))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    existing_user_id = session.get("user_id")
    if existing_user_id:
        existing_user = get_user_by_id(existing_user_id)
        if existing_user and int(existing_user["is_active"] or 0) == 1:
            session["role"] = existing_user["role"]
            session["full_name"] = existing_user["full_name"]
            return redirect(get_home_route_for_user(existing_user))
        session.clear()

    if request.method == "POST":
        client_ip = get_client_ip()
        if is_login_rate_limited(client_ip):
            flash(f"Too many login attempts. Please wait {LOGIN_WINDOW_MINUTES} minutes and try again.", "danger")
            return render_template("login.html"), 429

        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        user = fetchone("""
            SELECT * FROM users
            WHERE username = ? AND is_active = 1
        """, (username,))

        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            session["full_name"] = user["full_name"]
            clear_login_failures(client_ip)

            log_activity(user["id"], "LOGIN", f"{user['full_name']} logged in")

            if user["role"] == "admin":
                flash("Welcome Admin.", "success")
                return redirect(get_home_route_for_user(user))

            if user["role"] == "scanner":
                flash("Scanner kiosk ready.", "success")
                return redirect(url_for("scanner_kiosk"))

            flash("Login successful.", "success")
            return redirect(url_for("dashboard"))

        register_login_failure(client_ip)
        flash("Invalid username or password.", "danger")

    return render_template("login.html")


@app.route("/logout", methods=["POST"])
@login_required()
def logout():
    user_id = session.get("user_id")
    if user_id:
        log_activity(user_id, "LOGOUT", "User logged out")
    session.clear()
    flash("Logged out successfully.", "info")
    return redirect(url_for("login"))


# =========================
# EMPLOYEE
# =========================
@app.route("/dashboard")
@login_required(role="employee")
def dashboard():
    user = get_user_by_id(session["user_id"])
    if not user:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    today_attendance = get_current_attendance(user["id"])
    open_break = get_open_break(user["id"], today_attendance)

    notifications = fetchall("""
        SELECT * FROM notifications
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT 10
    """, (user["id"],))

    logs = fetchall("""
        SELECT * FROM activity_logs
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT 10
    """, (user["id"],))

    current_status = get_user_live_status(user["id"])
    override_status = get_employee_override_status_for_date(user["id"], today_str())
    if override_status:
        current_status = override_status["label"]
    todays_break_minutes = total_break_minutes(today_attendance["id"], include_open=True) if today_attendance else 0
    todays_work_minutes = total_work_minutes(today_attendance) if today_attendance else 0
    todays_break_sessions = build_break_sessions(today_attendance["id"]) if today_attendance else []
    break_limit_minutes = get_employee_break_limit(user)
    over_break_minutes = get_overbreak_minutes(todays_break_minutes, break_limit_minutes)
    leave_summary = get_leave_balance_summary(user)
    latest_payslip = None
    released_runs = get_employee_released_payroll_runs(user["id"], limit=1)
    if released_runs:
        latest_payslip = released_runs[0]

    return render_template(
        "employee_dashboard.html",
        user=user,
        today_attendance=today_attendance,
        notifications=notifications,
        current_status=current_status,
        override_status=override_status,
        todays_break_minutes=todays_break_minutes,
        todays_work_minutes=todays_work_minutes,
        todays_break_sessions=todays_break_sessions,
        leave_summary=leave_summary,
        break_limit_minutes=break_limit_minutes,
        over_break_minutes=over_break_minutes,
        minutes_to_hm=minutes_to_hm,
        latest_payslip=latest_payslip,
        current_calendar_year=now_dt().year,
        current_calendar_month=now_dt().month,
        current_calendar_label=now_dt().strftime("%B %Y")
    )


@app.route("/actions")
@login_required(role="employee")
def employee_actions():
    user = get_user_by_id(session["user_id"])
    if not user:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    manual_attendance_block_message = get_manual_attendance_block_message()
    if manual_attendance_block_message:
        flash(manual_attendance_block_message, "info")
        return redirect(url_for("dashboard"))

    return render_template(
        "employee_actions.html",
        user=user,
        manual_attendance_block_message=manual_attendance_block_message
    )


@app.route("/activity")
@login_required(role="employee")
def employee_activity():
    user = get_user_by_id(session["user_id"])
    if not user:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    logs = fetchall("""
        SELECT * FROM activity_logs
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT 50
    """, (user["id"],))

    return render_template("employee_activity.html", user=user, logs=logs)


@app.route("/notifications")
@login_required()
def notifications_page():
    user = get_user_by_id(session["user_id"])
    notifications = fetchall("""
        SELECT *
        FROM notifications
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT 100
    """, (session["user_id"],))

    return render_template(
        "notifications.html",
        user=user,
        notifications=notifications
    )


@app.route("/history")
@login_required(role="employee")
def employee_history():
    user = get_user_by_id(session["user_id"])
    records = build_employee_history_records(user, limit=60)
    return render_template("employee_history.html", records=records, minutes_to_hm=minutes_to_hm)


@app.route("/my-payroll")
@login_required(role="employee")
def employee_payroll_history():
    user = get_user_by_id(session["user_id"])
    if not user:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    payroll_runs = get_employee_released_payroll_runs(user["id"], limit=24)
    return render_template(
        "employee_payroll_history.html",
        user=user,
        payroll_runs=payroll_runs
    )


@app.route("/my-payroll/<int:payroll_run_id>")
@login_required(role="employee")
def employee_payslip(payroll_run_id):
    user = get_user_by_id(session["user_id"])
    if not user:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    payslip = get_employee_released_payroll_item(user["id"], payroll_run_id)
    if not payslip:
        flash("Released payslip not found for your account.", "warning")
        return redirect(url_for("employee_payroll_history"))

    return render_template(
        "employee_payslip.html",
        user=user,
        payslip=payslip
    )


@app.route("/my-payroll/<int:payroll_run_id>/download.pdf")
@login_required(role="employee")
def download_employee_payslip_pdf(payroll_run_id):
    user = get_user_by_id(session["user_id"])
    if not user:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    payslip = get_employee_released_payroll_item(user["id"], payroll_run_id)
    if not payslip:
        flash("Released payslip not found for your account.", "warning")
        return redirect(url_for("employee_payroll_history"))

    pdf_bytes = build_employee_payslip_pdf_bytes(
        payslip,
        printed_at_text=format_datetime_12h(now_str()),
    )
    filename = build_employee_payslip_pdf_filename(payslip)
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(filename)}",
            "Cache-Control": "no-store",
        },
    )


@app.route("/attendance-calendar")
@login_required(role="employee")
def employee_attendance_calendar():
    user = get_user_by_id(session["user_id"])
    if not user:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    year, month = parse_calendar_month(
        request.args.get("year", ""),
        request.args.get("month", "")
    )
    calendar_data = build_employee_attendance_calendar(user, year, month)
    return render_template(
        "employee_attendance_calendar.html",
        user=user,
        calendar_data=calendar_data
    )


@app.route("/profile", methods=["GET", "POST"])
@login_required(role="employee")
def employee_profile():
    user = get_user_by_id(session["user_id"])
    if not user:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        if not full_name:
            flash("Full name is required.", "danger")
            return redirect(url_for("employee_profile"))

        password = request.form.get("password", "").strip()

        profile_image = user["profile_image"]
        file = request.files.get("profile_image")
        if file and file.filename:
            saved = save_uploaded_file(file, prefix=f"profile_{user['id']}", allowed_exts=IMAGE_EXTENSIONS)
            if not saved:
                flash("Invalid profile image type.", "danger")
                return redirect(url_for("employee_profile"))
            profile_image = saved

        if password:
            execute_db("""
                UPDATE users
                SET full_name = ?, password_hash = ?, profile_image = ?
                WHERE id = ?
            """, (full_name, generate_password_hash(password), profile_image, user["id"]), commit=True)
        else:
            execute_db("""
                UPDATE users
                SET full_name = ?, profile_image = ?
                WHERE id = ?
            """, (full_name, profile_image, user["id"]), commit=True)

        session["full_name"] = full_name
        log_activity(user["id"], "UPDATE PROFILE", "Employee updated profile")
        flash("Profile updated successfully.", "success")
        return redirect(url_for("employee_profile"))

    return render_template("employee_profile.html", user=user)


@app.route("/admin/profile", methods=["GET", "POST"])
@login_required(role="admin")
def admin_profile():
    user = get_user_by_id(session["user_id"])
    if not user or user["role"] != "admin":
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    if request.method == "POST":
        form_action = request.form.get("form_action", "admin_profile").strip()

        if form_action == "create_admin_account":
            if not admin_has_permission(user, "settings"):
                flash("Your admin account cannot manage other admin accounts.", "danger")
                return redirect(url_for("admin_profile"))
            admin_full_name = (request.form.get("admin_full_name", "") or "").strip()
            admin_username = (request.form.get("admin_username", "") or "").strip()
            admin_password = (request.form.get("admin_password", "") or "").strip()
            admin_role_preset = normalize_admin_role_preset(request.form.get("admin_role_preset", ""))
            admin_permission_values = request.form.getlist("admin_permissions")
            if not admin_permission_values and admin_role_preset:
                admin_permission_values = list(ADMIN_ROLE_PRESETS[admin_role_preset]["permissions"])
            admin_role_preset, admin_permissions = sync_admin_role_preset(admin_role_preset, admin_permission_values)

            if not admin_full_name or not admin_username or not admin_password:
                flash("Admin name, username, and password are required.", "danger")
                return redirect(url_for("admin_profile"))
            if not admin_permissions:
                flash("Choose at least one permission for the new admin account.", "danger")
                return redirect(url_for("admin_profile"))
            if fetchone("SELECT id FROM users WHERE username = ?", (admin_username,)):
                flash("That username is already in use.", "warning")
                return redirect(url_for("admin_profile"))

            execute_db("""
                INSERT INTO users (
                    full_name, username, password_hash, role, department, position,
                    admin_permissions, admin_role_preset, shift_start, break_limit_minutes, is_active, created_at
                )
                VALUES (?, ?, ?, 'admin', ?, ?, ?, ?, ?, ?, 1, ?)
            """, (
                admin_full_name,
                admin_username,
                generate_password_hash(admin_password),
                "Stellar Seats",
                "Operations Admin",
                admin_permissions,
                admin_role_preset or None,
                DEFAULT_SHIFT_START,
                BREAK_LIMIT_MINUTES,
                now_str()
            ), commit=True)
            preset_meta = get_admin_role_preset_meta(preset_code=admin_role_preset, permission_values=admin_permissions)
            log_activity(session["user_id"], "CREATE ADMIN ACCOUNT", f"Created admin account {admin_username} with {preset_meta['label']} ({admin_permissions})")
            flash("Admin account created.", "success")
            return redirect(url_for("admin_profile"))

        if form_action == "update_admin_account":
            if not admin_has_permission(user, "settings"):
                flash("Your admin account cannot manage other admin accounts.", "danger")
                return redirect(url_for("admin_profile"))
            admin_id_raw = (request.form.get("admin_id", "") or "").strip()
            if not admin_id_raw.isdigit():
                flash("Admin account not found.", "warning")
                return redirect(url_for("admin_profile"))

            target_admin = fetchone("SELECT * FROM users WHERE id = ? AND role = 'admin'", (int(admin_id_raw),))
            if not target_admin:
                flash("Admin account not found.", "warning")
                return redirect(url_for("admin_profile"))

            if int(target_admin["id"]) == int(session["user_id"]):
                flash("Use the profile form above to update your own main account details.", "info")
                return redirect(url_for("admin_profile"))

            admin_full_name = (request.form.get("admin_full_name", "") or "").strip()
            admin_username = (request.form.get("admin_username", "") or "").strip()
            admin_password = (request.form.get("admin_password", "") or "").strip()
            admin_role_preset = normalize_admin_role_preset(request.form.get("admin_role_preset", ""))
            admin_permission_values = request.form.getlist("admin_permissions")
            if not admin_permission_values and admin_role_preset:
                admin_permission_values = list(ADMIN_ROLE_PRESETS[admin_role_preset]["permissions"])
            admin_role_preset, admin_permissions = sync_admin_role_preset(admin_role_preset, admin_permission_values)
            is_active = 1 if request.form.get("is_active") == "1" else 0

            if not admin_full_name or not admin_username:
                flash("Admin name and username are required.", "danger")
                return redirect(url_for("admin_profile"))
            if not admin_permissions:
                flash("Choose at least one permission for the admin account.", "danger")
                return redirect(url_for("admin_profile"))

            username_owner = fetchone("SELECT id FROM users WHERE username = ? AND id != ?", (admin_username, target_admin["id"]))
            if username_owner:
                flash("That username is already in use by another account.", "warning")
                return redirect(url_for("admin_profile"))

            if admin_password:
                execute_db("""
                    UPDATE users
                    SET full_name = ?, username = ?, password_hash = ?, admin_permissions = ?, admin_role_preset = ?, is_active = ?
                    WHERE id = ?
                """, (
                    admin_full_name,
                    admin_username,
                    generate_password_hash(admin_password),
                    admin_permissions,
                    admin_role_preset or None,
                    is_active,
                    target_admin["id"]
                ), commit=True)
            else:
                execute_db("""
                    UPDATE users
                    SET full_name = ?, username = ?, admin_permissions = ?, admin_role_preset = ?, is_active = ?
                    WHERE id = ?
                """, (
                    admin_full_name,
                    admin_username,
                    admin_permissions,
                    admin_role_preset or None,
                    is_active,
                    target_admin["id"]
                ), commit=True)
            preset_meta = get_admin_role_preset_meta(preset_code=admin_role_preset, permission_values=admin_permissions)
            log_activity(session["user_id"], "UPDATE ADMIN ACCOUNT", f"Updated admin account {admin_username} with {preset_meta['label']} ({admin_permissions})")
            flash("Admin account updated.", "success")
            return redirect(url_for("admin_profile"))

        if form_action == "scanner_account":
            if not admin_has_permission(user, "settings"):
                flash("Your admin account cannot change scanner settings.", "danger")
                return redirect(url_for("admin_profile"))
            scanner_full_name = request.form.get("scanner_full_name", "").strip() or "Scanner Kiosk"
            scanner_username = request.form.get("scanner_username", "").strip()
            scanner_password = request.form.get("scanner_password", "").strip()

            if not scanner_username:
                flash("Scanner username is required.", "danger")
                return redirect(url_for("admin_profile"))

            existing_scanner = get_scanner_account()
            username_owner = fetchone("SELECT id, role FROM users WHERE username = ?", (scanner_username,))
            if username_owner and (not existing_scanner or username_owner["id"] != existing_scanner["id"]):
                flash("That scanner username is already in use.", "warning")
                return redirect(url_for("admin_profile"))

            if existing_scanner:
                if scanner_password:
                    execute_db("""
                        UPDATE users
                        SET full_name = ?, username = ?, password_hash = ?
                        WHERE id = ?
                    """, (scanner_full_name, scanner_username, generate_password_hash(scanner_password), existing_scanner["id"]), commit=True)
                    log_activity(session["user_id"], "UPDATE SCANNER ACCOUNT", f"Updated scanner account {scanner_username} and reset password")
                else:
                    execute_db("""
                        UPDATE users
                        SET full_name = ?, username = ?
                        WHERE id = ?
                    """, (scanner_full_name, scanner_username, existing_scanner["id"]), commit=True)
                    log_activity(session["user_id"], "UPDATE SCANNER ACCOUNT", f"Updated scanner account {scanner_username}")
                flash("Scanner account updated.", "success")
            else:
                if not scanner_password:
                    flash("Scanner password is required when creating the scanner account.", "danger")
                    return redirect(url_for("admin_profile"))
                execute_db("""
                    INSERT INTO users (
                        full_name, username, password_hash, role, department, position,
                        break_limit_minutes, is_active, created_at
                    )
                    VALUES (?, ?, ?, 'scanner', ?, ?, ?, 1, ?)
                """, (
                    scanner_full_name,
                    scanner_username,
                    generate_password_hash(scanner_password),
                    "Kiosk",
                    "Scanner Only",
                    BREAK_LIMIT_MINUTES,
                    now_str()
                ), commit=True)
                log_activity(session["user_id"], "CREATE SCANNER ACCOUNT", f"Created scanner account {scanner_username}")
                flash("Scanner account created.", "success")
            return redirect(url_for("admin_profile"))

        if form_action == "attendance_settings":
            if not admin_has_permission(user, "settings"):
                flash("Your admin account cannot change attendance settings.", "danger")
                return redirect(url_for("admin_profile"))
            scanner_attendance_mode = 1 if request.form.get("scanner_attendance_mode") == "1" else 0
            scanner_lock_timeout_seconds = parse_positive_int(request.form.get("scanner_lock_timeout_seconds", "90"), 90)
            scanner_lock_timeout_seconds = max(min(scanner_lock_timeout_seconds, 900), 15)
            overtime_multiplier_raw = (request.form.get("overtime_multiplier", "") or "").strip()
            try:
                overtime_multiplier = float(overtime_multiplier_raw or 1.25)
            except ValueError:
                flash("Overtime multiplier must be a valid number.", "danger")
                return redirect(url_for("admin_profile"))
            overtime_multiplier = max(min(overtime_multiplier, 5.0), 1.0)

            scanner_exit_pin = (request.form.get("scanner_exit_pin", "") or "").strip()
            current_settings = get_company_settings()
            scanner_exit_pin_hash = current_settings.get("scanner_exit_pin_hash")
            if scanner_exit_pin:
                if len(scanner_exit_pin) < 4:
                    flash("Scanner kiosk PIN must be at least 4 characters.", "danger")
                    return redirect(url_for("admin_profile"))
                scanner_exit_pin_hash = generate_password_hash(scanner_exit_pin)
            elif request.form.get("clear_scanner_exit_pin") == "1":
                scanner_exit_pin_hash = None

            execute_db("""
                INSERT INTO company_settings (
                    id, id_signatory_name, id_signatory_title, id_signature_file,
                    scanner_attendance_mode, scanner_lock_timeout_seconds, scanner_exit_pin_hash, overtime_multiplier
                )
                VALUES (1, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(id) DO UPDATE SET
                    id_signatory_name = excluded.id_signatory_name,
                    id_signatory_title = excluded.id_signatory_title,
                    id_signature_file = excluded.id_signature_file,
                    scanner_attendance_mode = excluded.scanner_attendance_mode,
                    scanner_lock_timeout_seconds = excluded.scanner_lock_timeout_seconds,
                    scanner_exit_pin_hash = excluded.scanner_exit_pin_hash,
                    overtime_multiplier = excluded.overtime_multiplier
            """, (
                current_settings.get("id_signatory_name") or "Kirk Danny Fernandez",
                current_settings.get("id_signatory_title") or "Head Of Operations",
                current_settings.get("id_signature_file"),
                scanner_attendance_mode,
                scanner_lock_timeout_seconds,
                scanner_exit_pin_hash,
                overtime_multiplier
            ), commit=True)
            log_activity(session["user_id"], "UPDATE ATTENDANCE SETTINGS", f"Scanner mode {'enabled' if scanner_attendance_mode else 'disabled'}, kiosk timeout {scanner_lock_timeout_seconds}s, overtime multiplier {overtime_multiplier:.2f}x")
            flash("Attendance and kiosk settings updated.", "success")
            return redirect(url_for("admin_profile"))

        full_name = request.form.get("full_name", "").strip()
        if not full_name:
            flash("Full name is required.", "danger")
            return redirect(url_for("admin_profile"))

        password = request.form.get("password", "").strip()

        profile_image = user["profile_image"]
        file = request.files.get("profile_image")
        if file and file.filename:
            saved = save_uploaded_file(file, prefix=f"profile_{user['id']}", allowed_exts=IMAGE_EXTENSIONS)
            if not saved:
                flash("Invalid profile image type.", "danger")
                return redirect(url_for("admin_profile"))
            profile_image = saved

        if password:
            execute_db("""
                UPDATE users
                SET full_name = ?, password_hash = ?, profile_image = ?
                WHERE id = ?
            """, (full_name, generate_password_hash(password), profile_image, user["id"]), commit=True)
            log_activity(user["id"], "UPDATE ADMIN PROFILE", "Admin updated profile and password")
        else:
            execute_db("""
                UPDATE users
                SET full_name = ?, profile_image = ?
                WHERE id = ?
            """, (full_name, profile_image, user["id"]), commit=True)
            log_activity(user["id"], "UPDATE ADMIN PROFILE", "Admin updated profile")

        session["full_name"] = full_name
        flash("Admin profile updated successfully.", "success")
        return redirect(url_for("admin_profile"))

    return render_template(
        "admin_profile.html",
        user=user,
        admin_accounts=get_admin_accounts(),
        scanner_account=get_scanner_account(),
        company_settings=get_company_settings(),
        scanner_exit_pin_configured=has_scanner_exit_pin()
    )


@app.route("/corrections", methods=["GET", "POST"])
@login_required(role="employee")
def employee_corrections():
    user = get_user_by_id(session["user_id"])
    if not user:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    if request.method == "POST":
        request_type = request.form.get("request_type", "").strip()
        work_date = request.form.get("work_date", "").strip()
        end_work_date = request.form.get("end_work_date", "").strip()
        message = request.form.get("message", "").strip()
        requested_time_in = request.form.get("requested_time_in", "")
        requested_break_start = request.form.get("requested_break_start", "")
        requested_break_end = request.form.get("requested_break_end", "")
        requested_time_out = request.form.get("requested_time_out", "")

        if request_type not in ATTENDANCE_REQUEST_TYPES:
            flash("Please choose a valid correction type.", "danger")
            return redirect(url_for("employee_corrections"))

        if not work_date or not message:
            flash("Work date and details are required.", "danger")
            return redirect(url_for("employee_corrections"))

        try:
            requested_time_in = normalize_optional_clock_time(requested_time_in)
            requested_break_start = normalize_optional_clock_time(requested_break_start)
            requested_break_end = normalize_optional_clock_time(requested_break_end)
            requested_time_out = normalize_optional_clock_time(requested_time_out)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("employee_corrections"))

        if request_type == "Undertime" and not requested_time_out:
            flash("Requested time out is required for undertime requests.", "danger")
            return redirect(url_for("employee_corrections"))

        if request_type in LEAVE_REQUEST_TYPES:
            try:
                work_date, end_work_date = normalize_request_date_range(work_date, end_work_date or work_date)
            except ValueError as exc:
                flash(str(exc), "danger")
                return redirect(url_for("employee_corrections"))
            requested_time_in = ""
            requested_break_start = ""
            requested_break_end = ""
            requested_time_out = ""

            existing_leave = has_overlapping_leave_request(user["id"], request_type, work_date, end_work_date)
            if existing_leave:
                flash(f"A {request_type.lower()} request already exists in that date range.", "warning")
                return redirect(url_for("employee_corrections"))
        else:
            end_work_date = work_date

        attendance, break_row = get_attendance_context(user["id"], work_date)
        requested_time_in_dt, requested_break_start_dt, requested_break_end_dt, requested_time_out_dt = resolve_correction_datetimes(
            work_date,
            time_in_value=requested_time_in,
            break_start_value=requested_break_start,
            break_end_value=requested_break_end,
            time_out_value=requested_time_out,
            existing_time_in=attendance["time_in"] if attendance else None,
            existing_break_start=break_row["break_start"] if break_row else None,
            existing_break_end=break_row["break_end"] if break_row else None,
            existing_time_out=attendance["time_out"] if attendance else None,
            use_existing_values=False
        )

        execute_db("""
            INSERT INTO correction_requests (
                user_id, request_type, work_date, end_work_date, message,
                requested_time_in, requested_break_start, requested_break_end, requested_time_out,
                status, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending', ?)
        """, (
            user["id"],
            request_type,
            work_date,
            end_work_date,
            message,
            requested_time_in_dt,
            requested_break_start_dt,
            requested_break_end_dt,
            requested_time_out_dt,
            now_str()
        ), commit=True)

        log_activity(user["id"], "CORRECTION REQUEST", f"Submitted {request_type} request for {format_request_date_range(work_date, end_work_date)}")
        flash("Correction request submitted.", "success")
        return redirect(url_for("employee_corrections"))

    requests = get_correction_requests(user_id=user["id"])
    leave_summary = get_leave_balance_summary(user)
    request_summary = build_correction_request_summary(requests)
    return render_template(
        "employee_corrections.html",
        user=user,
        requests=requests,
        leave_summary=leave_summary,
        request_summary=request_summary
    )


@app.route("/time-in", methods=["POST"])
@login_required(role="employee")
def time_in():
    user_id = session["user_id"]
    user = get_user_by_id(user_id)

    manual_attendance_block_message = get_manual_attendance_block_message()
    if manual_attendance_block_message:
        flash(manual_attendance_block_message, "warning")
        return redirect(url_for("employee_actions"))

    if not user:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    existing = auto_close_stale_attendance(user, get_current_attendance(user_id), actor_id=user_id, source_label="Employee portal")
    override_status = get_employee_override_status_for_date(user_id, today_str())
    override_block_message = get_attendance_override_block_message(override_status, "time_in", existing)
    if override_block_message:
        flash(override_block_message, "danger")
        return redirect(url_for("dashboard"))

    if existing and existing["time_in"] and not existing["time_out"]:
        flash("You are already timed in.", "warning")
        return redirect(url_for("dashboard"))

    file = request.files.get("proof_file")
    proof_filename = None

    if file and file.filename:
        proof_filename = save_uploaded_file(file, prefix=f"proof_{user_id}", allowed_exts=IMAGE_EXTENSIONS | DOCUMENT_EXTENSIONS)
        if not proof_filename:
            flash("Invalid upload file type.", "danger")
            return redirect(url_for("dashboard"))

    execute_db("""
        INSERT INTO attendance (
            user_id, work_date, time_in, time_out, status, proof_file, notes,
            late_flag, late_minutes, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        user_id,
        today_str(),
        now_str(),
        None,
        "Timed In",
        proof_filename,
        request.form.get("notes", "").strip(),
        *calculate_late_info(now_str(), parse_shift_start(user["shift_start"] if user else DEFAULT_SHIFT_START)),
        now_str(),
        now_str()
    ), commit=True)

    shift_start = parse_shift_start(user["shift_start"] if user else DEFAULT_SHIFT_START)
    latest_attendance = get_current_attendance(user_id)

    if latest_attendance and latest_attendance["late_flag"]:
        create_notification(
            user_id,
            "Late Time-In",
            f"You timed in late by {latest_attendance['late_minutes']} minute(s). Shift start: {shift_start} ET."
        )
    else:
        create_notification(user_id, "Timed In", f"You timed in at {now_str()} ET.")

    log_activity(user_id, "TIME IN", f"Employee timed in. Shift: {shift_start}")
    invalidate_admin_employee_rows_cache()
    flash("Time in successful.", "success")
    return redirect(url_for("dashboard"))


@app.route("/start-break", methods=["POST"])
@login_required(role="employee")
def start_break():
    user_id = session["user_id"]
    user = get_user_by_id(user_id)
    manual_attendance_block_message = get_manual_attendance_block_message()
    if manual_attendance_block_message:
        flash(manual_attendance_block_message, "warning")
        return redirect(url_for("employee_actions"))
    attendance = auto_close_stale_attendance(user, get_current_attendance(user_id), actor_id=user_id, source_label="Employee portal")
    override_status = get_employee_override_status_for_date(user_id, today_str())
    override_block_message = get_attendance_override_block_message(override_status, "start_break", attendance)
    if override_block_message:
        flash(override_block_message, "danger")
        return redirect(url_for("dashboard"))

    if not attendance or not attendance["time_in"] or attendance["time_out"]:
        flash("You must be timed in first.", "danger")
        return redirect(url_for("dashboard"))

    open_break = get_open_break(user_id, attendance)
    if open_break:
        flash("You are already on break.", "warning")
        return redirect(url_for("dashboard"))

    break_limit_minutes = get_employee_break_limit(user)
    used_break_minutes = total_break_minutes(attendance["id"])
    if used_break_minutes >= break_limit_minutes:
        flash(f"Your daily break limit of {break_limit_minutes} minutes has already been used.", "warning")
        return redirect(url_for("dashboard"))

    execute_db("""
        INSERT INTO breaks (user_id, attendance_id, work_date, break_start, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (user_id, attendance["id"], attendance["work_date"], now_str(), now_str()), commit=True)

    execute_db("""
        UPDATE attendance
        SET status = ?, updated_at = ?
        WHERE id = ?
    """, ("On Break", now_str(), attendance["id"]), commit=True)

    remaining_break = max(break_limit_minutes - used_break_minutes, 0)
    create_notification(user_id, "Break Started", f"You started break at {now_str()} ET. Remaining break allowance: {remaining_break} minute(s).")
    log_activity(user_id, "BREAK START", "Employee started break")
    invalidate_admin_employee_rows_cache()
    flash("Break started.", "info")
    return redirect(url_for("dashboard"))


@app.route("/end-break", methods=["POST"])
@login_required(role="employee")
def end_break():
    user_id = session["user_id"]
    user = get_user_by_id(user_id)
    manual_attendance_block_message = get_manual_attendance_block_message()
    if manual_attendance_block_message:
        flash(manual_attendance_block_message, "warning")
        return redirect(url_for("employee_actions"))
    attendance = auto_close_stale_attendance(user, get_current_attendance(user_id), actor_id=user_id, source_label="Employee portal")
    override_status = get_employee_override_status_for_date(user_id, today_str())
    override_block_message = get_attendance_override_block_message(override_status, "end_break", attendance)
    if override_block_message:
        flash(override_block_message, "danger")
        return redirect(url_for("dashboard"))
    open_break = get_open_break(user_id, attendance)

    if not open_break:
        flash("No active break found.", "warning")
        return redirect(url_for("dashboard"))

    execute_db("""
        UPDATE breaks
        SET break_end = ?
        WHERE id = ?
    """, (now_str(), open_break["id"]), commit=True)

    if attendance:
        execute_db("""
            UPDATE attendance
            SET status = ?, updated_at = ?
            WHERE id = ?
    """, ("Timed In", now_str(), attendance["id"]), commit=True)

    create_notification(user_id, "Break Ended", f"You ended break at {now_str()} ET.")
    if attendance:
        total_break = total_break_minutes(attendance["id"])
        break_limit_minutes = get_employee_break_limit(user)
        if is_overbreak(total_break, break_limit_minutes):
            create_notification(
                user_id,
                "Break Limit Exceeded",
                f"Your total break time for today is {minutes_to_hm(total_break)}, which is over your {break_limit_minutes} minute limit."
            )
    log_activity(user_id, "BREAK END", "Employee ended break")
    invalidate_admin_employee_rows_cache()
    flash("Break ended.", "success")
    return redirect(url_for("dashboard"))


@app.route("/time-out", methods=["POST"])
@login_required(role="employee")
def time_out():
    user_id = session["user_id"]
    user = get_user_by_id(user_id)
    manual_attendance_block_message = get_manual_attendance_block_message()
    if manual_attendance_block_message:
        flash(manual_attendance_block_message, "warning")
        return redirect(url_for("employee_actions"))
    attendance = auto_close_stale_attendance(user, get_current_attendance(user_id), actor_id=user_id, source_label="Employee portal")
    override_status = get_employee_override_status_for_date(user_id, today_str())
    override_block_message = get_attendance_override_block_message(override_status, "time_out", attendance)
    if override_block_message:
        flash(override_block_message, "danger")
        return redirect(url_for("dashboard"))

    if not attendance or not attendance["time_in"]:
        flash("You are not timed in.", "danger")
        return redirect(url_for("dashboard"))

    if attendance["time_out"]:
        flash("You are already timed out.", "warning")
        return redirect(url_for("dashboard"))

    open_break = get_open_break(user_id, attendance)
    if open_break:
        execute_db("""
            UPDATE breaks
            SET break_end = ?
            WHERE id = ?
        """, (now_str(), open_break["id"]), commit=True)

    execute_db("""
        UPDATE attendance
        SET time_out = ?, status = ?, updated_at = ?
        WHERE id = ?
    """, (now_str(), "Timed Out", now_str(), attendance["id"]), commit=True)

    total_break = total_break_minutes(attendance["id"])
    break_limit_minutes = get_employee_break_limit(user)
    user_row = get_user_by_id(user_id)
    updated_attendance = get_attendance_by_id(attendance["id"])
    ok, msg = append_attendance_to_google_sheet(user_row, updated_attendance)

    create_notification(user_id, "Timed Out", f"You timed out at {now_str()} ET.")
    if attendance and is_overbreak(total_break, break_limit_minutes):
        create_notification(
            user_id,
            "Break Limit Exceeded",
            f"Your total break time for today is {minutes_to_hm(total_break)}, which is over your {break_limit_minutes} minute limit."
        )
    log_activity(user_id, "TIME OUT", f"Employee timed out. Sheets sync: {msg if ok else 'Skipped/Failed'}")
    invalidate_admin_employee_rows_cache()
    flash("Time out successful.", "success")
    return redirect(url_for("dashboard"))


@app.route("/notifications/read/<int:notif_id>", methods=["POST"])
@login_required()
def read_notification(notif_id):
    execute_db("""
        UPDATE notifications
        SET is_read = 1
        WHERE id = ? AND user_id = ?
    """, (notif_id, session["user_id"]), commit=True)

    if session.get("role") == "admin":
        return redirect(request.referrer or url_for("admin_dashboard"))
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/notifications/read-all", methods=["POST"])
@login_required()
def read_all_notifications():
    execute_db("""
        UPDATE notifications
        SET is_read = 1
        WHERE user_id = ? AND is_read = 0
    """, (session["user_id"],), commit=True)

    flash("All notifications marked as read.", "success")
    return redirect(request.referrer or url_for("notifications_page"))


@app.route("/uploads/<path:filename>")
@login_required()
def uploaded_file(filename):
    user = get_user_by_id(session["user_id"])
    if not can_access_uploaded_file(user, filename):
        abort(403)
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.route("/manifest.webmanifest")
def web_manifest():
    manifest_path = os.path.join(app.static_folder, "manifest.webmanifest")
    with open(manifest_path, "r", encoding="utf-8") as manifest_file:
        payload = manifest_file.read()
    response = Response(payload, mimetype="application/manifest+json")
    response.headers["Cache-Control"] = "no-cache"
    return response


@app.route("/service-worker.js")
def service_worker():
    response = send_from_directory(app.static_folder, "service-worker.js")
    response.headers["Cache-Control"] = "no-cache"
    return response


# =========================
# ADMIN
# =========================
def invalidate_admin_employee_rows_cache():
    _admin_employee_rows_cache["stamp"] = 0
    _admin_employee_rows_cache["rows"] = []
    invalidate_option_caches()
    invalidate_reports_cache()


def get_admin_current_attendance_map():
    attendance_map = {}
    for row in fetchall("""
        SELECT *
        FROM attendance
        WHERE time_in IS NOT NULL
          AND time_out IS NULL
        ORDER BY user_id ASC, id DESC
    """):
        row_dict = dict(row)
        attendance_map.setdefault(int(row_dict["user_id"]), row_dict)

    for row in fetchall("""
        SELECT *
        FROM attendance
        WHERE work_date = ?
        ORDER BY user_id ASC, id DESC
    """, (today_str(),)):
        row_dict = dict(row)
        attendance_map.setdefault(int(row_dict["user_id"]), row_dict)
    return attendance_map


def get_admin_open_break_map(attendance_ids):
    if not attendance_ids:
        return {}
    placeholders = ", ".join(["?"] * len(attendance_ids))
    rows = fetchall(f"""
        SELECT *
        FROM breaks
        WHERE attendance_id IN ({placeholders})
          AND break_end IS NULL
        ORDER BY attendance_id ASC, id DESC
    """, tuple(attendance_ids))
    result = {}
    for row in rows:
        row_dict = dict(row)
        result.setdefault(int(row_dict["attendance_id"]), row_dict)
    return result


def get_break_minutes_map(attendance_ids, include_open=False):
    if not attendance_ids:
        return {}
    placeholders = ", ".join(["?"] * len(attendance_ids))
    result = {}
    current_dt = parse_db_datetime(now_str())
    for row in fetchall(f"""
        SELECT attendance_id, break_start, break_end
        FROM breaks
        WHERE attendance_id IN ({placeholders})
    """, tuple(attendance_ids)):
        row_dict = dict(row)
        break_start_dt = parse_db_datetime(row_dict.get("break_start"))
        if not break_start_dt:
            continue
        break_end_dt = parse_db_datetime(row_dict.get("break_end"))
        if not break_end_dt:
            if not include_open:
                continue
            break_end_dt = current_dt
        break_minutes = max(int((break_end_dt - break_start_dt).total_seconds() // 60), 0)
        result[int(row_dict["attendance_id"])] = result.get(int(row_dict["attendance_id"]), 0) + break_minutes
    return result


def get_admin_break_minutes_map(attendance_ids):
    return get_break_minutes_map(attendance_ids, include_open=True)


def get_admin_override_maps():
    suspension_map = {}
    for row in fetchall("""
        SELECT *
        FROM disciplinary_actions
        WHERE action_type = 'Suspension'
          AND action_date <= ?
          AND COALESCE(end_date, action_date) >= ?
        ORDER BY action_date DESC, id DESC
    """, (today_str(), today_str())):
        row_dict = dict(row)
        suspension_map.setdefault(int(row_dict["user_id"]), row_dict)

    leave_map = {}
    for row in fetchall("""
        SELECT *
        FROM correction_requests
        WHERE status = 'Approved'
          AND request_type IN ('Sick Leave', 'Paid Leave')
          AND work_date <= ?
          AND COALESCE(end_work_date, work_date) >= ?
        ORDER BY work_date DESC, id DESC
    """, (today_str(), today_str())):
        row_dict = dict(row)
        leave_map.setdefault(int(row_dict["user_id"]), row_dict)
    return suspension_map, leave_map


def get_admin_open_overtime_map():
    overtime_map = {}
    rows = fetchall("""
        SELECT *
        FROM overtime_sessions
        WHERE overtime_end IS NULL
        ORDER BY user_id ASC, id DESC
    """)
    for row in rows:
        row_dict = dict(row)
        if row_dict.get("work_date") and row_dict["work_date"] < today_str():
            auto_close_stale_overtime_session(row_dict["user_id"], row_dict, source_label="Admin dashboard")
            continue
        overtime_map.setdefault(int(row_dict["user_id"]), row_dict)
    return overtime_map


def get_admin_live_status_label(user_id, attendance_row=None, open_break_row=None, open_overtime_row=None):
    if open_overtime_row:
        return "On Overtime"
    attendance = attendance_row or get_current_attendance(user_id)
    if not attendance:
        return "Offline"
    if attendance["time_in"] and not attendance["time_out"]:
        return "On Break" if open_break_row else "Timed In"
    if attendance["time_out"]:
        return "Timed Out"
    return "Offline"


def build_admin_employee_rows_snapshot():
    users = fetchall("""
        SELECT * FROM users
        WHERE role = 'employee'
        ORDER BY full_name ASC
    """)

    today_context_map = get_effective_employee_context_map(users, reference_date=today_str())
    attendance_map = get_admin_current_attendance_map()
    open_overtime_map = get_admin_open_overtime_map()
    attendance_ids = [int(row["id"]) for row in attendance_map.values()]
    open_break_map = get_admin_open_break_map(attendance_ids)
    break_minutes_map = get_admin_break_minutes_map(attendance_ids)
    suspension_map, leave_map = get_admin_override_maps()
    file_exists_cache = {}
    attendance_context_cache = {}

    def cached_file_exists(filename):
        cleaned = (filename or "").strip()
        if not cleaned:
            return False
        if cleaned not in file_exists_cache:
            file_exists_cache[cleaned] = uploaded_file_exists(cleaned)
        return file_exists_cache[cleaned]

    def get_attendance_context(user_row, attendance_row):
        if not attendance_row:
            return user_row
        reference_key = (
            int(attendance_row["user_id"]),
            normalize_history_reference(
                reference_datetime=get_attendance_reference_datetime(attendance_row),
                reference_date=attendance_row.get("work_date") if hasattr(attendance_row, "get") else attendance_row["work_date"]
            )
        )
        if reference_key not in attendance_context_cache:
            attendance_context_cache[reference_key] = get_effective_employee_context(
                user_row=user_row,
                reference_datetime=get_attendance_reference_datetime(attendance_row),
                reference_date=attendance_row["work_date"]
            )
        return attendance_context_cache[reference_key] or user_row

    employees = []
    for user in users:
        user_id = int(user["id"])
        display_user = today_context_map.get(user_id, dict(user))
        attendance = attendance_map.get(user_id)
        open_break = open_break_map.get(int(attendance["id"])) if attendance else None
        live_status = get_admin_live_status_label(user_id, attendance_row=attendance, open_break_row=open_break, open_overtime_row=open_overtime_map.get(user_id))
        scheduled_today = is_scheduled_on_date(display_user, today_str())
        override_status = None
        if user_id in suspension_map:
            suspension = suspension_map[user_id]
            override_status = {
                "type": "Suspension",
                "label": "Suspended",
                "details": suspension.get("details") or "",
                "end_date": suspension.get("end_date") or suspension.get("action_date") or today_str(),
            }
        elif user_id in leave_map:
            leave = leave_map[user_id]
            override_status = {
                "type": leave["request_type"],
                "label": leave["request_type"],
                "details": leave.get("message") or leave.get("admin_note") or "",
                "end_date": leave.get("end_work_date") or leave["work_date"],
            }
        suspension_today = override_status if override_status and override_status["type"] == "Suspension" else None
        leave_today = override_status if override_status and override_status["type"] in LEAVE_REQUEST_TYPES else None
        absent_today = False
        if user["is_active"] == 1 and not attendance and scheduled_today and not override_status:
            shift_dt, _ = get_shift_bounds_for_work_date(display_user, today_str())
            absent_today = now_dt() >= (shift_dt + timedelta(minutes=LATE_GRACE_MINUTES))

        attendance_context_user = display_user
        if attendance and attendance.get("work_date") and attendance["work_date"] != today_str():
            attendance_context_user = get_attendance_context(user, attendance)

        missing_timeout_today = False
        if attendance and user["is_active"] == 1 and attendance.get("time_in") and not attendance.get("time_out"):
            _, shift_end_dt = get_shift_bounds_for_work_date(attendance_context_user, attendance["work_date"])
            missing_timeout_today = now_dt() >= (shift_end_dt + timedelta(minutes=LATE_GRACE_MINUTES))

        undertime_today = 1 if is_undertime_record(
            attendance_context_user,
            {**dict(attendance), "source_type": "attendance"} if attendance else None
        ) else 0
        status_display = live_status

        if user["is_active"] != 1:
            status_display = "Inactive"
        elif override_status:
            status_display = override_status["label"]
        elif absent_today:
            status_display = "Absent"
        elif missing_timeout_today:
            status_display = "Missing Time Out"
        elif undertime_today:
            status_display = "Undertime"
        elif not scheduled_today and live_status == "Offline":
            status_display = "Off Day"

        row = {
            "id": display_user["id"],
            "full_name": display_user["full_name"],
            "username": display_user["username"],
            "department": display_user["department"],
            "position": display_user["position"],
            "schedule_days": display_user["schedule_days"] or DEFAULT_SCHEDULE_DAYS,
            "schedule_summary": get_schedule_summary(display_user["schedule_days"] or DEFAULT_SCHEDULE_DAYS),
            "scheduled_today": 1 if scheduled_today else 0,
            "absent_flag": 1 if absent_today else 0,
            "suspension_flag": 1 if suspension_today else 0,
            "leave_flag": 1 if leave_today else 0,
            "shift_start": display_user["shift_start"] or DEFAULT_SHIFT_START,
            "shift_end": display_user["shift_end"] or DEFAULT_SHIFT_END,
            "break_window_start": display_user["break_window_start"] or DEFAULT_BREAK_WINDOW_START,
            "break_window_end": display_user["break_window_end"] or DEFAULT_BREAK_WINDOW_END,
            "schedule_window_summary": f"{display_user['shift_start'] or DEFAULT_SHIFT_START} - {display_user['shift_end'] or DEFAULT_SHIFT_END}",
            "break_limit_minutes": parse_break_limit_minutes(display_user.get("break_limit_minutes") if hasattr(display_user, "get") else display_user["break_limit_minutes"]),
            "profile_image": display_user["profile_image"],
            "profile_image_available": 1 if cached_file_exists(display_user["profile_image"]) else 0,
            "is_active": user["is_active"],
            "status": live_status,
            "status_display": status_display,
            "time_in": attendance["time_in"] if attendance else None,
            "time_out": attendance["time_out"] if attendance else None,
            "proof_file": attendance["proof_file"] if attendance else None,
            "proof_file_available": 1 if attendance and cached_file_exists(attendance["proof_file"]) else 0,
            "late_flag": attendance["late_flag"] if attendance else 0,
            "late_minutes": attendance["late_minutes"] if attendance else 0,
            "break_minutes": break_minutes_map.get(int(attendance["id"]), 0) if attendance else 0
        }
        row["over_break_minutes"] = get_overbreak_minutes(row["break_minutes"], row["break_limit_minutes"])
        row["over_break_flag"] = 1 if row["over_break_minutes"] > 0 else 0
        row["missing_timeout_flag"] = 1 if missing_timeout_today else 0
        row["undertime_flag"] = undertime_today
        row["avatar_initials"] = get_avatar_initials(user["full_name"])
        row["attention_score"] = int(row["absent_flag"]) + int(row["late_flag"]) + int(row["over_break_flag"]) + int(row["missing_timeout_flag"]) + int(row["undertime_flag"])
        employees.append(row)

    return employees


def filter_admin_employee_rows(rows, status_filter="", search="", department_filter="", over_break_only=""):
    employees = []
    normalized_search = search.lower().strip()
    for row in rows:
        if status_filter and row["status_display"] != status_filter:
            continue

        if over_break_only == "1" and row["over_break_flag"] != 1:
            continue

        if normalized_search:
            hay = f"{row['full_name']} {row['username']} {row['department']} {row['position']} {row['shift_start']} {row['schedule_summary']}".lower()
            if normalized_search not in hay:
                continue

        if department_filter and (row["department"] or "").strip() != department_filter:
            continue

        employees.append(row)
    return employees


def get_admin_employee_rows(status_filter="", search="", department_filter="", over_break_only=""):
    cache_age = now_timestamp() - int(_admin_employee_rows_cache.get("stamp") or 0)
    if cache_age > ADMIN_STATUS_CACHE_TTL_SECONDS or not _admin_employee_rows_cache.get("rows"):
        _admin_employee_rows_cache["rows"] = build_admin_employee_rows_snapshot()
        _admin_employee_rows_cache["stamp"] = now_timestamp()

    return filter_admin_employee_rows(
        _admin_employee_rows_cache["rows"],
        status_filter=status_filter,
        search=search,
        department_filter=department_filter,
        over_break_only=over_break_only
    )


ADMIN_LIVE_STATUS_FIELDS = (
    "id",
    "full_name",
    "username",
    "department",
    "position",
    "schedule_summary",
    "schedule_window_summary",
    "scheduled_today",
    "profile_image",
    "profile_image_available",
    "avatar_initials",
    "status_display",
    "time_in",
    "time_out",
    "shift_end",
    "break_minutes",
    "break_limit_minutes",
    "proof_file",
    "proof_file_available",
    "absent_flag",
    "suspension_flag",
    "late_flag",
    "late_minutes",
    "undertime_flag",
    "over_break_flag",
    "over_break_minutes",
    "missing_timeout_flag",
)


def build_admin_live_status_payload(rows):
    payload_rows = []
    for row in rows:
        item = {field: row.get(field) for field in ADMIN_LIVE_STATUS_FIELDS}
        item["profile_image_url"] = (
            url_for("uploaded_file", filename=row["profile_image"])
            if row.get("profile_image") and row.get("profile_image_available")
            else ""
        )
        item["proof_file_url"] = (
            url_for("uploaded_file", filename=row["proof_file"])
            if row.get("proof_file") and row.get("proof_file_available")
            else ""
        )
        item["row_signature"] = "|".join(str(item.get(field, "")) for field in ADMIN_LIVE_STATUS_FIELDS)
        payload_rows.append(item)
    return payload_rows


def get_incident_reports(report_employee="", report_department="", report_type="", report_date_from="", report_date_to=""):
    report_sql = """
        SELECT r.*, u.full_name, u.department, reviewer.full_name AS reviewed_by_name,
               d.action_type AS linked_action_type,
               d.action_date AS linked_action_date,
               d.end_date AS linked_action_end_date
        FROM incident_reports r
        LEFT JOIN users u ON u.id = r.user_id
        LEFT JOIN users reviewer ON reviewer.id = r.reviewed_by
        LEFT JOIN disciplinary_actions d ON d.id = r.disciplinary_action_id
        WHERE 1=1
    """
    report_params = []

    if report_employee:
        report_sql += " AND r.user_id = ?"
        report_params.append(report_employee)

    if report_department:
        report_sql += " AND COALESCE(r.report_department, u.department, '') = ?"
        report_params.append(report_department)

    if report_type:
        report_sql += " AND LOWER(r.error_type) = ?"
        report_params.append(report_type.lower())

    if report_date_from:
        report_sql += " AND COALESCE(r.report_date, r.incident_date) >= ?"
        report_params.append(report_date_from)

    if report_date_to:
        report_sql += " AND COALESCE(r.report_date, r.incident_date) <= ?"
        report_params.append(report_date_to)

    report_sql += " ORDER BY r.id DESC LIMIT 100"
    rows = [dict(row) for row in fetchall(report_sql, report_params)]
    for row in rows:
        if not row.get("policy_incident_count"):
            row["policy_incident_count"] = count_repeated_incidents(row["user_id"], row["error_type"])
        if not row.get("incident_action"):
            row["incident_action"] = get_incident_policy_action(row["policy_incident_count"], exact_threshold=False)
    return rows


def get_exception_collections(employee_rows):
    absent = [row for row in employee_rows if row["absent_flag"] == 1]
    late = [row for row in employee_rows if row["late_flag"] == 1]
    over_break = [row for row in employee_rows if row["over_break_flag"] == 1]
    missing_timeout = [row for row in employee_rows if row["missing_timeout_flag"] == 1]
    undertime = [row for row in employee_rows if row["undertime_flag"] == 1]

    return {
        "absent": sorted(absent, key=lambda row: (row["department"] or "", row["full_name"] or "")),
        "late": sorted(late, key=lambda row: (-row["late_minutes"], row["full_name"] or "")),
        "over_break": sorted(over_break, key=lambda row: (-row["over_break_minutes"], row["full_name"] or "")),
        "missing_timeout": sorted(missing_timeout, key=lambda row: (row["shift_end"] or "", row["full_name"] or "")),
        "undertime": sorted(undertime, key=lambda row: (row["shift_end"] or "", row["full_name"] or "")),
    }


def get_suspicious_attendance_records(search="", limit=50):
    rows = fetchall("""
        SELECT a.*, u.full_name, u.username, u.department, u.position, u.shift_start, u.shift_end, u.break_limit_minutes
        FROM attendance a
        JOIN users u ON u.id = a.user_id
        WHERE u.role = 'employee'
        ORDER BY a.work_date DESC, a.id DESC
        LIMIT 300
    """)

    candidates = []
    for row in rows:
        item = dict(row)
        if search:
            hay = f"{item['full_name']} {item['username']} {item['work_date']}".lower()
            if search.lower() not in hay:
                continue
        issues, warnings = collect_attendance_diagnostics(item, item)
        if not issues and not warnings:
            continue
        break_rows = get_break_rows(item["id"])
        item["issue_summary"] = "; ".join(issues) if issues else "Needs review"
        item["issues"] = issues
        item["warnings"] = warnings
        item["severity"] = "error" if issues else "warning"
        item["warning_summary"] = "; ".join(warnings)
        item["break_count"] = len(break_rows)
        item["raw_work_minutes"] = total_work_minutes(item)
        item["break_summary"] = summarize_break_sessions(build_break_sessions(item["id"])) if break_rows else "No break sessions recorded."
        item["time_in_input"] = item["time_in"].replace(" ", "T")[:16] if item.get("time_in") else ""
        item["time_out_input"] = item["time_out"].replace(" ", "T")[:16] if item.get("time_out") else ""
        candidates.append(item)
        if len(candidates) >= limit:
            break
    return candidates


def update_attendance_record_by_admin(attendance_id, time_in_value="", time_out_value="", clear_breaks=False):
    attendance = get_attendance_by_id(attendance_id)
    if not attendance:
        raise ValueError("Attendance record not found.")

    user = get_user_by_id(attendance["user_id"])
    final_time_in = parse_datetime_local_input(time_in_value) if time_in_value else attendance["time_in"]
    final_time_out = parse_datetime_local_input(time_out_value) if time_out_value else attendance["time_out"]
    effective_user = get_effective_employee_context(
        user_row=user,
        reference_datetime=final_time_in or attendance["time_in"] or normalize_history_reference(reference_date=attendance["work_date"]),
        reference_date=attendance["work_date"]
    ) or user

    if final_time_in and final_time_out and final_time_out < final_time_in:
        shift_start = parse_shift_start(effective_user["shift_start"] if effective_user else DEFAULT_SHIFT_START)
        shift_end = parse_shift_end(effective_user["shift_end"] if effective_user else DEFAULT_SHIFT_END)
        overnight_shift = shift_end <= shift_start
        if overnight_shift:
            adjusted_time_out = parse_db_datetime(final_time_out) + timedelta(days=1)
            final_time_out = adjusted_time_out.strftime("%Y-%m-%d %H:%M:%S")

    if final_time_in and final_time_out and final_time_out < final_time_in:
        raise ValueError("Time out cannot be earlier than time in.")

    late_flag, late_minutes = calculate_late_info(final_time_in, parse_shift_start(effective_user["shift_start"] if effective_user else DEFAULT_SHIFT_START))
    open_break = get_open_break_for_attendance(attendance_id)
    if final_time_out:
        final_status = "Timed Out"
    elif open_break:
        final_status = "On Break"
    elif final_time_in:
        final_status = "Timed In"
    else:
        final_status = "Offline"

    execute_db("""
        UPDATE attendance
        SET time_in = ?, time_out = ?, status = ?, late_flag = ?, late_minutes = ?, updated_at = ?
        WHERE id = ?
    """, (
        final_time_in,
        final_time_out,
        final_status,
        late_flag,
        late_minutes,
        now_str(),
        attendance_id
    ), commit=True)

    if clear_breaks:
        execute_db("DELETE FROM breaks WHERE attendance_id = ?", (attendance_id,), commit=True)

    if final_time_out:
        execute_db("""
            UPDATE breaks
            SET break_end = CASE
                WHEN break_end IS NULL OR break_end > ? THEN ?
                ELSE break_end
            END
            WHERE attendance_id = ?
        """, (final_time_out, final_time_out, attendance_id), commit=True)

    invalidate_admin_employee_rows_cache()
    return get_attendance_by_id(attendance_id)


def notify_admins_for_exceptions(exception_groups):
    for row in exception_groups["absent"]:
        create_admin_alert_once(
            "Absent Today",
            f"{row['full_name']} is scheduled for today in {row['department'] or 'Unassigned'} and has not timed in yet."
        )

    for row in exception_groups["missing_timeout"]:
        create_admin_alert_once(
            "Missing Time Out",
            f"{row['full_name']} reached the scheduled shift end at {row['shift_end']} without timing out."
        )


def build_correction_time_details(item, current=False):
    prefix = "current_" if current else "requested_"
    entries = []

    field_specs = [
        ("time_in", "Time In"),
        ("break_start", "Break Start"),
        ("break_end", "Break End"),
        ("time_out", "Time Out"),
    ]
    for field_name, label in field_specs:
        raw_value = item.get(f"{prefix}{field_name}")
        if not raw_value:
            continue
        entries.append({
            "label": label,
            "value": format_time_12h(raw_value),
        })

    return entries


def build_correction_request_summary(rows):
    return {
        "total": len(rows),
        "pending": len([row for row in rows if row.get("status") == "Pending"]),
        "approved": len([row for row in rows if row.get("status") == "Approved"]),
        "rejected": len([row for row in rows if row.get("status") == "Rejected"]),
        "leave": len([row for row in rows if row.get("request_type") in LEAVE_REQUEST_TYPES]),
    }


def enrich_correction_request_tracking(item):
    item["submitted_display"] = format_datetime_12h(item.get("created_at")) if item.get("created_at") else ""
    item["reviewed_display"] = format_datetime_12h(item.get("reviewed_at")) if item.get("reviewed_at") else ""
    item["requested_time_details"] = build_correction_time_details(item, current=False)
    item["current_time_details"] = build_correction_time_details(item, current=True)
    item["has_requested_time_details"] = bool(item["requested_time_details"])
    item["has_current_time_details"] = bool(item["current_time_details"])
    item["is_leave_request"] = item.get("request_type") in LEAVE_REQUEST_TYPES
    item["status_tone"] = {
        "Pending": "yellow",
        "Approved": "green",
        "Rejected": "red",
    }.get(item.get("status"), "gray")
    item["status_chip_class"] = {
        "Pending": "status status-yellow",
        "Approved": "status status-green",
        "Rejected": "status status-red",
    }.get(item.get("status"), "status status-gray")
    item["status_headline"] = {
        "Pending": "Waiting for admin review",
        "Approved": "Approved and recorded",
        "Rejected": "Reviewed and declined",
    }.get(item.get("status"), "Request status unavailable")

    if item["status"] == "Approved":
        status_detail = item.get("applied_changes") or item.get("admin_note") or "Your request was approved."
    elif item["status"] == "Rejected":
        status_detail = item.get("admin_note") or "Admin rejected this request."
    else:
        status_detail = "Your request is queued for review."
    item["status_detail"] = status_detail

    review_actor = item.get("reviewed_by_name") or "Admin"
    item["timeline_steps"] = [
        {
            "label": "Submitted",
            "state": "done",
            "detail": item["submitted_display"] or "Request created",
        },
        {
            "label": "Under Review",
            "state": "current" if item.get("status") == "Pending" else "done",
            "detail": "Waiting for admin review" if item.get("status") == "Pending" else f"{review_actor} on {item['reviewed_display'] or 'reviewed'}",
        },
        {
            "label": "Final Decision",
            "state": "pending" if item.get("status") == "Pending" else ("approved" if item.get("status") == "Approved" else "rejected"),
            "detail": status_detail if item.get("status") != "Pending" else "No final decision yet.",
        },
    ]
    return item


def get_correction_requests(user_id=None, status="", date_from="", date_to=""):
    sql = """
        SELECT c.*, u.full_name, u.username, reviewer.full_name AS reviewed_by_name
        FROM correction_requests c
        JOIN users u ON u.id = c.user_id
        LEFT JOIN users reviewer ON reviewer.id = c.reviewed_by
        WHERE 1=1
    """
    params = []

    if user_id:
        sql += " AND c.user_id = ?"
        params.append(user_id)

    if status:
        sql += " AND c.status = ?"
        params.append(status)

    if date_from:
        sql += " AND COALESCE(c.end_work_date, c.work_date) >= ?"
        params.append(date_from)

    if date_to:
        sql += " AND c.work_date <= ?"
        params.append(date_to)

    sql += " ORDER BY c.id DESC LIMIT 200"
    rows = fetchall(sql, params)
    enriched_rows = []

    for row in rows:
        item = dict(row)
        item["display_date_range"] = format_request_date_range(item["work_date"], item.get("end_work_date"))
        item["requested_days"] = get_request_day_count(item)
        attendance = fetchone("""
            SELECT *
            FROM attendance
            WHERE user_id = ? AND work_date = ?
            ORDER BY id DESC LIMIT 1
        """, (item["user_id"], item["work_date"]))
        break_row = None

        if attendance:
            break_row = fetchone("""
                SELECT *
                FROM breaks
                WHERE attendance_id = ?
                ORDER BY id ASC LIMIT 1
            """, (attendance["id"],))

        item["current_time_in"] = attendance["time_in"] if attendance else None
        item["current_time_out"] = attendance["time_out"] if attendance else None
        item["current_break_start"] = break_row["break_start"] if break_row else None
        item["current_break_end"] = break_row["break_end"] if break_row else None
        item["requested_time_in_input"] = split_datetime_to_time(item.get("requested_time_in"))
        item["requested_break_start_input"] = split_datetime_to_time(item.get("requested_break_start"))
        item["requested_break_end_input"] = split_datetime_to_time(item.get("requested_break_end"))
        item["requested_time_out_input"] = split_datetime_to_time(item.get("requested_time_out"))
        item["current_time_in_input"] = split_datetime_to_time(item.get("current_time_in"))
        item["current_break_start_input"] = split_datetime_to_time(item.get("current_break_start"))
        item["current_break_end_input"] = split_datetime_to_time(item.get("current_break_end"))
        item["current_time_out_input"] = split_datetime_to_time(item.get("current_time_out"))
        enrich_correction_request_tracking(item)
        enriched_rows.append(item)

    return enriched_rows


def get_approved_special_requests(user_id=None, search="", department="", date_from="", date_to=""):
    sql = """
        SELECT c.*, u.full_name, u.username, u.break_limit_minutes
        FROM correction_requests c
        JOIN users u ON u.id = c.user_id
        WHERE c.status = 'Approved' AND c.request_type IN ('Undertime', 'Sick Leave', 'Paid Leave')
    """
    params = []

    if user_id:
        sql += " AND c.user_id = ?"
        params.append(user_id)

    if search:
        sql += " AND (LOWER(u.full_name) LIKE ? OR LOWER(u.username) LIKE ?)"
        s = f"%{search.lower()}%"
        params.extend([s, s])

    if department:
        sql += " AND COALESCE(u.department, '') = ?"
        params.append(department)

    if date_from:
        sql += " AND COALESCE(c.end_work_date, c.work_date) >= ?"
        params.append(date_from)

    if date_to:
        sql += " AND c.work_date <= ?"
        params.append(date_to)

    sql += " ORDER BY c.work_date DESC, c.id DESC"
    return fetchall(sql, params)


def enrich_history_record(row, break_limit_minutes, employee_row=None):
    row = dict(row)
    record_type = row.get("request_type") or "Attendance"
    is_attendance = row.get("source_type", "attendance") == "attendance"
    if employee_row:
        break_limit_minutes = get_employee_break_limit(
            employee_row,
            reference_datetime=row.get("time_in") or row.get("created_at"),
            reference_date=row.get("work_date")
        )
    break_minutes = total_break_minutes(row["id"]) if is_attendance and row.get("id") else 0
    work_row = dict(row)
    if record_type == "Undertime" and row.get("requested_time_out"):
        work_row["time_out"] = row.get("requested_time_out")
        row["time_out"] = row.get("requested_time_out")
    suspicious_work_flag = 1 if is_suspicious_work_duration(employee_row, work_row) else 0
    raw_work_minutes = total_work_minutes(work_row) if is_attendance else 0
    work_minutes = 0 if suspicious_work_flag else raw_work_minutes
    over_break_minutes = get_overbreak_minutes(break_minutes, break_limit_minutes)
    break_sessions = build_break_sessions(row["id"]) if is_attendance and row.get("id") else []
    undertime_flag = 1 if not suspicious_work_flag and is_undertime_record(employee_row, work_row) else 0
    if record_type == "Attendance" and undertime_flag:
        record_type = "Undertime"
    display_status = row.get("status") or ""
    if undertime_flag and is_attendance and row.get("time_out"):
        display_status = "Undertime"
    data_issue_note = "Suspicious work duration hidden. Please review this record." if suspicious_work_flag else ""

    return {
        "row": row,
        "break_minutes": break_minutes,
        "break_sessions": break_sessions,
        "break_sessions_summary": summarize_break_sessions(break_sessions),
        "work_minutes": work_minutes,
        "raw_work_minutes": raw_work_minutes,
        "over_break_minutes": over_break_minutes,
        "undertime_flag": undertime_flag,
        "suspicious_work_flag": suspicious_work_flag,
        "data_issue_note": data_issue_note,
        "display_status": display_status,
        "absent_flag": 1 if employee_row and is_attendance and is_absent_today(employee_row, row) else 0,
        "record_type": record_type,
        "request_note": " | ".join(part for part in [row.get("admin_note") or row.get("message") or "", data_issue_note] if part),
    }


def build_employee_history_records(user_row, limit=60):
    attendance_rows = [
        dict(row) for row in fetchall("""
            SELECT * FROM attendance
            WHERE user_id = ?
            ORDER BY work_date DESC, id DESC
            LIMIT ?
        """, (user_row["id"], limit))
    ]
    special_requests = [dict(row) for row in get_approved_special_requests(user_id=user_row["id"])]
    suspension_rows = [dict(row) for row in get_disciplinary_actions(action_type="Suspension", user_id=user_row["id"])]
    request_map = {}
    for row in special_requests:
        request_dates = expand_request_dates(row["work_date"], row.get("end_work_date")) if row["request_type"] in LEAVE_REQUEST_TYPES else [row["work_date"]]
        for request_work_date in request_dates:
            matched_attendance, _ = get_matching_attendance_context_for_request(
                row["user_id"],
                request_work_date,
                request_type=row["request_type"],
                requested_time_out=row.get("requested_time_out")
            )
            target_work_date = matched_attendance["work_date"] if matched_attendance else request_work_date
            request_map[(row["user_id"], target_work_date)] = row

    combined = []
    seen_keys = set()

    for row in attendance_rows:
        key = (row["user_id"], row["work_date"])
        special_request = request_map.get(key)
        row["source_type"] = "attendance"
        row["request_type"] = special_request["request_type"] if special_request else ""
        row["admin_note"] = special_request["admin_note"] if special_request else ""
        row["message"] = special_request["message"] if special_request else ""
        row["requested_time_out"] = special_request["requested_time_out"] if special_request else ""
        combined.append(enrich_history_record(row, get_employee_break_limit(user_row), employee_row=user_row))
        seen_keys.add(key)

    for row in special_requests:
        request_dates = expand_request_dates(row["work_date"], row.get("end_work_date")) if row["request_type"] in LEAVE_REQUEST_TYPES else [row["work_date"]]
        for request_work_date in request_dates:
            matched_attendance, _ = get_matching_attendance_context_for_request(
                row["user_id"],
                request_work_date,
                request_type=row["request_type"],
                requested_time_out=row.get("requested_time_out")
            )
            target_work_date = matched_attendance["work_date"] if matched_attendance else request_work_date
            key = (row["user_id"], target_work_date)
            if key in seen_keys or row["request_type"] not in (LEAVE_REQUEST_TYPES | {"Undertime"}):
                continue
            is_leave_request = row["request_type"] in LEAVE_REQUEST_TYPES
            synthetic_row = {
                "id": None,
                "user_id": row["user_id"],
                "work_date": target_work_date,
                "time_in": None,
                "time_out": None if is_leave_request else row.get("requested_time_out"),
                "status": row["request_type"] if is_leave_request else "Approved Undertime",
                "proof_file": None,
                "late_flag": 0,
                "late_minutes": 0,
                "request_type": row["request_type"],
                "admin_note": row["admin_note"],
                "message": row["message"],
                "source_type": "request",
            }
            combined.append(enrich_history_record(synthetic_row, get_employee_break_limit(user_row), employee_row=user_row))

    suspension_dates = set()
    for row in suspension_rows:
        for work_date in expand_suspension_dates(row):
            if (row["user_id"], work_date) in seen_keys:
                continue
            suspension_dates.add((row["user_id"], work_date))
            synthetic_row = {
                "id": None,
                "user_id": row["user_id"],
                "work_date": work_date,
                "time_in": None,
                "time_out": None,
                "status": "Suspension",
                "proof_file": None,
                "late_flag": 0,
                "late_minutes": 0,
                "request_type": "Suspension",
                "admin_note": row.get("details") or "",
                "message": row.get("details") or "",
                "source_type": "request",
            }
            combined.append(enrich_history_record(synthetic_row, get_employee_break_limit(user_row), employee_row=user_row))

    combined.sort(key=lambda item: (item["row"]["work_date"] or "", item["row"].get("id") or 0), reverse=True)
    return combined[:limit]


def build_admin_history_records(search="", department="", type_filter="", late_only="", absent_only="", over_break_only="", date_from="", date_to="", limit=200):
    sql = """
        SELECT a.*, u.full_name, u.username, u.break_limit_minutes
        FROM attendance a
        JOIN users u ON u.id = a.user_id
        WHERE 1=1
    """
    params = []

    if search:
        sql += " AND (LOWER(u.full_name) LIKE ? OR LOWER(u.username) LIKE ?)"
        s = f"%{search.lower()}%"
        params.extend([s, s])

    if department:
        sql += " AND COALESCE(u.department, '') = ?"
        params.append(department)

    if late_only == "1":
        sql += " AND a.late_flag = 1"

    if date_from:
        sql += " AND a.work_date >= ?"
        params.append(date_from)

    if date_to:
        sql += " AND a.work_date <= ?"
        params.append(date_to)

    sql += " ORDER BY a.work_date DESC, a.id DESC LIMIT ?"
    params.append(limit)

    attendance_rows = [dict(row) for row in fetchall(sql, params)]
    special_requests = [dict(row) for row in get_approved_special_requests(
        search=search,
        department=department,
        date_from=date_from,
        date_to=date_to
    )]
    suspension_rows = [dict(row) for row in get_disciplinary_actions(
        action_type="Suspension",
        department=department
    )]
    request_map = {}
    for row in special_requests:
        request_dates = expand_request_dates(row["work_date"], row.get("end_work_date")) if row["request_type"] in LEAVE_REQUEST_TYPES else [row["work_date"]]
        for request_work_date in request_dates:
            matched_attendance, _ = get_matching_attendance_context_for_request(
                row["user_id"],
                request_work_date,
                request_type=row["request_type"],
                requested_time_out=row.get("requested_time_out")
            )
            target_work_date = matched_attendance["work_date"] if matched_attendance else request_work_date
            request_map[(row["user_id"], target_work_date)] = row
    attendance_key_map = {(row["user_id"], row["work_date"]): row for row in attendance_rows}

    enriched = []
    seen_keys = set()

    for row in attendance_rows:
        employee = get_user_by_id(row["user_id"])
        key = (row["user_id"], row["work_date"])
        special_request = request_map.get(key)
        row["source_type"] = "attendance"
        row["request_type"] = special_request["request_type"] if special_request else ""
        row["admin_note"] = special_request["admin_note"] if special_request else ""
        row["message"] = special_request["message"] if special_request else ""
        row["requested_time_out"] = special_request["requested_time_out"] if special_request else ""

        item = enrich_history_record(row, parse_break_limit_minutes(row["break_limit_minutes"]), employee_row=employee)
        if type_filter and item["record_type"] != type_filter:
            continue
        if absent_only == "1" and item["absent_flag"] != 1:
            continue
        if over_break_only == "1" and item["over_break_minutes"] <= 0:
            continue
        enriched.append(item)
        seen_keys.add(key)

    for row in special_requests:
        request_dates = expand_request_dates(row["work_date"], row.get("end_work_date")) if row["request_type"] in LEAVE_REQUEST_TYPES else [row["work_date"]]
        for request_work_date in request_dates:
            matched_attendance, _ = get_matching_attendance_context_for_request(
                row["user_id"],
                request_work_date,
                request_type=row["request_type"],
                requested_time_out=row.get("requested_time_out")
            )
            target_work_date = matched_attendance["work_date"] if matched_attendance else request_work_date
            key = (row["user_id"], target_work_date)
            if key in seen_keys or row["request_type"] not in (LEAVE_REQUEST_TYPES | {"Undertime"}):
                continue
            is_leave_request = row["request_type"] in LEAVE_REQUEST_TYPES
            synthetic_row = {
                "id": None,
                "user_id": row["user_id"],
                "full_name": row["full_name"],
                "username": row["username"],
                "break_limit_minutes": row["break_limit_minutes"],
                "work_date": target_work_date,
                "time_in": None,
                "time_out": None if is_leave_request else row.get("requested_time_out"),
                "status": row["request_type"] if is_leave_request else "Approved Undertime",
                "proof_file": None,
                "late_flag": 0,
                "late_minutes": 0,
                "request_type": row["request_type"],
                "admin_note": row["admin_note"],
                "message": row["message"],
                "source_type": "request",
            }
            item = enrich_history_record(synthetic_row, parse_break_limit_minutes(row["break_limit_minutes"]))
            if type_filter and item["record_type"] != type_filter:
                continue
            enriched.append(item)

    suspension_date_keys = set()
    for row in suspension_rows:
        if search:
            hay = f"{row.get('full_name', '')} {row.get('username', '')}".lower()
            if search.lower() not in hay:
                continue
        for work_date in expand_suspension_dates(row):
            key = (row["user_id"], work_date)
            if key in seen_keys:
                continue
            if date_from and work_date < date_from:
                continue
            if date_to and work_date > date_to:
                continue
            suspension_date_keys.add(key)
            synthetic_row = {
                "id": None,
                "user_id": row["user_id"],
                "full_name": row["full_name"],
                "username": row["username"],
                "break_limit_minutes": row.get("break_limit_minutes", BREAK_LIMIT_MINUTES),
                "work_date": work_date,
                "time_in": None,
                "time_out": None,
                "status": "Suspension",
                "proof_file": None,
                "late_flag": 0,
                "late_minutes": 0,
                "request_type": "Suspension",
                "admin_note": row.get("details") or "",
                "message": row.get("details") or "",
                "source_type": "request",
            }
            item = enrich_history_record(synthetic_row, parse_break_limit_minutes(row.get("break_limit_minutes", BREAK_LIMIT_MINUTES)))
            if type_filter and item["record_type"] != type_filter:
                continue
            enriched.append(item)

    candidate_dates = []
    if date_from or date_to:
        try:
            start_date = datetime.strptime(date_from or date_to, "%Y-%m-%d").date()
            end_date = datetime.strptime(date_to or date_from, "%Y-%m-%d").date()
            if start_date <= end_date:
                total_days = (end_date - start_date).days
                if total_days <= 60:
                    candidate_dates = [
                        (start_date + timedelta(days=offset)).strftime("%Y-%m-%d")
                        for offset in range(total_days + 1)
                    ]
        except Exception:
            candidate_dates = []
    else:
        candidate_dates = [today_str()]

    if candidate_dates and (not type_filter or type_filter == "Absent"):
        employee_sql = """
            SELECT *
            FROM users
            WHERE role = 'employee'
        """
        employee_params = []

        if search:
            employee_sql += " AND (LOWER(full_name) LIKE ? OR LOWER(username) LIKE ?)"
            s = f"%{search.lower()}%"
            employee_params.extend([s, s])

        if department:
            employee_sql += " AND COALESCE(department, '') = ?"
            employee_params.append(department)

        employee_sql += " ORDER BY full_name ASC"
        employees = fetchall(employee_sql, employee_params)

        for employee in employees:
            if employee["is_active"] != 1:
                continue
            for work_date in candidate_dates:
                effective_employee = get_effective_employee_context(
                    user_row=employee,
                    reference_date=work_date
                ) or dict(employee)
                key = (employee["id"], work_date)
                special_request = request_map.get(key)
                if key in attendance_key_map:
                    continue
                if key in suspension_date_keys:
                    continue
                if special_request and special_request["request_type"] in LEAVE_REQUEST_TYPES:
                    continue
                if not is_scheduled_on_date(effective_employee, work_date):
                    continue

                synthetic_row = {
                    "id": None,
                    "user_id": employee["id"],
                    "full_name": employee["full_name"],
                    "username": employee["username"],
                    "break_limit_minutes": effective_employee["break_limit_minutes"],
                    "work_date": work_date,
                    "time_in": None,
                    "time_out": None,
                    "status": "Absent",
                    "proof_file": None,
                    "late_flag": 0,
                    "late_minutes": 0,
                    "request_type": "Absent",
                    "admin_note": "",
                    "message": "",
                    "source_type": "request",
                }
                item = enrich_history_record(synthetic_row, parse_break_limit_minutes(effective_employee["break_limit_minutes"]))
                if type_filter and item["record_type"] != type_filter:
                    continue
                if over_break_only == "1":
                    continue
                enriched.append(item)

    enriched.sort(key=lambda item: (item["row"]["work_date"] or "", item["row"].get("id") or 0), reverse=True)
    return enriched[:limit]


def build_attendance_audit_rows(date_from="", date_to="", employee_id="", source_filter="", limit=250):
    activity_target_expr = "a.user_id"
    if column_exists("activity_logs", "target_user_id"):
        activity_target_expr = "COALESCE(a.target_user_id, a.user_id)"

    params = ["AUTO CLOSE%"]
    where_employee = ""
    if employee_id.isdigit():
        where_employee = f"AND {activity_target_expr} = ?"
        params.append(int(employee_id))

    manual_rows = fetchall(f"""
        SELECT
            a.id,
            a.created_at,
            'manual' AS source_type,
            a.action AS event_action,
            a.details AS event_details,
            actor.full_name AS actor_name,
            actor.username AS actor_username,
            {activity_target_expr} AS target_user_id,
            employee.full_name AS employee_name,
            employee.department AS employee_department
        FROM activity_logs a
        LEFT JOIN users actor ON actor.id = a.user_id
        LEFT JOIN users employee ON employee.id = {activity_target_expr}
        WHERE (
            a.action IN ('TIME IN', 'BREAK START', 'BREAK END', 'TIME OUT', 'CORRECTION REQUEST', 'REVIEW CORRECTION')
            OR a.action LIKE ?
        )
        {where_employee}
    """, tuple(params))

    scanner_rows = []
    if table_exists("scanner_logs"):
        scanner_exprs = get_scanner_log_select_expressions()
        scanner_params = []
        scanner_where_employee = ""
        if employee_id.isdigit():
            scanner_where_employee = "AND sl.employee_user_id = ?"
            scanner_params.append(int(employee_id))

        scanner_rows = fetchall(f"""
            SELECT
                sl.id,
                sl.created_at,
                'scanner' AS source_type,
                sl.action_type AS event_action,
                sl.result_message AS event_details,
                {scanner_exprs['scanner_name']} AS actor_name,
                {scanner_exprs['scanner_username']} AS actor_username,
                sl.employee_user_id AS target_user_id,
                {scanner_exprs['employee_name']} AS employee_name,
                {scanner_exprs['employee_department']} AS employee_department,
                sl.result_status
            FROM scanner_logs sl
            LEFT JOIN users scanner ON scanner.id = sl.scanner_user_id
            LEFT JOIN users employee ON employee.id = sl.employee_user_id
            WHERE 1=1
            {scanner_where_employee}
        """, tuple(scanner_params))

    overtime_rows = []
    if table_exists("overtime_sessions"):
        overtime_params = []
        overtime_where_employee = ""
        if employee_id.isdigit():
            overtime_where_employee = "AND o.user_id = ?"
            overtime_params.append(int(employee_id))

        overtime_rows = fetchall(f"""
            SELECT
                o.id,
                o.overtime_start,
                o.overtime_end,
                o.created_at,
                u.full_name AS actor_name,
                u.username AS actor_username,
                o.user_id AS target_user_id,
                u.full_name AS employee_name,
                u.department AS employee_department
            FROM overtime_sessions o
            JOIN users u ON u.id = o.user_id
            WHERE 1=1
            {overtime_where_employee}
        """, tuple(overtime_params))

    combined = []
    for row in manual_rows:
        item = dict(row)
        item["source_label"] = "Manual / Admin"
        combined.append(item)
    for row in scanner_rows:
        item = dict(row)
        item["source_label"] = "Scanner Kiosk"
        combined.append(item)
    for row in overtime_rows:
        base = dict(row)
        combined.append({
            "id": f"ot-start-{base['id']}",
            "created_at": base.get("overtime_start") or base.get("created_at"),
            "source_type": "overtime",
            "event_action": "OVERTIME START",
            "event_details": "Overtime session started.",
            "actor_name": base.get("actor_name"),
            "actor_username": base.get("actor_username"),
            "target_user_id": base.get("target_user_id"),
            "employee_name": base.get("employee_name"),
            "employee_department": base.get("employee_department"),
            "source_label": "Overtime",
        })
        if base.get("overtime_end"):
            combined.append({
                "id": f"ot-end-{base['id']}",
                "created_at": base.get("overtime_end"),
                "source_type": "overtime",
                "event_action": "OVERTIME END",
                "event_details": "Overtime session ended.",
                "actor_name": base.get("actor_name"),
                "actor_username": base.get("actor_username"),
                "target_user_id": base.get("target_user_id"),
                "employee_name": base.get("employee_name"),
                "employee_department": base.get("employee_department"),
                "source_label": "Overtime",
            })

    if date_from:
        combined = [row for row in combined if (row.get("created_at") or "")[:10] >= date_from]
    if date_to:
        combined = [row for row in combined if (row.get("created_at") or "")[:10] <= date_to]
    if source_filter:
        combined = [row for row in combined if row.get("source_type") == source_filter]

    combined.sort(key=lambda item: ((item.get("created_at") or ""), str(item.get("id") or "")), reverse=True)
    return combined[:limit]


def apply_attendance_correction(user_id, work_date, time_in_value="", break_start_value="", break_end_value="", time_out_value=""):
    undertime_only_adjustment = bool(time_out_value and not (time_in_value or break_start_value or break_end_value))
    attendance, break_row = get_matching_attendance_context_for_request(
        user_id,
        work_date,
        request_type="Undertime" if undertime_only_adjustment else "",
        requested_time_out=combine_work_date_and_time(work_date, time_out_value) if time_out_value else None
    )
    target_work_date = attendance["work_date"] if attendance else work_date

    before_values = {
        "time_in": attendance["time_in"] if attendance else None,
        "break_start": break_row["break_start"] if break_row else None,
        "break_end": break_row["break_end"] if break_row else None,
        "time_out": attendance["time_out"] if attendance else None,
    }

    final_time_in, final_break_start, final_break_end, final_time_out = resolve_correction_datetimes(
        target_work_date,
        time_in_value=time_in_value,
        break_start_value=break_start_value,
        break_end_value=break_end_value,
        time_out_value=time_out_value,
        existing_time_in=attendance["time_in"] if attendance else None,
        existing_break_start=break_row["break_start"] if break_row else None,
        existing_break_end=break_row["break_end"] if break_row else None,
        existing_time_out=attendance["time_out"] if attendance else None
    )

    if undertime_only_adjustment and final_time_out:
        final_time_out = combine_work_date_and_time(
            target_work_date,
            time_out_value,
            not_before=attendance["time_in"] if attendance else None
        )
        if final_break_start and final_break_start > final_time_out:
            final_break_start = None
            final_break_end = None
        elif final_break_end and final_break_end > final_time_out:
            final_break_end = final_time_out

    if (time_in_value or (attendance and attendance["time_in"])) and final_time_in and final_time_out:
        if final_time_out < final_time_in:
            raise ValueError("Time out cannot be earlier than time in.")

    if (break_start_value or break_end_value or break_row) and not final_time_in:
        raise ValueError("Set time in before saving break corrections.")

    if final_break_start and final_break_end and final_break_end < final_break_start:
        raise ValueError("Break end cannot be earlier than break start.")

    if final_break_start and final_time_in and final_break_start < final_time_in:
        raise ValueError("Break start cannot be earlier than time in.")

    if final_break_end and final_time_out and final_break_end > final_time_out:
        raise ValueError("Break end cannot be later than time out.")

    user = get_user_by_id(user_id)
    effective_user = get_effective_employee_context(
        user_row=user,
        reference_datetime=final_time_in or normalize_history_reference(reference_date=target_work_date),
        reference_date=target_work_date
    )
    late_flag, late_minutes = calculate_late_info(final_time_in, parse_shift_start(effective_user["shift_start"] if effective_user else DEFAULT_SHIFT_START))

    if final_time_out:
        final_status = "Timed Out"
    elif final_break_start and not final_break_end:
        final_status = "On Break"
    elif final_time_in:
        final_status = "Timed In"
    else:
        final_status = "Offline"

    if attendance:
        execute_db("""
            UPDATE attendance
            SET time_in = ?, time_out = ?, status = ?, late_flag = ?, late_minutes = ?, updated_at = ?
            WHERE id = ?
        """, (
            final_time_in,
            final_time_out,
            final_status,
            late_flag,
            late_minutes,
            now_str(),
            attendance["id"]
        ), commit=True)
    elif final_time_in or final_time_out or final_break_start or final_break_end:
        execute_db("""
            INSERT INTO attendance (
                user_id, work_date, time_in, time_out, status, proof_file, notes,
                late_flag, late_minutes, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user_id,
            target_work_date,
            final_time_in,
            final_time_out,
            final_status,
            None,
            "Updated through correction request",
            late_flag,
            late_minutes,
            now_str(),
            now_str()
        ), commit=True)
        attendance = fetchone("""
            SELECT *
            FROM attendance
            WHERE user_id = ? AND work_date = ?
            ORDER BY id DESC LIMIT 1
        """, (user_id, target_work_date))

    if not attendance:
        return

    if break_row and (
        final_break_start != break_row["break_start"]
        or final_break_end != break_row["break_end"]
    ):
        execute_db("""
            UPDATE breaks
            SET break_start = ?, break_end = ?
            WHERE id = ?
        """, (final_break_start, final_break_end, break_row["id"]), commit=True)
    elif not break_row and (final_break_start or final_break_end):
        execute_db("""
            INSERT INTO breaks (user_id, attendance_id, work_date, break_start, break_end, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (user_id, attendance["id"], target_work_date, final_break_start, final_break_end, now_str()), commit=True)

    after_values = {
        "time_in": final_time_in,
        "break_start": final_break_start,
        "break_end": final_break_end,
        "time_out": final_time_out,
    }
    invalidate_admin_employee_rows_cache()
    return build_correction_change_summary(before_values, after_values)


@app.route("/admin")
@login_required(role="admin")
def admin_dashboard():
    current_admin = get_user_by_id(session["user_id"])
    if not current_admin:
        session.clear()
        flash("Your session expired. Please log in again.", "warning")
        return redirect(url_for("login"))

    status_filter = request.args.get("status", "").strip()
    search = request.args.get("search", "").strip()
    department_filter = request.args.get("department", "").strip()
    over_break_only = request.args.get("over_break_only", "").strip()
    page = parse_positive_int(request.args.get("page", "1"), 1)
    page_size = parse_positive_int(request.args.get("page_size", "25"), 25)

    all_employee_rows = get_admin_employee_rows()
    filtered_rows = filter_admin_employee_rows(
        all_employee_rows,
        status_filter=status_filter,
        search=search,
        department_filter=department_filter,
        over_break_only=over_break_only
    )
    pagination = paginate_items(filtered_rows, page, page_size)
    employees = pagination["items"]
    departments = get_department_options()

    logs = fetchall("""
        SELECT a.*, u.full_name
        FROM activity_logs a
        JOIN users u ON u.id = a.user_id
        ORDER BY a.id DESC
        LIMIT 25
    """)

    late_today_row = fetchone("""
        SELECT COUNT(*) AS cnt
        FROM attendance
        WHERE work_date = ? AND late_flag = 1
    """, (today_str(),))
    late_today = late_today_row["cnt"] if late_today_row else 0
    active_users = [row for row in all_employee_rows if row["is_active"] == 1]
    exception_groups = get_exception_collections(all_employee_rows)
    notify_admins_for_exceptions(exception_groups)
    notify_admins_for_leave_and_disciplinary_events()
    admin_notifications = fetchall("""
        SELECT *
        FROM notifications
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT 8
    """, (session["user_id"],))

    stats = {
        "total_employees": len(active_users),
        "scheduled_today": len([emp for emp in all_employee_rows if emp["scheduled_today"] == 1 and emp["is_active"] == 1]),
        "absent_today": len(exception_groups["absent"]),
        "timed_in": len([emp for emp in all_employee_rows if emp["status_display"] == "Timed In"]),
        "on_break": len([emp for emp in all_employee_rows if emp["status_display"] == "On Break"]),
        "timed_out": len([emp for emp in all_employee_rows if emp["status_display"] == "Timed Out"]),
        "late_today": late_today,
        "over_break_today": len(exception_groups["over_break"]),
        "missing_timeout": len(exception_groups["missing_timeout"]),
        "undertime_today": len(exception_groups["undertime"])
    }

    return render_template(
        "admin_dashboard.html",
        employees=employees,
        pagination=pagination,
        logs=logs,
        stats=stats,
        break_limit_minutes=BREAK_LIMIT_MINUTES,
        status_filter=status_filter,
        search=search,
        department_filter=department_filter,
        departments=departments,
        over_break_only=over_break_only,
        today_schedule_code=get_today_schedule_code(),
        exception_groups=exception_groups,
        admin_notifications=admin_notifications
    )


@app.route("/admin/live-status")
@login_required(role="admin")
def admin_live_status():
    page = parse_positive_int(request.args.get("page", "1"), 1)
    page_size = parse_positive_int(request.args.get("page_size", "25"), 25)
    rows = get_admin_employee_rows(
        status_filter=request.args.get("status", "").strip(),
        search=request.args.get("search", "").strip(),
        department_filter=request.args.get("department", "").strip(),
        over_break_only=request.args.get("over_break_only", "").strip()
    )
    pagination = paginate_items(rows, page, page_size)
    return jsonify({
        "rows": build_admin_live_status_payload(pagination["items"]),
        "pagination": {
            "page": pagination["page"],
            "page_size": pagination["page_size"],
            "total": pagination["total"],
            "total_pages": pagination["total_pages"],
            "has_prev": pagination["has_prev"],
            "has_next": pagination["has_next"],
            "start_index": pagination["start_index"],
            "end_index": pagination["end_index"],
        },
        "generated_at": now_str(),
    })


@app.route("/admin/payroll")
@login_required(role="admin")
def admin_payroll():
    filters = normalize_payroll_filters(
        request.args.get("period", "this_month"),
        request.args.get("date_from", "").strip(),
        request.args.get("date_to", "").strip(),
        request.args.get("department", "").strip(),
        request.args.get("employee_id", "").strip()
    )
    period = filters["period"]
    department_filter = filters["department_filter"]
    employee_filter = filters["employee_filter"]
    date_from = filters["date_from"]
    date_to = filters["date_to"]
    departments = get_department_options()
    employees = get_employee_options()
    payroll_rows = build_payroll_rows(
        date_from,
        date_to,
        department_filter=department_filter,
        employee_filter=employee_filter
    )
    stats = build_payroll_stats(payroll_rows)
    adjustments = get_payroll_adjustments(date_from, date_to, department_filter=department_filter, employee_filter=employee_filter)
    recurring_rules = get_payroll_recurring_rules(
        department_filter=department_filter,
        employee_filter=employee_filter,
        include_inactive=True
    )
    current_run = get_payroll_run(date_from, date_to, department_filter=department_filter, employee_filter=employee_filter)
    if current_run:
        current_run = enrich_admin_payroll_run(current_run)
    recent_runs = get_recent_payroll_runs()
    editing_recurring_rule = get_payroll_recurring_rule(request.args.get("edit_recurring_rule", ""))

    return render_template(
        "admin_payroll.html",
        payroll_rows=payroll_rows,
        departments=departments,
        employees=employees,
        department_filter=department_filter,
        employee_filter=employee_filter,
        period=period,
        date_from=date_from.strftime("%Y-%m-%d"),
        date_to=date_to.strftime("%Y-%m-%d"),
        stats=stats,
        adjustments=adjustments,
        recurring_rules=recurring_rules,
        current_run=current_run,
        recent_runs=recent_runs,
        employee_filter_label=get_payroll_employee_filter_label(employee_filter),
        editing_recurring_rule=editing_recurring_rule
    )


@app.route("/admin/payroll/adjustments", methods=["POST"])
@login_required(role="admin")
def add_payroll_adjustment():
    redirect_args = payroll_filter_redirect_args(request.form)
    filters = normalize_payroll_filters(
        request.form.get("period", "this_month"),
        request.form.get("date_from", ""),
        request.form.get("date_to", ""),
        request.form.get("department", ""),
        request.form.get("employee_id", "")
    )
    user_id_raw = (request.form.get("user_id", "") or "").strip()
    adjustment_type = (request.form.get("adjustment_type", "") or "").strip()
    label = (request.form.get("label", "") or "").strip()
    notes = (request.form.get("notes", "") or "").strip()
    amount = parse_money_value(request.form.get("amount", "0"))

    if not user_id_raw.isdigit():
        flash("Please choose an employee for the payroll adjustment.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    if adjustment_type not in {"Allowance", "Deduction"}:
        flash("Payroll adjustment type must be Allowance or Deduction.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    if not label:
        flash("Payroll adjustment label is required.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    if amount <= 0:
        flash("Payroll adjustment amount must be greater than zero.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    employee = fetchone("""
        SELECT id, full_name, department
        FROM users
        WHERE id = ? AND role = 'employee'
    """, (int(user_id_raw),))
    if not employee:
        flash("Selected employee was not found.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    execute_db("""
        INSERT INTO payroll_adjustments (
            user_id, date_from, date_to, adjustment_type, label,
            amount, notes, created_by, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        employee["id"],
        filters["date_from_text"],
        filters["date_to_text"],
        adjustment_type,
        label,
        amount,
        notes or None,
        session["user_id"],
        now_str()
    ), commit=True)

    log_activity(
        session["user_id"],
        "ADD PAYROLL ADJUSTMENT",
        f"{adjustment_type} {format_currency(amount)} for {employee['full_name']} ({label})"
    )
    invalidate_reports_cache()
    flash(
        f"{adjustment_type} added for {employee['full_name']}. Save or re-release this payroll period to update the payslip snapshot.",
        "success"
    )
    return redirect(url_for("admin_payroll", **redirect_args))


@app.route("/admin/payroll/adjustments/<int:adjustment_id>/delete", methods=["POST"])
@login_required(role="admin")
def delete_payroll_adjustment(adjustment_id):
    redirect_args = payroll_filter_redirect_args(request.form)
    adjustment = fetchone("""
        SELECT pa.*, u.full_name AS employee_name
        FROM payroll_adjustments pa
        JOIN users u ON u.id = pa.user_id
        WHERE pa.id = ?
    """, (adjustment_id,))
    if not adjustment:
        flash("Payroll adjustment not found.", "warning")
        return redirect(url_for("admin_payroll", **redirect_args))

    execute_db("DELETE FROM payroll_adjustments WHERE id = ?", (adjustment_id,), commit=True)
    log_activity(
        session["user_id"],
        "DELETE PAYROLL ADJUSTMENT",
        f"Removed {adjustment['adjustment_type']} {format_currency(adjustment['amount'])} for {adjustment['employee_name']} ({adjustment['label']})"
    )
    invalidate_reports_cache()
    flash("Payroll adjustment removed. Save or re-release this payroll period to refresh the payslip snapshot.", "info")
    return redirect(url_for("admin_payroll", **redirect_args))


@app.route("/admin/payroll/recurring-rules", methods=["POST"])
@login_required(role="admin")
def save_payroll_recurring_rule():
    redirect_args = payroll_filter_redirect_args(request.form)
    rule_id_raw = (request.form.get("rule_id", "") or "").strip()
    user_id_raw = (request.form.get("user_id", "") or "").strip()
    adjustment_type = (request.form.get("adjustment_type", "") or "").strip()
    label = (request.form.get("label", "") or "").strip()
    recurrence_type = (request.form.get("recurrence_type", "") or "Every Payroll").strip() or "Every Payroll"
    start_date = (request.form.get("start_date", "") or "").strip()
    end_date = (request.form.get("end_date", "") or "").strip()
    notes = (request.form.get("notes", "") or "").strip()
    amount = parse_money_value(request.form.get("amount", "0"))

    if not user_id_raw.isdigit():
        flash("Please choose an employee for the recurring payroll rule.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    if adjustment_type not in {"Allowance", "Deduction"}:
        flash("Recurring payroll rule type must be Allowance or Deduction.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    if recurrence_type not in {"Every Payroll", "Monthly"}:
        flash("Recurring payroll rule recurrence must be Every Payroll or Monthly.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    if not label:
        flash("Recurring payroll rule label is required.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    if amount <= 0:
        flash("Recurring payroll rule amount must be greater than zero.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    employee = fetchone("""
        SELECT id, full_name
        FROM users
        WHERE id = ? AND role = 'employee'
    """, (int(user_id_raw),))
    if not employee:
        flash("Selected employee was not found.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    start_date_value = parse_iso_date(start_date) if start_date else None
    end_date_value = parse_iso_date(end_date) if end_date else None
    if start_date and not start_date_value:
        flash("Recurring payroll rule start date is invalid.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))
    if end_date and not end_date_value:
        flash("Recurring payroll rule end date is invalid.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))
    if start_date_value and end_date_value and start_date_value > end_date_value:
        flash("Recurring payroll rule end date must not be earlier than the start date.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))
    if recurrence_type == "Monthly" and not start_date_value:
        flash("Monthly recurring rules need a start date so the monthly anchor day is clear.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    timestamp = now_str()
    if rule_id_raw.isdigit():
        existing = fetchone("""
            SELECT *
            FROM payroll_recurring_rules
            WHERE id = ?
        """, (int(rule_id_raw),))
        if not existing:
            flash("Recurring payroll rule not found.", "warning")
            return redirect(url_for("admin_payroll", **redirect_args))

        execute_db("""
            UPDATE payroll_recurring_rules
            SET user_id = ?, adjustment_type = ?, label = ?, amount = ?,
                recurrence_type = ?, start_date = ?, end_date = ?, notes = ?, updated_at = ?
            WHERE id = ?
        """, (
            employee["id"],
            adjustment_type,
            label,
            amount,
            recurrence_type,
            start_date or None,
            end_date or None,
            notes or None,
            timestamp,
            existing["id"]
        ), commit=True)
        log_activity(
            session["user_id"],
            "UPDATE PAYROLL RECURRING RULE",
            f"Updated {adjustment_type.lower()} rule {label} for {employee['full_name']}."
        )
        invalidate_reports_cache()
        flash("Recurring payroll rule updated.", "success")
    else:
        execute_db("""
            INSERT INTO payroll_recurring_rules (
                user_id, adjustment_type, label, amount, recurrence_type,
                start_date, end_date, notes, is_active, created_by, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            employee["id"],
            adjustment_type,
            label,
            amount,
            recurrence_type,
            start_date or None,
            end_date or None,
            notes or None,
            1,
            session["user_id"],
            timestamp,
            timestamp
        ), commit=True)
        log_activity(
            session["user_id"],
            "ADD PAYROLL RECURRING RULE",
            f"Added {adjustment_type.lower()} recurring rule {label} for {employee['full_name']}."
        )
        invalidate_reports_cache()
        flash("Recurring payroll rule added. It will be included automatically when the rule matches the payroll period.", "success")

    return redirect(url_for("admin_payroll", **redirect_args))


@app.route("/admin/payroll/recurring-rules/<int:rule_id>/toggle", methods=["POST"])
@login_required(role="admin")
def toggle_payroll_recurring_rule(rule_id):
    redirect_args = payroll_filter_redirect_args(request.form)
    rule = fetchone("""
        SELECT prr.*, u.full_name AS employee_name
        FROM payroll_recurring_rules prr
        JOIN users u ON u.id = prr.user_id
        WHERE prr.id = ?
    """, (rule_id,))
    if not rule:
        flash("Recurring payroll rule not found.", "warning")
        return redirect(url_for("admin_payroll", **redirect_args))

    new_status = 0 if int(rule.get("is_active") or 0) == 1 else 1
    execute_db("""
        UPDATE payroll_recurring_rules
        SET is_active = ?, updated_at = ?
        WHERE id = ?
    """, (new_status, now_str(), rule_id), commit=True)
    log_activity(
        session["user_id"],
        "TOGGLE PAYROLL RECURRING RULE",
        f"{'Activated' if new_status == 1 else 'Paused'} recurring rule {rule['label']} for {rule['employee_name']}."
    )
    invalidate_reports_cache()
    flash(
        f"Recurring payroll rule {'activated' if new_status == 1 else 'paused'}.",
        "success" if new_status == 1 else "info"
    )
    return redirect(url_for("admin_payroll", **redirect_args))


@app.route("/admin/payroll/recurring-rules/<int:rule_id>/delete", methods=["POST"])
@login_required(role="admin")
def delete_payroll_recurring_rule(rule_id):
    redirect_args = payroll_filter_redirect_args(request.form)
    rule = fetchone("""
        SELECT prr.*, u.full_name AS employee_name
        FROM payroll_recurring_rules prr
        JOIN users u ON u.id = prr.user_id
        WHERE prr.id = ?
    """, (rule_id,))
    if not rule:
        flash("Recurring payroll rule not found.", "warning")
        return redirect(url_for("admin_payroll", **redirect_args))

    execute_db("DELETE FROM payroll_recurring_rules WHERE id = ?", (rule_id,), commit=True)
    log_activity(
        session["user_id"],
        "DELETE PAYROLL RECURRING RULE",
        f"Deleted recurring rule {rule['label']} for {rule['employee_name']}."
    )
    invalidate_reports_cache()
    flash("Recurring payroll rule deleted.", "info")
    return redirect(url_for("admin_payroll", **redirect_args))


@app.route("/admin/payroll/run", methods=["POST"])
@login_required(role="admin")
def save_payroll_run():
    redirect_args = payroll_filter_redirect_args(request.form)
    filters = normalize_payroll_filters(
        request.form.get("period", "this_month"),
        request.form.get("date_from", ""),
        request.form.get("date_to", ""),
        request.form.get("department", ""),
        request.form.get("employee_id", "")
    )
    action = (request.form.get("run_action", "draft") or "draft").strip().lower()
    status = "Released" if action == "release" else "Draft"
    notes = (request.form.get("notes", "") or "").strip()
    payroll_rows = build_payroll_rows(
        filters["date_from"],
        filters["date_to"],
        department_filter=filters["department_filter"],
        employee_filter=filters["employee_filter"]
    )

    if not payroll_rows:
        flash("There are no payroll rows in this view to save.", "warning")
        return redirect(url_for("admin_payroll", **redirect_args))

    if status == "Released" and any(row["has_rate"] == 0 for row in payroll_rows):
        flash("Release is blocked until every employee in view has an hourly rate.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    try:
        payroll_run, _ = save_payroll_run_snapshot(
            filters["date_from"],
            filters["date_to"],
            department_filter=filters["department_filter"],
            employee_filter=filters["employee_filter"],
            status=status,
            notes=notes,
            actor_id=session["user_id"]
        )
    except ValueError as exc:
        flash(str(exc), "warning")
        return redirect(url_for("admin_payroll", **redirect_args))
    log_activity(
        session["user_id"],
        "RELEASE PAYROLL" if status == "Released" else "SAVE PAYROLL DRAFT",
        f"{status} payroll for {filters['date_from_text']} to {filters['date_to_text']} ({len(payroll_rows)} employee row(s))"
    )
    invalidate_reports_cache()
    flash(
        f"Payroll {status.lower()} saved for {filters['date_from_text']} to {filters['date_to_text']}.",
        "success"
    )
    return redirect(url_for("admin_payroll", **redirect_args))


@app.route("/admin/payroll/runs/<int:payroll_run_id>/delete", methods=["POST"])
@login_required(role="admin")
def delete_payroll_run(payroll_run_id):
    redirect_args = payroll_filter_redirect_args(request.form)
    payroll_run = fetchone("""
        SELECT pr.*, creator.full_name AS created_by_name
        FROM payroll_runs pr
        LEFT JOIN users creator ON creator.id = pr.created_by
        WHERE pr.id = ?
    """, (payroll_run_id,))

    if not payroll_run:
        flash("Payroll run not found.", "warning")
        return redirect(url_for("admin_payroll", **redirect_args))

    payroll_run = enrich_admin_payroll_run(payroll_run)
    if payroll_run["status"] != "Draft":
        flash("Only draft payroll snapshots can be deleted.", "danger")
        return redirect(url_for("admin_payroll", **redirect_args))

    db = get_db()
    try:
        execute_db("DELETE FROM payroll_run_item_adjustments WHERE payroll_run_id = ?", (payroll_run_id,))
        execute_db("DELETE FROM payroll_run_items WHERE payroll_run_id = ?", (payroll_run_id,))
        execute_db("DELETE FROM payroll_runs WHERE id = ?", (payroll_run_id,))
        db.commit()
    except Exception:
        db.rollback()
        raise

    log_activity(
        session["user_id"],
        "DELETE PAYROLL DRAFT",
        f"Deleted draft payroll snapshot {payroll_run['period_label']} ({payroll_run['item_count']} employee row(s))"
    )
    invalidate_reports_cache()
    flash(f"Deleted draft payroll snapshot for {payroll_run['period_label']}.", "info")
    return redirect(url_for("admin_payroll", **redirect_args))


@app.route("/admin/payroll/bulk-release", methods=["POST"])
@login_required(role="admin")
def bulk_release_payroll_runs():
    redirect_args = payroll_filter_redirect_args(request.form)
    raw_ids = request.form.getlist("payroll_run_ids")
    payroll_run_ids = []
    seen_ids = set()
    for raw_id in raw_ids:
        raw_text = str(raw_id or "").strip()
        if not raw_text.isdigit():
            continue
        run_id = int(raw_text)
        if run_id in seen_ids:
            continue
        seen_ids.add(run_id)
        payroll_run_ids.append(run_id)

    if not payroll_run_ids:
        flash("Select at least one draft payroll run to bulk release.", "warning")
        return redirect(url_for("admin_payroll", **redirect_args))

    placeholders = ", ".join(["?"] * len(payroll_run_ids))
    runs = [
        enrich_admin_payroll_run(row)
        for row in fetchall(f"""
            SELECT pr.*, creator.full_name AS created_by_name
            FROM payroll_runs pr
            LEFT JOIN users creator ON creator.id = pr.created_by
            WHERE pr.id IN ({placeholders})
            ORDER BY pr.id DESC
        """, tuple(payroll_run_ids))
    ]
    run_map = {int(run["id"]): run for run in runs}

    releasable_runs = []
    skipped_messages = []
    for requested_run_id in payroll_run_ids:
        run = run_map.get(requested_run_id)
        if not run:
            skipped_messages.append(f"Run #{requested_run_id} was not found.")
            continue
        if run["status"] == "Released":
            skipped_messages.append(f"{run['period_label']} was already released.")
            continue
        if not run["can_release"]:
            skipped_messages.append(f"{run['period_label']} was skipped: {run['release_block_reason']}")
            continue
        releasable_runs.append(run)

    if not releasable_runs:
        flash("None of the selected payroll runs were eligible for release.", "warning")
        for message in skipped_messages[:4]:
            flash(message, "info")
        return redirect(url_for("admin_payroll", **redirect_args))

    timestamp = now_str()
    db = get_db()
    try:
        for run in releasable_runs:
            execute_db("""
                UPDATE payroll_runs
                SET status = 'Released',
                    updated_at = ?,
                    released_at = ?
                WHERE id = ?
            """, (timestamp, timestamp, run["id"]))
        db.commit()
    except Exception:
        db.rollback()
        raise

    released_labels = ", ".join(run["period_label"] for run in releasable_runs[:3])
    if len(releasable_runs) > 3:
        released_labels += f", and {len(releasable_runs) - 3} more"
    log_activity(
        session["user_id"],
        "BULK RELEASE PAYROLL",
        f"Released {len(releasable_runs)} payroll run(s): {released_labels}"
    )
    invalidate_reports_cache()
    flash(f"Released {len(releasable_runs)} payroll run(s).", "success")
    for message in skipped_messages[:4]:
        flash(message, "info")
    return redirect(url_for("admin_payroll", **redirect_args))


@app.route("/admin/payroll/export.xlsx")
@login_required(role="admin")
def export_admin_payroll_excel():
    filters = normalize_payroll_filters(
        request.args.get("period", "this_month"),
        request.args.get("date_from", ""),
        request.args.get("date_to", ""),
        request.args.get("department", ""),
        request.args.get("employee_id", "")
    )
    payroll_rows = build_payroll_rows(
        filters["date_from"],
        filters["date_to"],
        department_filter=filters["department_filter"],
        employee_filter=filters["employee_filter"]
    )
    adjustments = build_effective_payroll_adjustments(
        filters["date_from"],
        filters["date_to"],
        department_filter=filters["department_filter"],
        employee_filter=filters["employee_filter"]
    )
    current_run = get_payroll_run(
        filters["date_from"],
        filters["date_to"],
        department_filter=filters["department_filter"],
        employee_filter=filters["employee_filter"]
    )
    stats = build_payroll_stats(payroll_rows)

    try:
        from openpyxl import Workbook
    except Exception:
        flash("Excel export requires openpyxl. Install dependencies and try again.", "danger")
        return redirect(url_for("admin_payroll", **payroll_filter_redirect_args(request.args)))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Payroll Summary"
    sheet.append(["Payroll Period", f"{filters['date_from_text']} to {filters['date_to_text']}"])
    sheet.append(["Department Filter", filters["department_filter"] or "All Departments"])
    sheet.append(["Employee Filter", get_payroll_employee_filter_label(filters["employee_filter"])])
    sheet.append(["Run Status", current_run["status"] if current_run else "Not Saved"])
    sheet.append(["Gross Payroll", stats["total_gross"]])
    sheet.append(["Overtime Pay", stats["total_overtime_pay"]])
    sheet.append(["Allowances", stats["total_allowances"]])
    sheet.append(["Deductions", stats["total_deductions"]])
    sheet.append(["Final Payroll", stats["total_final_pay"]])
    sheet.append([])
    sheet.append([
        "Employee", "Username", "Department", "Position", "Hourly Rate",
        "Days Worked", "Total Hours", "Overtime Hours", "Late Minutes",
        "Break Minutes", "Suspension Days", "Lost Pay Estimate",
        "Gross Pay", "Overtime Pay", "Allowances", "Deductions", "Final Pay", "Status"
    ])
    for row in payroll_rows:
        sheet.append([
            row["full_name"],
            row["username"],
            row["department"],
            row["position"],
            row["hourly_rate"],
            row["days_worked"],
            row["total_hours"],
            row["overtime_hours"],
            row["late_minutes"],
            row["break_minutes"],
            row["suspension_days"],
            row["suspension_pay"],
            row["gross_pay"],
            row["overtime_pay"],
            row["allowances"],
            row["deductions"],
            row["final_pay"],
            row["status_label"],
        ])

    adjustment_sheet = workbook.create_sheet(title="Adjustments")
    adjustment_sheet.append(["Employee", "Source", "Recurrence", "Type", "Label", "Amount", "Notes", "Created By", "Created At"])
    for adjustment in adjustments:
        adjustment_sheet.append([
            adjustment["employee_name"],
            adjustment.get("source_kind") or "Manual",
            adjustment.get("recurrence_type") or "",
            adjustment["adjustment_type"],
            adjustment["label"],
            adjustment["amount"],
            adjustment["notes"] or "",
            adjustment["created_by_name"] or "",
            adjustment["created_at"],
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="payroll-summary-{filters["date_from_text"]}-to-{filters["date_to_text"]}.xlsx"'
        }
    )


@app.route("/admin/payroll/print")
@login_required(role="admin")
def print_admin_payroll():
    filters = normalize_payroll_filters(
        request.args.get("period", "this_month"),
        request.args.get("date_from", ""),
        request.args.get("date_to", ""),
        request.args.get("department", ""),
        request.args.get("employee_id", "")
    )
    payroll_rows = build_payroll_rows(
        filters["date_from"],
        filters["date_to"],
        department_filter=filters["department_filter"],
        employee_filter=filters["employee_filter"]
    )
    stats = build_payroll_stats(payroll_rows)
    adjustments = build_effective_payroll_adjustments(
        filters["date_from"],
        filters["date_to"],
        department_filter=filters["department_filter"],
        employee_filter=filters["employee_filter"]
    )
    current_run = get_payroll_run(
        filters["date_from"],
        filters["date_to"],
        department_filter=filters["department_filter"],
        employee_filter=filters["employee_filter"]
    )

    return render_template(
        "admin_payroll_print.html",
        payroll_rows=payroll_rows,
        stats=stats,
        adjustments=adjustments,
        current_run=current_run,
        department_filter=filters["department_filter"],
        employee_filter_label=get_payroll_employee_filter_label(filters["employee_filter"]),
        date_from=filters["date_from_text"],
        date_to=filters["date_to_text"],
    )


@app.route("/admin/reports")
@login_required(role="admin")
def admin_reports():
    filters = normalize_admin_report_filters(
        request.args.get("date_from", "").strip(),
        request.args.get("date_to", "").strip(),
        request.args.get("department", "").strip(),
        request.args.get("period", "").strip()
    )
    report_data = get_cached_admin_reports_data(
        filters["date_from"],
        filters["date_to"],
        department_filter=filters["department_filter"]
    )
    return render_template(
        "admin_reports.html",
        departments=get_department_options(),
        period=filters["period"],
        period_label=filters["period_label"],
        date_from=filters["date_from_text"],
        date_to=filters["date_to_text"],
        department_filter=filters["department_filter"],
        report_data=report_data,
        minutes_to_hm=minutes_to_hm
    )


@app.route("/admin/reports/export.xlsx")
@login_required(role="admin")
def export_admin_reports_excel():
    filters = normalize_admin_report_filters(
        request.args.get("date_from", "").strip(),
        request.args.get("date_to", "").strip(),
        request.args.get("department", "").strip(),
        request.args.get("period", "").strip()
    )
    report_data = get_cached_admin_reports_data(
        filters["date_from"],
        filters["date_to"],
        department_filter=filters["department_filter"]
    )
    try:
        from openpyxl import Workbook
    except Exception:
        flash("Excel export requires openpyxl. Install dependencies and try again.", "danger")
        return redirect(url_for(
            "admin_reports",
            period=filters["period"],
            date_from=filters["date_from_text"],
            date_to=filters["date_to_text"],
            department=filters["department_filter"]
        ))

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Date From", filters["date_from_text"]])
    summary_sheet.append(["Date To", filters["date_to_text"]])
    summary_sheet.append(["Preset", filters["period_label"]])
    summary_sheet.append(["Department", filters["department_filter"] or "All Departments"])
    summary_sheet.append([])
    summary_sheet.append(["Metric", "Value"])
    for label, value in [
        ("Employees in scope", report_data["summary"]["employee_count"]),
        ("Attendance days", report_data["summary"]["attendance_days"]),
        ("Attendance hours", report_data["summary"]["attendance_hours"]),
        ("Average hours per active day", report_data["summary"]["avg_hours_per_day"]),
        ("Average hours per employee", report_data["summary"]["avg_hours_per_employee"]),
        ("Late punches", report_data["summary"]["late_punches"]),
        ("Late rate percent", report_data["summary"]["late_rate_percent"]),
        ("Overtime hours", report_data["summary"]["overtime_hours"]),
        ("Overtime share percent", report_data["summary"]["overtime_share_percent"]),
        ("Break minutes", report_data["summary"]["break_minutes"]),
        ("Leave requests", report_data["summary"]["leave_requests"]),
        ("Incident follow-ups", report_data["summary"]["incident_follow_ups"]),
        ("Pending corrections", report_data["summary"]["pending_corrections"]),
        ("Incident reports", report_data["summary"]["incident_reports"]),
        ("Released payroll runs", report_data["summary"]["released_payroll_runs"]),
        ("Released payroll rows", report_data["summary"]["released_payroll_rows"]),
        ("Released payroll total", report_data["summary"]["released_payroll_total"]),
    ]:
        summary_sheet.append([label, value])
    autosize_workbook_sheet(summary_sheet)

    append_workbook_rows(workbook, "Trend Highlights", report_data["trend_highlights"])
    append_workbook_rows(workbook, "Department Highlights", report_data["department_highlights"])
    append_workbook_rows(workbook, "Case Summary", report_data["case_rows"])
    append_workbook_rows(workbook, "Department Summary", report_data["department_rows"])
    append_workbook_rows(workbook, "Daily Trend", report_data["daily_rows"])
    append_workbook_rows(workbook, "Leave Summary", report_data["leave_rows"])
    append_workbook_rows(workbook, "Correction Summary", report_data["correction_rows"])
    append_workbook_rows(workbook, "Incident Summary", report_data["incident_rows"])
    append_workbook_rows(workbook, "Employee Leaders", report_data["top_employee_rows"])
    append_workbook_rows(workbook, "Payroll Summary", report_data["payroll_department_rows"])
    append_workbook_rows(workbook, "Released Payroll", report_data["released_runs"])
    filename = f"admin-reports-{filters['date_from_text']}-to-{filters['date_to_text']}.xlsx"
    return workbook_to_response(workbook, filename)


@app.route("/admin/leave")
@login_required(role="admin")
def admin_leave_dashboard():
    year = parse_positive_int(request.args.get("year", str(now_dt().year)), now_dt().year)
    department = request.args.get("department", "").strip()
    employee_id = request.args.get("employee_id", "").strip()
    departments = get_department_options()
    employee_options = get_employee_options()
    if department:
        employee_options = [row for row in employee_options if (row.get("department") or "") == department]
    leave_rows = build_leave_dashboard_rows(year=year, department=department, user_id=employee_id or None)
    pending_requests = get_pending_leave_requests(user_id=employee_id or None, department=department, year=year)

    stats = {
        "employees": len(leave_rows),
        "sick_used": sum(row["sick_used"] for row in leave_rows),
        "paid_used": sum(row["paid_used"] for row in leave_rows),
        "sick_remaining": sum(row["sick_remaining"] for row in leave_rows),
        "paid_remaining": sum(row["paid_remaining"] for row in leave_rows),
        "pending_total": sum(row["pending_total"] for row in leave_rows),
        "sick_exhausted": len([row for row in leave_rows if row["sick_remaining"] <= 0]),
        "paid_exhausted": len([row for row in leave_rows if row["paid_remaining"] <= 0]),
        "overdue_pending": len([row for row in pending_requests if int(row.get("age_days") or 0) >= 3]),
    }

    return render_template(
        "admin_leave_dashboard.html",
        leave_rows=leave_rows,
        pending_requests=pending_requests,
        departments=departments,
        employees=employee_options,
        department=department,
        employee_id=employee_id,
        year=year,
        stats=stats
    )


@app.route("/admin/leave/export.xlsx")
@login_required(role="admin")
def export_admin_leave_dashboard_excel():
    year = parse_positive_int(request.args.get("year", str(now_dt().year)), now_dt().year)
    department = request.args.get("department", "").strip()
    employee_id = request.args.get("employee_id", "").strip()
    leave_rows = build_leave_dashboard_rows(year=year, department=department, user_id=employee_id or None)

    try:
        from openpyxl import Workbook
    except Exception:
        flash("Excel export requires openpyxl. Install dependencies and try again.", "danger")
        return redirect(url_for("admin_leave_dashboard", year=year, department=department, employee_id=employee_id))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leave Dashboard"
    selected_employee_name = ""
    if employee_id:
        selected_employee = next((row for row in get_employee_options() if str(row["id"]) == employee_id), None)
        selected_employee_name = selected_employee["full_name"] if selected_employee else employee_id
    sheet.append(["Year", year])
    sheet.append(["Department", department or "All Departments"])
    sheet.append(["Employee", selected_employee_name or "All Employees"])
    sheet.append([])
    sheet.append([
        "Employee", "Username", "Department", "Sick Allotment", "Sick Used", "Sick Remaining",
        "Paid Allotment", "Paid Used", "Paid Remaining", "Pending Sick", "Pending Paid",
        "Manual Sick Used", "Manual Paid Used", "Approved Sick Used", "Approved Paid Used"
    ])
    for row in leave_rows:
        sheet.append([
            row["full_name"], row["username"], row["department"], row["sick_total"], row["sick_used"], row["sick_remaining"],
            row["paid_total"], row["paid_used"], row["paid_remaining"], row["pending_sick"], row["pending_paid"],
            row["sick_used_manual"], row["paid_used_manual"], row["sick_used_in_app"], row["paid_used_in_app"]
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="leave-dashboard-{year}.xlsx"'}
    )


@app.route("/admin/history")
@login_required(role="admin")
def admin_history():
    search = request.args.get("search", "").strip()
    department = request.args.get("department", "").strip()
    type_filter = request.args.get("type_filter", "").strip()
    late_only = request.args.get("late_only", "").strip()
    absent_only = request.args.get("absent_only", "").strip()
    over_break_only = request.args.get("over_break_only", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    departments = get_department_options()
    enriched = build_admin_history_records(
        search=search,
        department=department,
        type_filter=type_filter,
        late_only=late_only,
        absent_only=absent_only,
        over_break_only=over_break_only,
        date_from=date_from,
        date_to=date_to,
        limit=200
    )

    return render_template(
        "admin_history.html",
        records=enriched,
        search=search,
        department=department,
        type_filter=type_filter,
        departments=departments,
        late_only=late_only,
        absent_only=absent_only,
        over_break_only=over_break_only,
        date_from=date_from,
        date_to=date_to,
        minutes_to_hm=minutes_to_hm
    )


@app.route("/admin/data-tools", methods=["GET", "POST"])
@login_required(role="admin")
def admin_data_tools():
    search = request.args.get("search", "").strip()
    cleanup_from = request.args.get("cleanup_from", "").strip()
    cleanup_to = request.args.get("cleanup_to", "").strip()

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "create_backup":
            try:
                backup_path = create_sqlite_backup()
                log_activity(session["user_id"], "CREATE BACKUP", f"Created SQLite backup at {backup_path}")
                flash(f"Backup created: {os.path.basename(backup_path)}", "success")
            except ValueError as exc:
                flash(str(exc), "danger")
            return redirect(url_for("admin_data_tools", search=search))
        if action == "record_external_backup":
            result = record_external_backup_marker(
                note=request.form.get("external_backup_note", ""),
                actor_id=session.get("user_id"),
            )
            log_activity(
                session["user_id"],
                "RECORD EXTERNAL BACKUP",
                f"Marked external provider backup verified at {result['backup_at']}."
            )
            flash("External Render/Postgres backup note saved. You now have a visible reminder before reset or cleanup.", "success")
            return redirect(url_for("admin_data_tools", search=search, cleanup_from=cleanup_from, cleanup_to=cleanup_to))
        if action in {"cleanup_logs_weekly", "cleanup_logs_monthly"}:
            retention_days = 7 if action == "cleanup_logs_weekly" else 30
            try:
                result = perform_log_retention_cleanup(retention_days)
                log_activity(
                    session["user_id"],
                    "CLEANUP LOGS",
                    f"Removed {result['total_removed']} old log row(s) older than {retention_days} day(s)."
                )
                flash(
                    "Cleanup completed. Removed "
                    f"{result['activity_logs']} activity log(s), "
                    f"{result['scanner_logs']} scanner log(s), "
                    f"{result['login_attempts']} login attempt row(s), and "
                    f"{result['read_notifications']} read notification(s) older than {retention_days} day(s).",
                    "success"
                )
            except ValueError as exc:
                flash(str(exc), "danger")
            return redirect(url_for("admin_data_tools", search=search))
        if action == "cleanup_logs_custom":
            cleanup_from = request.form.get("cleanup_from", "").strip()
            cleanup_to = request.form.get("cleanup_to", "").strip()
            try:
                result = perform_log_cleanup_for_date_range(cleanup_from, cleanup_to)
                log_activity(
                    session["user_id"],
                    "CLEANUP LOGS",
                    f"Removed {result['total_removed']} log row(s) for {result['date_from']} to {result['date_to']}."
                )
                flash(
                    "Cleanup completed for "
                    f"{result['date_from']} to {result['date_to']}. Removed "
                    f"{result['activity_logs']} activity log(s), "
                    f"{result['scanner_logs']} scanner log(s), "
                    f"{result['login_attempts']} login attempt row(s), and "
                    f"{result['read_notifications']} read notification(s).",
                    "success"
                )
                cleanup_from = result["date_from"]
                cleanup_to = result["date_to"]
            except ValueError as exc:
                flash(str(exc), "danger")
            return redirect(url_for("admin_data_tools", search=search, cleanup_from=cleanup_from, cleanup_to=cleanup_to))
        if action == "go_live_reset":
            confirmation = request.form.get("confirmation_text", "").strip().upper()
            if confirmation != "RESET":
                flash("Type RESET exactly before running the go-live reset.", "danger")
                return redirect(url_for("admin_data_tools", search=search, cleanup_from=cleanup_from, cleanup_to=cleanup_to))
            if using_postgres() and request.form.get("confirm_no_backup") != "1":
                flash("On Postgres/Render, no automatic database backup is created by this reset. Confirm that you understand before continuing.", "danger")
                return redirect(url_for("admin_data_tools", search=search, cleanup_from=cleanup_from, cleanup_to=cleanup_to))
            try:
                result = perform_go_live_reset()
                backup_note = f" Backup: {os.path.basename(result['backup_path'])}." if result.get("backup_path") else ""
                if not result.get("backup_supported"):
                    backup_note = " No automatic database backup was created for this Postgres reset."
                upload_note = f" Removed {result['removed_uploads']} orphaned proof uploads." if result.get("removed_uploads") else ""
                log_activity(session["user_id"], "GO-LIVE RESET", "Cleared operational attendance data for go-live.")
                flash(f"Go-live reset completed.{backup_note}{upload_note}", "success")
            except ValueError as exc:
                flash(str(exc), "danger")
            return redirect(url_for("admin_data_tools", search=search, cleanup_from=cleanup_from, cleanup_to=cleanup_to))

        attendance_id = request.form.get("attendance_id", "").strip()
        if not attendance_id:
            flash("Attendance record is required.", "danger")
            return redirect(url_for("admin_data_tools", search=search, cleanup_from=cleanup_from, cleanup_to=cleanup_to))

        try:
            updated_row = update_attendance_record_by_admin(
                attendance_id=int(attendance_id),
                time_in_value=request.form.get("time_in", "").strip(),
                time_out_value=request.form.get("time_out", "").strip(),
                clear_breaks=request.form.get("clear_breaks", "").strip() == "1",
            )
            employee = get_user_by_id(updated_row["user_id"]) if updated_row else None
            employee_name = employee["full_name"] if employee else f"Attendance #{attendance_id}"
            log_activity(session["user_id"], "FIX ATTENDANCE", f"Updated suspicious row #{attendance_id} for {employee_name}")
            flash("Attendance record updated.", "success")
        except ValueError as exc:
            flash(str(exc), "danger")

        return redirect(url_for("admin_data_tools", search=search, cleanup_from=cleanup_from, cleanup_to=cleanup_to))

    candidates = get_suspicious_attendance_records(search=search, limit=60)
    backups = get_backup_files(limit=12)
    cleanup_summary = get_log_cleanup_summary()
    recovery_snapshot = get_backup_recovery_snapshot()
    return render_template(
        "admin_data_tools.html",
        candidates=candidates,
        backups=backups,
        search=search,
        cleanup_from=cleanup_from,
        cleanup_to=cleanup_to,
        cleanup_summary=cleanup_summary,
        recovery_snapshot=recovery_snapshot,
        using_postgres_reset=using_postgres(),
        format_datetime_12h=format_datetime_12h,
        minutes_to_hm=minutes_to_hm
    )


@app.route("/admin/data-tools/recovery-pack.xlsx")
@login_required(role="admin")
def download_recovery_pack():
    try:
        workbook = build_recovery_pack_workbook()
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("admin_data_tools"))
    log_activity(session["user_id"], "DOWNLOAD RECOVERY PACK", "Downloaded the operational recovery workbook.")
    return workbook_to_response(workbook, f"recovery-pack-{today_str()}.xlsx")


@app.route("/admin/exceptions/export.xlsx")
@login_required(role="admin")
def export_admin_exceptions_excel():
    exception_type = request.args.get("type", "absent").strip().lower()
    department = request.args.get("department", "").strip()

    exception_groups = get_exception_collections(get_admin_employee_rows(department_filter=department))
    selected_rows = exception_groups.get(exception_type, [])
    titles = {
        "absent": "Absent Today",
        "late": "Late Today",
        "over_break": "Over Break",
        "missing_timeout": "Missing Time Out",
        "undertime": "Undertime Today",
    }

    try:
        from openpyxl import Workbook
    except Exception:
        flash("Excel export requires openpyxl. Install dependencies and try again.", "danger")
        return redirect(url_for("admin_dashboard", department=department))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = titles.get(exception_type, "Exceptions")
    sheet.append([
        "Employee",
        "Username",
        "Department",
        "Position",
        "Schedule",
        "Shift Start",
        "Shift End",
        "Status",
        "Time In",
        "Time Out",
        "Late Minutes",
        "Break Minutes",
        "Over Break Minutes"
    ])

    for row in selected_rows:
        sheet.append([
            row["full_name"] or "",
            row["username"] or "",
            row["department"] or "",
            row["position"] or "",
            row["schedule_summary"] or "",
            row["shift_start"] or "",
            row["shift_end"] or "",
            row["status_display"] or "",
            row["time_in"] or "",
            row["time_out"] or "",
            row["late_minutes"] if row["late_flag"] else 0,
            row["break_minutes"] or 0,
            row["over_break_minutes"] or 0,
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{exception_type}-exceptions-{today_str()}.xlsx"'}
    )


@app.route("/admin/corrections")
@login_required(role="admin")
def admin_corrections():
    status = request.args.get("status", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    requests = get_correction_requests(status=status, date_from=date_from, date_to=date_to)

    return render_template(
        "admin_corrections.html",
        requests=requests,
        status=status,
        date_from=date_from,
        date_to=date_to
    )


@app.route("/admin/corrections/<int:request_id>/update", methods=["POST"])
@login_required(role="admin")
def update_correction_request(request_id):
    status = request.form.get("status", "").strip()
    admin_note = request.form.get("admin_note", "").strip()
    requested_time_in = request.form.get("requested_time_in", "")
    requested_break_start = request.form.get("requested_break_start", "")
    requested_break_end = request.form.get("requested_break_end", "")
    requested_time_out = request.form.get("requested_time_out", "")

    if status not in {"Pending", "Approved", "Rejected"}:
        flash("Invalid correction status.", "danger")
        return redirect(url_for("admin_corrections"))

    correction = fetchone("""
        SELECT c.*, u.full_name
        FROM correction_requests c
        JOIN users u ON u.id = c.user_id
        WHERE c.id = ?
    """, (request_id,))

    if not correction:
        flash("Correction request not found.", "danger")
        return redirect(url_for("admin_corrections"))

    current_status = (correction["status"] or "").strip()
    review_locked = current_status in {"Approved", "Rejected"}
    if review_locked and status != current_status:
        flash(
            f"This correction request is already {current_status.lower()}. "
            "Update the note only, or create a new request if another correction is needed.",
            "warning",
        )
        return redirect(url_for("admin_corrections"))

    if review_locked:
        execute_db("""
            UPDATE correction_requests
            SET admin_note = ?
            WHERE id = ?
        """, (
            admin_note,
            request_id
        ), commit=True)
        flash("Reviewed correction note updated.", "success")
        return redirect(url_for("admin_corrections"))

    applied_changes = correction["applied_changes"] if "applied_changes" in correction.keys() else None

    try:
        requested_time_in = normalize_optional_clock_time(requested_time_in)
        requested_break_start = normalize_optional_clock_time(requested_break_start)
        requested_break_end = normalize_optional_clock_time(requested_break_end)
        requested_time_out = normalize_optional_clock_time(requested_time_out)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("admin_corrections"))

    if correction["request_type"] in LEAVE_REQUEST_TYPES:
        requested_time_in = ""
        requested_break_start = ""
        requested_break_end = ""
        requested_time_out = ""
    elif correction["request_type"] == "Undertime" and status == "Approved" and not requested_time_out:
        flash("Requested time out is required to approve an undertime request.", "danger")
        return redirect(url_for("admin_corrections"))

    if status == "Approved":
        if correction["request_type"] in (LEAVE_REQUEST_TYPES | {"Undertime"}):
            requested_time_out_dt = combine_work_date_and_time(correction["work_date"], requested_time_out) if requested_time_out else None
            attendance, _ = get_matching_attendance_context_for_request(
                correction["user_id"],
                correction["work_date"],
                request_type=correction["request_type"],
                requested_time_out=requested_time_out_dt
            )
            if correction["request_type"] == "Undertime" and attendance and requested_time_out:
                try:
                    applied_changes = apply_attendance_correction(
                        correction["user_id"],
                        attendance["work_date"],
                        time_out_value=requested_time_out
                    )
                except ValueError as exc:
                    flash(str(exc), "danger")
                    return redirect(url_for("admin_corrections"))
            else:
                applied_changes = describe_request_review_result(
                    correction["request_type"],
                    format_request_date_range(correction["work_date"], correction.get("end_work_date")),
                    requested_time_out
                )
        else:
            try:
                applied_changes = apply_attendance_correction(
                    correction["user_id"],
                    correction["work_date"],
                    time_in_value=requested_time_in,
                    break_start_value=requested_break_start,
                    break_end_value=requested_break_end,
                    time_out_value=requested_time_out
                )
            except ValueError as exc:
                flash(str(exc), "danger")
                return redirect(url_for("admin_corrections"))

        reviewed_at = now_str() if status in {"Approved", "Rejected"} else None
    reviewed_by = session["user_id"] if status in {"Approved", "Rejected"} else None

    preview_request_type = correction["request_type"] if correction["request_type"] == "Undertime" else ""
    preview_requested_time_out = combine_work_date_and_time(correction["work_date"], requested_time_out) if requested_time_out else None
    attendance, break_row = get_matching_attendance_context_for_request(
        correction["user_id"],
        correction["work_date"],
        request_type=preview_request_type,
        requested_time_out=preview_requested_time_out
    )
    preview_work_date = attendance["work_date"] if attendance and correction["request_type"] == "Undertime" else correction["work_date"]

    if correction["request_type"] == "Undertime":
        requested_break_start = ""
        requested_break_end = ""

    requested_time_in_dt, requested_break_start_dt, requested_break_end_dt, requested_time_out_dt = resolve_correction_datetimes(
        preview_work_date,
        time_in_value=requested_time_in,
        break_start_value=requested_break_start,
        break_end_value=requested_break_end,
        time_out_value=requested_time_out,
        existing_time_in=attendance["time_in"] if attendance else None,
        existing_break_start=break_row["break_start"] if break_row else None,
        existing_break_end=break_row["break_end"] if break_row else None,
        existing_time_out=attendance["time_out"] if attendance else None,
        use_existing_values=False
    )

    execute_db("""
        UPDATE correction_requests
        SET status = ?, admin_note = ?, reviewed_by = ?, reviewed_at = ?,
            requested_time_in = ?, requested_break_start = ?, requested_break_end = ?, requested_time_out = ?,
            applied_changes = ?
        WHERE id = ?
    """, (
        status,
        admin_note,
        reviewed_by,
        reviewed_at,
        requested_time_in_dt,
        requested_break_start_dt,
        requested_break_end_dt,
        requested_time_out_dt,
        applied_changes if status == "Approved" else None,
        request_id
    ), commit=True)

    notification_message = f"Your {correction['request_type']} request for {correction['work_date']} is now {status}."
    if correction["request_type"] in LEAVE_REQUEST_TYPES:
        notification_message = f"Your {correction['request_type']} request for {format_request_date_range(correction['work_date'], correction.get('end_work_date'))} is now {status}."
    if status == "Approved" and applied_changes:
        notification_message = f"{notification_message} Applied: {applied_changes}"

    create_notification(
        correction["user_id"],
        "Correction Request Updated",
        notification_message
    )
    log_details = f"{status} correction request #{request_id} for {correction['full_name']}"
    if status == "Approved" and applied_changes:
        log_details = f"{log_details} | {applied_changes}"
    log_activity(session["user_id"], "REVIEW CORRECTION", log_details, target_user_id=correction["user_id"])
    flash("Correction request updated.", "success")
    return redirect(url_for("admin_corrections"))


@app.route("/admin/history/export.xlsx")
@login_required(role="admin")
def export_admin_history_excel():
    search = request.args.get("search", "").strip()
    department = request.args.get("department", "").strip()
    type_filter = request.args.get("type_filter", "").strip()
    late_only = request.args.get("late_only", "").strip()
    absent_only = request.args.get("absent_only", "").strip()
    over_break_only = request.args.get("over_break_only", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    records = build_admin_history_records(
        search=search,
        department=department,
        type_filter=type_filter,
        late_only=late_only,
        absent_only=absent_only,
        over_break_only=over_break_only,
        date_from=date_from,
        date_to=date_to,
        limit=500
    )

    try:
        from openpyxl import Workbook
    except Exception:
        flash("Excel export requires openpyxl. Install dependencies and try again.", "danger")
        return redirect(url_for(
            "admin_history",
            search=search,
            department=department,
            type_filter=type_filter,
            late_only=late_only,
            absent_only=absent_only,
            over_break_only=over_break_only,
            date_from=date_from,
            date_to=date_to
        ))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance History"
    sheet.append([
        "Employee",
        "Username",
        "Work Date",
        "Record Type",
        "Time In",
        "Time Out",
        "Status",
        "Late Minutes",
        "Break Limit",
        "Break Minutes",
        "Overbreak Minutes",
        "Work Hours",
        "Proof File",
        "Admin Note"
    ])

    for item in records:
        row = item["row"]
        sheet.append([
            row["full_name"] or "",
            row["username"] or "",
            row["work_date"] or "",
            item["record_type"] or "Attendance",
            row["time_in"] or "",
            row["time_out"] or "",
            item["display_status"] or row["status"] or "",
            row["late_minutes"] if row["late_flag"] else 0,
            parse_break_limit_minutes(row["break_limit_minutes"]) if row.get("break_limit_minutes") is not None else 0,
            item["break_minutes"],
            item["over_break_minutes"],
            minutes_to_decimal_hours(item["work_minutes"]),
            row["proof_file"] or "",
            item["request_note"]
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"attendance-history-{today_str()}.xlsx\"'}
    )


@app.route("/admin/error-reports")
@login_required(role="admin")
def admin_error_reports():
    report_employee = request.args.get("report_employee", "").strip()
    report_department = request.args.get("report_department", "").strip()
    report_type = request.args.get("report_type", "").strip()
    report_date_from = request.args.get("report_date_from", "").strip()
    report_date_to = request.args.get("report_date_to", "").strip()

    employees = get_employee_options()
    departments = get_department_options()
    error_types = get_incident_error_type_options()
    reports = get_incident_reports(
        report_employee=report_employee,
        report_department=report_department,
        report_type=report_type,
        report_date_from=report_date_from,
        report_date_to=report_date_to
    )

    return render_template(
        "admin_error_reports.html",
        employees=employees,
        departments=departments,
        reports=reports,
        report_employee=report_employee,
        report_department=report_department,
        report_type=report_type,
        report_date_from=report_date_from,
        report_date_to=report_date_to,
        error_types=error_types
    )


@app.route("/admin/error-reports/export.xlsx")
@login_required(role="admin")
def export_admin_error_reports_excel():
    report_employee = request.args.get("report_employee", "").strip()
    report_department = request.args.get("report_department", "").strip()
    report_type = request.args.get("report_type", "").strip()
    report_date_from = request.args.get("report_date_from", "").strip()
    report_date_to = request.args.get("report_date_to", "").strip()

    reports = get_incident_reports(
        report_employee=report_employee,
        report_department=report_department,
        report_type=report_type,
        report_date_from=report_date_from,
        report_date_to=report_date_to
    )

    try:
        from openpyxl import Workbook
    except Exception:
        flash("Excel export requires openpyxl. Install dependencies and try again.", "danger")
        return redirect(url_for(
            "admin_error_reports",
            report_employee=report_employee,
            report_department=report_department,
            report_type=report_type,
            report_date_from=report_date_from,
            report_date_to=report_date_to
        ))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Error Reports"
    sheet.append([
        "Employee",
        "Department",
        "Error Type",
        "Report Date",
        "Message",
        "Status",
        "Policy Count",
        "Policy Step",
        "Linked Disciplinary",
        "Admin Note",
        "Created At",
        "Reviewed At",
        "Reviewed By"
    ])

    for report in reports:
        sheet.append([
            report["full_name"] if report["full_name"] else report["employee_name"] if report["employee_name"] else "Unknown",
            report["report_department"] or report["department"] or "",
            report["error_type"] or "",
            report["report_date"] if report["report_date"] else report["incident_date"] if report["incident_date"] else "",
            report["message"] or "",
            report["status"] or "Open",
            report["policy_incident_count"] or "",
            report["incident_action"] or "",
            f"#{report['disciplinary_action_id']} {report['linked_action_type']}" if report.get("disciplinary_action_id") else "",
            report["admin_note"] or "",
            report["created_at"] or "",
            report["reviewed_at"] or "",
            report["reviewed_by_name"] or ""
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="error-reports-{today_str()}.xlsx"'}
    )


@app.route("/admin/incident-report")
@login_required(role="admin")
def admin_incident_report():
    employees = fetchall("""
        SELECT id, full_name
        FROM users
        WHERE role = 'employee'
        ORDER BY full_name ASC
    """)
    return render_template(
        "admin_incident_report.html",
        employees=employees,
        error_types=get_incident_error_type_options(),
        incident_policy=INCIDENT_DISCIPLINARY_POLICY
    )


@app.route("/admin/disciplinary")
@login_required(role="admin")
def admin_disciplinary_dashboard():
    action_type = request.args.get("action_type", "").strip()
    user_id = request.args.get("user_id", "").strip()
    department = request.args.get("department", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    employees = get_employee_options()
    departments = get_department_options()
    actions = get_disciplinary_actions(
        action_type=action_type,
        user_id=user_id,
        department=department,
        date_from=date_from,
        date_to=date_to
    )

    summary = {
        "coaching": len([row for row in actions if row["action_type"] == "Coaching"]),
        "nte": len([row for row in actions if row["action_type"] == "NTE"]),
        "suspension": len([row for row in actions if row["action_type"] == "Suspension"]),
        "termination": len([row for row in actions if row["action_type"] == "Termination"]),
        "active_suspensions": len([row for row in actions if row["action_type"] == "Suspension" and row["status_label"] == "Active"]),
        "upcoming_suspensions": len([row for row in actions if row["action_type"] == "Suspension" and row["status_label"] == "Upcoming"]),
        "starts_today": len([row for row in actions if row["action_type"] == "Suspension" and row["action_date"] == today_str()]),
    }

    return render_template(
        "admin_disciplinary_dashboard.html",
        employees=employees,
        departments=departments,
        disciplinary_types=DISCIPLINARY_ACTION_TYPES,
        actions=actions,
        action_type=action_type,
        user_id=user_id,
        department=department,
        date_from=date_from,
        date_to=date_to,
        summary=summary
    )


@app.route("/admin/disciplinary/export.xlsx")
@login_required(role="admin")
def export_admin_disciplinary_excel():
    action_type = request.args.get("action_type", "").strip()
    user_id = request.args.get("user_id", "").strip()
    department = request.args.get("department", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    actions = get_disciplinary_actions(
        action_type=action_type,
        user_id=user_id,
        department=department,
        date_from=date_from,
        date_to=date_to
    )

    try:
        from openpyxl import Workbook
    except Exception:
        flash("Excel export requires openpyxl. Install dependencies and try again.", "danger")
        return redirect(url_for(
            "admin_disciplinary_dashboard",
            action_type=action_type,
            user_id=user_id,
            department=department,
            date_from=date_from,
            date_to=date_to
        ))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Disciplinary"
    sheet.append(["Employee", "Username", "Department", "Type", "Start Date", "Duration Days", "End Date", "Status", "Incident #", "Error Type", "Details"])
    for row in actions:
        sheet.append([
            row["full_name"], row["username"], row["department"], row["action_type"], row["action_date"],
            row["duration_days"] if row["action_type"] == "Suspension" else "",
            row["end_date"] or row["action_date"], row["status_label"],
            row.get("incident_report_id") or "",
            row.get("error_type") or row.get("incident_error_type") or "",
            row["details"] or ""
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="disciplinary-records-{today_str()}.xlsx"'}
    )


@app.route("/admin/error-reports/<int:report_id>/update", methods=["POST"])
@login_required(role="admin")
def update_incident_report(report_id):
    status = request.form.get("status", "").strip()
    admin_note = request.form.get("admin_note", "").strip()

    if status not in {"Open", "Reviewed", "Resolved"}:
        flash("Invalid report status.", "danger")
        return redirect(url_for("admin_error_reports"))

    report = fetchone("""
        SELECT r.*, u.full_name
        FROM incident_reports r
        LEFT JOIN users u ON u.id = r.user_id
        WHERE r.id = ?
    """, (report_id,))

    if not report:
        flash("Report not found.", "danger")
        return redirect(url_for("admin_error_reports"))

    reviewed_at = now_str() if status in {"Reviewed", "Resolved"} else None
    reviewed_by = session["user_id"] if status in {"Reviewed", "Resolved"} else None

    execute_db("""
        UPDATE incident_reports
        SET status = ?, admin_note = ?, reviewed_by = ?, reviewed_at = ?
        WHERE id = ?
    """, (status, admin_note, reviewed_by, reviewed_at, report_id), commit=True)

    employee_name = report["full_name"] if report["full_name"] else report["employee_name"] or f"User {report['user_id']}"
    log_activity(
        session["user_id"],
        "UPDATE INCIDENT",
        f"Incident #{report_id} for {employee_name} marked as {status}"
    )

    flash("Incident report updated.", "success")
    return redirect(url_for("admin_error_reports"))


@app.route("/admin/error-reports/<int:report_id>/edit", methods=["POST"])
@login_required(role="admin")
def edit_incident_report(report_id):
    report = fetchone("""
        SELECT *
        FROM incident_reports
        WHERE id = ?
    """, (report_id,))
    if not report:
        flash("Report not found.", "danger")
        return redirect(url_for("admin_error_reports"))

    error_type = normalize_incident_error_type(
        request.form.get("error_type", ""),
        request.form.get("new_error_type", "")
    )
    report_date = request.form.get("report_date", "").strip()
    report_department = request.form.get("report_department", "").strip()
    message = request.form.get("message", "").strip()

    if not error_type or not report_date:
        flash("Error type and report date are required.", "danger")
        return redirect(url_for("admin_error_reports"))

    execute_db("""
        UPDATE incident_reports
        SET error_type = ?, report_date = ?, incident_date = ?, report_department = ?, message = ?,
            policy_incident_count = 0
        WHERE id = ?
    """, (error_type, report_date, report_date, report_department, message, report_id), commit=True)
    sync_result = sync_incident_policy(report_id, session["user_id"], allow_create=not bool(report["disciplinary_action_id"]))

    employee = get_user_by_id(report["user_id"])
    employee_name = employee["full_name"] if employee else report.get("employee_name") or f"User {report['user_id']}"
    log_activity(session["user_id"], "EDIT INCIDENT", f"Edited incident #{report_id} for {employee_name}")
    if sync_result.get("created_action"):
        flash(f"Incident report updated. Linked {sync_result['policy_action']} record created.", "success")
    else:
        flash("Incident report updated.", "success")
    return redirect(url_for("admin_error_reports"))


@app.route("/admin/error-reports/<int:report_id>/delete", methods=["POST"])
@login_required(role="admin")
def delete_incident_report(report_id):
    report = fetchone("""
        SELECT *
        FROM incident_reports
        WHERE id = ?
    """, (report_id,))
    if not report:
        flash("Report not found.", "danger")
        return redirect(url_for("admin_error_reports"))

    employee = get_user_by_id(report["user_id"])
    employee_name = employee["full_name"] if employee else report.get("employee_name") or f"User {report['user_id']}"
    execute_db("DELETE FROM incident_reports WHERE id = ?", (report_id,), commit=True)
    log_activity(session["user_id"], "DELETE INCIDENT", f"Deleted incident #{report_id} for {employee_name}")
    flash("Incident report deleted.", "info")
    return redirect(url_for("admin_error_reports"))


def normalize_employee_id_dates(id_issue_date_raw, id_expiration_date_raw):
    id_issue_date = (id_issue_date_raw or "").strip()
    id_expiration_date = (id_expiration_date_raw or "").strip()

    issue_date_obj = None
    expiration_date_obj = None

    if id_issue_date:
        try:
            issue_date_obj = datetime.strptime(id_issue_date, "%Y-%m-%d").date()
            id_issue_date = issue_date_obj.strftime("%Y-%m-%d")
        except ValueError:
            return None, None, "ID Issue Date must be a valid date."

    if id_expiration_date:
        try:
            expiration_date_obj = datetime.strptime(id_expiration_date, "%Y-%m-%d").date()
            id_expiration_date = expiration_date_obj.strftime("%Y-%m-%d")
        except ValueError:
            return None, None, "ID Expiration Date must be a valid date."

    if issue_date_obj and expiration_date_obj and expiration_date_obj < issue_date_obj:
        return None, None, "ID Expiration Date cannot be earlier than ID Issue Date."

    return id_issue_date, id_expiration_date, None


@app.route("/admin/employees", methods=["GET", "POST"])
@login_required(role="admin")
def manage_employees():
    employee_search = request.values.get("search", "").strip()

    if request.method == "POST":
        form_action = (request.form.get("form_action", "add_employee") or "add_employee").strip()

        if form_action == "save_schedule_special_rule":
            special_date = (request.form.get("special_date", "") or "").strip()
            rule_type = normalize_schedule_special_rule_type(request.form.get("rule_type", "holiday"))
            label = build_schedule_special_rule_label(rule_type, request.form.get("special_rule_label", ""))
            notes = (request.form.get("special_rule_notes", "") or "").strip()
            special_date_value = parse_iso_date(special_date)
            if not special_date_value:
                flash("Choose a valid holiday or rest-day date.", "danger")
                return redirect(url_for("manage_employees", search=employee_search))
            if special_date_value < now_dt().date():
                flash("Holiday and rest-day rules can only be created for today or future dates.", "danger")
                return redirect(url_for("manage_employees", search=employee_search))

            existing_rule = fetchone("""
                SELECT id
                FROM schedule_special_dates
                WHERE special_date = ?
                LIMIT 1
            """, (special_date_value.strftime("%Y-%m-%d"),))
            if existing_rule:
                execute_db("""
                    UPDATE schedule_special_dates
                    SET rule_type = ?, label = ?, notes = ?, created_by = ?, updated_at = ?
                    WHERE id = ?
                """, (
                    rule_type,
                    label,
                    notes or None,
                    session["user_id"],
                    now_str(),
                    existing_rule["id"]
                ), commit=True)
                log_activity(session["user_id"], "UPDATE SCHEDULE RULE", f"Updated {SCHEDULE_SPECIAL_RULE_LABELS[rule_type]} rule for {special_date_value.strftime('%Y-%m-%d')}.")
                flash(f"Updated the {SCHEDULE_SPECIAL_RULE_LABELS[rule_type].lower()} rule for {special_date_value.strftime('%Y-%m-%d')}.", "success")
            else:
                execute_db("""
                    INSERT INTO schedule_special_dates (
                        special_date, rule_type, label, notes, created_by, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    special_date_value.strftime("%Y-%m-%d"),
                    rule_type,
                    label,
                    notes or None,
                    session["user_id"],
                    now_str(),
                    now_str()
                ), commit=True)
                log_activity(session["user_id"], "CREATE SCHEDULE RULE", f"Created {SCHEDULE_SPECIAL_RULE_LABELS[rule_type]} rule for {special_date_value.strftime('%Y-%m-%d')}.")
                flash(f"Saved the {SCHEDULE_SPECIAL_RULE_LABELS[rule_type].lower()} rule for {special_date_value.strftime('%Y-%m-%d')}.", "success")

            invalidate_schedule_special_rule_cache()
            invalidate_admin_employee_rows_cache()
            invalidate_reports_cache()
            return redirect(url_for("manage_employees", search=employee_search))

        if form_action == "delete_schedule_special_rule":
            raw_rule_id = (request.form.get("schedule_special_rule_id", "") or "").strip()
            if not raw_rule_id.isdigit():
                flash("Schedule rule not found.", "warning")
                return redirect(url_for("manage_employees", search=employee_search))
            existing_rule = fetchone("""
                SELECT *
                FROM schedule_special_dates
                WHERE id = ?
                LIMIT 1
            """, (int(raw_rule_id),))
            if not existing_rule:
                flash("Schedule rule not found.", "warning")
                return redirect(url_for("manage_employees", search=employee_search))
            existing_rule = dict(existing_rule)
            execute_db("DELETE FROM schedule_special_dates WHERE id = ?", (int(raw_rule_id),), commit=True)
            invalidate_schedule_special_rule_cache()
            invalidate_admin_employee_rows_cache()
            invalidate_reports_cache()
            log_activity(session["user_id"], "DELETE SCHEDULE RULE", f"Deleted {build_schedule_special_rule_label(existing_rule.get('rule_type'), existing_rule.get('label'))} on {existing_rule.get('special_date')}.")
            flash("Schedule rule deleted.", "info")
            return redirect(url_for("manage_employees", search=employee_search))

        if form_action == "create_schedule_preset":
            preset_name = (request.form.get("preset_name", "") or "").strip()
            department_scope = (request.form.get("preset_department_scope", "") or "").strip()
            notes = (request.form.get("preset_notes", "") or "").strip()
            schedule_days = normalize_schedule_days(request.form.getlist("preset_schedule_days"))
            shift_start = parse_shift_start(request.form.get("preset_shift_start", DEFAULT_SHIFT_START))
            shift_end = parse_shift_end(request.form.get("preset_shift_end", DEFAULT_SHIFT_END))
            break_limit_minutes = parse_break_limit_minutes(request.form.get("preset_break_limit_minutes", BREAK_LIMIT_MINUTES))

            if not preset_name:
                flash("Schedule preset name is required.", "danger")
                return redirect(url_for("manage_employees", search=employee_search))

            existing_preset = fetchone("SELECT id FROM schedule_presets WHERE LOWER(name) = LOWER(?)", (preset_name,))
            if existing_preset:
                flash("A schedule preset with that name already exists.", "warning")
                return redirect(url_for("manage_employees", search=employee_search))

            execute_db("""
                INSERT INTO schedule_presets (
                    name, department_scope, schedule_days, shift_start, shift_end,
                    break_limit_minutes, notes, created_by, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                preset_name,
                department_scope or None,
                schedule_days,
                shift_start,
                shift_end,
                break_limit_minutes,
                notes or None,
                session["user_id"],
                now_str(),
                now_str()
            ), commit=True)
            log_activity(session["user_id"], "CREATE SCHEDULE PRESET", f"Created preset {preset_name} ({schedule_days} | {shift_start}-{shift_end} | break {break_limit_minutes}m)")
            flash("Schedule preset created.", "success")
            return redirect(url_for("manage_employees", search=employee_search))

        if form_action == "delete_schedule_preset":
            preset = get_schedule_preset(request.form.get("schedule_preset_id", ""))
            if not preset:
                flash("Schedule preset not found.", "warning")
                return redirect(url_for("manage_employees", search=employee_search))

            affected_employees = [
                dict(row)
                for row in fetchall(
                    "SELECT * FROM users WHERE role = 'employee' AND schedule_preset_id = ?",
                    (preset["id"],)
                )
            ]

            db = get_db()
            try:
                execute_db("UPDATE users SET schedule_preset_id = NULL WHERE schedule_preset_id = ?", (preset["id"],))
                for employee in affected_employees:
                    employee["schedule_preset_id"] = None
                    record_employee_schedule_history(employee, actor_id=session["user_id"])
                execute_db("DELETE FROM schedule_presets WHERE id = ?", (preset["id"],))
                db.commit()
            except Exception:
                db.rollback()
                raise

            log_activity(session["user_id"], "DELETE SCHEDULE PRESET", f"Deleted preset {preset['name']}")
            flash("Schedule preset deleted. Employees keep their last saved schedule values.", "info")
            return redirect(url_for("manage_employees", search=employee_search))

        if form_action == "apply_schedule_preset_bulk":
            preset = get_schedule_preset(request.form.get("bulk_schedule_preset_id", ""))
            effective_date = (request.form.get("bulk_effective_date", "") or "").strip()
            change_notes = (request.form.get("bulk_schedule_notes", "") or "").strip()
            employee_ids = []
            for raw_id in request.form.getlist("employee_ids"):
                raw_text = str(raw_id or "").strip()
                if raw_text.isdigit():
                    employee_ids.append(int(raw_text))
            employee_ids = sorted(set(employee_ids))

            if not preset:
                flash("Choose a valid schedule preset to apply.", "danger")
                return redirect(url_for("manage_employees", search=employee_search))
            if not employee_ids:
                flash("Select at least one employee before applying a schedule preset.", "warning")
                return redirect(url_for("manage_employees", search=employee_search))

            placeholders = ", ".join(["?"] * len(employee_ids))
            existing_ids = [
                int(row["id"])
                for row in fetchall(
                    f"SELECT id FROM users WHERE role = 'employee' AND id IN ({placeholders})",
                    tuple(employee_ids)
                )
            ]
            if not existing_ids:
                flash("No valid employees were selected.", "warning")
                return redirect(url_for("manage_employees", search=employee_search))

            update_placeholders = ", ".join(["?"] * len(existing_ids))
            selected_employees = [
                dict(row)
                for row in fetchall(
                    f"SELECT * FROM users WHERE role = 'employee' AND id IN ({update_placeholders})",
                    tuple(existing_ids)
                )
            ]
            scope_conflicts = [
                employee["full_name"]
                for employee in selected_employees
                if not schedule_preset_matches_department(preset, employee.get("department"))
            ]
            if scope_conflicts:
                flash(
                    f"Preset {preset['name']} is scoped to {preset['department_scope']} and cannot be applied to: {', '.join(scope_conflicts[:3])}" +
                    ("..." if len(scope_conflicts) > 3 else ""),
                    "danger"
                )
                return redirect(url_for("manage_employees", search=employee_search))

            effective_date_value = parse_iso_date(effective_date) if effective_date else None
            if effective_date and not effective_date_value:
                flash("Choose a valid effective date for the bulk schedule change.", "danger")
                return redirect(url_for("manage_employees", search=employee_search))
            if effective_date_value and effective_date_value <= now_dt().date():
                flash("Leave Effective Date blank to apply immediately, or choose a future date to queue the rollout.", "warning")
                return redirect(url_for("manage_employees", search=employee_search))

            if effective_date_value:
                db = get_db()
                queued_count = 0
                try:
                    for employee in selected_employees:
                        queue_future_schedule_change(
                            employee,
                            {
                                "schedule_preset_id": preset["id"],
                                "schedule_days": preset["schedule_days"],
                                "shift_start": preset["shift_start"],
                                "shift_end": preset["shift_end"],
                                "break_limit_minutes": preset["break_limit_minutes"],
                            },
                            effective_date_value.strftime("%Y-%m-%d"),
                            actor_id=session["user_id"],
                            notes=change_notes
                        )
                        queued_count += 1
                    db.commit()
                except Exception:
                    db.rollback()
                    raise

                invalidate_reports_cache()
                log_activity(
                    session["user_id"],
                    "QUEUE SCHEDULE PRESET",
                    f"Queued preset {preset['name']} for {queued_count} employee(s) effective {effective_date_value.strftime('%Y-%m-%d')}"
                )
                flash(
                    f"Queued {preset['name']} for {queued_count} employee(s) starting {effective_date_value.strftime('%Y-%m-%d')}.",
                    "success"
                )
                return redirect(url_for("manage_employees", search=employee_search))

            db = get_db()
            try:
                execute_db(f"""
                    UPDATE users
                    SET schedule_preset_id = ?,
                        schedule_days = ?,
                        shift_start = ?,
                        shift_end = ?,
                        break_limit_minutes = ?
                    WHERE role = 'employee' AND id IN ({update_placeholders})
                """, (
                    preset["id"],
                    preset["schedule_days"],
                    preset["shift_start"],
                    preset["shift_end"],
                    preset["break_limit_minutes"],
                    *existing_ids
                ))
                for employee in selected_employees:
                    employee["schedule_preset_id"] = preset["id"]
                    employee["schedule_days"] = preset["schedule_days"]
                    employee["shift_start"] = preset["shift_start"]
                    employee["shift_end"] = preset["shift_end"]
                    employee["break_limit_minutes"] = preset["break_limit_minutes"]
                    record_employee_schedule_history(employee, actor_id=session["user_id"])
                db.commit()
            except Exception:
                db.rollback()
                raise

            invalidate_admin_employee_rows_cache()
            invalidate_reports_cache()
            log_activity(session["user_id"], "BULK APPLY SCHEDULE PRESET", f"Applied preset {preset['name']} to {len(existing_ids)} employee(s)")
            flash(f"Applied {preset['name']} to {len(existing_ids)} employee(s).", "success")
            return redirect(url_for("manage_employees", search=employee_search))

        full_name = request.form.get("full_name", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        department = request.form.get("department", "").strip() or "Stellar Seats"
        position = request.form.get("position", "").strip() or "Employee"
        emergency_contact_name = request.form.get("emergency_contact_name", "").strip()
        emergency_contact_phone = request.form.get("emergency_contact_phone", "").strip()
        id_issue_date, id_expiration_date, id_date_error = normalize_employee_id_dates(
            request.form.get("id_issue_date", ""),
            request.form.get("id_expiration_date", "")
        )
        barcode_id = request.form.get("barcode_id", "").strip()
        hourly_rate = parse_money_value(request.form.get("hourly_rate", "0"))
        sick_leave_days = parse_non_negative_int(request.form.get("sick_leave_days", DEFAULT_SICK_LEAVE_DAYS), DEFAULT_SICK_LEAVE_DAYS)
        paid_leave_days = parse_non_negative_int(request.form.get("paid_leave_days", DEFAULT_PAID_LEAVE_DAYS), DEFAULT_PAID_LEAVE_DAYS)
        sick_leave_used_manual = parse_non_negative_int(request.form.get("sick_leave_used_manual", "0"), 0)
        paid_leave_used_manual = parse_non_negative_int(request.form.get("paid_leave_used_manual", "0"), 0)
        schedule_assignment = resolve_schedule_assignment(request.form)
        schedule_days = schedule_assignment["schedule_days"]
        shift_start = schedule_assignment["shift_start"]
        shift_end = schedule_assignment["shift_end"]
        break_limit_minutes = schedule_assignment["break_limit_minutes"]
        schedule_preset_id = schedule_assignment["schedule_preset_id"]
        selected_preset = get_schedule_preset(schedule_preset_id) if schedule_preset_id else None

        if not full_name or not username or not password:
            flash("Full name, username, and password are required.", "danger")
            return redirect(url_for("manage_employees"))

        existing = fetchone("SELECT id FROM users WHERE username = ?", (username,))
        if existing:
            flash("Username already exists.", "warning")
            return redirect(url_for("manage_employees"))

        if id_date_error:
            flash(id_date_error, "danger")
            return redirect(url_for("manage_employees"))

        if selected_preset and not schedule_preset_matches_department(selected_preset, department):
            flash(f"Preset {selected_preset['name']} is scoped to {selected_preset['department_scope']}. Choose a matching department or use a custom schedule.", "danger")
            return redirect(url_for("manage_employees"))

        if barcode_id:
            existing_barcode = fetchone("SELECT id FROM users WHERE TRIM(COALESCE(barcode_id, '')) = ?", (barcode_id,))
            if existing_barcode:
                flash("Barcode ID already exists.", "warning")
                return redirect(url_for("manage_employees"))

        profile_image = None
        file = request.files.get("profile_image")
        if file and file.filename:
            profile_image = save_uploaded_file(file, prefix="profile", allowed_exts=IMAGE_EXTENSIONS)
            if not profile_image:
                flash("Invalid profile image file type.", "danger")
                return redirect(url_for("manage_employees"))

        execute_db("""
            INSERT INTO users (
                full_name, username, password_hash, role, profile_image,
                department, position, emergency_contact_name, emergency_contact_phone, id_issue_date, id_expiration_date, barcode_id, hourly_rate, sick_leave_days, paid_leave_days, sick_leave_used_manual, paid_leave_used_manual, schedule_days, shift_start, shift_end, schedule_preset_id, break_window_start, break_window_end, break_limit_minutes, is_active, created_at
            )
            VALUES (?, ?, ?, 'employee', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
        """, (
            full_name,
            username,
            generate_password_hash(password),
            profile_image,
            department,
            position,
            emergency_contact_name,
            emergency_contact_phone,
            id_issue_date,
            id_expiration_date,
            barcode_id,
            hourly_rate,
            sick_leave_days,
            paid_leave_days,
            sick_leave_used_manual,
            paid_leave_used_manual,
            schedule_days,
            shift_start,
            shift_end,
            schedule_preset_id,
            DEFAULT_BREAK_WINDOW_START,
            DEFAULT_BREAK_WINDOW_END,
            break_limit_minutes,
            now_str()
        ), commit=True)

        new_user = fetchone("SELECT id FROM users WHERE username = ?", (username,))
        if new_user:
            created_employee = get_user_by_id(new_user["id"])
            if created_employee:
                record_employee_schedule_history(
                    created_employee,
                    actor_id=session["user_id"],
                    effective_at=created_employee["created_at"] if created_employee["created_at"] else now_str(),
                    commit=True
                )
            create_notification(new_user["id"], "Account Created", "Your account has been created by admin.")
            log_activity(
                session["user_id"],
                "ADD EMPLOYEE",
                f"Added employee: {full_name} | " + summarize_employee_admin_changes(None, {
                    "department": department,
                    "position": position,
                    "emergency_contact_name": emergency_contact_name or "(not set)",
                    "emergency_contact_phone": emergency_contact_phone or "(not set)",
                    "id_issue_date": id_issue_date or "(auto)",
                    "id_expiration_date": id_expiration_date or "(auto)",
                    "barcode_id": barcode_id or "(not set)",
                    "hourly_rate": hourly_rate,
                    "sick_leave_days": sick_leave_days,
                    "paid_leave_days": paid_leave_days,
                    "sick_leave_used_manual": sick_leave_used_manual,
                        "paid_leave_used_manual": paid_leave_used_manual,
                        "shift_start": shift_start,
                        "shift_end": shift_end,
                        "schedule_preset_id": schedule_preset_id or "(custom)",
                        "break_limit_minutes": break_limit_minutes,
                        "is_active": 1,
                    })
                )

        invalidate_admin_employee_rows_cache()
        invalidate_reports_cache()
        flash("Employee added successfully.", "success")
        return redirect(url_for("manage_employees"))

    sql = """
        SELECT u.*, sp.name AS schedule_preset_name
        FROM users u
        LEFT JOIN schedule_presets sp ON sp.id = u.schedule_preset_id
        WHERE role = 'employee'
    """
    params = []

    if employee_search:
        sql += """
            AND (
                LOWER(full_name) LIKE ?
                OR LOWER(username) LIKE ?
                OR LOWER(COALESCE(department, '')) LIKE ?
                OR LOWER(COALESCE(position, '')) LIKE ?
            )
        """
        search_like = f"%{employee_search.lower()}%"
        params.extend([search_like, search_like, search_like, search_like])

    sql += " ORDER BY u.id DESC"
    employees = fetchall(sql, params)
    schedule_presets = []
    preset_assignment_counts = {}
    for row in fetchall("""
        SELECT schedule_preset_id, COUNT(*) AS cnt
        FROM users
        WHERE role = 'employee' AND schedule_preset_id IS NOT NULL
        GROUP BY schedule_preset_id
    """):
        preset_assignment_counts[int(row["schedule_preset_id"])] = int(row["cnt"] or 0)
    for preset in get_schedule_presets():
        item = dict(preset)
        item["schedule_summary"] = get_schedule_summary(item["schedule_days"] or DEFAULT_SCHEDULE_DAYS)
        item["window_summary"] = f"{item['shift_start'] or DEFAULT_SHIFT_START} - {item['shift_end'] or DEFAULT_SHIFT_END}"
        item["assigned_count"] = preset_assignment_counts.get(int(item["id"]), 0)
        schedule_presets.append(item)
    future_schedule_changes = get_future_schedule_changes(limit=18)
    employee_future_change_map = build_future_schedule_change_map([int(emp["id"]) for emp in employees])
    future_schedule_summary = {
        "queued_total": len(future_schedule_changes),
        "employee_count": len({int(change["user_id"]) for change in future_schedule_changes}),
        "next_effective_date": future_schedule_changes[0]["effective_date"] if future_schedule_changes else "",
    }
    schedule_special_rules = get_schedule_special_dates(limit=18)
    schedule_special_rule_summary = {
        "total": len(schedule_special_rules),
        "holiday_count": len([rule for rule in schedule_special_rules if rule["rule_type"] == "holiday"]),
        "rest_day_count": len([rule for rule in schedule_special_rules if rule["rule_type"] == "rest_day"]),
        "next_date": schedule_special_rules[0]["special_date"] if schedule_special_rules else "",
    }

    return render_template(
        "manage_employees.html",
        employees=employees,
        weekday_options=WEEKDAY_OPTIONS,
        employee_search=employee_search,
        schedule_presets=schedule_presets,
        future_schedule_changes=future_schedule_changes,
        future_schedule_summary=future_schedule_summary,
        employee_future_change_map=employee_future_change_map,
        schedule_special_rules=schedule_special_rules,
        schedule_special_rule_summary=schedule_special_rule_summary,
        schedule_special_rule_options=SCHEDULE_SPECIAL_RULE_OPTIONS,
        today_date=today_str(),
        tomorrow_date=(now_dt().date() + timedelta(days=1)).strftime("%Y-%m-%d")
    )


@app.route("/admin/future-schedule-changes/<int:change_id>/delete", methods=["POST"])
@login_required(role="admin")
def delete_future_schedule_change(change_id):
    change = fetchone("""
        SELECT fsc.*, u.full_name
        FROM employee_future_schedule_changes fsc
        LEFT JOIN users u ON u.id = fsc.user_id
        WHERE fsc.id = ?
    """, (change_id,))
    if not change:
        flash("Queued schedule change not found.", "warning")
        return redirect(request.referrer or url_for("manage_employees"))
    change = dict(change)
    if change.get("applied_at"):
        flash("This schedule change has already been applied and can no longer be deleted.", "danger")
        return redirect(request.referrer or url_for("manage_employees"))

    execute_db("DELETE FROM employee_future_schedule_changes WHERE id = ?", (change_id,), commit=True)
    invalidate_schedule_change_apply_state()
    invalidate_reports_cache()
    employee_label = change.get("full_name") or f"User {change['user_id']}"
    log_activity(
        session["user_id"],
        "DELETE FUTURE SCHEDULE CHANGE",
        f"Removed queued schedule change for {employee_label} effective {change['effective_date']}."
    )
    flash("Queued schedule change deleted.", "info")
    return redirect(request.referrer or url_for("manage_employees"))


@app.route("/admin/edit-employee/<int:user_id>", methods=["GET", "POST"])
@login_required(role="admin")
def edit_employee(user_id):
    user = fetchone("""
        SELECT * FROM users
        WHERE id = ? AND role = 'employee'
    """, (user_id,))

    if not user:
        flash("Employee not found.", "danger")
        return redirect(url_for("manage_employees"))

    if request.method == "POST":
        form_action = (request.form.get("form_action", "edit_employee") or "edit_employee").strip()
        if form_action == "schedule_future_change":
            effective_date = (request.form.get("effective_date", "") or "").strip()
            change_notes = (request.form.get("future_schedule_notes", "") or "").strip()
            schedule_assignment = resolve_schedule_assignment(request.form, fallback_user=user)
            selected_preset = get_schedule_preset(schedule_assignment["schedule_preset_id"]) if schedule_assignment["schedule_preset_id"] else None
            if selected_preset and not schedule_preset_matches_department(selected_preset, user["department"]):
                flash(f"Preset {selected_preset['name']} is scoped to {selected_preset['department_scope']}. Choose a matching department or use a custom schedule.", "danger")
                return redirect(url_for("edit_employee", user_id=user_id))
            try:
                queue_future_schedule_change(
                    user,
                    schedule_assignment,
                    effective_date,
                    actor_id=session["user_id"],
                    notes=change_notes
                )
                get_db().commit()
            except ValueError as exc:
                flash(str(exc), "danger")
                return redirect(url_for("edit_employee", user_id=user_id))

            invalidate_reports_cache()
            log_activity(
                session["user_id"],
                "QUEUE EMPLOYEE SCHEDULE CHANGE",
                f"Queued a schedule change for {user['full_name']} effective {effective_date}."
            )
            flash(f"Future schedule change queued for {user['full_name']} starting {effective_date}.", "success")
            return redirect(url_for("edit_employee", user_id=user_id))

        original_user = user
        full_name = request.form.get("full_name", "").strip()
        username = request.form.get("username", "").strip()
        department = request.form.get("department", "").strip() or "Stellar Seats"
        position = request.form.get("position", "").strip() or "Employee"
        emergency_contact_name = request.form.get("emergency_contact_name", "").strip()
        emergency_contact_phone = request.form.get("emergency_contact_phone", "").strip()
        id_issue_date, id_expiration_date, id_date_error = normalize_employee_id_dates(
            request.form.get("id_issue_date", ""),
            request.form.get("id_expiration_date", "")
        )
        barcode_id = request.form.get("barcode_id", "").strip()
        hourly_rate = parse_money_value(request.form.get("hourly_rate", user["hourly_rate"] or 0))
        sick_leave_days = parse_non_negative_int(request.form.get("sick_leave_days", user["sick_leave_days"] if user["sick_leave_days"] is not None else DEFAULT_SICK_LEAVE_DAYS), DEFAULT_SICK_LEAVE_DAYS)
        paid_leave_days = parse_non_negative_int(request.form.get("paid_leave_days", user["paid_leave_days"] if user["paid_leave_days"] is not None else DEFAULT_PAID_LEAVE_DAYS), DEFAULT_PAID_LEAVE_DAYS)
        sick_leave_used_manual = parse_non_negative_int(request.form.get("sick_leave_used_manual", user["sick_leave_used_manual"] if user["sick_leave_used_manual"] is not None else 0), 0)
        paid_leave_used_manual = parse_non_negative_int(request.form.get("paid_leave_used_manual", user["paid_leave_used_manual"] if user["paid_leave_used_manual"] is not None else 0), 0)
        schedule_assignment = resolve_schedule_assignment(request.form, fallback_user=user)
        schedule_days = schedule_assignment["schedule_days"]
        shift_start = schedule_assignment["shift_start"]
        shift_end = schedule_assignment["shift_end"]
        break_limit_minutes = schedule_assignment["break_limit_minutes"]
        schedule_preset_id = schedule_assignment["schedule_preset_id"]
        selected_preset = get_schedule_preset(schedule_preset_id) if schedule_preset_id else None
        is_active = 1 if request.form.get("is_active") == "1" else 0
        password = request.form.get("password", "").strip()

        if not full_name or not username:
            flash("Full name and username are required.", "danger")
            return redirect(url_for("edit_employee", user_id=user_id))

        existing = fetchone("""
            SELECT id FROM users
            WHERE username = ? AND id != ?
        """, (username, user_id))

        if existing:
            flash("Username already used by another employee.", "warning")
            return redirect(url_for("edit_employee", user_id=user_id))

        if id_date_error:
            flash(id_date_error, "danger")
            return redirect(url_for("edit_employee", user_id=user_id))

        if selected_preset and not schedule_preset_matches_department(selected_preset, department):
            flash(f"Preset {selected_preset['name']} is scoped to {selected_preset['department_scope']}. Choose a matching department or use a custom schedule.", "danger")
            return redirect(url_for("edit_employee", user_id=user_id))

        if barcode_id:
            existing_barcode = fetchone("""
                SELECT id FROM users
                WHERE TRIM(COALESCE(barcode_id, '')) = ? AND id != ?
            """, (barcode_id, user_id))
            if existing_barcode:
                flash("Barcode ID already used by another employee.", "warning")
                return redirect(url_for("edit_employee", user_id=user_id))

        profile_image = user["profile_image"]
        file = request.files.get("profile_image")
        if file and file.filename:
            saved = save_uploaded_file(file, prefix=f"profile_{user_id}", allowed_exts=IMAGE_EXTENSIONS)
            if not saved:
                flash("Invalid profile image file type.", "danger")
                return redirect(url_for("edit_employee", user_id=user_id))
            profile_image = saved

        if password:
            execute_db("""
                UPDATE users
                SET full_name = ?, username = ?, password_hash = ?, profile_image = ?,
                    department = ?, position = ?, emergency_contact_name = ?, emergency_contact_phone = ?, id_issue_date = ?, id_expiration_date = ?, barcode_id = ?, hourly_rate = ?, sick_leave_days = ?, paid_leave_days = ?, sick_leave_used_manual = ?, paid_leave_used_manual = ?, schedule_days = ?, shift_start = ?, shift_end = ?, schedule_preset_id = ?, break_window_start = ?, break_window_end = ?, break_limit_minutes = ?, is_active = ?
                WHERE id = ?
            """, (
                full_name, username, generate_password_hash(password), profile_image,
                department, position, emergency_contact_name, emergency_contact_phone, id_issue_date, id_expiration_date, barcode_id, hourly_rate, sick_leave_days, paid_leave_days, sick_leave_used_manual, paid_leave_used_manual, schedule_days, shift_start, shift_end, schedule_preset_id, user["break_window_start"] or DEFAULT_BREAK_WINDOW_START, user["break_window_end"] or DEFAULT_BREAK_WINDOW_END, break_limit_minutes, is_active, user_id
            ), commit=True)
        else:
            execute_db("""
                UPDATE users
                SET full_name = ?, username = ?, profile_image = ?,
                    department = ?, position = ?, emergency_contact_name = ?, emergency_contact_phone = ?, id_issue_date = ?, id_expiration_date = ?, barcode_id = ?, hourly_rate = ?, sick_leave_days = ?, paid_leave_days = ?, sick_leave_used_manual = ?, paid_leave_used_manual = ?, schedule_days = ?, shift_start = ?, shift_end = ?, schedule_preset_id = ?, break_window_start = ?, break_window_end = ?, break_limit_minutes = ?, is_active = ?
                WHERE id = ?
            """, (
                full_name, username, profile_image,
                department, position, emergency_contact_name, emergency_contact_phone, id_issue_date, id_expiration_date, barcode_id, hourly_rate, sick_leave_days, paid_leave_days, sick_leave_used_manual, paid_leave_used_manual, schedule_days, shift_start, shift_end, schedule_preset_id, user["break_window_start"] or DEFAULT_BREAK_WINDOW_START, user["break_window_end"] or DEFAULT_BREAK_WINDOW_END, break_limit_minutes, is_active, user_id
            ), commit=True)

        log_activity(
            session["user_id"],
            "EDIT EMPLOYEE",
            f"Edited employee: {full_name} | " + summarize_employee_admin_changes(original_user, {
                "department": department,
                "position": position,
                "emergency_contact_name": emergency_contact_name or "(not set)",
                "emergency_contact_phone": emergency_contact_phone or "(not set)",
                "id_issue_date": id_issue_date or "(auto)",
                "id_expiration_date": id_expiration_date or "(auto)",
                "barcode_id": barcode_id or "(not set)",
                "hourly_rate": hourly_rate,
                "sick_leave_days": sick_leave_days,
                "paid_leave_days": paid_leave_days,
                "sick_leave_used_manual": sick_leave_used_manual,
                "paid_leave_used_manual": paid_leave_used_manual,
                "shift_start": shift_start,
                "shift_end": shift_end,
                "schedule_preset_id": schedule_preset_id or "(custom)",
                "break_limit_minutes": break_limit_minutes,
                "is_active": is_active,
            })
        )
        updated_user = get_user_by_id(user_id)
        if updated_user:
            record_employee_schedule_history(updated_user, actor_id=session["user_id"], commit=True)
        invalidate_admin_employee_rows_cache()
        invalidate_reports_cache()
        flash("Employee updated successfully.", "success")
        return redirect(url_for("manage_employees"))

    return render_template(
        "edit_employee.html",
        employee=user,
        weekday_options=WEEKDAY_OPTIONS,
        employee_schedule_days=get_schedule_day_codes(user["schedule_days"] if user["schedule_days"] else DEFAULT_SCHEDULE_DAYS),
        schedule_presets=get_schedule_presets(),
        schedule_history_rows=get_recent_employee_schedule_history(user_id, limit=8),
        future_schedule_changes=get_future_schedule_changes(user_id=user_id, limit=8),
        tomorrow_date=(now_dt().date() + timedelta(days=1)).strftime("%Y-%m-%d")
    )


@app.route("/admin/employee-id/<int:user_id>")
@login_required(role="admin")
def print_employee_id(user_id):
    employee = fetchone("""
        SELECT *
        FROM users
        WHERE id = ? AND role = 'employee'
    """, (user_id,))

    if not employee:
        flash("Employee not found.", "danger")
        return redirect(url_for("manage_employees"))
    employee = dict(employee)

    card_number = get_employee_card_number(employee)
    barcode_value = (employee["barcode_id"] or card_number).strip()
    company_settings = get_company_settings()
    if employee["id_issue_date"]:
        issue_date_value = employee["id_issue_date"]
    else:
        try:
            issue_dt = datetime.strptime((employee["created_at"] or now_str())[:19], "%Y-%m-%d %H:%M:%S").date()
        except Exception:
            issue_dt = now_dt().date()
        issue_date_value = issue_dt.strftime("%Y-%m-%d")

    if employee["id_expiration_date"]:
        expiration_date_value = employee["id_expiration_date"]
    else:
        try:
            base_issue_dt = datetime.strptime(issue_date_value, "%Y-%m-%d").date()
        except Exception:
            base_issue_dt = now_dt().date()
        expiration_date_value = (base_issue_dt + timedelta(days=365)).strftime("%Y-%m-%d")

    try:
        formatted_issue_date = datetime.strptime(issue_date_value, "%Y-%m-%d").strftime("%m/%d/%Y")
    except Exception:
        formatted_issue_date = issue_date_value
    try:
        formatted_expiration_date = datetime.strptime(expiration_date_value, "%Y-%m-%d").strftime("%m/%d/%Y")
    except Exception:
        formatted_expiration_date = expiration_date_value
    return render_template(
        "admin_employee_id_card.html",
        employee=employee,
        card_number=card_number,
        barcode_value=barcode_value,
        company_settings=company_settings,
        issue_date=formatted_issue_date,
        expiration_date=formatted_expiration_date
    )


@app.route("/admin/employee-id/<int:user_id>/barcode")
@login_required(role="admin")
def download_employee_barcode(user_id):
    employee = fetchone("""
        SELECT *
        FROM users
        WHERE id = ? AND role = 'employee'
    """, (user_id,))

    if not employee:
        flash("Employee not found.", "danger")
        return redirect(url_for("manage_employees"))
    employee = dict(employee)

    card_number = get_employee_card_number(employee)
    barcode_value = (employee["barcode_id"] or card_number).strip()
    svg_markup = generate_code128_svg_markup(barcode_value)
    if not svg_markup:
        flash("Barcode is not available for this employee yet.", "warning")
        return redirect(url_for("print_employee_id", user_id=user_id))

    safe_name = secure_filename(employee.get("full_name") or f"employee-{user_id}") or f"employee-{user_id}"
    return Response(
        svg_markup,
        mimetype="image/svg+xml",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}-barcode.svg"'
        },
    )


@app.route("/scanner")
@login_required(role="scanner")
def scanner_kiosk():
    return render_template("scanner.html")


@app.route("/scanner/unlock", methods=["POST"])
@login_required(role="scanner")
def scanner_kiosk_unlock():
    pin_value = (request.form.get("pin", "") or "").strip()
    if not has_scanner_exit_pin():
        return jsonify({"ok": True, "message": "Scanner unlocked."})
    if verify_scanner_exit_pin(pin_value):
        return jsonify({"ok": True, "message": "Scanner unlocked."})
    return jsonify({"ok": False, "message": "Incorrect kiosk PIN."}), 403


@app.route("/admin/scanner-logs")
@login_required(role="admin")
def admin_scanner_logs():
    date_from = (request.args.get("date_from", "") or "").strip()
    date_to = (request.args.get("date_to", "") or "").strip()
    action_type = (request.args.get("action_type", "") or "").strip()
    result_status = (request.args.get("result_status", "") or "").strip()
    employee_id = (request.args.get("employee_id", "") or "").strip()

    if date_from and date_to:
        try:
            start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
            end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
            if start_date > end_date:
                date_from, date_to = date_to, date_from
        except ValueError:
            pass

    where_clauses = []
    params = []

    if date_from:
        where_clauses.append("sl.created_at >= ?")
        params.append(f"{date_from} 00:00:00")
    if date_to:
        where_clauses.append("sl.created_at <= ?")
        params.append(f"{date_to} 23:59:59")
    if action_type:
        where_clauses.append("sl.action_type = ?")
        params.append(action_type)
    if result_status:
        where_clauses.append("sl.result_status = ?")
        params.append(result_status)
    if employee_id.isdigit():
        where_clauses.append("sl.employee_user_id = ?")
        params.append(int(employee_id))

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
    scanner_exprs = get_scanner_log_select_expressions()

    rows = fetchall(f"""
        SELECT
            sl.*,
            {scanner_exprs['scanner_name']} AS scanner_name,
            {scanner_exprs['scanner_username']} AS scanner_username,
            {scanner_exprs['employee_name']} AS employee_name,
            {scanner_exprs['employee_department']} AS employee_department,
            {scanner_exprs['employee_position']} AS employee_position
        FROM scanner_logs sl
        LEFT JOIN users scanner ON scanner.id = sl.scanner_user_id
        LEFT JOIN users employee ON employee.id = sl.employee_user_id
        {where_sql}
        ORDER BY sl.created_at DESC, sl.id DESC
        LIMIT 300
    """, tuple(params))

    stats = fetchone(f"""
        SELECT
            COUNT(*) AS total_scans,
            SUM(CASE WHEN sl.result_status = 'success' THEN 1 ELSE 0 END) AS success_count,
            SUM(CASE WHEN sl.result_status = 'error' THEN 1 ELSE 0 END) AS error_count,
            SUM(CASE WHEN substr(sl.created_at, 1, 10) = ? THEN 1 ELSE 0 END) AS today_count
        FROM scanner_logs sl
        {where_sql}
    """, tuple([today_str(), *params])) or {}

    employees = fetchall("""
        SELECT id, full_name, department
        FROM users
        WHERE role = 'employee'
        ORDER BY full_name
    """)

    return render_template(
        "admin_scanner_logs.html",
        scanner_logs=rows,
        stats=stats,
        date_from=date_from,
        date_to=date_to,
        action_type=action_type,
        result_status=result_status,
        employee_id=employee_id,
        employees=employees
    )


@app.route("/admin/attendance-audit")
@login_required(role="admin")
def admin_attendance_audit():
    date_from = (request.args.get("date_from", "") or "").strip()
    date_to = (request.args.get("date_to", "") or "").strip()
    employee_id = (request.args.get("employee_id", "") or "").strip()
    source_filter = (request.args.get("source", "") or "").strip()

    if date_from and date_to:
        try:
            start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
            end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
            if start_date > end_date:
                date_from, date_to = date_to, date_from
        except ValueError:
            pass

    rows = build_attendance_audit_rows(
        date_from=date_from,
        date_to=date_to,
        employee_id=employee_id,
        source_filter=source_filter
    )
    employees = fetchall("""
        SELECT id, full_name, department
        FROM users
        WHERE role = 'employee'
        ORDER BY full_name
    """)
    return render_template(
        "admin_attendance_audit.html",
        audit_rows=rows,
        date_from=date_from,
        date_to=date_to,
        employee_id=employee_id,
        source_filter=source_filter,
        employees=employees
    )


@app.route("/scanner/scan", methods=["POST"])
@login_required(role="scanner")
def scanner_kiosk_scan():
    action_type = (request.form.get("action_type", "") or "").strip()
    barcode_value = (request.form.get("barcode_value", "") or "").strip()
    scanner_user_id = session.get("user_id")
    scanner_user = get_user_by_id(scanner_user_id)
    source_label = "Tablet kiosk"
    device_label = "Tablet camera kiosk"
    ip_address = get_client_ip()
    user_agent = request.headers.get("User-Agent", "")
    scanner_log_kwargs = {
        "source_label": source_label,
        "device_label": device_label,
        "ip_address": ip_address,
        "user_agent": user_agent,
        "scanner_name_snapshot": row_get(scanner_user, "full_name"),
        "scanner_username_snapshot": row_get(scanner_user, "username"),
    }

    if action_type not in {"time_in", "start_break", "end_break", "time_out", "overtime_start", "overtime_end"}:
        log_scanner_activity(
            scanner_user_id, action_type, barcode_value, "error",
            "Please choose a valid attendance action.",
            **scanner_log_kwargs
        )
        return jsonify({"ok": False, "message": "Please choose a valid attendance action."}), 400

    if not barcode_value:
        log_scanner_activity(
            scanner_user_id, action_type, barcode_value, "error",
            "Scan or enter a barcode first.",
            **scanner_log_kwargs
        )
        return jsonify({"ok": False, "message": "Scan or enter a barcode first."}), 400

    barcode_lookup = find_employee_barcode_matches(barcode_value)
    if barcode_lookup["is_duplicate"]:
        log_scanner_activity(
            scanner_user_id, action_type, barcode_value, "error",
            "This barcode is assigned to multiple employees. Fix the duplicate Barcode ID records first.",
            **scanner_log_kwargs
        )
        return jsonify({
            "ok": False,
            "message": "This barcode is assigned to multiple employees. Fix the duplicate Barcode ID records first."
        }), 409

    employee = barcode_lookup["matches"][0] if barcode_lookup["matches"] else None
    if not employee:
        log_scanner_activity(
            scanner_user_id, action_type, barcode_value, "error",
            "No employee matched that barcode. Check the employee Barcode ID first.",
            **scanner_log_kwargs
        )
        return jsonify({"ok": False, "message": "No employee matched that barcode. Check the employee Barcode ID first."}), 404

    ok, message, employee_row = perform_attendance_action(
        employee["id"],
        action_type,
        actor_id=scanner_user_id,
        source_label=source_label
    )
    employee_for_payload = employee_row or employee
    attendance = get_current_attendance(employee["id"])
    overtime_session = get_open_overtime_session(employee["id"])
    break_minutes = total_break_minutes(attendance["id"], include_open=True) if attendance else 0
    log_scanner_activity(
        scanner_user_id,
        action_type,
        barcode_value,
        "success" if ok else "error",
        message,
        employee_user_id=employee["id"],
        employee_name_snapshot=row_get(employee_for_payload, "full_name"),
        employee_department_snapshot=row_get(employee_for_payload, "department"),
        employee_position_snapshot=row_get(employee_for_payload, "position"),
        **scanner_log_kwargs
    )

    return jsonify({
        "ok": ok,
        "message": message,
        "employee_name": employee_for_payload["full_name"],
        "department": employee_for_payload["department"] or "",
        "position": employee_for_payload["position"] or "",
        "avatar_initials": get_avatar_initials(employee_for_payload["full_name"]),
        "profile_image_url": url_for("uploaded_file", filename=employee_for_payload["profile_image"]) if employee_for_payload["profile_image"] and uploaded_file_exists(employee_for_payload["profile_image"]) else None,
        "barcode_value": barcode_value,
        "status": "On Overtime" if overtime_session else (attendance["status"] if attendance else "Offline"),
        "time_in": attendance["time_in"] if attendance else None,
        "time_out": attendance["time_out"] if attendance else None,
        "break_minutes": break_minutes,
        "action_type": action_type
    }), (200 if ok else 400)


@app.route("/admin/employee-id/signatory", methods=["POST"])
@login_required(role="admin")
def update_employee_id_signatory():
    signatory_name = request.form.get("id_signatory_name", "").strip() or "Kirk Danny Fernandez"
    signatory_title = request.form.get("id_signatory_title", "").strip() or "Head Of Operations"
    hr_signatory_name = request.form.get("hr_signatory_name", "").strip()
    hr_signatory_title = request.form.get("hr_signatory_title", "").strip() or "Human Resources Manager"
    current_settings = get_company_settings()
    signature_file = current_settings.get("id_signature_file") if current_settings else None
    hr_signature_file = current_settings.get("hr_signature_file") if current_settings else None

    file = request.files.get("id_signature_file")
    if file and file.filename:
        saved = save_uploaded_file(file, prefix="id_signature", allowed_exts=IMAGE_EXTENSIONS)
        if not saved:
            flash("Invalid signature image file type.", "danger")
            employee_id = request.form.get("employee_id", "").strip()
            if employee_id:
                return redirect(url_for("print_employee_id", user_id=employee_id))
            return redirect(url_for("manage_employees"))
        signature_file = saved

    hr_file = request.files.get("hr_signature_file")
    if hr_file and hr_file.filename:
        saved_hr = save_uploaded_file(hr_file, prefix="hr_signature", allowed_exts=IMAGE_EXTENSIONS)
        if not saved_hr:
            flash("Invalid Human Resources signature image file type.", "danger")
            employee_id = request.form.get("employee_id", "").strip()
            if employee_id:
                return redirect(url_for("print_employee_id", user_id=employee_id))
            return redirect(url_for("manage_employees"))
        hr_signature_file = saved_hr

    execute_db("""
        INSERT INTO company_settings (
            id, id_signatory_name, id_signatory_title, id_signature_file,
            hr_signatory_name, hr_signatory_title, hr_signature_file
        )
        VALUES (1, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            id_signatory_name = excluded.id_signatory_name,
            id_signatory_title = excluded.id_signatory_title,
            id_signature_file = excluded.id_signature_file,
            hr_signatory_name = excluded.hr_signatory_name,
            hr_signatory_title = excluded.hr_signatory_title,
            hr_signature_file = excluded.hr_signature_file
    """, (
        signatory_name,
        signatory_title,
        signature_file,
        hr_signatory_name,
        hr_signatory_title,
        hr_signature_file,
    ), commit=True)

    hr_log_name = hr_signatory_name or "Human Resources Manager"
    log_activity(
        session["user_id"],
        "UPDATE ID SIGNATORY",
        f"Updated ID signatories to {signatory_name} | {signatory_title} and {hr_log_name} | {hr_signatory_title}"
    )
    flash("ID signatory details updated.", "success")
    employee_id = request.form.get("employee_id", "").strip()
    if employee_id:
        return redirect(url_for("print_employee_id", user_id=employee_id))
    return redirect(url_for("manage_employees"))


@app.route("/admin/delete-employee/<int:user_id>", methods=["POST"])
@login_required(role="admin")
def delete_employee(user_id):
    user = fetchone("""
        SELECT * FROM users
        WHERE id = ? AND role = 'employee'
    """, (user_id,))

    if not user:
        flash("Employee not found.", "danger")
        return redirect(url_for("manage_employees"))

    upload_filenames = []
    if user["profile_image"]:
        upload_filenames.append(user["profile_image"])
    upload_filenames.extend([
        row["proof_file"]
        for row in fetchall("""
            SELECT proof_file
            FROM attendance
            WHERE user_id = ? AND proof_file IS NOT NULL AND TRIM(proof_file) != ''
        """, (user_id,))
    ])

    def count_rows(query, params):
        row = fetchone(query, params)
        if not row:
            return 0
        return int(row["cnt"] or 0)

    protected_history_counts = {
        "attendance": count_rows("SELECT COUNT(*) AS cnt FROM attendance WHERE user_id = ?", (user_id,)),
        "scanner_logs": count_rows("SELECT COUNT(*) AS cnt FROM scanner_logs WHERE employee_user_id = ? OR scanner_user_id = ?", (user_id, user_id)),
        "payroll_items": count_rows("SELECT COUNT(*) AS cnt FROM payroll_run_items WHERE user_id = ?", (user_id,)),
        "overtime": count_rows("SELECT COUNT(*) AS cnt FROM overtime_sessions WHERE user_id = ?", (user_id,)),
        "corrections": count_rows("SELECT COUNT(*) AS cnt FROM correction_requests WHERE user_id = ?", (user_id,)),
        "incidents": count_rows("SELECT COUNT(*) AS cnt FROM incident_reports WHERE user_id = ?", (user_id,)),
        "disciplinary": count_rows("SELECT COUNT(*) AS cnt FROM disciplinary_actions WHERE user_id = ?", (user_id,)),
    }
    if any(protected_history_counts.values()):
        flash("This employee has historical attendance, payroll, or audit records. Deactivate the account instead of deleting it.", "warning")
        return redirect(url_for("manage_employees"))

    db = get_db()
    try:
        execute_db("DELETE FROM notifications WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM employee_future_schedule_changes WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM employee_schedule_history WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM breaks WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM attendance WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM overtime_sessions WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM activity_logs WHERE user_id = ? OR target_user_id = ?", (user_id, user_id))
        execute_db("DELETE FROM correction_requests WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM scanner_logs WHERE employee_user_id = ? OR scanner_user_id = ?", (user_id, user_id))
        execute_db("DELETE FROM payroll_adjustments WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM payroll_recurring_rules WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM payroll_run_item_adjustments WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM incident_reports WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM disciplinary_actions WHERE user_id = ?", (user_id,))
        execute_db("DELETE FROM users WHERE id = ?", (user_id,))
        db.commit()
    except Exception:
        db.rollback()
        raise

    for filename in upload_filenames:
        delete_uploaded_file_if_unused(filename)

    log_activity(session["user_id"], "DELETE EMPLOYEE", f"Deleted employee: {user['full_name']}")
    invalidate_admin_employee_rows_cache()
    flash("Employee deleted successfully.", "info")
    return redirect(url_for("manage_employees"))


@app.route("/admin/send-notification", methods=["POST"])
@login_required(role="admin")
def send_admin_notification():
    user_id = request.form.get("user_id")
    title = request.form.get("title", "").strip()
    message = request.form.get("message", "").strip()

    if not user_id or not title or not message:
        flash("All notification fields are required.", "danger")
        return redirect(url_for("admin_dashboard"))

    create_notification(user_id, title, message)
    log_activity(session["user_id"], "SEND NOTIFICATION", f"Sent notification to user ID {user_id}")
    flash("Notification sent successfully.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/create-incident", methods=["POST"])
@login_required(role="admin")
def create_incident_route():
    user_id = request.form.get("user_id")
    error_type = normalize_incident_error_type(
        request.form.get("error_type", ""),
        request.form.get("new_error_type", "")
    )
    report_date = request.form.get("report_date", "").strip()
    message = request.form.get("message", "").strip()

    if not user_id or not error_type or not report_date:
        flash("All fields are required.", "danger")
        return redirect(url_for("admin_incident_report"))

    employee = get_user_by_id(user_id)

    incident = create_incident(
        user_id=user_id,
        error_type=error_type,
        report_date=report_date,
        message=message,
        admin_id=session["user_id"],
        incident_action="",
        report_department=employee["department"] if employee else ""
    )
    sync_result = sync_incident_policy(incident["id"], session["user_id"]) if incident else {}

    employee_name = employee["full_name"] if employee else f"User {user_id}"
    log_activity(
        session["user_id"],
        "CREATE INCIDENT",
        f"{error_type} report created for {employee_name}"
    )

    if sync_result.get("created_action"):
        flash(
            f"Incident report created. Policy step #{sync_result['incident_count']} created a linked {sync_result['policy_action']} record.",
            "success"
        )
    elif sync_result.get("policy_action") and sync_result.get("message"):
        flash(
            f"Incident report created. Policy recommends {sync_result['policy_action']}, but no disciplinary record was auto-created: {sync_result['message']}",
            "warning"
        )
    else:
        flash("Incident report created.", "success")
    return redirect(url_for("admin_incident_report"))


@app.route("/admin/disciplinary/create", methods=["POST"])
@login_required(role="admin")
def create_disciplinary_action_route():
    user_id = request.form.get("user_id", "").strip()
    action_type = request.form.get("action_type", "").strip()
    action_date = request.form.get("action_date", "").strip()
    duration_days = parse_non_negative_int(request.form.get("duration_days", "1"), 1)
    details = request.form.get("details", "").strip()

    if not user_id or not action_type or not action_date:
        flash("Employee, action type, and action date are required.", "danger")
        return redirect(url_for("admin_disciplinary_dashboard"))

    if action_type not in DISCIPLINARY_ACTION_TYPES:
        flash("Invalid disciplinary action type.", "danger")
        return redirect(url_for("admin_disciplinary_dashboard"))

    if action_type == "Suspension" and duration_days <= 0:
        flash("Suspension days must be at least 1.", "danger")
        return redirect(url_for("admin_disciplinary_dashboard"))

    if action_type != "Suspension":
        duration_days = 1

    employee = get_user_by_id(user_id)
    conflict = find_conflicting_disciplinary_action(user_id, action_type, action_date, duration_days)
    if conflict:
        flash(conflict["conflict_reason"], "warning")
        return redirect(url_for("admin_disciplinary_dashboard"))
    create_disciplinary_action(
        user_id=user_id,
        action_type=action_type,
        action_date=action_date,
        details=details,
        created_by=session["user_id"],
        duration_days=duration_days
    )

    employee_name = employee["full_name"] if employee else f"User {user_id}"
    log_activity(
        session["user_id"],
        "CREATE DISCIPLINARY ACTION",
        f"{action_type} created for {employee_name}" + (f" for {duration_days} day(s)" if action_type == "Suspension" else "")
    )
    flash(f"{action_type} record created.", "success")
    return redirect(url_for("admin_disciplinary_dashboard"))


@app.route("/admin/disciplinary/<int:action_id>/update", methods=["POST"])
@login_required(role="admin")
def update_disciplinary_action_route(action_id):
    action = get_disciplinary_action_by_id(action_id)
    if not action:
        flash("Disciplinary record not found.", "danger")
        return redirect(url_for("admin_disciplinary_dashboard"))

    action_type = request.form.get("action_type", "").strip()
    action_date = request.form.get("action_date", "").strip()
    duration_days = parse_non_negative_int(request.form.get("duration_days", "1"), 1)
    details = request.form.get("details", "").strip()

    if not action_type or not action_date:
        flash("Action type and action date are required.", "danger")
        return redirect(url_for("admin_disciplinary_dashboard"))
    if action_type not in DISCIPLINARY_ACTION_TYPES:
        flash("Invalid disciplinary action type.", "danger")
        return redirect(url_for("admin_disciplinary_dashboard"))

    if action_type != "Suspension":
        duration_days = 1
    conflict = find_conflicting_disciplinary_action(action["user_id"], action_type, action_date, duration_days, exclude_id=action_id)
    if conflict:
        flash(conflict["conflict_reason"], "warning")
        return redirect(url_for("admin_disciplinary_dashboard"))
    end_date = calculate_suspension_end_date(action_date, duration_days) if action_type == "Suspension" else action_date

    execute_db("""
        UPDATE disciplinary_actions
        SET action_type = ?, action_date = ?, duration_days = ?, end_date = ?, details = ?
        WHERE id = ?
    """, (action_type, action_date, duration_days, end_date, details, action_id), commit=True)

    employee = get_user_by_id(action["user_id"])
    employee_name = employee["full_name"] if employee else f"User {action['user_id']}"
    log_activity(session["user_id"], "UPDATE DISCIPLINARY ACTION", f"Updated {action_type} for {employee_name}")
    flash("Disciplinary record updated.", "success")
    return redirect(url_for("admin_disciplinary_dashboard"))


@app.route("/admin/disciplinary/<int:action_id>/delete", methods=["POST"])
@login_required(role="admin")
def delete_disciplinary_action_route(action_id):
    action = get_disciplinary_action_by_id(action_id)
    if not action:
        flash("Disciplinary record not found.", "danger")
        return redirect(url_for("admin_disciplinary_dashboard"))

    employee = get_user_by_id(action["user_id"])
    employee_name = employee["full_name"] if employee else f"User {action['user_id']}"
    execute_db("DELETE FROM disciplinary_actions WHERE id = ?", (action_id,), commit=True)
    log_activity(session["user_id"], "DELETE DISCIPLINARY ACTION", f"Deleted {action['action_type']} for {employee_name}")
    flash("Disciplinary record deleted.", "info")
    return redirect(url_for("admin_disciplinary_dashboard"))


# =========================
# STARTUP
# =========================
with app.app_context():
    init_db()


# =========================
# MAIN
# =========================
if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=int(os.environ.get("PORT", "5000")),
        debug=os.environ.get("FLASK_DEBUG", "").strip() == "1"
    )
