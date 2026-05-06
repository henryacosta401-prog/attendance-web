from io import BytesIO

from flask import Response


_deps = {}


def configure_recovery_pack_services(deps):
    _deps.update(deps)


def _dep(name):
    return _deps[name]


def workbook_to_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def autosize_workbook_sheet(sheet):
    width_map = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width_map[cell.column_letter] = max(width_map.get(cell.column_letter, 0), len(str(cell.value)))
    for column_letter, width in width_map.items():
        sheet.column_dimensions[column_letter].width = min(max(width + 2, 12), 40)


def append_workbook_rows(workbook, title, rows):
    sheet = workbook.create_sheet(title=title[:31])
    normalized_rows = [dict(row) for row in (rows or [])]
    if not normalized_rows:
        sheet.append(["Message"])
        sheet.append(["No rows exported for this sheet."])
        autosize_workbook_sheet(sheet)
        return sheet

    headers = list(normalized_rows[0].keys())
    sheet.append(headers)
    for row in normalized_rows:
        sheet.append([row.get(header) for header in headers])
    autosize_workbook_sheet(sheet)
    return sheet


def get_recovery_pack_company_settings():
    settings = dict(_dep("get_company_settings")() or {})
    scanner_pin_hash = settings.pop("scanner_exit_pin_hash", None)
    settings["scanner_exit_pin_configured"] = 1 if scanner_pin_hash else 0
    return settings


def build_recovery_pack_workbook():
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise ValueError("Recovery pack export requires openpyxl.") from exc

    fetchall = _dep("fetchall")
    recovery_snapshot = _dep("get_backup_recovery_snapshot")()
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Recovery Pack", "Stellar Seats Attendance"])
    overview.append(["Generated At", _dep("now_str")()])
    overview.append(["Environment", "Postgres" if _dep("using_postgres")() else "SQLite"])
    overview.append(["Purpose", "Operational recovery reference and export workbook"])
    overview.append(["Note", "This workbook is not a one-click restore. Keep it with upload/file backups."])
    overview.append(["Last External Backup Noted", recovery_snapshot["external_backup"]["last_at"] or "Not noted in app"])
    overview.append(["External Backup Note", recovery_snapshot["external_backup"]["note"] or ""])
    overview.append(["Employee Accounts", recovery_snapshot["counts"]["employee_accounts"]])
    overview.append(["Admin Accounts", recovery_snapshot["counts"]["admin_accounts"]])
    overview.append(["Attendance Rows", recovery_snapshot["counts"]["attendance_rows"]])
    overview.append(["Scanner Logs", recovery_snapshot["counts"]["scanner_logs"]])
    overview.append(["Payroll Runs", recovery_snapshot["counts"]["payroll_runs"]])
    overview.append(["Pending Schedule Changes", recovery_snapshot["counts"]["future_schedule_changes"]])
    overview.append(["Holiday / Rest-Day Rules", recovery_snapshot["counts"]["schedule_special_rules"]])
    overview.append(["Uploaded Files", recovery_snapshot["upload_count"]])
    autosize_workbook_sheet(overview)

    guide = workbook.create_sheet(title="Recovery Guide")
    guide.append(["Step", "Action"])
    guide.append(["1", "Download the recovery pack before any major cleanup, reset, or policy change."])
    guide.append(["2", "Create an external Postgres backup or provider snapshot if production is using Render/Postgres."])
    guide.append(["3", "Keep uploaded files, proof uploads, and barcode assets together with this workbook."])
    guide.append(["4", "If rebuilding production, restore users and settings first, then attendance/payroll/log data."])
    guide.append(["5", "Use the workbook sheets as the source of truth for schedule presets, payroll rules, and historical workflows."])
    autosize_workbook_sheet(guide)

    append_workbook_rows(workbook, "Users", fetchall("""
        SELECT id, full_name, username, role, department, position, barcode_id,
               hourly_rate, schedule_days, shift_start, shift_end, schedule_preset_id,
               admin_permissions, admin_role_preset, is_active, created_at
        FROM users
        ORDER BY role ASC, full_name ASC
    """))
    append_workbook_rows(workbook, "Company Settings", [get_recovery_pack_company_settings()])
    append_workbook_rows(workbook, "Schedule Presets", fetchall("""
        SELECT sp.*, creator.full_name AS created_by_name
        FROM schedule_presets sp
        LEFT JOIN users creator ON creator.id = sp.created_by
        ORDER BY sp.name ASC
    """))
    append_workbook_rows(workbook, "Future Schedule Changes", fetchall("""
        SELECT fsc.*, u.full_name, preset.name AS preset_name, creator.full_name AS created_by_name
        FROM employee_future_schedule_changes fsc
        LEFT JOIN users u ON u.id = fsc.user_id
        LEFT JOIN schedule_presets preset ON preset.id = fsc.schedule_preset_id
        LEFT JOIN users creator ON creator.id = fsc.created_by
        ORDER BY fsc.effective_date ASC, fsc.id ASC
    """))
    append_workbook_rows(workbook, "Schedule Special Dates", fetchall("""
        SELECT ssd.*, creator.full_name AS created_by_name
        FROM schedule_special_dates ssd
        LEFT JOIN users creator ON creator.id = ssd.created_by
        ORDER BY ssd.special_date ASC, ssd.id ASC
    """))
    append_workbook_rows(workbook, "Attendance", fetchall("SELECT * FROM attendance ORDER BY work_date DESC, id DESC"))
    append_workbook_rows(workbook, "Breaks", fetchall("SELECT * FROM breaks ORDER BY work_date DESC, id DESC"))
    append_workbook_rows(workbook, "Corrections", fetchall("SELECT * FROM correction_requests ORDER BY created_at DESC, id DESC"))
    append_workbook_rows(workbook, "Scanner Logs", fetchall("SELECT * FROM scanner_logs ORDER BY created_at DESC, id DESC"))
    append_workbook_rows(workbook, "Overtime", fetchall("SELECT * FROM overtime_sessions ORDER BY created_at DESC, id DESC"))
    append_workbook_rows(workbook, "Payroll Runs", fetchall("SELECT * FROM payroll_runs ORDER BY updated_at DESC, id DESC"))
    append_workbook_rows(workbook, "Payroll Items", fetchall("SELECT * FROM payroll_run_items ORDER BY payroll_run_id DESC, id DESC"))
    append_workbook_rows(workbook, "Payroll Item Adjustments", fetchall("SELECT * FROM payroll_run_item_adjustments ORDER BY payroll_run_id DESC, id DESC"))
    append_workbook_rows(workbook, "Payroll Adjustments", fetchall("SELECT * FROM payroll_adjustments ORDER BY created_at DESC, id DESC"))
    append_workbook_rows(workbook, "Recurring Rules", fetchall("SELECT * FROM payroll_recurring_rules ORDER BY updated_at DESC, id DESC"))
    append_workbook_rows(workbook, "Incident Reports", fetchall("SELECT * FROM incident_reports ORDER BY created_at DESC, id DESC"))
    append_workbook_rows(workbook, "Disciplinary", fetchall("SELECT * FROM disciplinary_actions ORDER BY created_at DESC, id DESC"))
    return workbook
